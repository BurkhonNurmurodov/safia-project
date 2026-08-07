"""
Excel export of the /quality dashboard — the report, not the raw register.

The page computes every number it shows (KPIs, the per-supervisor matrix, the
trend buckets, the breakdowns, the seasonality matrix, the filtered rows) and
sends them here already labelled in the viewer's language. This module is a pure
formatter: it never re-derives a figure and never re-implements the four-language
dictionary — whatever the operator is looking at is what lands in the workbook.

Layout mirrors the dashboard's reading order across five tabs:

    Обзор        banner · applied scope · KPI cards · status-by-supervisor
                 (or, for a brigadir profile, the closure ribbon + aging buckets)
    Динамика     the trend buckets as a table + a stacked area chart
    Аналитика    type mix · foreign objects · hotspots · cells at fault ·
                 accountability, each with data bars and a native chart
    Сезонность   the share-of-period matrix under a colour scale
    Реестр       the filtered rows, in the on-screen sort order

Every sheet drops gridlines and draws its own rules, so the file opens as a
report rather than a dump. Colours are the platform's: brand gold for chrome and
the traffic-light statuses (green resolved / red open / amber waiting / orange
recurring) wherever a figure carries a status.
"""
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, DoughnutChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── palette (the dashboard's own) ────────────────────────────────────────────
BRAND       = "C8973F"
BRAND_DEEP  = "A87B2C"
BRAND_SOFT  = "FAF2E3"
CREAM       = "F6E7CC"
INK         = "1F2937"
INK_SOFT    = "6B7280"
INK_FAINT   = "9AA3AF"
LINE        = "E4E7EB"
BAND        = "F8FAFC"
PANEL       = "FFFFFF"
GREEN       = "22C55E"
RED         = "EF4444"
AMBER       = "EAB308"
ORANGE      = "F97316"
PURPLE      = "8B5CF6"
SLATE       = "94A3B8"

# Soft tints for status cells — the page's 13%-alpha chips, flattened onto white.
TINT = {
    GREEN: "E7F8EE", RED: "FDEAEA", AMBER: "FBF4DA",
    ORANGE: "FEEEE2", PURPLE: "EFE9FD", SLATE: "EFF2F6", BRAND: "F7EEDC",
}
STATUS_COLOR = {
    "done": GREEN, "open": RED, "waiting": AMBER,
    "repeat": ORANGE, "not_required": SLATE,
}

FONT = "Segoe UI"
NUM = "#,##0"
PCT1 = '0.0"%"'
PCT2 = '0.00"%"'


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _side(color: str = LINE, style: str = "thin") -> Side:
    return Side(style=style, color=color)


BOX = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
HAIR = Border(bottom=_side(LINE))
LEFT = Alignment(horizontal="left", vertical="center", indent=1)
RIGHT = Alignment(horizontal="right", vertical="center", indent=1)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _block(
    ws: Worksheet, r1: int, c1: int, r2: int, c2: int, value: Any = None, *,
    font: Optional[Font] = None, fill: Optional[PatternFill] = None,
    align: Optional[Alignment] = None, border: Optional[Border] = None,
    fmt: Optional[str] = None,
):
    """Style every cell of a rectangle, then merge it. Styling first is what makes
    a border wrap the whole merged range instead of only its anchor."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
    if (r1, c1) != (r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    top = ws.cell(r1, c1)
    if value is not None:
        top.value = value
    top.font = font or Font(name=FONT, size=10, color=INK)
    top.alignment = align or LEFT
    if fmt:
        top.number_format = fmt
    return top


def _sheet(wb: Workbook, title: str, widths: dict[int, float], landscape: bool = False) -> Worksheet:
    ws = wb.create_sheet(title[:31] or "Sheet")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BRAND
    ws.column_dimensions["A"].width = 2.2
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


def _banner(ws: Worksheet, row: int, c1: int, c2: int, title: str, subtitle: str = "") -> int:
    """Gold title band — the report's letterhead, repeated at the top of each tab."""
    ws.row_dimensions[row].height = 30
    _block(ws, row, c1, row, c2, title, fill=_fill(BRAND),
           font=Font(name=FONT, size=15, bold=True, color="FFFFFF"),
           align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[row + 1].height = 17
    _block(ws, row + 1, c1, row + 1, c2, subtitle or " ", fill=_fill(BRAND),
           font=Font(name=FONT, size=9.5, color=CREAM),
           align=Alignment(horizontal="left", vertical="top", indent=1))
    return row + 3


def _section(ws: Worksheet, row: int, c1: int, c2: int, title: str, subtitle: str = "") -> int:
    """Uppercase band with a gold rule under it — the card headers of the page."""
    ws.row_dimensions[row].height = 21
    split = c1 + max(0, (c2 - c1) * 2 // 3)
    band = Border(bottom=_side(BRAND, "medium"))
    _block(ws, row, c1, row, split, (title or "").upper(), fill=_fill(BRAND_SOFT), border=band,
           font=Font(name=FONT, size=10, bold=True, color=INK))
    _block(ws, row, split + 1, row, c2, subtitle or " ", fill=_fill(BRAND_SOFT), border=band,
           font=Font(name=FONT, size=9, color=INK_SOFT),
           align=Alignment(horizontal="right", vertical="center", indent=1))
    return row + 2


def _delta_text(kpi: dict) -> tuple[str, str]:
    """The KPI's movement chip: ▲/▼ + value, green when the move is the good one."""
    d = kpi.get("delta")
    if d in (None, "", 0):
        return "", INK_FAINT
    arrow = "▲" if (isinstance(d, (int, float)) and d > 0) else "▼"
    good = kpi.get("deltaGood")
    color = INK_SOFT if good is None else (GREEN if good else RED)
    sign = "+" if isinstance(d, (int, float)) and d > 0 else ""
    return f"{arrow} {sign}{d}{kpi.get('deltaSuffix', '%')}", color


def _kpi_cards(ws: Worksheet, row: int, c1: int, kpis: list[dict]) -> int:
    """Six cards across a 12-column grid, two columns each — the page's KPI strip.
    A card carries a coloured top rule (its icon tint), the value, the label and
    the small hint line, with the period-over-period move parked top-right."""
    ws.row_dimensions[row].height = 24
    ws.row_dimensions[row + 1].height = 15
    ws.row_dimensions[row + 2].height = 14
    for i, k in enumerate(kpis[:6]):
        a = c1 + i * 2
        b = a + 1
        accent = (k.get("color") or BRAND).lstrip("#").upper()
        card = Border(left=_side(), right=_side(), top=_side(accent, "medium"))
        for c in (a, b):
            ws.cell(row, c).border = card
            ws.cell(row, c).fill = _fill(PANEL)
        ws.cell(row, a).value = k.get("value")
        ws.cell(row, a).font = Font(name=FONT, size=16, bold=True, color=INK)
        ws.cell(row, a).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        txt, dcolor = _delta_text(k)
        ws.cell(row, b).value = txt or None
        ws.cell(row, b).font = Font(name=FONT, size=9, bold=True, color=dcolor)
        ws.cell(row, b).alignment = Alignment(horizontal="right", vertical="center", indent=1)

        _block(ws, row + 1, a, row + 1, b, k.get("label") or "",
               fill=_fill(PANEL), border=Border(left=_side(), right=_side()),
               font=Font(name=FONT, size=9.5, color=INK_SOFT))
        _block(ws, row + 2, a, row + 2, b, k.get("hint") or " ",
               fill=_fill(PANEL), border=Border(left=_side(), right=_side(), bottom=_side()),
               font=Font(name=FONT, size=8.5, color=INK_FAINT))
    return row + 4


def _meta_strip(ws: Worksheet, row: int, c1: int, c2: int, items: list[dict]) -> int:
    """The scope the numbers were taken under (period, view, filters, sync time) —
    four per line, so a reader can always tell what the file is a snapshot of."""
    if not items:
        return row
    span = max(1, (c2 - c1 + 1) // 4)
    per_line = max(1, (c2 - c1 + 1) // span)
    for i, it in enumerate(items):
        line, slot = divmod(i, per_line)
        r = row + line * 2
        a = c1 + slot * span
        b = min(c2, a + span - 1)
        ws.row_dimensions[r].height = 13
        ws.row_dimensions[r + 1].height = 15
        _block(ws, r, a, r, b, (it.get("label") or "").upper(), fill=_fill(BAND),
               border=Border(left=_side(), right=_side(), top=_side()),
               font=Font(name=FONT, size=8, bold=True, color=INK_FAINT))
        _block(ws, r + 1, a, r + 1, b, it.get("value") or "—", fill=_fill(BAND),
               border=Border(left=_side(), right=_side(), bottom=_side()),
               font=Font(name=FONT, size=10, bold=True, color=INK))
    lines = (len(items) + per_line - 1) // per_line
    return row + lines * 2 + 1


def _head_row(ws: Worksheet, row: int, c1: int, labels: list[str], *, span: int = 1,
              first_span: Optional[int] = None, height: float = 30):
    """Header band of a table — dark ink on the inner-panel grey, wrapped."""
    ws.row_dimensions[row].height = height
    c = c1
    for i, label in enumerate(labels):
        w = (first_span or span) if i == 0 else span
        align = LEFT if i == 0 else CENTER
        _block(ws, row, c, row, c + w - 1, label, fill=_fill(BAND), border=BOX,
               font=Font(name=FONT, size=9, bold=True, color=INK_SOFT),
               align=Alignment(horizontal="left" if i == 0 else "center",
                               vertical="center", wrap_text=True, indent=1 if i == 0 else 0))
        c += w
    return c


def _overview(wb: Workbook, p: dict) -> None:
    """Tab 1 — letterhead, the scope the report was taken under, the KPI strip and
    the status matrix that sits directly under it on screen."""
    tabs = p.get("sheets") or {}
    lbl = p.get("labels") or {}
    ws = _sheet(wb, tabs.get("overview", "Overview"), {c: 13.0 for c in range(2, 14)}, landscape=True)
    C1, C2 = 2, 13

    row = _banner(ws, 2, C1, C2, p.get("title") or "", p.get("subtitle") or "")
    row = _meta_strip(ws, row, C1, C2, p.get("meta") or [])

    kpis = p.get("kpis") or []
    if kpis:
        row = _section(ws, row, C1, C2, lbl.get("kpi", "Key figures"), lbl.get("vsPrev", ""))
        row = _kpi_cards(ws, row, C1, kpis)

    sup = p.get("sup_status")
    if sup and sup.get("rows"):
        row = _section(ws, row, C1, C2, sup.get("title", ""), sup.get("subtitle", ""))
        cols = sup.get("columns") or []
        pct = sup.get("mode") == "pct"
        _head_row(ws, row, C1, cols, span=2, first_span=2)
        row += 1
        colors = [GREEN, RED, ORANGE, PURPLE]
        first = row
        for i, r in enumerate(sup["rows"]):
            ws.row_dimensions[row].height = 17
            bg = _fill(PANEL if i % 2 == 0 else BAND)
            _block(ws, row, C1, row, C1 + 1, r.get("name") or "", fill=bg, border=BOX,
                   font=Font(name=FONT, size=10, color=INK))
            vals = r.get("values") or []
            for j, v in enumerate(vals):
                c = C1 + 2 + j * 2
                last = j == len(vals) - 1
                color = INK if last else (colors[j] if v else INK_FAINT)
                _block(ws, row, c, row, c + 1, v, fill=bg, border=BOX, align=RIGHT,
                       fmt=(NUM if last or not pct else PCT2),
                       font=Font(name=FONT, size=10, bold=True, color=color))
            row += 1
        total = sup.get("total")
        if total:
            ws.row_dimensions[row].height = 18
            top = Border(left=_side(), right=_side(), top=_side(BRAND, "medium"), bottom=_side())
            _block(ws, row, C1, row, C1 + 1, total.get("name") or lbl.get("total", "Total"),
                   fill=_fill(BRAND_SOFT), border=top,
                   font=Font(name=FONT, size=10, bold=True, color=INK))
            for j, v in enumerate(total.get("values") or []):
                c = C1 + 2 + j * 2
                _block(ws, row, c, row, c + 1, v, fill=_fill(BRAND_SOFT), border=top, align=RIGHT,
                       fmt=(NUM if j == len(total["values"]) - 1 or not pct else PCT2),
                       font=Font(name=FONT, size=10, bold=True, color=INK))
            row += 1
        ws.freeze_panes = ws.cell(first, C1)
        row += 1

    row = _closure(ws, row, C1, C2, p)
    ws.print_title_rows = "1:3"


def _closure(ws: Worksheet, row: int, c1: int, c2: int, p: dict) -> int:
    """A brigadir profile sees one row of the matrix, so the page swaps it for the
    closure ribbon + the aging of what is still open — both come across here."""
    lbl = p.get("labels") or {}
    cl = p.get("closure")
    if cl:
        row = _section(ws, row, c1, c2, cl.get("title", ""), cl.get("subtitle", ""))
        ws.row_dimensions[row].height = 30
        _block(ws, row, c1, row, c1 + 5, cl.get("rate"), fill=_fill(PANEL), border=BOX,
               fmt=PCT1, font=Font(name=FONT, size=20, bold=True, color=GREEN))
        _block(ws, row, c1 + 6, row, c2, cl.get("total"), fill=_fill(PANEL), border=BOX,
               align=RIGHT, fmt=NUM, font=Font(name=FONT, size=20, bold=True, color=INK))
        ws.row_dimensions[row + 1].height = 14
        _block(ws, row + 1, c1, row + 1, c1 + 5, cl.get("rateLabel") or "", fill=_fill(PANEL),
               border=Border(left=_side(), right=_side(), bottom=_side()),
               font=Font(name=FONT, size=9, color=INK_SOFT))
        _block(ws, row + 1, c1 + 6, row + 1, c2, cl.get("totalLabel") or "", fill=_fill(PANEL),
               border=Border(left=_side(), right=_side(), bottom=_side()), align=RIGHT,
               font=Font(name=FONT, size=9, color=INK_SOFT))
        row += 3

        legs = cl.get("legs") or []
        span = max(1, (c2 - c1 + 1) // max(1, len(legs)))
        ws.row_dimensions[row].height = 13
        ws.row_dimensions[row + 1].height = 20
        ws.row_dimensions[row + 2].height = 14
        for i, leg in enumerate(legs):
            a = c1 + i * span
            b = c2 if i == len(legs) - 1 else a + span - 1
            color = (leg.get("color") or SLATE).lstrip("#").upper()
            _block(ws, row, a, row, b, leg.get("label") or "", fill=_fill(TINT.get(color, BAND)),
                   border=Border(left=_side(), right=_side(), top=_side(color, "medium")),
                   font=Font(name=FONT, size=9, bold=True, color=INK_SOFT))
            _block(ws, row + 1, a, row + 1, b, leg.get("value"), fill=_fill(PANEL),
                   border=Border(left=_side(), right=_side()), fmt=NUM,
                   font=Font(name=FONT, size=14, bold=True, color=INK))
            share = f"{leg.get('share', 0)}%"
            txt, dcolor = _delta_text({**leg, "deltaSuffix": ""})
            _block(ws, row + 2, a, row + 2, b, f"{share}   {txt}".strip(), fill=_fill(PANEL),
                   border=Border(left=_side(), right=_side(), bottom=_side()),
                   font=Font(name=FONT, size=9, color=dcolor if txt else INK_FAINT))
        row += 4

    ag = p.get("aging")
    if ag and ag.get("buckets"):
        row = _section(ws, row, c1, c2, ag.get("title", ""), ag.get("subtitle", ""))
        _head_row(ws, row, c1, [ag.get("colBucket", ""), ag.get("colCount", "")],
                  span=4, first_span=8, height=20)
        row += 1
        for i, b in enumerate(ag["buckets"]):
            ws.row_dimensions[row].height = 17
            bg = _fill(PANEL if i % 2 == 0 else BAND)
            color = (b.get("color") or SLATE).lstrip("#").upper()
            _block(ws, row, c1, row, c1 + 7, b.get("label") or "", fill=bg, border=BOX,
                   font=Font(name=FONT, size=10, color=INK))
            _block(ws, row, c1 + 8, row, c2, b.get("n"), fill=bg, border=BOX, align=RIGHT, fmt=NUM,
                   font=Font(name=FONT, size=10, bold=True, color=color if b.get("n") else INK_FAINT))
            row += 1
        if ag.get("oldest"):
            _block(ws, row, c1, row, c2, f"{ag.get('oldestLabel', '')}: {ag['oldest']}",
                   font=Font(name=FONT, size=9, italic=True, color=INK_FAINT))
            row += 1
        row += 1
    return row


def _trend(wb: Workbook, p: dict) -> None:
    """Tab 2 — the dynamics card: one row per bucket, one column per band, plus a
    native stacked area chart in the page's own series colours."""
    tr = p.get("trend")
    if not tr or not tr.get("labels"):
        return
    tabs = p.get("sheets") or {}
    lbl = p.get("labels") or {}
    series = tr.get("series") or []
    ncol = 1 + len(series) + 1
    widths = {2: 16.0}
    widths.update({c: 15.0 for c in range(3, 2 + ncol)})
    ws = _sheet(wb, tabs.get("trend", "Trend"), widths, landscape=True)
    C1 = 2
    C2 = C1 + ncol - 1

    row = _banner(ws, 2, C1, C2, tr.get("title") or "", tr.get("subtitle") or "")
    row = _section(ws, row, C1, C2, tr.get("title", ""), tr.get("granLabel", ""))

    head = [tr.get("colPeriod", "")] + [s.get("name", "") for s in series] + [lbl.get("total", "Total")]
    _head_row(ws, row, C1, head, height=26)
    head_row = row
    row += 1
    first = row
    for i, label in enumerate(tr["labels"]):
        ws.row_dimensions[row].height = 17
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        _block(ws, row, C1, row, C1, label, fill=bg, border=BOX,
               font=Font(name=FONT, size=10, color=INK))
        tot = 0
        for j, s in enumerate(series):
            v = (s.get("data") or [])[i] if i < len(s.get("data") or []) else 0
            tot += v or 0
            _block(ws, row, C1 + 1 + j, row, C1 + 1 + j, v, fill=bg, border=BOX, align=RIGHT,
                   fmt=NUM, font=Font(name=FONT, size=10, color=INK if v else INK_FAINT))
        _block(ws, row, C2, row, C2, tot, fill=bg, border=BOX, align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=10, bold=True, color=INK))
        row += 1
    last = row - 1

    ws.freeze_panes = ws.cell(first, C1 + 1)

    chart = AreaChart()
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.height, chart.width = 9.5, 26
    chart.style = 2
    chart.y_axis.majorGridlines = None
    for j, s in enumerate(series):
        ref = Reference(ws, min_col=C1 + 1 + j, min_row=head_row, max_row=last)
        chart.add_data(ref, titles_from_data=True)
        color = (s.get("color") or SLATE).lstrip("#").upper()
        sr = chart.series[-1]
        sr.graphicalProperties.solidFill = color
        sr.graphicalProperties.line.solidFill = "FFFFFF"
        sr.graphicalProperties.line.width = 12700
    chart.set_categories(Reference(ws, min_col=C1, min_row=first, max_row=last))
    ws.add_chart(chart, f"{get_column_letter(C1)}{row + 1}")


def _bar_table(ws: Worksheet, row: int, c1: int, block: dict, lbl: dict, *,
               color: str = BRAND, chart_at: Optional[str] = None, top_n: int = 0) -> int:
    """One breakdown card: label · count · share, with in-cell data bars. `chart_at`
    anchors a native bar chart beside it when the block deserves a picture."""
    rows = block.get("rows") or []
    if not rows:
        return row
    row = _section(ws, row, c1, c1 + 3, block.get("title", ""), block.get("subtitle", ""))
    _head_row(ws, row, c1, [block.get("colLabel", ""), lbl.get("count", "Count"),
                            lbl.get("share", "Share")], height=20)
    head_row = row
    row += 1
    first = row
    total = sum((r.get("n") or 0) for r in rows) or 1
    for i, r in enumerate(rows):
        ws.row_dimensions[row].height = 17
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        dot = (r.get("color") or "").lstrip("#").upper()
        _block(ws, row, c1, row, c1, r.get("label") or "", fill=bg, border=BOX,
               font=Font(name=FONT, size=10, color=INK))
        _block(ws, row, c1 + 1, row, c1 + 1, r.get("n"), fill=bg, border=BOX, align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=10, bold=True, color=dot or INK))
        _block(ws, row, c1 + 2, row, c1 + 2, round((r.get("n") or 0) * 100.0 / total, 1),
               fill=bg, border=BOX, align=RIGHT, fmt=PCT1,
               font=Font(name=FONT, size=10, color=INK_SOFT))
        row += 1
    last = row - 1
    col = get_column_letter(c1 + 1)
    ws.conditional_formatting.add(
        f"{col}{first}:{col}{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=color, showValue=True),
    )

    if chart_at:
        chart = BarChart()
        chart.type = "bar"
        chart.height, chart.width = max(6.0, 0.62 * len(rows[:top_n or len(rows)]) + 2), 11.5
        chart.legend = None
        chart.y_axis.majorGridlines = None
        chart.add_data(Reference(ws, min_col=c1 + 1, min_row=head_row, max_row=last),
                       titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=c1, min_row=first, max_row=last))
        chart.series[0].graphicalProperties.solidFill = color
        ws.add_chart(chart, chart_at)
    return row + 1


def _analytics(wb: Workbook, p: dict) -> None:
    """Tab 3 — every breakdown card on the page, stacked in the on-screen order."""
    blocks = [(k, p.get(k)) for k in ("types", "cats", "hotspots", "cells")]
    acc = p.get("acc")
    if not any(b and b.get("rows") for _, b in blocks) and not (acc and acc.get("rows")):
        return
    tabs = p.get("sheets") or {}
    lbl = p.get("labels") or {}
    ws = _sheet(wb, tabs.get("analytics", "Breakdowns"),
                {2: 40.0, 3: 12.0, 4: 11.0, 5: 3.0, 6: 13.0, 7: 13.0, 8: 13.0, 9: 13.0,
                 10: 11.0, 11: 10.0}, landscape=True)
    C1 = 2

    row = _banner(ws, 2, C1, 11, p.get("title") or "", p.get("subtitle") or "")

    types = p.get("types")
    if types and types.get("rows"):
        start = row
        row = _bar_table(ws, row, C1, types, lbl, color=RED)
        # the page's donut, as a real chart parked beside the table
        n = len(types["rows"])
        chart = DoughnutChart(holeSize=62)
        chart.height, chart.width = 8.4, 11.5
        chart.legend.position = "r"
        chart.add_data(Reference(ws, min_col=C1 + 1, min_row=start + 2, max_row=start + 1 + n),
                       titles_from_data=False)
        chart.set_categories(Reference(ws, min_col=C1, min_row=start + 2, max_row=start + 1 + n))
        pts = []
        for i, r in enumerate(types["rows"]):
            dp = DataPoint(idx=i)
            dp.graphicalProperties.solidFill = (r.get("color") or SLATE).lstrip("#").upper()
            dp.graphicalProperties.line.solidFill = "FFFFFF"
            pts.append(dp)
        chart.series[0].data_points = pts
        ws.add_chart(chart, f"{get_column_letter(C1 + 4)}{start + 1}")

    for key, color in (("cats", GREEN), ("hotspots", BRAND_DEEP)):
        b = p.get(key)
        if b and b.get("rows"):
            row = _bar_table(ws, row, C1, b, lbl, color=color)

    cells = p.get("cells")
    if cells and cells.get("rows"):
        start = row
        row = _bar_table(ws, row, C1, cells, lbl, color=RED,
                         chart_at=f"{get_column_letter(C1 + 4)}{start + 1}",
                         top_n=len(cells["rows"]))

    if acc and acc.get("rows"):
        row = _section(ws, row, C1, 11, acc.get("title", ""), acc.get("subtitle", ""))
        head = [acc.get("colName", "")] + (acc.get("statuses") or []) + \
               [lbl.get("total", "Total"), acc.get("colRate", "%")]
        _head_row(ws, row, C1, head, height=30)
        row += 1
        first = row
        colors = [GREEN, RED, AMBER, ORANGE]
        for i, r in enumerate(acc["rows"]):
            ws.row_dimensions[row].height = 17
            bg = _fill(PANEL if i % 2 == 0 else BAND)
            _block(ws, row, C1, row, C1, r.get("name") or "", fill=bg, border=BOX,
                   font=Font(name=FONT, size=10, color=INK))
            vals = r.get("values") or []
            for j, v in enumerate(vals):
                _block(ws, row, C1 + 1 + j, row, C1 + 1 + j, v, fill=bg, border=BOX, align=RIGHT,
                       fmt=NUM, font=Font(name=FONT, size=10, bold=True,
                                          color=(colors[j] if j < len(colors) and v else INK_FAINT)))
            _block(ws, row, C1 + 1 + len(vals), row, C1 + 1 + len(vals), sum(vals), fill=bg,
                   border=BOX, align=RIGHT, fmt=NUM, font=Font(name=FONT, size=10, bold=True, color=INK))
            _block(ws, row, C1 + 2 + len(vals), row, C1 + 2 + len(vals), r.get("rate"), fill=bg,
                   border=BOX, align=RIGHT, fmt=PCT1, font=Font(name=FONT, size=10, color=INK_SOFT))
            row += 1
        col = get_column_letter(C1 + 2 + len(acc.get("statuses") or []))
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{row - 1}",
            ColorScaleRule(start_type="num", start_value=0, start_color=TINT[RED],
                           mid_type="num", mid_value=50, mid_color=TINT[AMBER],
                           end_type="num", end_value=100, end_color=TINT[GREEN]),
        )


def _season(wb: Workbook, p: dict) -> None:
    """Tab 4 — the seasonality grid: type × period share, under the same red→green
    colour scale the on-screen heatmap uses, with the period totals on top."""
    s = p.get("season")
    if not s or not s.get("rows"):
        return
    tabs = p.get("sheets") or {}
    lbl = p.get("labels") or {}
    labels = s.get("labels") or []
    ncol = 1 + len(labels)
    widths = {2: 26.0}
    widths.update({c: 8.6 for c in range(3, 2 + ncol)})
    ws = _sheet(wb, tabs.get("season", "Seasonality"), widths, landscape=True)
    C1 = 2
    C2 = C1 + ncol - 1

    row = _banner(ws, 2, C1, C2, s.get("title") or "", s.get("subtitle") or "")
    row = _section(ws, row, C1, C2, s.get("title", ""), s.get("scopeLabel", ""))

    _head_row(ws, row, C1, [s.get("firstColLabel", "")] + labels, height=28)
    row += 1

    # column totals — the row the heatmap prints above the grid
    ws.row_dimensions[row].height = 17
    _block(ws, row, C1, row, C1, lbl.get("total", "Total"), fill=_fill(BRAND_SOFT), border=BOX,
           font=Font(name=FONT, size=9.5, bold=True, color=INK))
    for j, v in enumerate(s.get("colTotals") or []):
        _block(ws, row, C1 + 1 + j, row, C1 + 1 + j, v, fill=_fill(BRAND_SOFT), border=BOX,
               align=Alignment(horizontal="center", vertical="center"), fmt=NUM,
               font=Font(name=FONT, size=9.5, bold=True, color=INK if v else INK_FAINT))
    row += 1

    first = row
    for r in s["rows"]:
        ws.row_dimensions[row].height = 17
        _block(ws, row, C1, row, C1, r.get("label") or "", fill=_fill(PANEL), border=BOX,
               font=Font(name=FONT, size=10, color=INK))
        for j, v in enumerate(r.get("data") or []):
            _block(ws, row, C1 + 1 + j, row, C1 + 1 + j, v, fill=_fill(PANEL), border=BOX,
                   align=Alignment(horizontal="center", vertical="center"), fmt=PCT1,
                   font=Font(name=FONT, size=9.5, color=INK))
        row += 1

    rng = f"{get_column_letter(C1 + 1)}{first}:{get_column_letter(C2)}{row - 1}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        mid_type="percentile", mid_value=60, mid_color=TINT[AMBER],
        end_type="max", end_color="F3A5A5",
    ))
    ws.freeze_panes = ws.cell(first, C1 + 1)


def _register(wb: Workbook, p: dict) -> None:
    """Tab 5 — the register itself, in the order and column set the page shows,
    with the status column carrying its traffic-light tint."""
    reg = p.get("register")
    if not reg or not reg.get("columns"):
        return
    tabs = p.get("sheets") or {}
    cols = reg["columns"]
    rows = reg.get("rows") or []
    date_idx = reg.get("dateIdx")

    # auto-fit: the longest cell in each column, held to a sane band
    widths = {}
    for i, c in enumerate(cols):
        longest = len(str(c))
        for r in rows[:600]:
            v = (r.get("c") or [])[i] if i < len(r.get("c") or []) else ""
            longest = max(longest, len(str(v or "")))
        widths[2 + i] = min(38.0, max(11.0, longest + 3.0))
    ws = _sheet(wb, tabs.get("register", "Register"), widths, landscape=True)
    C1 = 2
    C2 = C1 + len(cols) - 1

    row = _banner(ws, 2, C1, C2, reg.get("title") or "", reg.get("subtitle") or "")
    row = _section(ws, row, C1, C2, reg.get("title", ""), reg.get("countLabel", ""))

    _head_row(ws, row, C1, cols, height=26)
    head_row = row
    row += 1
    first = row
    for i, r in enumerate(rows):
        ws.row_dimensions[row].height = 16
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        cells = r.get("c") or []
        scolor = STATUS_COLOR.get(r.get("st") or "")
        for j, v in enumerate(cells[:len(cols)]):
            c = C1 + j
            fmt = None
            align = LEFT
            font = Font(name=FONT, size=9.5, color=INK)
            fill = bg
            if j == date_idx and isinstance(v, str) and len(v) == 10:
                try:
                    v = datetime.strptime(v, "%Y-%m-%d").date()
                    fmt = "DD.MM.YYYY"
                    align = Alignment(horizontal="center", vertical="center")
                except ValueError:
                    pass
            if j == reg.get("statusIdx") and scolor:
                fill = _fill(TINT.get(scolor, BAND))
                font = Font(name=FONT, size=9.5, bold=True, color=scolor)
                align = Alignment(horizontal="center", vertical="center")
            elif j in (reg.get("returnIdx"), reg.get("statusIdx")):
                align = Alignment(horizontal="center", vertical="center")
            _block(ws, row, c, row, c, v if v not in (None, "") else "—", fill=fill, border=BOX,
                   align=align, font=font, fmt=fmt)
        row += 1

    ws.auto_filter.ref = f"{get_column_letter(C1)}{head_row}:{get_column_letter(C2)}{max(first, row - 1)}"
    ws.freeze_panes = ws.cell(first, C1 + 1)
    ws.print_title_rows = f"{head_row}:{head_row}"


def build_quality_workbook(p: dict) -> BytesIO:
    """Assemble the five tabs and hand back the saved workbook."""
    wb = Workbook()
    wb.remove(wb.active)
    _overview(wb, p)
    _trend(wb, p)
    _analytics(wb, p)
    _season(wb, p)
    _register(wb, p)
    if not wb.sheetnames:                       # nothing on screen matched the filters
        ws = _sheet(wb, (p.get("sheets") or {}).get("overview", "Overview"),
                    {c: 13.0 for c in range(2, 14)})
        _banner(ws, 2, 2, 13, p.get("title") or "", p.get("subtitle") or "")
    wb.properties.title = p.get("title") or "Quality"
    wb.properties.creator = "Safia Dashboard"
    wb.properties.created = datetime.now()
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
