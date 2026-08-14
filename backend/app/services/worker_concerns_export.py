"""
Excel export of /worker-concerns — the leaders' KPI report over the concerns
workers file to their cell leader.

Unlike the Quality export (whose page computes everything client-side and posts
finished numbers), this page's register is server-paginated and its aggregates
are server-computed — so the ROUTER re-queries the data through the very same
``_apply_scope_and_filters`` every page endpoint uses (viewer locks included)
and hands the numbers here. The frontend contributes only presentation: every
label, sheet name and meta line arrives already in the viewer's language. This
module is a pure formatter and never re-derives a figure.

It speaks the platform's one workbook language — the primitives are imported
from quality_export rather than copied, so the two reports can never drift
apart visually: gold letterhead, the scope strip, KPI cards, zebra tables,
traffic-light status tints, native charts, no gridlines.

Workbook (mirrors the page's three views):

    Obzor          letterhead · scope · KPI cards · daily dynamics (table +
                   stacked columns) · by-brigadir matrix · top open cells (+bar)
    Liderlar KPI   band legend · one row per registered leader, % tinted by the
                   admin-set bands · «unassigned» bucket · totals row
    Reyestr        every matching row (paging is a screen affordance, not a
                   narrowing) · tinted status chips · auto-filter · frozen head
"""
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.quality_export import (
    AMBER, BAND, BOX, BRAND, BRAND_SOFT, FONT, GREEN, INK, INK_FAINT, INK_SOFT,
    NUM, PANEL, PCT1, RED, RIGHT, SLATE, TINT,
    _banner, _block, _fill, _head_row, _kpi_cards, _meta_strip, _section,
    _sheet, _side,
)

# The page's own status palette (semantic traffic-light, never brand gold).
BLUE = "3B82F6"
OTHER_INK = "64748B"
WC_COLOR = {"done": GREEN, "doing": AMBER, "todo": RED,
            "deferred": SLATE, "other": OTHER_INK}
WC_TINT = {**TINT, OTHER_INK: "E8ECF1"}
DATE_FMT = "DD.MM.YYYY"


def _pd(s: Any) -> Optional[date]:
    try:
        return date.fromisoformat(s) if s else None
    except (ValueError, TypeError):
        return None


def _n(v: int) -> str:
    return f"{v:,}".replace(",", " ")


def _st_label(p: dict, st: str, raw: str = "") -> str:
    """A status as the viewer reads it; «other» keeps the sheet's raw wording
    when there is one (the page's StChip does exactly this)."""
    lbl = (p.get("status_labels") or {}).get(st) or st
    return (raw or lbl) if st == "other" else lbl


def _head_spans(ws: Worksheet, row: int, c1: int, cols: list[tuple[str, int]],
                height: float = 26) -> int:
    """A header band whose columns each carry their own span — _head_row only
    lets the FIRST column differ, and the Obzor tables need more shapes."""
    ws.row_dimensions[row].height = height
    c = c1
    for i, (label, span) in enumerate(cols):
        _block(ws, row, c, row, c + span - 1, label, fill=_fill(BAND), border=BOX,
               font=Font(name=FONT, size=9, bold=True, color=INK_SOFT),
               align=Alignment(horizontal="left" if i == 0 else "center",
                               vertical="center", wrap_text=True,
                               indent=1 if i == 0 else 0))
        c += span
    return c


# ── Obzor ────────────────────────────────────────────────────────────────────

def _daily_block(ws: Worksheet, row: int, c1: int, p: dict, sts: list[str]) -> int:
    """The daily dynamics: one row per day of the EXACT range (zero days kept —
    a gap is data), plus the page's stacked columns as a native chart."""
    lbl = p.get("labels") or {}
    daily = p.get("daily") or {}
    days = daily.get("days") or []
    per = daily.get("rows") or {}
    if not days:
        return row
    c2 = c1 + 2 + len(sts)                       # date(2) + statuses + total
    row = _section(ws, row, c1, c2, lbl.get("secDaily", ""), lbl.get("secDailySub", ""))
    head = [(lbl.get("colDate", ""), 2)] \
        + [(_st_label(p, s), 1) for s in sts] + [(lbl.get("colTotal", ""), 1)]
    _head_spans(ws, row, c1, head)
    head_row = row
    row += 1
    first = row
    for i, d in enumerate(days):
        v = per.get(d) or {}
        ws.row_dimensions[row].height = 15
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        dd = _pd(d)
        _block(ws, row, c1, row, c1 + 1, dd or d, fill=bg, border=BOX,
               fmt=DATE_FMT if dd else None,
               font=Font(name=FONT, size=9.5, color=INK))
        tot = 0
        for j, s in enumerate(sts):
            n = v.get(s) or 0
            tot += n
            _block(ws, row, c1 + 2 + j, row, c1 + 2 + j, n, fill=bg, border=BOX,
                   align=RIGHT, fmt=NUM,
                   font=Font(name=FONT, size=9.5,
                             color=WC_COLOR[s] if n else INK_FAINT))
        _block(ws, row, c2, row, c2, tot, fill=bg, border=BOX, align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=9.5, bold=True,
                         color=INK if tot else INK_FAINT))
        row += 1
    last = row - 1

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.height, chart.width = 8.6, 24.0
    chart.gapWidth = 40
    chart.legend.position = "t"
    chart.y_axis.majorGridlines = None
    for j, s in enumerate(sts):
        chart.add_data(Reference(ws, min_col=c1 + 2 + j, min_row=head_row,
                                 max_row=last), titles_from_data=True)
        chart.series[-1].graphicalProperties.solidFill = WC_COLOR[s]
    chart.set_categories(Reference(ws, min_col=c1, min_row=first, max_row=last))
    chart.x_axis.number_format = "DD.MM"
    ws.add_chart(chart, f"{get_column_letter(c1)}{row + 1}")
    return row + 19                              # room the chart occupies


def _brig_block(ws: Worksheet, row: int, c1: int, p: dict, sts: list[str]) -> int:
    """Per-brigadir status matrix with a data bar on the totals and the KPI
    bands' red→amber→green scale on the resolution %."""
    lbl = p.get("labels") or {}
    brig = p.get("brigadirs") or []
    if not brig:
        return row
    bands = (p.get("leaders") or {}).get("bands") or {"green": 80, "yellow": 50}
    c2 = c1 + 3 + len(sts) + 1                   # name(3) + statuses + total + %
    row = _section(ws, row, c1, c2, lbl.get("secBrig", ""), lbl.get("secBrigSub", ""))
    head = [(lbl.get("colBrig", ""), 3)] + [(_st_label(p, s), 1) for s in sts] \
        + [(lbl.get("colTotal", ""), 1), (lbl.get("colPct", ""), 1)]
    _head_spans(ws, row, c1, head)
    row += 1
    first = row
    for i, b in enumerate(brig):
        ws.row_dimensions[row].height = 16
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        _block(ws, row, c1, row, c1 + 2, b.get("name") or "—", fill=bg, border=BOX,
               font=Font(name=FONT, size=9.5, color=INK))
        for j, s in enumerate(sts):
            v = b.get(s) or 0
            _block(ws, row, c1 + 3 + j, row, c1 + 3 + j, v, fill=bg, border=BOX,
                   align=RIGHT, fmt=NUM,
                   font=Font(name=FONT, size=9.5, bold=True,
                             color=WC_COLOR[s] if v else INK_FAINT))
        _block(ws, row, c1 + 3 + len(sts), row, c1 + 3 + len(sts), b.get("total"),
               fill=bg, border=BOX, align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=9.5, bold=True, color=INK))
        _block(ws, row, c2, row, c2, b.get("pct"), fill=bg, border=BOX, align=RIGHT,
               fmt=PCT1, font=Font(name=FONT, size=9.5, color=INK_SOFT))
        row += 1
    last = row - 1
    tot_col = get_column_letter(c1 + 3 + len(sts))
    ws.conditional_formatting.add(
        f"{tot_col}{first}:{tot_col}{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color=BRAND, showValue=True))
    pct_col = get_column_letter(c2)
    ws.conditional_formatting.add(
        f"{pct_col}{first}:{pct_col}{last}",
        ColorScaleRule(start_type="num", start_value=0, start_color=TINT[RED],
                       mid_type="num", mid_value=bands["yellow"], mid_color=TINT[AMBER],
                       end_type="num", end_value=bands["green"], end_color=TINT[GREEN]))
    return row + 1


def _cells_block(ws: Worksheet, row: int, c1: int, p: dict) -> int:
    """Cells with the most open concerns — the «go fix this first» list, with a
    native horizontal bar chart parked beside it."""
    lbl = p.get("labels") or {}
    cells = p.get("top_cells") or []
    if not cells:
        return row
    c2 = c1 + 7
    start = row
    row = _section(ws, row, c1, c2, lbl.get("secCells", ""), lbl.get("secCellsSub", ""))
    _head_spans(ws, row, c1, [(lbl.get("colCell", ""), 2), (lbl.get("colLeader", ""), 4),
                              (lbl.get("colOpen", ""), 1), (lbl.get("colTotal", ""), 1)])
    head_row = row
    row += 1
    first = row
    for i, c in enumerate(cells):
        ws.row_dimensions[row].height = 16
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        _block(ws, row, c1, row, c1 + 1, c.get("code") or "—", fill=bg, border=BOX,
               font=Font(name=FONT, size=9.5, bold=True, color=INK))
        _block(ws, row, c1 + 2, row, c1 + 5, c.get("leader") or "—", fill=bg, border=BOX,
               font=Font(name=FONT, size=9.5, color=INK_SOFT))
        _block(ws, row, c1 + 6, row, c1 + 6, c.get("open"), fill=bg, border=BOX,
               align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=9.5, bold=True,
                         color=RED if c.get("open") else INK_FAINT))
        _block(ws, row, c2, row, c2, c.get("total"), fill=bg, border=BOX,
               align=RIGHT, fmt=NUM, font=Font(name=FONT, size=9.5, color=INK))
        row += 1
    last = row - 1
    open_col = get_column_letter(c1 + 6)
    ws.conditional_formatting.add(
        f"{open_col}{first}:{open_col}{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color=RED, showValue=True))

    chart = BarChart()
    chart.type = "bar"
    chart.height, chart.width = max(6.0, 0.62 * len(cells) + 2), 10.5
    chart.legend = None
    chart.y_axis.majorGridlines = None
    chart.add_data(Reference(ws, min_col=c1 + 6, min_row=head_row, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=c1, min_row=first, max_row=last))
    chart.series[0].graphicalProperties.solidFill = RED
    ws.add_chart(chart, f"{get_column_letter(c2 + 2)}{start + 2}")
    return row + 1


def _obzor(wb: Workbook, p: dict, sts: list[str]) -> None:
    """Tab 1 — letterhead, the scope the report was taken under, the KPI strip
    and the three Obzor breakdowns in on-screen reading order."""
    tabs = p.get("sheets") or {}
    lbl = p.get("labels") or {}
    ws = _sheet(wb, tabs.get("overview", "Overview"),
                {c: 13.0 for c in range(2, 14)}, landscape=True)
    C1, C2 = 2, 13

    row = _banner(ws, 2, C1, C2, p.get("title") or "", p.get("subtitle") or "")
    row = _meta_strip(ws, row, C1, C2, p.get("meta") or [])

    k = p.get("kpi") or {}
    row = _section(ws, row, C1, C2, lbl.get("secKpi", ""), "")
    row = _kpi_cards(ws, row, C1, [
        {"value": k.get("total", 0), "label": lbl.get("kTotal", ""), "color": BRAND},
        {"value": k.get("done", 0), "label": lbl.get("kResolved", ""),
         "hint": f'{k["pct"]}%' if k.get("pct") is not None else "", "color": GREEN},
        {"value": k.get("doing", 0), "label": lbl.get("kDoing", ""), "color": AMBER},
        {"value": k.get("open", 0), "label": lbl.get("kOpen", ""),
         "hint": lbl.get("kOpenHint", ""), "color": RED},
        {"value": k.get("workers", 0), "label": lbl.get("kWorkers", ""),
         "hint": lbl.get("kWorkersHint", ""), "color": BLUE},
    ])
    if k.get("undated"):
        # Rows whose filing date is unreadable — outside every date-bound
        # figure above; the file must say so exactly as the page does.
        _block(ws, row, C1, row, C2, f'{_n(k["undated"])} {lbl.get("undatedNote", "")}',
               font=Font(name=FONT, size=9, italic=True, color=INK_SOFT))
        row += 2

    row = _daily_block(ws, row, C1, p, sts)
    row = _brig_block(ws, row, C1, p, sts)
    _cells_block(ws, row, C1, p)
    ws.print_title_rows = "1:3"


# ── Liderlar KPI ─────────────────────────────────────────────────────────────

def _band_of(r: dict, bands: dict) -> str:
    if not r.get("ranked") or r.get("pct") is None:
        return "low"
    if r["pct"] >= bands["green"]:
        return "green"
    if r["pct"] >= bands["yellow"]:
        return "yellow"
    return "red"


_BAND_COLOR = {"green": GREEN, "yellow": AMBER, "red": RED, "low": SLATE}


def _leaders_sheet(wb: Workbook, p: dict, sts: list[str]) -> None:
    """Tab 2 — the KPI table itself: the grading legend first, then one row per
    REGISTERED leader with the % cell tinted by the admin-set bands, the
    explicit «unassigned» bucket and a totals row. No merged cells in the table
    body, so Excel's own filter buttons stay usable."""
    tabs = p.get("sheets") or {}
    lbl = p.get("labels") or {}
    ld = p.get("leaders") or {}
    rows = ld.get("rows") or []
    bands = ld.get("bands") or {"green": 80, "yellow": 50}

    ncols = 4 + len(sts) + 2                     # leader/brig/cells/total + sts + open/pct
    widths = {2: 27.0, 3: 30.0, 4: 22.0, 5: 10.0}
    widths.update({5 + i: 11.5 for i in range(1, len(sts) + 1)})
    widths[5 + len(sts) + 1] = 11.5
    widths[5 + len(sts) + 2] = 12.5
    ws = _sheet(wb, tabs.get("leaders", "Leaders KPI"), widths, landscape=True)
    C1 = 2
    C2 = C1 + ncols - 1

    row = _banner(ws, 2, C1, C2, p.get("title") or "", p.get("subtitle") or "")

    # The grading scale, spelled out before any number is read against it.
    leg = lbl.get("bandLegend") or {}
    chips = [(leg.get("green", ""), GREEN), (leg.get("yellow", ""), AMBER),
             (leg.get("red", ""), RED), (leg.get("low", ""), SLATE)]
    span = max(1, ncols // 4)
    ws.row_dimensions[row].height = 18
    for i, (text, color) in enumerate(chips):
        a = C1 + i * span
        b = C2 if i == 3 else a + span - 1
        _block(ws, row, a, row, b, text, fill=_fill(WC_TINT.get(color, BAND)),
               border=Border(top=_side(color, "medium"), left=_side(),
                             right=_side(), bottom=_side()),
               align=Alignment(horizontal="center", vertical="center"),
               font=Font(name=FONT, size=9, bold=True,
                         color=color if color != SLATE else INK_SOFT))
    row += 1
    _block(ws, row, C1, row, C2, lbl.get("lowNHint", ""),
           font=Font(name=FONT, size=8.5, italic=True, color=INK_FAINT))
    row += 2

    row = _section(ws, row, C1, C2, lbl.get("secLeaders", ""),
                   f'{lbl.get("secLeadersSub", "")} · {_n(len(rows))} {lbl.get("leadersWord", "")}')
    head = [lbl.get("colLeader", ""), lbl.get("colBrig", ""), lbl.get("colCells", ""),
            lbl.get("colTotal", "")] + [_st_label(p, s) for s in sts] \
        + [lbl.get("colOpen", ""), lbl.get("colPct", "")]
    _head_row(ws, row, C1, head, span=1)
    head_row = row
    row += 1
    first = row

    if not rows:
        _block(ws, row, C1, row, C2, lbl.get("noMatch", "—"), border=BOX,
               align=Alignment(horizontal="center", vertical="center"),
               font=Font(name=FONT, size=10, color=INK_FAINT))
        row += 1
    for i, r in enumerate(rows):
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        _block(ws, row, C1, row, C1, r.get("leader") or "—", fill=bg, border=BOX,
               align=Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1),
               font=Font(name=FONT, size=9.5, bold=True, color=INK))
        _block(ws, row, C1 + 1, row, C1 + 1, ", ".join(r.get("brigadirs") or []) or "—",
               fill=bg, border=BOX,
               align=Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1),
               font=Font(name=FONT, size=9, color=INK_SOFT))
        _block(ws, row, C1 + 2, row, C1 + 2, ", ".join(r.get("cells") or []) or "—",
               fill=bg, border=BOX,
               align=Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1),
               font=Font(name=FONT, size=9, color=INK_SOFT))
        _block(ws, row, C1 + 3, row, C1 + 3, r.get("total"), fill=bg, border=BOX,
               align=RIGHT, fmt=NUM, font=Font(name=FONT, size=9.5, bold=True, color=INK))
        for j, s in enumerate(sts):
            v = r.get(s) or 0
            _block(ws, row, C1 + 4 + j, row, C1 + 4 + j, v, fill=bg, border=BOX,
                   align=RIGHT, fmt=NUM,
                   font=Font(name=FONT, size=9.5,
                             color=WC_COLOR[s] if v else INK_FAINT))
        _block(ws, row, C1 + 4 + len(sts), row, C1 + 4 + len(sts), r.get("open"),
               fill=bg, border=BOX, align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=9.5, bold=True,
                         color=RED if r.get("open") else INK_FAINT))
        band = _band_of(r, bands)
        bc = _BAND_COLOR[band]
        _block(ws, row, C2, row, C2, r.get("pct"), fill=_fill(WC_TINT.get(bc, BAND)),
               border=BOX, align=Alignment(horizontal="center", vertical="center"),
               fmt=PCT1,
               font=Font(name=FONT, size=9.5, bold=True,
                         color=bc if band != "low" else INK_SOFT))
        row += 1
    last_data = row - 1

    ua = ld.get("unassigned")
    if ua:
        # Cells with no registered leader — visible, never ranked (page rule).
        bg = _fill(WC_TINT[SLATE])
        _block(ws, row, C1, row, C1, lbl.get("unassignedRow", ""), fill=bg, border=BOX,
               align=Alignment(horizontal="left", vertical="center", indent=1),
               font=Font(name=FONT, size=9.5, italic=True, color=INK_SOFT))
        for c in (C1 + 1, C1 + 2):
            _block(ws, row, c, row, c, "—", fill=bg, border=BOX,
                   font=Font(name=FONT, size=9.5, color=INK_FAINT))
        _block(ws, row, C1 + 3, row, C1 + 3, ua.get("total"), fill=bg, border=BOX,
               align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=9.5, bold=True, italic=True, color=INK_SOFT))
        for j, s in enumerate(sts):
            _block(ws, row, C1 + 4 + j, row, C1 + 4 + j, ua.get(s) or 0, fill=bg,
                   border=BOX, align=RIGHT, fmt=NUM,
                   font=Font(name=FONT, size=9.5, italic=True, color=INK_SOFT))
        _block(ws, row, C1 + 4 + len(sts), row, C1 + 4 + len(sts),
               (ua.get("total") or 0) - (ua.get("done") or 0), fill=bg, border=BOX,
               align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=9.5, italic=True, color=INK_SOFT))
        _block(ws, row, C2, row, C2, "—", fill=bg, border=BOX,
               align=Alignment(horizontal="center", vertical="center"),
               font=Font(name=FONT, size=9.5, color=INK_FAINT))
        row += 1

    if rows or ua:
        top = Border(left=_side(), right=_side(),
                     top=_side(BRAND, "medium"), bottom=_side())
        tot = sum(r.get("total") or 0 for r in rows) + ((ua or {}).get("total") or 0)
        done = sum(r.get("done") or 0 for r in rows) + ((ua or {}).get("done") or 0)
        sums = {s: sum(r.get(s) or 0 for r in rows) + ((ua or {}).get(s) or 0)
                for s in sts}
        bg = _fill(BRAND_SOFT)
        bold = Font(name=FONT, size=10, bold=True, color=INK)
        _block(ws, row, C1, row, C1, lbl.get("colTotal", ""), fill=bg, border=top,
               align=Alignment(horizontal="left", vertical="center", indent=1), font=bold)
        for c in (C1 + 1, C1 + 2):
            _block(ws, row, c, row, c, "", fill=bg, border=top)
        _block(ws, row, C1 + 3, row, C1 + 3, tot, fill=bg, border=top,
               align=RIGHT, fmt=NUM, font=bold)
        for j, s in enumerate(sts):
            _block(ws, row, C1 + 4 + j, row, C1 + 4 + j, sums[s], fill=bg, border=top,
                   align=RIGHT, fmt=NUM, font=bold)
        _block(ws, row, C1 + 4 + len(sts), row, C1 + 4 + len(sts), tot - done,
               fill=bg, border=top, align=RIGHT, fmt=NUM, font=bold)
        _block(ws, row, C2, row, C2,
               round(done * 100 / tot, 1) if tot else None, fill=bg, border=top,
               align=Alignment(horizontal="center", vertical="center"),
               fmt=PCT1, font=bold)
        row += 1

    if ld.get("undated"):
        _block(ws, row, C1, row, C2, f'{_n(ld["undated"])} {lbl.get("undatedNote", "")}',
               font=Font(name=FONT, size=9, italic=True, color=INK_SOFT))
        row += 1

    if rows:
        ws.auto_filter.ref = (f"{get_column_letter(C1)}{head_row}:"
                              f"{get_column_letter(C2)}{last_data}")
    # A coordinate string, not ws.cell(): on an empty sheet that cell sits
    # inside the merged «no match» row and a MergedCell cannot anchor a freeze.
    ws.freeze_panes = f"{get_column_letter(C1 + 1)}{first}"
    ws.print_title_rows = f"{head_row}:{head_row}"


# ── Reyestr ──────────────────────────────────────────────────────────────────

def _register_sheet(wb: Workbook, p: dict) -> None:
    """Tab 3 — every row the chosen scope matches, in the on-screen sort order.
    Undated rows keep their place with «—», exactly as the page shows them."""
    tabs = p.get("sheets") or {}
    lbl = p.get("labels") or {}
    reg = p.get("register") or []
    widths = {2: 11.5, 3: 11.0, 4: 24.0, 5: 26.0, 6: 26.0, 7: 62.0, 8: 15.0}
    ws = _sheet(wb, tabs.get("register", "Register"), widths, landscape=True)
    C1, C2 = 2, 8

    row = _banner(ws, 2, C1, C2, p.get("title") or "", p.get("subtitle") or "")
    row = _section(ws, row, C1, C2, lbl.get("secRegister", ""),
                   f'{_n(len(reg))} {lbl.get("rows", "")}')
    head = [lbl.get("colDate", ""), lbl.get("colCell", ""), lbl.get("colBrig", ""),
            lbl.get("colLeader", ""), lbl.get("colOwner", ""),
            lbl.get("colText", ""), lbl.get("colStatus", "")]
    _head_row(ws, row, C1, head, span=1, height=24)
    head_row = row
    row += 1
    first = row

    if not reg:
        _block(ws, row, C1, row, C2, lbl.get("noMatch", "—"), border=BOX,
               align=Alignment(horizontal="center", vertical="center"),
               font=Font(name=FONT, size=10, color=INK_FAINT))
        row += 1
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", indent=1)
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for i, r in enumerate(reg):
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        d = _pd(r.get("d"))
        _block(ws, row, C1, row, C1, d or "—", fill=bg, border=BOX, align=center,
               fmt=DATE_FMT if d else None,
               font=Font(name=FONT, size=9.5, color=INK if d else INK_FAINT))
        _block(ws, row, C1 + 1, row, C1 + 1, r.get("cell") or "—", fill=bg,
               border=BOX, align=center, font=Font(name=FONT, size=9.5, color=INK))
        _block(ws, row, C1 + 2, row, C1 + 2, r.get("brigadir") or "—", fill=bg,
               border=BOX, align=left, font=Font(name=FONT, size=9, color=INK_SOFT))
        _block(ws, row, C1 + 3, row, C1 + 3, r.get("leader") or "—", fill=bg,
               border=BOX, align=left, font=Font(name=FONT, size=9, color=INK))
        _block(ws, row, C1 + 4, row, C1 + 4, r.get("owner") or "—", fill=bg,
               border=BOX, align=left, font=Font(name=FONT, size=9, color=INK_SOFT))
        _block(ws, row, C1 + 5, row, C1 + 5, r.get("text") or "—", fill=bg,
               border=BOX, align=wrap, font=Font(name=FONT, size=9.5, color=INK))
        st = r.get("st") or "other"
        sc = WC_COLOR.get(st, OTHER_INK)
        _block(ws, row, C2, row, C2, _st_label(p, st, r.get("straw") or ""),
               fill=_fill(WC_TINT.get(sc, BAND)), border=BOX, align=center,
               font=Font(name=FONT, size=9, bold=True,
                         color=sc if sc != SLATE else INK_SOFT))
        row += 1

    if reg:
        ws.auto_filter.ref = (f"{get_column_letter(C1)}{head_row}:"
                              f"{get_column_letter(C2)}{row - 1}")
    ws.freeze_panes = f"{get_column_letter(C1 + 1)}{first}"
    ws.print_title_rows = f"{head_row}:{head_row}"


def build_worker_concerns_workbook(p: dict) -> BytesIO:
    """Assemble the three tabs and hand back the saved workbook."""
    # Column set for the wide tables: the three core statuses always, deferred /
    # other only when the scope actually contains them (they are rare).
    sts = ["done", "doing", "todo"] \
        + [s for s in ("deferred", "other") if (p.get("status_counts") or {}).get(s)]
    wb = Workbook()
    wb.remove(wb.active)
    _obzor(wb, p, sts)
    _leaders_sheet(wb, p, sts)
    _register_sheet(wb, p)
    wb.properties.title = p.get("title") or "Worker concerns"
    wb.properties.creator = "Safia Dashboard"
    wb.properties.created = datetime.now()
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
