"""
Excel export of the /arc register.

A pure formatter: the router re-queries the rows through the same filters and
derived expressions the page uses and hands them here already serialised
(ISO timestamps, derived flags), together with the visible column keys in
on-screen order and (optionally) their headers in the viewer's language.
This module never re-derives a figure — it lays out what it is given.

Two things arrive as codes rather than words, because that is what the source
ships: the ticket ``status`` (an integer) and the source flag ``is_bot``. The
page sends the words for both; the English fallbacks here exist only so a
direct API call still produces a readable file.

Timestamps arrive as UTC ISO strings and are written as NAIVE Tashkent
datetimes with a number format, so Excel sorts and filters them as dates
instead of as text and the reader sees the plant's wall clock.
"""
from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Any, Callable, Optional

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_TASHKENT = timezone(timedelta(hours=5))
DT_FMT = "DD.MM.YYYY HH:MM"


def _xl_text(v: Any) -> Any:
    """Third-party ticket text goes into a cell verbatim, so two guards: a
    control character makes openpyxl raise (the whole export 500s), and a
    leading = + - @ turns a description into a formula when the file opens.
    Prefixing an apostrophe is what Excel itself does when a user types one."""
    if v is None or not isinstance(v, str):
        return v
    v = ILLEGAL_CHARACTERS_RE.sub("", v)
    if v[:1] in ("=", "+", "-", "@"):
        v = "'" + v
    return v


# Fallback headers — the page normally sends translated ones in ``labels``.
_DEFAULT_LABELS = {
    "num": "№",
    "created": "Created",
    "division": "Division",
    "category": "Category",
    "urgent": "Urgent",
    "description": "Description",
    "author": "Author",
    "brigada": "Brigade",
    "status": "Status",
    "due": "Due",
    "overdue": "Overdue",
    "started": "Started",
    "closed": "Closed",
    "hours": "Hours to close",
    "response": "Hours to start",
    "source": "Source",
    "files": "Files",
    "phone": "Phone",
    "manager": "Division manager",
    "deny_reason": "Deny reason",
    "state": "State",
    "cell": "Cell",
    "sup": "Brigadir",
    "leader": "Leader",
}

# Fallback status words, by the API's own code. The page sends the viewer's
# language in ``status_labels``; this is what a direct API call gets.
_DEFAULT_STATUS = {
    0: "Created",
    1: "In progress",
    3: "Completed",
    4: "Denied",
    6: "Handled",
}

# The order used when the page sends no column list at all.
_DEFAULT_COLUMNS = ("num", "created", "division", "cell", "category",
                    "description", "author", "brigada", "status", "due",
                    "started", "closed", "hours", "source", "files")


def _to_local(iso: Optional[str]) -> Optional[datetime]:
    """UTC ISO string → naive Tashkent datetime (what Excel wants)."""
    if not iso:
        return None
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(_TASHKENT).replace(tzinfo=None)


def _yn(v: Any) -> str:
    return "✓" if v else ""


def _state(r: dict) -> str:
    if r.get("is_cancelled"):
        return "cancelled"
    if r.get("is_open"):
        return "open"
    return "done"


def _files(r: dict) -> Any:
    fs = r.get("files")
    return len(fs) if isinstance(fs, list) and fs else None


# key → (value getter, kind, width). kind ∈ {"text", "dt", "num", "int"} and
# picks the cell's number format / alignment. A getter takes the row and, for
# the two coded columns, the label maps the page sent.
_COLS: dict[str, tuple[Callable[[dict, dict], Any], str, int]] = {
    "num":         (lambda r, L: r.get("request_num"), "int", 9),
    "created":     (lambda r, L: _to_local(r.get("created_at")), "dt", 17),
    "division":    (lambda r, L: r.get("division_name"), "text", 26),
    "category":    (lambda r, L: r.get("category_name"), "text", 24),
    "urgent":      (lambda r, L: _yn(r.get("category_urgent")), "text", 8),
    "description": (lambda r, L: r.get("description"), "text", 48),
    "author":      (lambda r, L: r.get("user_name"), "text", 24),
    "brigada":     (lambda r, L: r.get("brigada_name"), "text", 22),
    "status":      (lambda r, L: L["status"].get(str(r.get("status")))
                    or _DEFAULT_STATUS.get(r.get("status")), "text", 18),
    "due":         (lambda r, L: _to_local(r.get("due")), "dt", 17),
    "overdue":     (lambda r, L: _yn(r.get("overdue_now") or r.get("late")), "text", 9),
    "started":     (lambda r, L: _to_local(r.get("started_at")), "dt", 17),
    "closed":      (lambda r, L: _to_local(r.get("closed_at")), "dt", 17),
    "hours":       (lambda r, L: r.get("hours_to_close"), "num", 10),
    "response":    (lambda r, L: r.get("hours_to_start"), "num", 10),
    "source":      (lambda r, L: L["source"].get("bot" if r.get("is_bot") else "app"), "text", 10),
    "files":       (lambda r, L: _files(r), "int", 8),
    "phone":       (lambda r, L: r.get("user_phone"), "text", 16),
    "manager":     (lambda r, L: r.get("manager_name"), "text", 22),
    "deny_reason": (lambda r, L: r.get("deny_reason"), "text", 32),
    "state":       (lambda r, L: _state(r), "text", 11),
    # The cell the division NAMES (services/arc_cells.py) — its CODE, which is
    # the whole label on screen too (frontend utils/cellName.js: a cell is never
    # written out by its workshop name). Nothing at all where the division names
    # no cell; the two owner columns beside this one say whose it is.
    "cell":        (lambda r, L: r.get("cell_code"), "text", 12),
    # The cell's OWNERS on this platform's org chart, reached through that same
    # code — resolved by the router off the one cells map, never per row here.
    # Blank where the division names no cell, or names one the registry has
    # never heard of: an unreachable owner is not an empty one.
    "sup":         (lambda r, L: r.get("sup_name"), "text", 22),
    "leader":      (lambda r, L: r.get("leader_name"), "text", 22),
}

_HEAD_FILL = PatternFill("solid", fgColor="F1F5F9")
_HEAD_FONT = Font(bold=True, size=10)
_BODY_FONT = Font(size=10)
_RIGHT = Alignment(horizontal="right", vertical="top")
_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="top")


def build_arc_workbook(rows: list[dict], columns: Optional[list[str]] = None,
                       labels: Optional[dict[str, str]] = None,
                       status_labels: Optional[dict[str, str]] = None) -> BytesIO:
    """One sheet «ARC»: a bold header row, frozen at A2, one row per ticket
    in the page's column order. Unknown keys are skipped rather than failing
    the export — the page's catalog may grow ahead of this list."""
    keys = [k for k in (columns or _DEFAULT_COLUMNS) if k in _COLS]
    if not keys:
        keys = [k for k in _DEFAULT_COLUMNS if k in _COLS]
    labels = labels or {}
    # The two coded columns' words, in the viewer's language when the page sent
    # them. «source» rides in the same map under its two values.
    L = {
        "status": {k: v for k, v in (status_labels or {}).items()},
        "source": {"bot": labels.get("_bot") or "Bot", "app": labels.get("_app") or "App"},
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "ARC"

    for i, key in enumerate(keys, 1):
        c = ws.cell(row=1, column=i, value=_xl_text(labels.get(key) or _DEFAULT_LABELS.get(key, key)))
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = _COLS[key][2]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, r in enumerate(rows, 2):
        for ci, key in enumerate(keys, 1):
            getter, kind, _w = _COLS[key]
            v = getter(r, L)
            if kind == "text":
                v = _xl_text(v)
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = _BODY_FONT
            if kind == "dt":
                c.number_format = DT_FMT
                c.alignment = _RIGHT
            elif kind == "num":
                c.number_format = "0.0"
                c.alignment = _RIGHT
            elif kind == "int":
                c.number_format = "0"
                c.alignment = _RIGHT
            elif key in ("urgent", "overdue"):
                c.alignment = _CENTER
            else:
                c.alignment = _LEFT

    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(keys))}{len(rows) + 1}"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
