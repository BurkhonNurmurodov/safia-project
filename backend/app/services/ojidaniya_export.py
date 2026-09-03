"""
Excel export of the /downtime (Ojidaniya) page — the report, not the raw dump.

What people do to a raw export before they can use it — freeze the header, bold
it, widen the columns, give the numbers a format, sort by the biggest, add a
totals row, zebra the rows, add filters, turn the codes into words, pull the
per-person and per-category summaries out into their own tables, colour the
cells over the threshold, add the chart, name the sheets, and put a title with
the period and the filters at the top so the file explains itself when it is
forwarded — is done HERE, once, so the file opens ready to read and ready to
print.

The router computes every figure through the same code the page reads
(`_downtime` + `_cell_detail` in routers/downtime.py) and hands them here with
the viewer's labels; this module is a pure formatter and never re-derives a
number. Layout follows the page's reading order across five tabs:

    Umumiy       banner · the scope the numbers were taken under · KPI cards ·
                 per-brigadir table with data bars + bar chart · category share
                 with a doughnut in the platform's category colours
    Kunlik       brigadir × date matrix (the page's own bands: green at 0,
                 red past the 50-min flag) with
                 the fleet-per-day trend beneath it
    Reyestr      one row per brigadir-day, every category a column, filterable
    Yacheykalar  one row per EVENT the cells filed (cell, leader, clock, note);
                 a day still read off the shift report is one row per category,
                 marked as such, so the file is never shorter than the screen
    Izoh         what every category means and whether the загрузка counts it

Chrome is the platform's: brand gold, ink greys, and the traffic-light red for
anything over the 50-minute flag; category colours are the dashboard's own
(sent by the client off `catColor`), so a slice here is the slice on screen.
"""
from datetime import datetime
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.quality_export import (
    BAND, BOX, BRAND, BRAND_SOFT, FONT, INK, INK_FAINT, INK_SOFT,
    NUM, PANEL, PCT1, RED, RIGHT, SLATE, TINT,
    _banner, _block, _fill, _head_row, _kpi_cards, _meta_strip, _section,
    _sheet, _side,
)

INDIGO = "6366F1"            # the page's under-threshold bar
MIN = "#,##0.#"              # minutes: a decimal only when there is one
HRS = "0.0"
DATE_FMT = "DD.MM.YYYY"
THRESHOLD = 50
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)


def _xl(v: Any) -> Any:
    """Free text (a leader's note, a name) goes in verbatim, so two guards: a
    control character makes openpyxl raise, and a leading = + - @ turns the
    text into a formula when it opens."""
    if v is None or not isinstance(v, str):
        return v
    v = ILLEGAL_CHARACTERS_RE.sub("", v)
    if v[:1] in ("=", "+", "-", "@"):
        v = "'" + v
    return v


def _hex(c: Optional[str], fallback: str = SLATE) -> str:
    c = (c or "").lstrip("#").upper()
    return c if len(c) == 6 else fallback


def _iso(d: str):
    try:
        return datetime.strptime(d, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return d


def _cat_color(p: dict, name: str) -> str:
    return _hex(((p.get("cat_meta") or {}).get(name) or {}).get("color"))


def _cat_label(p: dict, name: str) -> str:
    return ((p.get("cat_meta") or {}).get(name) or {}).get("label") or ""


# ── tab 1: overview ───────────────────────────────────────────────────────────
def _overview(wb: Workbook, p: dict) -> None:
    L = p.get("labels") or {}
    tabs = p.get("sheets") or {}
    ws = _sheet(wb, tabs.get("overview", "Overview"), {c: 12.6 for c in range(2, 14)}, landscape=True)
    C1, C2 = 2, 13

    row = _banner(ws, 2, C1, C2, p.get("title") or "", p.get("subtitle") or "")
    row = _meta_strip(ws, row, C1, C2, p.get("meta") or [])

    kpis = p.get("kpis") or []
    if kpis:
        row = _section(ws, row, C1, C2, L.get("kpi", "Key figures"), "")
        row = _kpi_cards(ws, row, C1, kpis)

    # ── per-brigadir table ──
    sup = p.get("summary") or []
    cats = p.get("cats") or []
    if sup:
        row = _section(ws, row, C1, C2, L.get("bySup", ""), L.get("bySupSub", ""))
        head = [L.get("supervisor", ""), L.get("shift", ""), L.get("totalMin", ""),
                L.get("hours", ""), L.get("days", ""), L.get("flaggedDays", ""),
                L.get("avgDay", ""), L.get("topCat", ""), L.get("share", "")]
        _head_row(ws, row, C1, head, span=1, first_span=4, height=30)
        head_row = row
        row += 1
        first = row
        grand = sum(s.get("total") or 0 for s in sup) or 1
        for i, s in enumerate(sup):
            ws.row_dimensions[row].height = 17
            bg = _fill(PANEL if i % 2 == 0 else BAND)
            hot = (s.get("flagged_days") or 0) > 0
            _block(ws, row, C1, row, C1 + 3, _xl(s.get("name") or ""), fill=bg, border=BOX,
                   font=Font(name=FONT, size=10, color=INK))
            c = C1 + 4
            _block(ws, row, c, row, c, f"S{s['shift']}" if s.get("shift") else "—",
                   fill=bg, border=BOX, align=CENTER,
                   font=Font(name=FONT, size=9.5, color=INK_SOFT))
            _block(ws, row, c + 1, row, c + 1, s.get("total") or 0, fill=bg, border=BOX,
                   align=RIGHT, fmt=MIN,
                   font=Font(name=FONT, size=10, bold=True, color=RED if hot else INK))
            _block(ws, row, c + 2, row, c + 2, round((s.get("total") or 0) / 60.0, 1),
                   fill=bg, border=BOX, align=RIGHT, fmt=HRS,
                   font=Font(name=FONT, size=10, color=INK_SOFT))
            _block(ws, row, c + 3, row, c + 3, s.get("days") or 0, fill=bg, border=BOX,
                   align=RIGHT, fmt=NUM, font=Font(name=FONT, size=10, color=INK))
            fd = s.get("flagged_days") or 0
            _block(ws, row, c + 4, row, c + 4, fd, fill=_fill(TINT[RED]) if hot else bg,
                   border=BOX, align=RIGHT, fmt=NUM,
                   font=Font(name=FONT, size=10, bold=hot, color=RED if hot else INK_FAINT))
            _block(ws, row, c + 5, row, c + 5, s.get("avg") or 0, fill=bg, border=BOX,
                   align=RIGHT, fmt=MIN, font=Font(name=FONT, size=10, color=INK))
            top = s.get("top_cat") or ""
            _block(ws, row, c + 6, row, c + 6, top or "—", fill=bg, border=BOX, align=CENTER,
                   font=Font(name=FONT, size=10, bold=bool(top),
                             color=_cat_color(p, top) if top else INK_FAINT))
            _block(ws, row, c + 7, row, c + 7, round((s.get("total") or 0) * 100.0 / grand, 1),
                   fill=bg, border=BOX, align=RIGHT, fmt=PCT1,
                   font=Font(name=FONT, size=10, color=INK_SOFT))
            row += 1
        last = row - 1
        col = get_column_letter(C1 + 5)
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color=INDIGO, showValue=True),
        )
        # totals
        ws.row_dimensions[row].height = 18
        top_b = Border(left=_side(), right=_side(), top=_side(BRAND, "medium"), bottom=_side())
        tot = p.get("totals") or {}
        bold = Font(name=FONT, size=10, bold=True, color=INK)
        _block(ws, row, C1, row, C1 + 3, L.get("total", "Total"), fill=_fill(BRAND_SOFT),
               border=top_b, font=bold)
        c = C1 + 4
        _block(ws, row, c, row, c, "", fill=_fill(BRAND_SOFT), border=top_b)
        _block(ws, row, c + 1, row, c + 1, tot.get("total") or 0, fill=_fill(BRAND_SOFT),
               border=top_b, align=RIGHT, fmt=MIN, font=bold)
        _block(ws, row, c + 2, row, c + 2, round((tot.get("total") or 0) / 60.0, 1),
               fill=_fill(BRAND_SOFT), border=top_b, align=RIGHT, fmt=HRS, font=bold)
        _block(ws, row, c + 3, row, c + 3, tot.get("days") or 0, fill=_fill(BRAND_SOFT),
               border=top_b, align=RIGHT, fmt=NUM, font=bold)
        _block(ws, row, c + 4, row, c + 4, tot.get("flagged_days") or 0, fill=_fill(BRAND_SOFT),
               border=top_b, align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=10, bold=True,
                         color=RED if (tot.get("flagged_days") or 0) else INK))
        _block(ws, row, c + 5, row, c + 5, tot.get("avg") or 0, fill=_fill(BRAND_SOFT),
               border=top_b, align=RIGHT, fmt=MIN, font=bold)
        tc = tot.get("top_cat") or ""
        _block(ws, row, c + 6, row, c + 6, tc or "—", fill=_fill(BRAND_SOFT), border=top_b,
               align=CENTER, font=Font(name=FONT, size=10, bold=True,
                                       color=_cat_color(p, tc) if tc else INK_FAINT))
        _block(ws, row, c + 7, row, c + 7, 100, fill=_fill(BRAND_SOFT), border=top_b,
               align=RIGHT, fmt=PCT1, font=bold)
        row += 2

        # the page's bar chart: one bar per brigadir, red once the unit carries a
        # flagged day, indigo otherwise — the same two colours the screen uses
        chart = BarChart()
        chart.type = "bar"
        chart.style = 2
        chart.legend = None
        chart.y_axis.majorGridlines = None
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.height = max(6.0, 0.55 * len(sup) + 2.2)
        chart.width = 24
        chart.title = L.get("bySup") or None
        chart.add_data(Reference(ws, min_col=C1 + 5, min_row=head_row, max_row=last),
                       titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=C1, min_row=first, max_row=last))
        chart.x_axis.scaling.orientation = "maxMin"      # biggest on top, as on screen
        pts = []
        for i, s in enumerate(sup):
            dp = DataPoint(idx=i)
            dp.graphicalProperties.solidFill = RED if (s.get("flagged_days") or 0) > 0 else INDIGO
            dp.graphicalProperties.line.noFill = True
            pts.append(dp)
        chart.series[0].data_points = pts
        chart.series[0].graphicalProperties.solidFill = INDIGO
        chart.gapWidth = 60
        ws.add_chart(chart, f"{get_column_letter(C1)}{row}")
        row += int(chart.height * 1.95) + 2

    # ── category share ──
    share = p.get("cat_share") or []
    if share:
        start = row
        row = _section(ws, row, C1, C2, L.get("catShare", ""), L.get("catShareSub", ""))
        # header — built by hand because the name column spans four cells
        ws.row_dimensions[row].height = 26
        hfont = Font(name=FONT, size=9, bold=True, color=INK_SOFT)
        _block(ws, row, C1, row, C1, L.get("cat", ""), fill=_fill(BAND), border=BOX,
               font=hfont, align=CENTER)
        _block(ws, row, C1 + 1, row, C1 + 4, L.get("catName", ""), fill=_fill(BAND), border=BOX,
               font=hfont, align=Alignment(horizontal="left", vertical="center", indent=1))
        for j, lab in enumerate((L.get("minutes", ""), L.get("share", ""), L.get("counted", ""))):
            _block(ws, row, C1 + 5 + j, row, C1 + 5 + j, lab, fill=_fill(BAND), border=BOX,
                   font=hfont, align=CENTER)
        row += 1
        first = row
        total = sum(r.get("minutes") or 0 for r in share) or 1
        for i, r in enumerate(share):
            ws.row_dimensions[row].height = 17
            bg = _fill(PANEL if i % 2 == 0 else BAND)
            color = _cat_color(p, r.get("name") or "")
            _block(ws, row, C1, row, C1, r.get("name") or "", fill=bg, border=BOX, align=CENTER,
                   font=Font(name=FONT, size=10, bold=True, color=color))
            _block(ws, row, C1 + 1, row, C1 + 4, _xl(r.get("label") or ""), fill=bg, border=BOX,
                   font=Font(name=FONT, size=9.5, color=INK))
            _block(ws, row, C1 + 5, row, C1 + 5, r.get("minutes") or 0, fill=bg, border=BOX,
                   align=RIGHT, fmt=MIN, font=Font(name=FONT, size=10, bold=True, color=INK))
            _block(ws, row, C1 + 6, row, C1 + 6, round((r.get("minutes") or 0) * 100.0 / total, 1),
                   fill=bg, border=BOX, align=RIGHT, fmt=PCT1,
                   font=Font(name=FONT, size=10, color=INK_SOFT))
            counted = bool(r.get("counted"))
            _block(ws, row, C1 + 7, row, C1 + 7, L.get("yes", "Yes") if counted else L.get("no", "No"),
                   fill=bg, border=BOX, align=CENTER,
                   font=Font(name=FONT, size=9.5, color=INK if counted else INK_FAINT))
            row += 1
        last = row - 1
        col = get_column_letter(C1 + 5)
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color=BRAND, showValue=True),
        )
        ws.row_dimensions[row].height = 18
        top_b = Border(left=_side(), right=_side(), top=_side(BRAND, "medium"), bottom=_side())
        _block(ws, row, C1, row, C1 + 4, L.get("total", "Total"), fill=_fill(BRAND_SOFT),
               border=top_b, font=Font(name=FONT, size=10, bold=True, color=INK))
        _block(ws, row, C1 + 5, row, C1 + 5, total if total != 1 or share else 0,
               fill=_fill(BRAND_SOFT), border=top_b, align=RIGHT, fmt=MIN,
               font=Font(name=FONT, size=10, bold=True, color=INK))
        _block(ws, row, C1 + 6, row, C1 + 6, 100, fill=_fill(BRAND_SOFT), border=top_b,
               align=RIGHT, fmt=PCT1, font=Font(name=FONT, size=10, bold=True, color=INK))
        _block(ws, row, C1 + 7, row, C1 + 7, "", fill=_fill(BRAND_SOFT), border=top_b)
        row += 2

        # the doughnut, parked beside the table in the page's own colours
        if any((r.get("minutes") or 0) > 0 for r in share):
            chart = DoughnutChart(holeSize=58)
            chart.height, chart.width = max(7.5, 0.5 * len(share) + 4), 12
            chart.legend.position = "r"
            chart.add_data(Reference(ws, min_col=C1 + 5, min_row=first, max_row=last),
                           titles_from_data=False)
            chart.set_categories(Reference(ws, min_col=C1, min_row=first, max_row=last))
            pts = []
            for i, r in enumerate(share):
                dp = DataPoint(idx=i)
                dp.graphicalProperties.solidFill = _cat_color(p, r.get("name") or "")
                dp.graphicalProperties.line.solidFill = "FFFFFF"
                pts.append(dp)
            chart.series[0].data_points = pts
            ws.add_chart(chart, f"{get_column_letter(C1 + 8)}{start + 1}")
            row = max(row, start + int(chart.height * 1.95) + 3)

    ws.print_title_rows = "1:3"


# The screen's ramp for this same matrix (frontend components/idle/OjidaniyaMatrix
# .jsx, IDLE_SEGMENTS) as discrete BANDS, so a cell means the same thing in the
# file as it does on the page: 0 minutes is the best a day can go and reads
# green, the hue turns at the 50-minute flag, and 100+ is its own darkest red.
# It replaced a continuous white→amber→red colour scale, under which 0 and 12
# minutes were indistinguishable and nothing marked the threshold but the ink.
#
# Rendered in the workbook's own idiom — a light fill under band-coloured ink —
# rather than the screen's saturated fills: this sheet is landscape and meant to
# be printed, and a page of solid dark green costs the toner and the legibility
# without adding a fact. The BANDS are what has to match, and they do.
#
# Guarded on LEN() where the band touches zero: an empty cell is a day nobody
# reported, and Excel reads it as 0 — without the guard every unreported day
# would be filled as though the unit had reported a perfect one.
_MX_BANDS = [
    ("AND(LEN({c})>0,{c}=0)",        "D7F0DF", "15803D", True),   # nothing waited
    ("AND(LEN({c})>0,{c}>0,{c}<15)", "E7F8EE", "15803D", False),  # under 15 min
    ("AND({c}>=15,{c}<30)",          "EFF7D8", "4D7C0F", False),  # 15–29
    ("AND({c}>=30,{c}<50)",          "FBF4DA", "854D0E", False),  # 30–49
    ("AND({c}>=50,{c}<100)",         "FDEAEA", RED,      True),   # over the flag
    ("{c}>=100",                     "F7D5D5", "B91C1C", True),   # 100+
]


def _band_matrix(ws: Worksheet, rng: str, col: str, row: int) -> None:
    """Paint the brigadir × day range in the page's own bands."""
    anchor = f"{col}{row}"                      # rules are relative to the range's first cell
    for formula, fill, ink, bold in _MX_BANDS:
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[formula.format(c=anchor)],
            # NOT `_fill()`: that writes fgColor alone, which is what a normal
            # cell style wants and what a DIFFERENTIAL one (a conditional
            # format) reads as the pattern colour rather than the background —
            # Excel then paints nothing and the whole matrix comes out white,
            # with the file otherwise perfectly valid. Setting both ends emits
            # fgColor AND bgColor, which every reader renders the same way.
            fill=PatternFill("solid", start_color=fill, end_color=fill),
            font=Font(name=FONT, size=9.5, bold=bold, color=ink),
        ))


# ── tab 2: brigadir × date matrix + trend ────────────────────────────────────
def _daily(wb: Workbook, p: dict) -> None:
    L = p.get("labels") or {}
    tabs = p.get("sheets") or {}
    dates = p.get("dates") or []
    sup = p.get("summary") or []
    if not dates or not sup:
        return
    ncol = 3 + len(dates)
    widths = {2: 26.0, 3: 7.5, 4: 11.0}
    widths.update({c: 7.4 for c in range(5, 2 + ncol)})
    ws = _sheet(wb, tabs.get("daily", "Daily"), widths, landscape=True)
    C1 = 2
    C2 = C1 + ncol - 1

    row = _banner(ws, 2, C1, C2, p.get("title") or "", p.get("subtitle") or "")
    row = _section(ws, row, C1, C2, L.get("matrix", ""), L.get("matrixSub", ""))

    head = [L.get("supervisor", ""), L.get("shift", ""), L.get("total", "")] + \
           [_iso(d).strftime("%d.%m") if hasattr(_iso(d), "strftime") else d for d in dates]
    _head_row(ws, row, C1, head, height=26)
    # weekday initials under the dates read faster than a second header row
    row += 1
    first = row
    matrix = p.get("matrix") or {}
    for i, s in enumerate(sup):
        ws.row_dimensions[row].height = 17
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        hot = (s.get("flagged_days") or 0) > 0
        _block(ws, row, C1, row, C1, _xl(s.get("name") or ""), fill=bg, border=BOX,
               font=Font(name=FONT, size=10, color=INK))
        _block(ws, row, C1 + 1, row, C1 + 1, f"S{s['shift']}" if s.get("shift") else "—",
               fill=bg, border=BOX, align=CENTER, font=Font(name=FONT, size=9.5, color=INK_SOFT))
        _block(ws, row, C1 + 2, row, C1 + 2, s.get("total") or 0, fill=bg, border=BOX,
               align=RIGHT, fmt=MIN,
               font=Font(name=FONT, size=10, bold=True, color=RED if hot else INK))
        per = matrix.get(s.get("key")) or {}
        for j, d in enumerate(dates):
            v = per.get(d)
            _block(ws, row, C1 + 3 + j, row, C1 + 3 + j, v, fill=bg, border=BOX, align=CENTER,
                   fmt=MIN, font=Font(name=FONT, size=9.5, color=INK if v else INK_FAINT))
        row += 1
    last = row - 1

    # fleet per day — the row the trend is drawn from
    fleet = p.get("fleet_by_day") or {}
    ws.row_dimensions[row].height = 18
    top_b = Border(left=_side(), right=_side(), top=_side(BRAND, "medium"), bottom=_side())
    bold = Font(name=FONT, size=10, bold=True, color=INK)
    _block(ws, row, C1, row, C1 + 1, L.get("fleetTotal", "Total"), fill=_fill(BRAND_SOFT),
           border=top_b, font=bold)
    _block(ws, row, C1 + 2, row, C1 + 2, (p.get("totals") or {}).get("total") or 0,
           fill=_fill(BRAND_SOFT), border=top_b, align=RIGHT, fmt=MIN, font=bold)
    for j, d in enumerate(dates):
        v = fleet.get(d)
        _block(ws, row, C1 + 3 + j, row, C1 + 3 + j, v, fill=_fill(BRAND_SOFT), border=top_b,
               align=CENTER, fmt=MIN,
               font=Font(name=FONT, size=9.5, bold=True, color=INK if v else INK_FAINT))
    fleet_row = row
    row += 2

    rng = f"{get_column_letter(C1 + 3)}{first}:{get_column_letter(C2)}{last}"
    _band_matrix(ws, rng, get_column_letter(C1 + 3), first)
    ws.freeze_panes = ws.cell(first, C1 + 3)

    # ── trend: one small table (date · total · threshold) and a line chart ──
    row = _section(ws, row, C1, min(C2, C1 + 6), L.get("trend", ""), L.get("trendSub", ""))
    _head_row(ws, row, C1, [L.get("date", ""), L.get("fleetTotal", ""), L.get("threshold", "")],
              height=22)
    thead = row
    row += 1
    tfirst = row
    for i, d in enumerate(dates):
        ws.row_dimensions[row].height = 15
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        dv = _iso(d)
        _block(ws, row, C1, row, C1, dv, fill=bg, border=BOX, align=CENTER, fmt=DATE_FMT,
               font=Font(name=FONT, size=9.5, color=INK))
        v = fleet.get(d) or 0
        _block(ws, row, C1 + 1, row, C1 + 1, v, fill=bg, border=BOX, align=RIGHT, fmt=MIN,
               font=Font(name=FONT, size=9.5, bold=v > THRESHOLD, color=RED if v > THRESHOLD else INK))
        _block(ws, row, C1 + 2, row, C1 + 2, THRESHOLD, fill=bg, border=BOX, align=RIGHT, fmt=NUM,
               font=Font(name=FONT, size=9, color=INK_FAINT))
        row += 1
    tlast = row - 1

    chart = LineChart()
    chart.style = 2
    chart.height, chart.width = 8.5, 24
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.number_format = "dd.mm"
    chart.add_data(Reference(ws, min_col=C1 + 1, min_row=thead, max_row=tlast), titles_from_data=True)
    chart.add_data(Reference(ws, min_col=C1 + 2, min_row=thead, max_row=tlast), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=C1, min_row=tfirst, max_row=tlast))
    s0 = chart.series[0]
    s0.graphicalProperties.line.solidFill = INDIGO
    s0.graphicalProperties.line.width = 22000
    s0.smooth = False
    s1 = chart.series[1]
    s1.graphicalProperties.line.solidFill = RED
    s1.graphicalProperties.line.dashStyle = "dash"
    s1.graphicalProperties.line.width = 12000
    chart.legend.position = "t"
    ws.add_chart(chart, f"{get_column_letter(C1 + 4)}{thead}")
    ws.print_title_rows = f"{first - 1}:{first - 1}"
    _ = fleet_row


# ── tab 3: the daily register ────────────────────────────────────────────────
def _register(wb: Workbook, p: dict) -> None:
    L = p.get("labels") or {}
    tabs = p.get("sheets") or {}
    rows = p.get("daily_rows") or []
    cats = p.get("cats") or []
    if not rows:
        return
    fixed = [L.get("date", ""), L.get("supervisor", ""), L.get("shift", ""),
             L.get("totalMin", ""), L.get("flagged", "")]
    cols = fixed + list(cats) + [L.get("source", "")]
    widths = {2: 12.0, 3: 26.0, 4: 7.5, 5: 12.0, 6: 12.0}
    widths.update({7 + i: 9.5 for i in range(len(cats))})
    widths[7 + len(cats)] = 17.0
    ws = _sheet(wb, tabs.get("register", "Register"), widths, landscape=True)
    C1 = 2
    C2 = C1 + len(cols) - 1

    row = _banner(ws, 2, C1, C2, p.get("title") or "", p.get("subtitle") or "")
    row = _section(ws, row, C1, C2, L.get("register", ""),
                   f"{len(rows)} {L.get('rows', '')}".strip())
    _head_row(ws, row, C1, cols, height=28)
    # category headers carry their own colour so the eye finds a column fast
    for i, cat in enumerate(cats):
        cell = ws.cell(row, C1 + 5 + i)
        cell.font = Font(name=FONT, size=9, bold=True, color=_cat_color(p, cat))
    head_row = row
    row += 1
    first = row
    for i, r in enumerate(rows):
        ws.row_dimensions[row].height = 16
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        hot = bool(r.get("flagged"))
        _block(ws, row, C1, row, C1, _iso(r.get("date")), fill=bg, border=BOX, align=CENTER,
               fmt=DATE_FMT, font=Font(name=FONT, size=9.5, color=INK))
        _block(ws, row, C1 + 1, row, C1 + 1, _xl(r.get("name") or ""), fill=bg, border=BOX,
               font=Font(name=FONT, size=9.5, color=INK))
        _block(ws, row, C1 + 2, row, C1 + 2, f"S{r['shift']}" if r.get("shift") else "—",
               fill=bg, border=BOX, align=CENTER, font=Font(name=FONT, size=9, color=INK_SOFT))
        _block(ws, row, C1 + 3, row, C1 + 3, r.get("total") or 0,
               fill=_fill(TINT[RED]) if hot else bg, border=BOX, align=RIGHT, fmt=MIN,
               font=Font(name=FONT, size=10, bold=True, color=RED if hot else INK))
        _block(ws, row, C1 + 4, row, C1 + 4, L.get("yes", "Yes") if hot else "—",
               fill=_fill(TINT[RED]) if hot else bg, border=BOX, align=CENTER,
               font=Font(name=FONT, size=9, bold=hot, color=RED if hot else INK_FAINT))
        by = r.get("by_cat") or {}
        for j, cat in enumerate(cats):
            v = by.get(cat) or 0
            _block(ws, row, C1 + 5 + j, row, C1 + 5 + j, v, fill=bg, border=BOX, align=RIGHT,
                   fmt=MIN, font=Font(name=FONT, size=9.5, color=INK if v else INK_FAINT))
        src = r.get("source")
        _block(ws, row, C2, row, C2,
               L.get("srcCells", "") if src == "cells" else L.get("srcSheet", ""),
               fill=bg, border=BOX, font=Font(name=FONT, size=9, color=INK_SOFT))
        row += 1
    ws.auto_filter.ref = f"{get_column_letter(C1)}{head_row}:{get_column_letter(C2)}{row - 1}"
    ws.freeze_panes = ws.cell(first, C1 + 2)
    ws.print_title_rows = f"{head_row}:{head_row}"


# ── tab 4: the events the cells filed ────────────────────────────────────────
def _events(wb: Workbook, p: dict) -> None:
    L = p.get("labels") or {}
    tabs = p.get("sheets") or {}
    rows = p.get("events") or []
    cols = [L.get("date", ""), L.get("supervisor", ""), L.get("cell", ""), L.get("leader", ""),
            L.get("cat", ""), L.get("catName", ""), L.get("start", ""), L.get("end", ""),
            L.get("minutes", ""), L.get("status", ""), L.get("note", ""), L.get("source", "")]
    widths = {2: 12.0, 3: 24.0, 4: 10.0, 5: 24.0, 6: 8.5, 7: 36.0, 8: 10.5, 9: 10.5,
              10: 10.0, 11: 15.0, 12: 46.0, 13: 17.0}
    ws = _sheet(wb, tabs.get("events", "Events"), widths, landscape=True)
    C1, C2 = 2, 13

    row = _banner(ws, 2, C1, C2, p.get("title") or "", p.get("subtitle") or "")
    row = _section(ws, row, C1, C2, L.get("events", ""),
                   f"{len(rows)} {L.get('rows', '')}".strip() if rows else "")
    if not rows:
        _block(ws, row, C1, row, C2, L.get("noEvents", ""), fill=_fill(BAND), border=BOX,
               font=Font(name=FONT, size=10, italic=True, color=INK_SOFT))
        return
    _head_row(ws, row, C1, cols, height=28)
    head_row = row
    row += 1
    first = row
    for i, e in enumerate(rows):
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        cat = e.get("category") or ""
        stopped = e.get("stopped")
        _block(ws, row, C1, row, C1, _iso(e.get("date")), fill=bg, border=BOX, align=CENTER,
               fmt=DATE_FMT, font=Font(name=FONT, size=9.5, color=INK))
        _block(ws, row, C1 + 1, row, C1 + 1, _xl(e.get("name") or ""), fill=bg, border=BOX,
               font=Font(name=FONT, size=9.5, color=INK))
        _block(ws, row, C1 + 2, row, C1 + 2, _xl(e.get("cell") or "") or "—", fill=bg, border=BOX,
               align=CENTER, font=Font(name=FONT, size=9.5, bold=bool(e.get("cell")),
                                       color=INK if e.get("cell") else INK_FAINT))
        _block(ws, row, C1 + 3, row, C1 + 3, _xl(e.get("leader") or "") or "—", fill=bg, border=BOX,
               font=Font(name=FONT, size=9.5, color=INK if e.get("leader") else INK_FAINT))
        _block(ws, row, C1 + 4, row, C1 + 4, cat, fill=bg, border=BOX, align=CENTER,
               font=Font(name=FONT, size=9.5, bold=True, color=_cat_color(p, cat)))
        _block(ws, row, C1 + 5, row, C1 + 5, _xl(_cat_label(p, cat)) or "—", fill=bg, border=BOX,
               font=Font(name=FONT, size=9, color=INK_SOFT))
        _block(ws, row, C1 + 6, row, C1 + 6, e.get("start") or "—", fill=bg, border=BOX,
               align=CENTER, font=Font(name=FONT, size=9.5, color=INK if e.get("start") else INK_FAINT))
        _block(ws, row, C1 + 7, row, C1 + 7, e.get("end") or "—", fill=bg, border=BOX,
               align=CENTER, font=Font(name=FONT, size=9.5, color=INK if e.get("end") else INK_FAINT))
        _block(ws, row, C1 + 8, row, C1 + 8, e.get("minutes") or 0, fill=bg, border=BOX,
               align=RIGHT, fmt=MIN, font=Font(name=FONT, size=10, bold=True, color=INK))
        if stopped is None:
            st, sc = "—", INK_FAINT
        elif stopped:
            st, sc = L.get("stoppedYes", ""), RED
        else:
            st, sc = L.get("stoppedNo", ""), INK_SOFT
        _block(ws, row, C1 + 9, row, C1 + 9, st, fill=bg, border=BOX, align=CENTER,
               font=Font(name=FONT, size=9, bold=stopped is True, color=sc))
        note = _xl(e.get("note") or "")
        _block(ws, row, C1 + 10, row, C1 + 10, note or "", fill=bg, border=BOX, align=WRAP,
               font=Font(name=FONT, size=9, color=INK if note else INK_FAINT))
        src = e.get("source")
        _block(ws, row, C1 + 11, row, C1 + 11,
               L.get("srcCells", "") if src == "cells" else L.get("srcSheet", ""),
               fill=bg, border=BOX, font=Font(name=FONT, size=9, color=INK_SOFT))
        row += 1
    ws.auto_filter.ref = f"{get_column_letter(C1)}{head_row}:{get_column_letter(C2)}{row - 1}"
    ws.freeze_panes = ws.cell(first, C1 + 2)
    ws.print_title_rows = f"{head_row}:{head_row}"


# ── tab 5: what the categories mean ──────────────────────────────────────────
def _legend(wb: Workbook, p: dict) -> None:
    L = p.get("labels") or {}
    tabs = p.get("sheets") or {}
    meta = p.get("cat_meta") or {}
    order = p.get("cat_order") or list(meta.keys())
    if not order:
        return
    ws = _sheet(wb, tabs.get("legend", "Legend"), {2: 9.0, 3: 40.0, 4: 84.0, 5: 15.0}, landscape=True)
    C1, C2 = 2, 5
    row = _banner(ws, 2, C1, C2, L.get("legendTitle", ""), L.get("legendSub", ""))
    _head_row(ws, row, C1, [L.get("cat", ""), L.get("catName", ""), L.get("catNote", ""),
                            L.get("counted", "")], height=24)
    row += 1
    for i, name in enumerate(order):
        m = meta.get(name) or {}
        bg = _fill(PANEL if i % 2 == 0 else BAND)
        _block(ws, row, C1, row, C1, name, fill=bg, border=BOX, align=CENTER,
               font=Font(name=FONT, size=10, bold=True, color=_hex(m.get("color"))))
        _block(ws, row, C1 + 1, row, C1 + 1, _xl(m.get("label") or ""), fill=bg, border=BOX,
               align=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1),
               font=Font(name=FONT, size=10, bold=True, color=INK))
        _block(ws, row, C1 + 2, row, C1 + 2, _xl(m.get("note") or ""), fill=bg, border=BOX,
               align=WRAP, font=Font(name=FONT, size=9.5, color=INK_SOFT))
        counted = bool(m.get("counted"))
        _block(ws, row, C1 + 3, row, C1 + 3, L.get("yes", "Yes") if counted else L.get("no", "No"),
               fill=bg, border=BOX, align=Alignment(horizontal="center", vertical="top"),
               font=Font(name=FONT, size=9.5, bold=counted, color=INK if counted else INK_FAINT))
        # two lines of note per row is the usual; Excel will not autofit a row
        # whose height we set, so estimate from the text instead
        lines = max(1, (len(m.get("note") or "") + 84) // 85)
        ws.row_dimensions[row].height = 14 * lines + 6
        row += 1


def build_ojidaniya_workbook(p: dict) -> BytesIO:
    """Assemble the five tabs and hand back the saved workbook."""
    wb = Workbook()
    wb.remove(wb.active)
    _overview(wb, p)
    _daily(wb, p)
    _register(wb, p)
    _events(wb, p)
    _legend(wb, p)
    if not wb.sheetnames:
        ws = _sheet(wb, (p.get("sheets") or {}).get("overview", "Overview"),
                    {c: 13.0 for c in range(2, 14)})
        _banner(ws, 2, 2, 13, p.get("title") or "", p.get("subtitle") or "")
    wb.properties.title = p.get("title") or "Ojidaniya"
    wb.properties.creator = "Safia Dashboard"
    wb.properties.created = datetime.now()
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
