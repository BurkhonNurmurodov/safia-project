"""
Parsers for the two daily SAP exports.

  • «фаза»     (faza ….xlsx) — operations: one row per production order + operation
                step. Carries the work center, planned qty and confirmed output,
                but NOT the material/SKU. Header row 1, data from row 2.
  • «заголовок» (zaga ….xlsx) — order headers: maps each order → material (SKU)
                + name. Header row 1, data from row 2.

The dashboard is keyed by (SKU, work center), so the two files must be JOINED on
the order number: zaga gives order→SKU, faza gives the per-order operations.

Real column layout (1-based) — фаза:
    A Заказ(order) · B Операция · C Ресурс(work center) · D Кратк.текст(name)
    E Кол-во операции(PLAN) · F unit · G Системный статус · H start date
    I СамРанЗплДатаКнцВыплн(DATE filter) · J ПодтвВыходПрод(ACTUAL) · K work-place
Real column layout (1-based) — заголовок:
    A Заказ(order) · B Материал(SKU) · D Завод · E Кол-во заказа · F ПоставлКол-во
    G подтв.выход · J БазисСрокКонца(date) · L Краткий текст материала(name) · M статус
"""
from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

# фаза (operations)
FZ_ORDER, FZ_OP, FZ_WC, FZ_NAME, FZ_PLAN, FZ_STATUS, FZ_DATE, FZ_CONF = 1, 2, 3, 4, 5, 7, 9, 10
# заголовок (order headers)
ZG_ORDER, ZG_SKU, ZG_PLANT, ZG_ORDQTY, ZG_DELIV, ZG_CONF, ZG_DATE, ZG_NAME, ZG_STATUS = 1, 2, 4, 5, 6, 7, 10, 12, 13

# render-ready headers for the view switcher (faza shows the resolved SKU)
FAZA_COLUMNS = ["Заказ", "Опер.", "Команда", "SKU", "Наименование", "План", "Статус", "Дата", "Подтв."]
ZAGA_COLUMNS = ["Заказ", "SKU", "Завод", "Кол-во заказа", "Поставлено", "Подтв.", "Дата", "Наименование", "Статус"]

_SAP_RE = re.compile(r"^[A-Za-z]\d{5,}$")    # material / SKU: F00002310, S00000101
_WC4_RE = re.compile(r"^[A-Za-z]\d{4}$")      # work center: A1431, A2682 (plant W001 = 3 digits, excluded)


def _str(v) -> str:
    """Cell's string form, robust to SAP exports storing numbers as text."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _num(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", ".").strip())
        except ValueError:
            return 0.0
    return 0.0


def _get(row, idx):
    return row[idx - 1] if len(row) >= idx else None


def _classify(ws) -> str | None:
    """'faza' | 'zaga' | None — by sheet name, then content. Only фаза carries
    work-center codes (letter+4 digits); a sheet with SAP/material codes but no
    work centers is заголовок. Type/position independent."""
    title = (ws.title or "").lower()
    if "фаза" in title or "faza" in title:
        return "faza"
    if "заголов" in title or "zaga" in title:
        return "zaga"
    sap_seen = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 200:
            break
        for cell in row:
            s = _str(cell)
            if not s:
                continue
            if _WC4_RE.match(s):
                return "faza"
            if _SAP_RE.match(s):
                sap_seen = True
    return "zaga" if sap_seen else None


def _best_sheet(wb, kind: str):
    key = "фаза" if kind == "faza" else "заголов"
    for ws in wb.worksheets:
        if key in (ws.title or "").lower():
            return ws
    return wb.worksheets[0]


def _extract_faza(ws, target_date: date | None, own_wcs: set[str]) -> dict:
    """Raw operation rows (SKU resolved later by the caller via the order map)."""
    raw: list[dict] = []
    dates: set[date] = set()
    seen = 0
    for row in ws.iter_rows(values_only=True):
        wc = _str(_get(row, FZ_WC))
        if not _WC4_RE.match(wc):           # skips header + non-operation rows
            continue
        order = _str(_get(row, FZ_ORDER))
        if not order:
            continue
        d = _to_date(_get(row, FZ_DATE))
        if d is not None:
            dates.add(d)
        seen += 1
        if target_date is not None and d != target_date:
            continue
        if own_wcs and wc not in own_wcs:
            continue
        raw.append({
            "order": order,
            "op": _str(_get(row, FZ_OP)),
            "wc": wc,
            "name": _str(_get(row, FZ_NAME)),
            "plan": _num(_get(row, FZ_PLAN)),
            "status": _str(_get(row, FZ_STATUS)),
            "date": d.isoformat() if d else "",
            "conf": _num(_get(row, FZ_CONF)),
        })
    return {"raw": raw, "dates": sorted(dates), "seen": seen}


def _extract_zaga(ws, catalog_skus: set[str]) -> dict:
    """order→SKU map (ALL orders, for the join) + display rows (catalog SKUs)."""
    order_sku: dict[str, str] = {}
    rows: list[list] = []
    seen = 0
    for row in ws.iter_rows(values_only=True):
        sku = _str(_get(row, ZG_SKU))
        if not _SAP_RE.match(sku):          # skips header + blank rows
            continue
        order = _str(_get(row, ZG_ORDER))
        if order:
            order_sku[order] = sku
        seen += 1
        if catalog_skus and sku not in catalog_skus:
            continue
        rows.append([
            order, sku, _str(_get(row, ZG_PLANT)), _num(_get(row, ZG_ORDQTY)),
            _num(_get(row, ZG_DELIV)), _num(_get(row, ZG_CONF)),
            (_to_date(_get(row, ZG_DATE)).isoformat() if _to_date(_get(row, ZG_DATE)) else ""),
            _str(_get(row, ZG_NAME)), _str(_get(row, ZG_STATUS)),
        ])
    return {"order_sku": order_sku, "columns": ZAGA_COLUMNS, "rows": rows, "seen": seen}


# catalog sheet ("Sheet1 …"): main table A=SKU B=name C=labor D=WC; staffing
# block M=work center, S=capacity, W=штатка (columns 13 / 19 / 23).
CAT_SKU, CAT_NAME, CAT_LABOR, CAT_WC = 1, 2, 3, 4
CAT_WCM, CAT_CAP, CAT_SHT = 13, 19, 23


def _pick_catalog_sheet(wb, sheet_name: str | None):
    if sheet_name:
        for ws in wb.worksheets:
            if ws.title.strip().lower() == sheet_name.strip().lower():
                return ws
    # else: first sheet that looks like a dashboard (header row mentions Трудоёмкость)
    for ws in wb.worksheets:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 4:
                break
            if any("трудо" in _str(c).lower() for c in row):
                return ws
    return None


def parse_catalog_workbook(content: bytes, sheet_name: str | None = None) -> dict:
    """Parse a brigadir's catalog from a 'Sheet1 …' dashboard sheet:
    products (SKU, name, labor, work center) + work centers (штатка, capacity).
    Junk rows (blank / '0' SAP code) are dropped."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        ws = _pick_catalog_sheet(wb, sheet_name)
        if ws is None:
            return {"products": [], "work_centers": [], "sheet": None}

        # header row = the one mentioning Трудоёмкость; data starts after it
        header_i = 0
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows[:6]):
            if any("трудо" in _str(c).lower() for c in row):
                header_i = i
                break

        products = []
        order = 0
        blanks = 0
        for row in rows[header_i + 1:]:
            sku = _str(_get(row, CAT_SKU))
            if not sku:
                blanks += 1
                if blanks > 4:
                    break
                continue
            blanks = 0
            if sku == "0":          # junk row with no real SAP code
                continue
            labor = _get(row, CAT_LABOR)
            products.append({
                "sap_code": sku,
                "name": _str(_get(row, CAT_NAME)),
                "work_center": _str(_get(row, CAT_WC)),
                "labor_time": (_num(labor) if isinstance(labor, (int, float)) else None),
                "sort_order": order,
            })
            order += 1

        work_centers = []
        so = 0
        for row in rows:
            code = _str(_get(row, CAT_WCM))
            if not _WC4_RE.match(code):
                continue
            cap = _get(row, CAT_CAP)
            sht = _get(row, CAT_SHT)
            work_centers.append({
                "code": code,
                "shtatka": int(_num(sht)) if isinstance(sht, (int, float)) else 0,
                "capacity": (_num(cap) if isinstance(cap, (int, float)) else None),
                "sort_order": so,
            })
            so += 1

        return {"products": products, "work_centers": work_centers, "sheet": ws.title}
    finally:
        wb.close()


def read_workbook_slices(content: bytes, target_date: date | None,
                         own_wcs: set[str], catalog_skus: set[str],
                         force_type: str | None = None) -> dict:
    """Read an uploaded workbook once; return whichever of faza/zaga it holds.

    faza → {"raw": [op dicts], "dates": [...]}  (SKU resolved by the caller)
    zaga → {"order_sku": {...}, "rows": [...], "columns": [...]}
    force_type ('faza'|'zaga') skips auto-detection."""
    import openpyxl  # lazy — heavy import

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    out: dict = {"faza": None, "zaga": None}
    try:
        if force_type == "faza":
            out["faza"] = _extract_faza(_best_sheet(wb, "faza"), target_date, own_wcs)
        elif force_type == "zaga":
            out["zaga"] = _extract_zaga(_best_sheet(wb, "zaga"), catalog_skus)
        else:
            for ws in wb.worksheets:
                kind = _classify(ws)
                if kind == "faza" and out["faza"] is None:
                    out["faza"] = _extract_faza(ws, target_date, own_wcs)
                elif kind == "zaga" and out["zaga"] is None:
                    out["zaga"] = _extract_zaga(ws, catalog_skus)
    finally:
        wb.close()
    return out
