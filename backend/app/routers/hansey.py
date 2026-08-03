"""
«Hansey» production-problem register API.

A Hansey row is one production problem that cost a cell time: what happened,
which department caused it, when it started, when it was closed, and the
reflection on it (comment / answers / countermeasure). Duration is the whole
point of the page, so it is computed by the server on every write and never
accepted from the request.

Every row hangs off a CELL, and the cell is what drives access — the cells
registry already knows which supervisor unit a cell belongs to and which leader
(if any) owns it:

- leader        — the cells they own. They log problems on those cells, manage
                  what sits on them (including rows their supervisor logged
                  there) and close them. They may delete only their OWN rows,
                  and only while still open: once a problem is closed its
                  duration is evidence, and removing it is the supervisor's call.
- supervisor    — every cell of their unit, INCLUDING cells with no leader
                  assigned (they log on those directly). Full manage over
                  everything in the unit + the unit-wide analytics.
- shift-manager — every unit on their shift, same rights as the supervisor.
- admin         — everything.
- top-manager   — reads everything, writes nothing.

There is deliberately no escalation chain here (that is what /concerns is for)
and no notifications: this is a register plus its analytics.

Access is gated by the ``hansey`` page in the access matrix; a personal
``page.view.hansey`` grant at "all" widens READING only.
"""
from datetime import date as date_cls, datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import and_, or_
from sqlalchemy.orm import Session

from app.capabilities import page_scope_is_all
from app.database import get_db
from app.models import Admin, Cell, HanseyProblem, Manager, RoleProfile
from app.permissions import require_page

router = APIRouter(prefix="/api/hansey", tags=["hansey"])

# Departments a problem can be blamed on — the SAME whitelist the Concerns page
# uses, so a department chip means the same thing on both pages and needs no
# second set of translations. The client renders each key as
# concerns.category.<key>; the backend only validates membership.
DEPARTMENTS = {
    "ars", "inventory", "warehouse", "fridge", "procurement", "logistics",
    "it", "washing", "plan", "hr", "technologist", "raw_material",
}

# Roles that may write at all. top-manager reads everything and writes nothing.
WRITE_ROLES = ("admin", "shift-manager", "supervisor", "leader")


# ── time helpers ─────────────────────────────────────────────────────────────

def _parse_dt(value, field: str) -> Optional[datetime]:
    """Naive factory wall-clock datetime from an ISO string ("2026-08-03T14:30").
    A trailing Z/offset is dropped rather than converted: the value means a time
    on the shop floor, and shifting it by the caller's timezone would silently
    move problems across days."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    try:
        txt = str(value).strip().replace("Z", "")
        # Strip an explicit offset ("+05:00") — see the docstring.
        if len(txt) > 19 and txt[19] in "+-":
            txt = txt[:19]
        return datetime.fromisoformat(txt).replace(tzinfo=None)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"Invalid {field}")


def _duration(started: Optional[datetime], closed: Optional[datetime]) -> Optional[int]:
    """Minutes between start and close — the ONLY place duration is produced.
    NULL while the problem is open, so "open" and "took 0 minutes" never blur."""
    if started is None or closed is None:
        return None
    return max(0, int(round((closed - started).total_seconds() / 60)))


# ── scope helpers ────────────────────────────────────────────────────────────

def _own_leader_profile(db: Session, payload: dict) -> Optional[RoleProfile]:
    """The viewing leader's pre-created profile. A leader role row points at the
    UNIT (role_id = managers.id) and binds to a profile by (unit, name)."""
    return db.query(RoleProfile).filter_by(
        role="leader",
        manager_id=payload.get("role_id"),
        name=payload.get("full_name"),
    ).first()


def _viewer_shift(db: Session, payload: dict) -> Optional[int]:
    """A shift-manager's shift (1|2) — the JWT carries no shift field, so it is
    resolved from their claimed profile."""
    prof = db.query(RoleProfile).filter(
        RoleProfile.id == payload.get("role_id"),
        RoleProfile.role == "shift-manager",
    ).first()
    return prof.shift if prof else None


def _shift_unit_ids(db: Session, shift: Optional[int]) -> list[int]:
    if shift is None:
        return []
    return [mid for (mid,) in db.query(Manager.id).filter(Manager.shift == shift).all()]


def _own_cell_ids(db: Session, leader_profile_id: Optional[int]) -> list[int]:
    if not leader_profile_id:
        return []
    return [cid for (cid,) in db.query(Cell.id).filter(Cell.leader_id == leader_profile_id).all()]


def _owner_filter(ctx: dict):
    """Rows the caller created, matched on the identity _creator_identity stamps.

    Keyed on ctx["owner_id"], NOT on the JWT's role_id — for a leader those are
    different numbers: a leader role row's role_id is their UNIT (managers.id)
    while the row stores their role_profiles.id, so comparing role_id here would
    silently match nothing (or, worse, a same-numbered unrelated profile)."""
    if ctx.get("owner_id") is None:
        return HanseyProblem.id.is_(None)   # unresolvable identity → owns nothing
    return and_(
        HanseyProblem.owner_role == ctx["role"],
        HanseyProblem.owner_profile_id == ctx["owner_id"],
    )


def _viewer_ctx(db: Session, payload: dict) -> dict:
    """Everything the rights checks need about the caller, resolved once per
    request and reused across every row of the list."""
    role = payload.get("role")
    ctx = {
        "role": role,
        "role_id": payload.get("role_id"),
        "profile_id": None,     # leader: their role_profiles.id
        # The id ownership is stamped with (see _creator_identity): managers.id
        # for a supervisor — the manager row IS their profile — and the
        # role_profiles.id for every other role.
        "owner_id": payload.get("role_id"),
        "cell_ids": set(),      # leader: the cells they currently own
        "shift_units": set(),   # shift-manager: units on their shift
    }
    if role == "leader":
        prof = _own_leader_profile(db, payload)
        ctx["profile_id"] = prof.id if prof else None
        ctx["owner_id"] = ctx["profile_id"]
        ctx["cell_ids"] = set(_own_cell_ids(db, ctx["profile_id"]))
    elif role == "shift-manager":
        ctx["shift_units"] = set(_shift_unit_ids(db, _viewer_shift(db, payload)))
    return ctx


def _scope_query(query, payload: dict, db: Session, ctx: Optional[dict] = None):
    """Restrict a HanseyProblem query to what the caller may SEE.

    A leader sees everything on the cells they own — whoever logged it — plus
    rows carrying their own profile snapshot (cells they used to own) and
    anything they created. That is what keeps a leader's numbers reconcilable
    with their supervisor's: both sides count the same rows for the same cell.

    A personal ``page.view.hansey`` grant at "all" lifts the filter entirely.
    It widens READING only — _can_edit still decides who may write."""
    role = payload.get("role")
    if role in ("admin", "top-manager") or page_scope_is_all(db, payload, "hansey"):
        return query
    ctx = ctx or _viewer_ctx(db, payload)
    if role == "shift-manager":
        return query.filter(or_(
            HanseyProblem.manager_id.in_(ctx["shift_units"] or [0]),
            _owner_filter(payload),
        ))
    if role == "supervisor":
        return query.filter(or_(
            HanseyProblem.manager_id == payload.get("role_id"),
            _owner_filter(payload),
        ))
    if role == "leader":
        conds = [_owner_filter(payload)]
        if ctx["cell_ids"]:
            conds.append(HanseyProblem.cell_id.in_(list(ctx["cell_ids"])))
        if ctx["profile_id"]:
            conds.append(HanseyProblem.leader_profile_id == ctx["profile_id"])
        return query.filter(or_(*conds))
    return query.filter(HanseyProblem.id.is_(None))  # unknown role → nothing


def _is_owner(ctx: dict, p: HanseyProblem) -> bool:
    return (
        p.owner_role == ctx["role"]
        and p.owner_profile_id is not None
        and p.owner_profile_id == ctx["role_id"]
    )


def _can_edit(ctx: dict, p: HanseyProblem) -> bool:
    """Manage rights — edit any field, close, reopen. Supervisors have full
    manage over their unit (that includes rows their leaders wrote); a leader
    manages what sits on their own cells plus anything they created."""
    role = ctx["role"]
    if role == "admin":
        return True
    if role == "top-manager":
        return False
    if role == "shift-manager":
        return p.manager_id in ctx["shift_units"] or _is_owner(ctx, p)
    if role == "supervisor":
        return p.manager_id == ctx["role_id"] or _is_owner(ctx, p)
    if role == "leader":
        return (
            p.cell_id in ctx["cell_ids"]
            or (ctx["profile_id"] is not None and p.leader_profile_id == ctx["profile_id"])
            or _is_owner(ctx, p)
        )
    return False


def _can_delete(ctx: dict, p: HanseyProblem) -> bool:
    """A leader may remove only their OWN row and only while it is still open —
    a mistyped entry is fixable, a closed problem is evidence and removing it is
    the supervisor's call. Everyone above deletes whatever they can manage."""
    if ctx["role"] == "leader":
        return _is_owner(ctx, p) and p.closed_at is None
    return _can_edit(ctx, p)


def _assert_can_edit(ctx: dict, p: HanseyProblem):
    if _can_edit(ctx, p):
        return
    role = ctx["role"]
    if role == "top-manager":
        raise HTTPException(status_code=403, detail="Read-only access")
    if role == "supervisor":
        raise HTTPException(status_code=403, detail="This problem is outside your unit")
    if role == "shift-manager":
        raise HTTPException(status_code=403, detail="This problem is outside your shift")
    raise HTTPException(status_code=403, detail="You can only manage problems on your own cells")


# ── display helpers ──────────────────────────────────────────────────────────

def _cell_map(db: Session, rows) -> dict:
    """cell_id → live cell info (code, per-language workshop names, CURRENT
    leader, unit). Resolved live so a renamed cell or a re-assigned leader shows
    up everywhere at once, while the row's own snapshots keep history stable."""
    ids = {r.cell_id for r in rows if r.cell_id}
    if not ids:
        return {}
    out = {}
    q = (
        db.query(Cell, RoleProfile.name, Manager.name, Manager.shift)
        .outerjoin(RoleProfile, RoleProfile.id == Cell.leader_id)
        .outerjoin(Manager, Manager.id == Cell.manager_id)
        .filter(Cell.id.in_(ids))
    )
    for cell, leader_name, mgr_name, shift in q.all():
        out[cell.id] = {
            "verifix_code": cell.verifix_code,
            "sap_code": cell.sap_code,
            "name_workshop_uz": cell.name_workshop_uz,
            "name_workshop_uz_cyrl": cell.name_workshop_uz_cyrl,
            "name_workshop_ru": cell.name_workshop_ru,
            "name_workshop_en": cell.name_workshop_en,
            "leader_id": cell.leader_id,
            "leader": leader_name,
            "manager_id": cell.manager_id,
            "supervisor": mgr_name,
            "shift": shift,
        }
    return out


def _owner_names(db: Session, rows) -> dict:
    """(owner_role, owner_profile_id) → CURRENT profile name, batch-resolved so
    a canonical rename is reflected on every historical row. Supervisor profiles
    live in managers, every other role in role_profiles."""
    keys = {(r.owner_role, r.owner_profile_id) for r in rows if r.owner_role and r.owner_profile_id}
    if not keys:
        return {}
    out: dict = {}
    sup_ids = [pid for role, pid in keys if role == "supervisor"]
    prof_ids = [pid for role, pid in keys if role != "supervisor"]
    if sup_ids:
        for mid, name in db.query(Manager.id, Manager.name).filter(Manager.id.in_(sup_ids)).all():
            out[("supervisor", mid)] = name
    if prof_ids:
        for p in db.query(RoleProfile).filter(RoleProfile.id.in_(prof_ids)).all():
            out[(p.role, p.id)] = p.name
    return out


def _unit_names(db: Session, rows) -> dict:
    ids = {r.manager_id for r in rows if r.manager_id}
    if not ids:
        return {}
    return {
        mid: (name, shift)
        for mid, name, shift in db.query(Manager.id, Manager.name, Manager.shift)
        .filter(Manager.id.in_(ids)).all()
    }


def _iso(dt: Optional[datetime]) -> Optional[str]:
    return dt.isoformat(timespec="minutes") if dt else None


def _serialize(p: HanseyProblem, ctx=None, cells=None, owners=None, units=None) -> dict:
    cell = (cells or {}).get(p.cell_id) or {}
    unit = (units or {}).get(p.manager_id) or (None, None)
    out = {
        "id": p.id,
        "cell_id": p.cell_id,
        "cell_code": cell.get("verifix_code") or p.cell_code,
        "cell_name_uz": cell.get("name_workshop_uz"),
        "cell_name_uz_cyrl": cell.get("name_workshop_uz_cyrl"),
        "cell_name_ru": cell.get("name_workshop_ru"),
        "cell_name_en": cell.get("name_workshop_en"),
        # The cell's CURRENT leader; the row's own snapshot is the fallback so a
        # released cell doesn't blank out its history.
        "leader_id": cell.get("leader_id") or p.leader_profile_id,
        "leader_name": cell.get("leader"),
        "manager_id": p.manager_id,
        "supervisor_name": unit[0] or cell.get("supervisor"),
        "shift": unit[1] if unit[1] is not None else cell.get("shift"),
        "department": p.department,
        "problem": p.problem,
        "comment": p.comment,
        "answers": p.answers,
        "countermeasure": p.countermeasure,
        "started_at": _iso(p.started_at),
        "closed_at": _iso(p.closed_at),
        "duration_minutes": p.duration_minutes,
        "date": p.date.isoformat() if p.date else None,
        "status": "closed" if p.closed_at else "open",
        "owner_role": p.owner_role,
        "owner_name": (owners or {}).get((p.owner_role, p.owner_profile_id)) or p.owner_name,
        "created_at": p.created_at.isoformat() if p.created_at else None,
    }
    if ctx is not None:
        out["can_edit"] = _can_edit(ctx, p)
        out["can_delete"] = _can_delete(ctx, p)
    return out


# ── payloads ─────────────────────────────────────────────────────────────────

class ProblemIn(BaseModel):
    cell_id: int
    department: str
    problem: str
    comment: str
    answers: str
    countermeasure: str
    started_at: str
    closed_at: Optional[str] = None


def _validate(body: ProblemIn) -> dict:
    """Shared field validation. All four texts are required on every row — the
    register is only worth analysing if the reflection is filled in."""
    if body.department not in DEPARTMENTS:
        raise HTTPException(status_code=400, detail="A department is required")
    fields = {}
    for name in ("problem", "comment", "answers", "countermeasure"):
        value = (getattr(body, name) or "").strip()
        if not value:
            raise HTTPException(status_code=400, detail=f"{name} is required")
        if len(value) > 5000:
            raise HTTPException(status_code=400, detail=f"{name} is too long")
        fields[name] = value

    started = _parse_dt(body.started_at, "started_at")
    if started is None:
        raise HTTPException(status_code=400, detail="A start time is required")
    closed = _parse_dt(body.closed_at, "closed_at")
    if closed is not None and closed < started:
        raise HTTPException(status_code=400, detail="The close time cannot be before the start time")

    fields["started_at"] = started
    fields["closed_at"] = closed
    fields["duration_minutes"] = _duration(started, closed)
    fields["date"] = started.date()
    fields["department"] = body.department
    return fields


def _assert_can_write_cell(db: Session, payload: dict, ctx: dict, cell_id: int) -> Cell:
    """The cell a NEW row may be logged on, per role. A supervisor may log on any
    cell of their unit — including one with no leader assigned, which is exactly
    why leader_profile_id is nullable."""
    role = payload.get("role")
    if role not in WRITE_ROLES:
        raise HTTPException(status_code=403, detail="Read-only access")
    cell = db.query(Cell).filter(Cell.id == cell_id).first()
    if not cell:
        raise HTTPException(status_code=404, detail="Cell not found")
    if role == "admin":
        return cell
    if role == "shift-manager":
        if cell.manager_id not in ctx["shift_units"]:
            raise HTTPException(status_code=403, detail="This cell is outside your shift")
    elif role == "supervisor":
        if cell.manager_id != payload.get("role_id"):
            raise HTTPException(status_code=403, detail="This cell is outside your unit")
    elif role == "leader":
        if cell.id not in ctx["cell_ids"]:
            raise HTTPException(status_code=403, detail="You can only log problems on your own cells")
    return cell


def _creator_identity(db: Session, payload: dict):
    """(owner_role, owner_profile_id, name snapshot) for the authenticated
    creator — profile ids follow the bell's convention: managers.id for a
    supervisor, role_profiles.id for everyone else."""
    role = payload.get("role")
    if role == "supervisor":
        mgr = db.query(Manager).filter_by(id=payload.get("role_id")).first()
        return "supervisor", payload.get("role_id"), (mgr.name if mgr else payload.get("full_name"))
    prof = None
    if role == "leader":
        prof = _own_leader_profile(db, payload)
    elif role == "shift-manager":
        prof = db.query(RoleProfile).filter_by(id=payload.get("role_id"), role="shift-manager").first()
    elif role == "admin":
        a = db.query(Admin).filter_by(telegram_id=int(payload["sub"])).first()
        if a and a.profile_id:
            prof = db.query(RoleProfile).filter_by(id=a.profile_id, role="admin").first()
    return role, (prof.id if prof else None), (prof.name if prof else payload.get("full_name"))


# ── endpoints ────────────────────────────────────────────────────────────────

@router.get("")
def list_problems(
    date_from: Optional[str] = Query(default=None),
    date_to: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("hansey")),
):
    """Every problem in the caller's scope for the period. Filtering beyond the
    period happens on the client — the whole page (register AND both analytics
    boards) is computed from this one payload, so every filter reshapes the
    charts live without a refetch."""
    ctx = _viewer_ctx(db, payload)
    query = _scope_query(db.query(HanseyProblem), payload, db, ctx)

    if date_from:
        try:
            query = query.filter(HanseyProblem.date >= date_cls.fromisoformat(date_from))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from")
    if date_to:
        try:
            query = query.filter(HanseyProblem.date <= date_cls.fromisoformat(date_to))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to")

    rows = query.order_by(HanseyProblem.date.desc(), HanseyProblem.started_at.desc()).all()
    cells = _cell_map(db, rows)
    owners = _owner_names(db, rows)
    units = _unit_names(db, rows)
    role = payload.get("role")
    return {
        "role": role,
        # The leader board is a personal log; every other role gets the
        # unit-wide board with the by-leader / by-cell comparisons.
        "analytics": "leader" if role == "leader" else "unit",
        "can_create": role in WRITE_ROLES,
        "read_only": role == "top-manager",
        "data": [_serialize(p, ctx, cells, owners, units) for p in rows],
    }


@router.get("/cells")
def list_cells(
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("hansey")),
):
    """Cells the caller may log a problem on, for the form's cell picker and the
    page's cell/leader filters. Leaderless cells are included for supervisors
    and above — those are theirs to log on directly."""
    role = payload.get("role")
    ctx = _viewer_ctx(db, payload)
    q = (
        db.query(Cell, RoleProfile.name, Manager.id, Manager.name, Manager.shift)
        .outerjoin(RoleProfile, and_(RoleProfile.id == Cell.leader_id, RoleProfile.role == "leader"))
        .outerjoin(Manager, Manager.id == Cell.manager_id)
    )
    if role == "shift-manager":
        q = q.filter(Cell.manager_id.in_(ctx["shift_units"] or [0]))
    elif role == "supervisor":
        q = q.filter(Cell.manager_id == payload.get("role_id"))
    elif role == "leader":
        q = q.filter(Cell.id.in_(list(ctx["cell_ids"]) or [0]))
    elif role == "top-manager":
        return []

    rows = q.all()
    return [
        {
            "cell_id": cell.id,
            "verifix_code": cell.verifix_code,
            "sap_code": cell.sap_code,
            "name_workshop_uz": cell.name_workshop_uz,
            "name_workshop_uz_cyrl": cell.name_workshop_uz_cyrl,
            "name_workshop_ru": cell.name_workshop_ru,
            "name_workshop_en": cell.name_workshop_en,
            "leader_id": cell.leader_id,
            "leader": leader_name,
            "manager_id": mid,
            "supervisor": mname,
            "shift": shift,
        }
        for cell, leader_name, mid, mname, shift in sorted(
            rows, key=lambda r: (r[0].verifix_code or "").lower()
        )
    ]


@router.post("")
def create_problem(
    body: ProblemIn,
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("hansey")),
):
    ctx = _viewer_ctx(db, payload)
    cell = _assert_can_write_cell(db, payload, ctx, body.cell_id)
    fields = _validate(body)
    owner_role, owner_profile_id, owner_name = _creator_identity(db, payload)

    p = HanseyProblem(
        cell_id=cell.id,
        cell_code=cell.verifix_code,
        manager_id=cell.manager_id,
        leader_profile_id=cell.leader_id,
        owner_role=owner_role,
        owner_profile_id=owner_profile_id,
        owner_name=(owner_name or "").strip() or None,
        created_by=int(payload["sub"]),
        problem=fields.pop("problem"),
        comment=fields.pop("comment"),
        answers=fields.pop("answers"),
        countermeasure=fields.pop("countermeasure"),
        **fields,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return _serialize(p, ctx, _cell_map(db, [p]), _owner_names(db, [p]), _unit_names(db, [p]))


@router.put("/{problem_id}")
def update_problem(
    problem_id: int,
    body: ProblemIn,
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("hansey")),
):
    p = db.query(HanseyProblem).filter(HanseyProblem.id == problem_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Problem not found")
    ctx = _viewer_ctx(db, payload)
    _assert_can_edit(ctx, p)
    fields = _validate(body)

    # Moving a row to another cell re-homes it: the unit snapshot has to follow,
    # or the row would stay in the old supervisor's scope forever.
    if body.cell_id and body.cell_id != p.cell_id:
        cell = _assert_can_write_cell(db, payload, ctx, body.cell_id)
        p.cell_id = cell.id
        p.cell_code = cell.verifix_code
        p.manager_id = cell.manager_id
        p.leader_profile_id = cell.leader_id

    p.problem = fields.pop("problem")
    p.comment = fields.pop("comment")
    p.answers = fields.pop("answers")
    p.countermeasure = fields.pop("countermeasure")
    for key, value in fields.items():
        setattr(p, key, value)

    db.commit()
    db.refresh(p)
    return _serialize(p, ctx, _cell_map(db, [p]), _owner_names(db, [p]), _unit_names(db, [p]))


@router.delete("/{problem_id}", status_code=204)
def delete_problem(
    problem_id: int,
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("hansey")),
):
    p = db.query(HanseyProblem).filter(HanseyProblem.id == problem_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Problem not found")
    ctx = _viewer_ctx(db, payload)
    if not _can_delete(ctx, p):
        if ctx["role"] == "leader" and p.closed_at is not None:
            raise HTTPException(
                status_code=403,
                detail="A closed problem can only be deleted by the supervisor",
            )
        _assert_can_edit(ctx, p)
        raise HTTPException(status_code=403, detail="You cannot delete this problem")
    db.delete(p)
    db.commit()


# ── Excel export ─────────────────────────────────────────────────────────────

# Column header per language, in export order. Departments reuse the Concerns
# vocabulary; anything the sheet needs beyond that is spelled out here so the
# export never depends on the frontend's i18n bundle.
_XL_HEAD = {
    "uz": ["Sana", "Yacheyka", "Sex", "Lider", "Brigadir", "Smena", "Bo'lim", "Muammo",
           "Izoh", "Javoblar", "Qarshi chora", "Boshlandi", "Yopildi", "Davomiylik (daq)",
           "Davomiylik (soat)", "Holat", "Kim kiritdi"],
    "uz_cyrl": ["Сана", "Ячейка", "Сех", "Лидер", "Бригадир", "Смена", "Бўлим", "Муаммо",
                "Изоҳ", "Жавоблар", "Қарши чора", "Бошланди", "Ёпилди", "Давомийлиги (дақ)",
                "Давомийлиги (соат)", "Ҳолат", "Ким киритди"],
    "ru": ["Дата", "Ячейка", "Цех", "Лидер", "Бригадир", "Смена", "Отдел", "Проблема",
           "Комментарий", "Ответы", "Контрмера", "Начало", "Закрытие", "Длительность (мин)",
           "Длительность (ч)", "Статус", "Автор"],
    "en": ["Date", "Cell", "Workshop", "Leader", "Supervisor", "Shift", "Department", "Problem",
           "Comment", "Answers", "Countermeasure", "Started", "Closed", "Duration (min)",
           "Duration (h)", "Status", "Logged by"],
}

_XL_DEPT = {
    "uz": {"ars": "ARS", "inventory": "Inventar", "warehouse": "Ombor", "fridge": "Muzlatkich",
           "procurement": "Xarid", "logistics": "Logistika", "it": "IT", "washing": "Yuvish",
           "plan": "Reja", "hr": "HR", "technologist": "Texnolog", "raw_material": "Keles (xomashyo)"},
    "uz_cyrl": {"ars": "АРС", "inventory": "Инвентарь", "warehouse": "Омбор", "fridge": "Музлаткич",
                "procurement": "Харид", "logistics": "Логистика", "it": "IT", "washing": "Ювиш",
                "plan": "Режа", "hr": "HR", "technologist": "Технолог", "raw_material": "Келес (хомашё)"},
    "ru": {"ars": "АРС", "inventory": "Инвентарь", "warehouse": "Склад", "fridge": "Холодильник",
           "procurement": "Снабжение", "logistics": "Логистика", "it": "IT", "washing": "Мойка",
           "plan": "Планирование", "hr": "HR", "technologist": "Технологи", "raw_material": "Сырьё (Келес)"},
    "en": {"ars": "ARS", "inventory": "Inventory", "warehouse": "Warehouse", "fridge": "Fridge",
           "procurement": "Procurement", "logistics": "Logistics", "it": "IT", "washing": "Washing",
           "plan": "Planning", "hr": "HR", "technologist": "Technologist", "raw_material": "Raw material"},
}

_XL_STATUS = {
    "uz": ("Ochiq", "Yopilgan"), "uz_cyrl": ("Очиқ", "Ёпилган"),
    "ru": ("Открыта", "Закрыта"), "en": ("Open", "Closed"),
}


@router.get("/export.xlsx")
def export_problems(
    date_from: Optional[str] = Query(default=None),
    date_to: Optional[str] = Query(default=None),
    lang: str = Query(default="ru"),
    send: int = Query(default=0),
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("hansey")),
):
    """The register as a workbook, delivered to the caller's Telegram DM (the
    platform convention for exports). Always scope-filtered by the same
    _scope_query the page uses, so the file can never hold a row the caller
    cannot already read on screen."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    lang = lang if lang in _XL_HEAD else "ru"
    ctx = _viewer_ctx(db, payload)
    query = _scope_query(db.query(HanseyProblem), payload, db, ctx)
    if date_from:
        try:
            query = query.filter(HanseyProblem.date >= date_cls.fromisoformat(date_from))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from")
    if date_to:
        try:
            query = query.filter(HanseyProblem.date <= date_cls.fromisoformat(date_to))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to")
    rows = query.order_by(HanseyProblem.date.desc(), HanseyProblem.started_at.desc()).all()

    cells = _cell_map(db, rows)
    owners = _owner_names(db, rows)
    units = _unit_names(db, rows)
    depts = _XL_DEPT[lang]
    open_lbl, closed_lbl = _XL_STATUS[lang]
    name_key = f"name_workshop_{lang}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Hansey"

    head_fill = PatternFill("solid", fgColor="C8973F")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    headers = _XL_HEAD[lang]
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font, c.border = head_fill, head_font, border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for p in rows:
        cell = cells.get(p.cell_id) or {}
        unit = units.get(p.manager_id) or (None, None)
        ws.append([
            p.date.isoformat() if p.date else "",
            cell.get("verifix_code") or p.cell_code or "",
            cell.get(name_key) or cell.get("name_workshop_ru") or "",
            cell.get("leader") or "",
            unit[0] or cell.get("supervisor") or "",
            unit[1] if unit[1] is not None else (cell.get("shift") or ""),
            depts.get(p.department, p.department),
            p.problem or "",
            p.comment or "",
            p.answers or "",
            p.countermeasure or "",
            p.started_at.strftime("%d.%m.%Y %H:%M") if p.started_at else "",
            p.closed_at.strftime("%d.%m.%Y %H:%M") if p.closed_at else "",
            p.duration_minutes if p.duration_minutes is not None else "",
            round(p.duration_minutes / 60, 2) if p.duration_minutes is not None else "",
            closed_lbl if p.closed_at else open_lbl,
            owners.get((p.owner_role, p.owner_profile_id)) or p.owner_name or "",
        ])

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border, c.alignment = border, wrap

    for col, width in zip(
        "ABCDEFGHIJKLMNOPQ",
        [11, 10, 20, 20, 20, 7, 16, 42, 32, 32, 32, 17, 17, 15, 15, 12, 20],
    ):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Q{max(1, ws.max_row)}"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    period = f"{date_from or '…'}_{date_to or '…'}"
    fname = f"hansey_{period}.xlsx"
    if send:
        from app.telegram_bot import bot
        caption = f"📋 Hansey — {date_from or '…'} → {date_to or '…'}  •  {len(rows)}"
        try:
            bot.send_document(chat_id=int(payload["sub"]), document=(fname, bio.read()), caption=caption)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Telegram send failed: {e}")
        return {"ok": True, "rows": len(rows)}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
