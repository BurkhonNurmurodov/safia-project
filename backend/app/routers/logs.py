"""«Jurnal» — reading the action register.

The write side is `app/services/action_log.py`. This is the read side: the admin
tab's list, its category counts, its filter facets and its Excel export.

Three rules this router is built on.

**It ships KEYS, never sentences.** ``category`` and ``action`` come back exactly
as they were stored, and the tab renders them through the frontend's own
4-language table (``logs.cat.*`` / ``logs.act.*``). A Russian string baked in
here would be a second translation surface to keep in step with the first, and
the one thing this platform has learned twice is that two lists drift.

**It counts on the server.** The register grows forever, so the tab must never
be a "fetch the window and filter in the browser" table (what `DocAudit.jsx`
does over a few hundred rows and could not do over a hundred thousand).
Pagination, the category rail's counts and every facet are SQL.

**Its counts describe what the reader is looking at.** The rail's per-category
count is computed with every filter applied EXCEPT the category — otherwise
selecting «Attendance» would show «Attendance 88» beside «Identity 0» and read
as "nothing happened in Identity today", which is false. Same rule for the
action facet.

Admin-only, deliberately, and not grantable: this register says who reopened
whose day, who deleted a profile and who revealed a browser password. It is also
strictly read-only — there is no delete or purge endpoint anywhere, because an
audit trail with an erase button is not one.
"""
import logging
from datetime import date as date_t, datetime, timedelta, timezone
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import String, and_, cast, func, or_, select
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import ActionLog
from app.routers.admin import verify_admin
from app.services import action_log
from app.xlsx_delivery import deliver_xlsx

router = APIRouter(prefix="/api/admin/logs", tags=["logs"])

log = logging.getLogger(__name__)

TASHKENT = timezone(timedelta(hours=5))

MAX_PAGE_SIZE = 200
DEFAULT_PAGE_SIZE = 50


# ── filters ───────────────────────────────────────────────────────────────────

def _parse_date(s: Optional[str]) -> Optional[date_t]:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid date: {s}")


def _csv(v: Optional[str]) -> list[str]:
    return [p for p in (v or "").split(",") if p.strip()]


def _window(date_from: Optional[str], date_to: Optional[str]) -> tuple[datetime, datetime]:
    """The [from, to] day range as Tashkent-local instants.

    The register is read in wall-clock days — "what happened today" — so the
    bounds are built in Tashkent and compared against a timestamptz column,
    which Postgres handles as the same instant. Using UTC midnight instead would
    silently shift every boundary five hours and put the small hours of a
    shift-2 night on the wrong day.
    """
    d_to = _parse_date(date_to) or datetime.now(TASHKENT).date()
    d_from = _parse_date(date_from) or d_to
    if d_from > d_to:
        raise HTTPException(status_code=400, detail="from is after to")
    start = datetime.combine(d_from, datetime.min.time(), tzinfo=TASHKENT)
    end = datetime.combine(d_to + timedelta(days=1), datetime.min.time(), tzinfo=TASHKENT)
    return start, end


class _F:
    """One parsed filter set. `apply` takes an `except_` so a facet can leave
    out the dimension it is itself counting."""

    def __init__(self, date_from, date_to, category, action, outcome, source,
                 actor, unit_id, q, enriched):
        self.start, self.end = _window(date_from, date_to)
        self.category = category or None
        self.action = action or None
        self.outcome = _csv(outcome)
        self.source = _csv(source)
        self.actor = actor or None
        self.unit_id = unit_id
        self.q = (q or "").strip()
        self.enriched = enriched

    def apply(self, query, except_: str = ""):
        query = query.filter(ActionLog.created_at >= self.start,
                             ActionLog.created_at < self.end)
        if self.category and except_ != "category":
            query = query.filter(ActionLog.category == self.category)
        if self.action and except_ != "action":
            query = query.filter(ActionLog.action == self.action)
        if self.outcome and except_ != "outcome":
            query = query.filter(ActionLog.outcome.in_(self.outcome))
        if self.source and except_ != "source":
            query = query.filter(ActionLog.source.in_(self.source))
        if self.actor and except_ != "actor":
            # One control, one person. A person is their PROFILE — they may hold
            # two Telegram accounts, and "everything Aripova did" must return
            # both — but some of their rows carry no profile key at all (a
            # legacy unbound admin, a stale leader JWT, a resolution that
            # failed), and dropping those would make the register look like it
            # missed work it actually recorded.
            #
            # So a profile pick also claims the un-keyed rows of the ACCOUNTS
            # that profile was seen acting from. Bare `telegram_id == actor`
            # would over-claim in the other direction: it would hand back rows
            # belonging to a DIFFERENT profile the same account also holds.
            if ":" in self.actor:
                tids = select(ActionLog.actor_telegram_id).where(
                    ActionLog.actor_profile_key == self.actor,
                    ActionLog.actor_telegram_id.isnot(None)).distinct()
                query = query.filter(or_(
                    ActionLog.actor_profile_key == self.actor,
                    and_(ActionLog.actor_profile_key.is_(None),
                         ActionLog.actor_telegram_id.in_(tids))))
            else:
                # A numeric key is only ever offered when NO profile was ever
                # resolved for that account, so the account IS the identity.
                query = query.filter(
                    and_(ActionLog.actor_profile_key.is_(None),
                         cast(ActionLog.actor_telegram_id, String) == self.actor))
        if self.unit_id and except_ != "unit":
            query = query.filter(ActionLog.unit_id == self.unit_id)
        if self.enriched == "auto" and except_ != "enriched":
            query = query.filter(ActionLog.enriched.is_(False))
        elif self.enriched == "rich" and except_ != "enriched":
            query = query.filter(ActionLog.enriched.is_(True))
        if self.q and except_ != "q":
            like = f"%{self.q.lower()}%"
            query = query.filter(or_(
                func.lower(ActionLog.actor_name).like(like),
                func.lower(ActionLog.target_name).like(like),
                func.lower(ActionLog.unit_name).like(like),
                func.lower(ActionLog.path).like(like),
                func.lower(ActionLog.action).like(like),
                func.lower(ActionLog.reason).like(like),
                cast(ActionLog.target_id, String).like(like),
            ))
        return query


def _filters(
    date_from: Optional[str] = Query(None, alias="from"),
    date_to: Optional[str] = Query(None, alias="to"),
    category: Optional[str] = Query(None),
    action: Optional[str] = Query(None),
    outcome: Optional[str] = Query(None, description="csv"),
    source: Optional[str] = Query(None, description="csv"),
    actor: Optional[str] = Query(None, description="profile key or telegram id"),
    unit_id: Optional[int] = Query(None),
    q: Optional[str] = Query(None),
    enriched: Optional[str] = Query(None, description="auto | rich"),
) -> _F:
    return _F(date_from, date_to, category, action, outcome, source, actor,
              unit_id, q, enriched)


# ── rendering ─────────────────────────────────────────────────────────────────

def _row(r: ActionLog) -> dict:
    at = r.created_at.astimezone(TASHKENT) if r.created_at else None
    return {
        "id": r.id,
        "at": at.isoformat() if at else None,
        "time": at.strftime("%H:%M:%S") if at else None,
        "date": at.strftime("%Y-%m-%d") if at else None,
        "category": r.category,
        "action": r.action,
        "outcome": r.outcome,
        "source": r.source,
        "actor": r.actor_name,
        "actor_role": r.actor_role,
        "actor_key": r.actor_profile_key or (str(r.actor_telegram_id) if r.actor_telegram_id else None),
        "telegram_id": r.actor_telegram_id,
        "via_capability": r.via_capability,
        "ghost": bool(r.ghost),
        "target_kind": r.target_kind,
        "target_id": r.target_id,
        "target": r.target_name,
        "unit_id": r.unit_id,
        "unit": r.unit_name,
        "day": r.day.isoformat() if r.day else None,
        "details": [{"k": k, "v": v} for k, v in (r.details or [])],
        "changes": [{"f": f, "old": o, "new": n} for f, o, n in (r.changes or [])],
        "reason": r.reason,
        "enriched": bool(r.enriched),
        "method": r.method,
        "path": r.path,
        "status": r.status,
        "ms": r.duration_ms,
        "ip": r.ip,
        "version": r.app_version,
    }


# ── the register ──────────────────────────────────────────────────────────────

@router.get("")
def list_logs(
    f: _F = Depends(_filters),
    page: int = Query(1, ge=1),
    page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=MAX_PAGE_SIZE),
    db: Session = Depends(get_db),
    _: dict = Depends(verify_admin),
):
    """One page of the register, newest first."""
    base = f.apply(db.query(ActionLog))
    total = base.with_entities(func.count(ActionLog.id)).scalar() or 0
    rows = (base.order_by(ActionLog.created_at.desc(), ActionLog.id.desc())
            .offset((page - 1) * page_size).limit(page_size).all())
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "rows": [_row(r) for r in rows],
    }


@router.get("/summary")
def summary(
    f: _F = Depends(_filters),
    db: Session = Depends(get_db),
    _: dict = Depends(verify_admin),
):
    """The category rail's counts and the four stat tiles.

    Category counts leave the CATEGORY filter out (see the module docstring):
    they exist to tell the reader where the rest of the activity is, and a rail
    that zeroed every other row the moment one was selected would say the
    opposite of the truth.
    """
    cat_q = f.apply(db.query(ActionLog.category, func.count(ActionLog.id)),
                    except_="category").group_by(ActionLog.category)
    counts = {c: n for c, n in cat_q.all()}

    scoped = f.apply(db.query(ActionLog))
    total = scoped.with_entities(func.count(ActionLog.id)).scalar() or 0
    # Every filter applies, the outcome one included. The tile is rendered as a
    # SHARE of `total` ("15 of 100 — denied"), and two numbers taken from
    # different row sets cannot be read as a ratio: with Outcome = «done»
    # selected it would otherwise print 15 denied above a table holding none.
    denied = (f.apply(db.query(func.count(ActionLog.id)))
              .filter(ActionLog.outcome.in_(("denied", "refused", "error"))).scalar() or 0)
    actors = (scoped.with_entities(
        func.count(func.distinct(func.coalesce(
            ActionLog.actor_profile_key,
            cast(ActionLog.actor_telegram_id, String))))).scalar() or 0)
    last = scoped.order_by(ActionLog.created_at.desc()).first()

    # Outcome and source tallies feed the filter panel's option labels, so a
    # reader can see there ARE denied rows before opening the filter.
    out_counts = dict(f.apply(db.query(ActionLog.outcome, func.count(ActionLog.id)),
                              except_="outcome").group_by(ActionLog.outcome).all())
    src_counts = dict(f.apply(db.query(ActionLog.source, func.count(ActionLog.id)),
                              except_="source").group_by(ActionLog.source).all())

    return {
        "total": total,
        "denied": denied,
        "actors": actors,
        "auto": (f.apply(db.query(func.count(ActionLog.id)), except_="enriched")
                 .filter(ActionLog.enriched.is_(False)).scalar() or 0),
        "last": _row(last) if last else None,
        "categories": [{"key": c, "count": counts.get(c, 0)} for c in action_log.CATEGORIES],
        "outcomes": [{"key": k, "count": out_counts.get(k, 0)} for k in action_log.OUTCOMES],
        "sources": [{"key": k, "count": src_counts.get(k, 0)} for k in action_log.SOURCES],
    }


@router.get("/facets")
def facets(
    f: _F = Depends(_filters),
    db: Session = Depends(get_db),
    _: dict = Depends(verify_admin),
):
    """The people, units and action keys present in the current window.

    Every list is built from the rows the reader can actually reach, so a filter
    never offers a value that would return nothing — the same discipline the
    Quality org chain follows. The action list narrows to the selected category
    and the payload says so (`actions_narrowed`), because a shortened list must
    never be mistaken for missing data.
    """
    # ONE ROW PER PERSON. Grouping on coalesce(profile, account) alone splits
    # somebody whose profile resolved for most of their rows and not for the
    # rest into two entries carrying the same name — «Aripova · 40» beside
    # «Aripova · 3», neither of which is the answer to "what did Aripova do".
    # So the un-keyed rows of an account are folded into the profile that
    # account was seen acting as, matching exactly what `_F.apply` selects.
    #
    # No LIMIT. The list is already bounded by the window (as `units` and
    # `actions` beside it are), and the picker searches — while a silent cap
    # would let the client drop a pick that fell off the end and quietly show
    # the reader everybody's actions in place of one person's.
    raw = (f.apply(db.query(
        ActionLog.actor_profile_key.label("pk"),
        ActionLog.actor_telegram_id.label("tid"),
        func.max(ActionLog.actor_name).label("name"),
        func.max(ActionLog.actor_role).label("role"),
        func.count(ActionLog.id).label("n")), except_="actor")
        .group_by(ActionLog.actor_profile_key, ActionLog.actor_telegram_id).all())

    owner = {tid: pk for pk, tid, _n, _r, _c in raw if pk and tid}
    folded: dict[str, dict] = {}
    for pk, tid, name, role, n in raw:
        key = pk or owner.get(tid) or (str(tid) if tid else None)
        if not key:
            continue
        e = folded.setdefault(key, {"key": key, "name": name, "role": role, "count": 0})
        e["count"] += n
        # A profile-keyed row's name/role is the better label: an un-keyed row
        # may predate the binding that gave the person their current post.
        if pk:
            e["name"], e["role"] = name or e["name"], role or e["role"]
    actors = sorted(folded.values(), key=lambda a: -a["count"])

    units = (f.apply(db.query(ActionLog.unit_id, func.max(ActionLog.unit_name),
                              func.count(ActionLog.id)), except_="unit")
             .filter(ActionLog.unit_id.isnot(None))
             .group_by(ActionLog.unit_id)
             .order_by(func.count(ActionLog.id).desc()).all())

    actions = (f.apply(db.query(ActionLog.action, ActionLog.category,
                                func.count(ActionLog.id)), except_="action")
               .group_by(ActionLog.action, ActionLog.category)
               .order_by(func.count(ActionLog.id).desc()).all())

    return {
        "actors": actors,
        "units": [{"id": i, "name": n, "count": c} for i, n, c in units],
        "actions": [{"key": a, "category": cat, "count": c} for a, cat, c in actions],
        "actions_narrowed": bool(f.category),
    }


@router.get("/{log_id}")
def one(log_id: int, db: Session = Depends(get_db), _: dict = Depends(verify_admin)):
    """A single row — for a deep link out of a notification or another page."""
    r = db.query(ActionLog).filter_by(id=log_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Not found")
    return _row(r)


# ── export ────────────────────────────────────────────────────────────────────

_HEAD_FONT = Font(bold=True, size=10, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_BODY_FONT = Font(size=10)
_LEFT = Alignment(horizontal="left", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# (payload key, fallback header, width). The client sends its own translated
# headers in `labels`, so the sheet reads in the language the tab was read in.
_COLS = [
    ("date",      "Sana",      12),
    ("time",      "Vaqt",      10),
    ("category",  "Bo'lim",    18),
    ("action",    "Amal",      30),
    ("actor",     "Kim",       24),
    ("actor_role", "Rol",      14),
    ("source",    "Manba",     12),
    ("outcome",   "Natija",    12),
    ("unit",      "Brigada",   20),
    ("target",    "Obyekt",    26),
    ("day",       "Kun",       12),
    ("changes",   "O'zgarish", 46),
    ("reason",    "Sabab",     30),
    ("path",      "So'rov",    38),
    ("status",    "Status",    9),
]

EXPORT_CAP = 20_000


def _xl(v) -> str:
    """openpyxl raises on control characters, and a leading '=' makes Excel
    treat a logged path as a formula."""
    s = "" if v is None else str(v)
    s = ILLEGAL_CHARACTERS_RE.sub("", s)
    return ("'" + s) if s[:1] in ("=", "+", "-", "@") else s


class ExportBody(BaseModel):
    labels: dict[str, str] = {}
    filename: str = "actions.xlsx"
    caption: str = ""


@router.post("/export.xlsx")
def export_xlsx(
    request: Request,
    body: ExportBody,
    f: _F = Depends(_filters),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_admin),
):
    """Excel of the register AS FILTERED on screen.

    The client sends the labels it is displaying, so the workbook reads in the
    viewer's language and names the same categories the tab does. A browser
    session downloads it; inside Telegram it lands in the caller's chat
    (app/xlsx_delivery.py). Capped, and the cap is STATED in the sheet rather
    than silently truncating — a partial export that looks complete is worse
    than no export.
    """
    base = f.apply(db.query(ActionLog))
    total = base.with_entities(func.count(ActionLog.id)).scalar() or 0
    rows = (base.order_by(ActionLog.created_at.desc(), ActionLog.id.desc())
            .limit(EXPORT_CAP).all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Actions"

    for i, (key, fallback, width) in enumerate(_COLS, 1):
        c = ws.cell(row=1, column=i, value=_xl(body.labels.get(f"col.{key}") or fallback))
        c.font, c.fill, c.alignment = _HEAD_FONT, _HEAD_FILL, _CENTER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, r in enumerate(rows, 2):
        d = _row(r)
        vals = {
            **d,
            "category": body.labels.get(f"cat.{d['category']}") or d["category"],
            "action": body.labels.get(f"act.{d['action']}") or d["action"],
            "outcome": body.labels.get(f"out.{d['outcome']}") or d["outcome"],
            "source": body.labels.get(f"src.{d['source']}") or d["source"],
            "changes": " · ".join(
                f"{body.labels.get('f.' + str(c['f'])) or c['f']}: {c['old']} → {c['new']}"
                for c in d["changes"]),
        }
        for ci, (key, _fb, _w) in enumerate(_COLS, 1):
            c = ws.cell(row=ri, column=ci,
                        value=vals.get(key) if key == "status" else _xl(vals.get(key)))
            c.font, c.alignment = _BODY_FONT, _LEFT
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}{len(rows) + 1}"
    if total > len(rows):
        note = ws.cell(row=len(rows) + 3, column=1,
                       value=_xl((body.labels.get("capped") or
                                  "Only the first {n} of {total} rows are in this file")
                                 .replace("{n}", str(len(rows)))
                                 .replace("{total}", str(total))))
        note.font = Font(size=10, bold=True, color="B45309")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return deliver_xlsx(request, payload, body.filename, bio.read(), body.caption)
