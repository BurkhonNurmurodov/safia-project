"""
Staff (Xodimlar) — attendance editing & approval workflow.

Access:
  supervisor    → view workers, creates edit/delete requests
  shift-manager → sees requests for their shift, approves/rejects
  admin         → view/edit/delete directly, actions logged as processed requests
"""
import logging
import string
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from html import escape as _html_escape
from io import BytesIO
from typing import Annotated, List, Optional
from uuid import uuid4

import jwt
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.security import OAuth2PasswordBearer
from jwt import PyJWTError as JWTError
from pydantic import BaseModel
from sqlalchemy import distinct, func, or_, and_
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from app import identity
from app.capability_alerts import alert_grant_use, tv, unit_name
from app.capabilities import (
    CAP_ATTENDANCE_DELETE, CAP_ATTENDANCE_EDIT, CAP_CLEANUP, CAP_DAY_REOPEN,
    CAP_DOCUMENTS_APPROVE, CAP_REQUESTS_APPROVE,
    cap_scope, capability_pages, caller_denied_pages, has_cap, page_scope_is_all,
    profile_unit_ids, scope_is_all,
)
from app.config import settings
from app.database import get_db
from app.notify_ctx import notifications_suppressed
from app.permissions import get_page_access, role_can_access, require_page
from app.translit import transliterate
from app.models import (
    Admin, Attendance, Cell, DayApproval, EditRequest, ExchangeTask, HrDocument,
    HrDocumentHistory, Manager, Notification, RoleProfile, TelegramUser,
    TelegramUserRole,
)
from app.services import action_log
from app.services.day_state import confirmed_pairs, day_state
from app.xlsx_delivery import deliver_xlsx

router = APIRouter(prefix="/api/staff", tags=["staff"])

logger = logging.getLogger(__name__)

_oauth2 = OAuth2PasswordBearer(tokenUrl="/api/auth/webapp")

STAFF_ROLES = {"admin", "supervisor", "shift-manager"}


def _sm_shift(db: Session, role_id: Optional[int]) -> Optional[int]:
    """shift-manager role_id (a shift-manager RoleProfile id) → its shift.
    Pre-profile JWTs still carry the old fixed slot numbers 1-4; the rollout
    backfill created the slot profiles under those very ids, so the lookup
    covers them, with the historic 1/2→shift-1, 3/4→shift-2 mapping as a last
    resort for tokens issued before the migration ran."""
    if not role_id:
        return None
    p = db.query(RoleProfile).filter_by(id=role_id, role="shift-manager").first()
    if p and p.shift in (1, 2):
        return p.shift
    return 1 if role_id in (1, 2) else 2


def _sm_role_ids_for_shift(db: Session, shift: Optional[int]) -> list[int]:
    """All shift-manager profile ids working the given shift — the role_id
    values to notify when something happens on that shift."""
    return [
        p.id for p in db.query(RoleProfile)
        .filter_by(role="shift-manager", shift=shift).all()
    ]

# Roles that exist only as verifix-imported job titles and may NOT be chosen as
# the target of a Role Change document — staff can only acquire them via verifix
# uploads. This restriction applies ONLY to the role-change target picker; these
# roles still import freely and show everywhere else (staff filter, etc.).
# Bare "Кондитер" is intentionally assignable — only its "Кондитер/…"
# sub-department composites are blocked (handled by the prefix below).
VERIFIX_ONLY_TARGET_ROLES = {
    "Скульптор",
    "Фасовщик",
    "Бригадир",
    "Разработчик",
    "Оператор",
    "Оператор производственного оборудования",
    "Просеивальщик",
    "Заготовитель продуктов и сырья",
    "Специалист по отгрузке",
    "Холодильщик пищевой продукции",
}
_KONDITER_COMPOSITE_PREFIX = "Кондитер/"


def is_assignable_target_role(job_title: str) -> bool:
    """True if `job_title` may be selected as the target of a Role Change document.

    Blocked: the verifix-only base roles above, plus any "Кондитер/…" composite
    (a sub-department). Bare "Кондитер" stays assignable.
    """
    jt = (job_title or "").strip()
    if not jt:
        return False
    if jt in VERIFIX_ONLY_TARGET_ROLES:
        return False
    if jt.startswith(_KONDITER_COMPOSITE_PREFIX):
        return False
    return True


# ── Auth helpers ───────────────────────────────────────────────────────────────

def _get_caller(token: Annotated[str, Depends(_oauth2)]):
    try:
        return jwt.decode(token, settings.secret_key, algorithms=[settings.algorithm])
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")


def _require_staff(caller=Depends(_get_caller), db: Session = Depends(get_db)):
    # Staff endpoints back both the Staff and Daily pages; allow the caller if
    # their role may access either page (admin always passes) — or if a personal
    # capability grant unlocks one of them, so a granted approver can reach the
    # queue they were given without opening /staff for their whole role.
    if not role_can_access(caller.get("role"), ["staff", "daily"], get_page_access(db),
                           capability_pages(db, caller), caller_denied_pages(db, caller)):
        raise HTTPException(status_code=403, detail="Access denied")
    return caller


def _cap_covers_unit(caller, db: Session, capability: str, manager_id: int | None) -> bool:
    """True if the caller may perform an otherwise admin-only action on one
    unit — as a real admin, or through a capability grant that reaches it.

    "all" reaches every unit; "own" only the units the profile's normal row
    scoping already covers (see ``_caller_unit_ids``)."""
    if caller.get("role") == "admin":
        return True
    scope = cap_scope(db, caller, capability)
    if scope is None:
        return False
    if scope == "all":
        return True
    units = _caller_unit_ids(caller, db)
    return units is None or manager_id in units


def _require_cap_over_unit(caller, db: Session, capability: str, manager_id: int | None) -> None:
    """``_cap_covers_unit`` as a guard. Keeps the original 403 wording so a
    caller with no grant sees exactly what they saw before."""
    if not _cap_covers_unit(caller, db, capability, manager_id):
        raise HTTPException(status_code=403, detail="Admin only")


# ── Notification helpers ───────────────────────────────────────────────────────

_MONTHS = {
    "uz": ["yanvar","fevral","mart","aprel","may","iyun","iyul","avgust","sentabr","oktabr","noyabr","dekabr"],
    "uz_cyrl": ["январ","феврал","март","апрел","май","июн","июл","август","сентабр","октабр","ноябр","декабр"],
    "ru": ["января","февраля","марта","апреля","мая","июня","июля","августа","сентября","октября","ноября","декабря"],
    "en": ["January","February","March","April","May","June","July","August","September","October","November","December"],
}

_NOTIF_STRINGS: dict[str, dict[str, tuple[str, str]]] = {
    "day_closed": {
        "uz": ("Kun yopildi", "Sana: {date} | Yopdi: {closer_name}"),
        "uz_cyrl": ("Кун ёпилди", "Сана: {date} | Ёпди: {closer_name}"),
        "ru": ("День закрыт", "Дата: {date} | Закрыл(а): {closer_name}"),
        "en": ("Day closed", "Date: {date} | Closed by: {closer_name}"),
    },
    "day_reopened": {
        "uz": ("Kun qayta ochildi: {reopener_name}", "Sana: {date} — kun yana ochiq, ma'lumotlar yopilgunga qadar ko'rinmaydi"),
        "uz_cyrl": ("Кун қайта очилди: {reopener_name}", "Сана: {date} — кун яна очиқ, маълумотлар ёпилгунга қадар кўринмайди"),
        "ru": ("День переоткрыт: {reopener_name}", "Дата: {date} — день снова открыт, данные скрыты до закрытия"),
        "en": ("Day re-opened by {reopener_name}", "Date: {date} — the day is open again, data is hidden until it is closed"),
    },
    # Per-cell ojidaniya entries (/idle-cell). A leader's entry counts the
    # moment it is saved (from 2026-08-22 — no approval step); the cell's
    # brigadir is TOLD of each one so the register is reviewed before the day
    # is closed. The old approved/rejected/batch-approved templates went with
    # the queue — nothing sends them.
    "idle_request_new": {
        "uz": ("Yangi kutish kiritildi", "{cell} · {category} · {time} | Sana: {date} | Kiritdi: {leader_name}"),
        "uz_cyrl": ("Янги кутиш киритилди", "{cell} · {category} · {time} | Сана: {date} | Киритди: {leader_name}"),
        "ru": ("Новое ожидание внесено", "{cell} · {category} · {time} | Дата: {date} | Внёс(ла): {leader_name}"),
        "en": ("New idle time entered", "{cell} · {category} · {time} | Date: {date} | Entered by: {leader_name}"),
    },
    "verifix_uploaded": {
        "uz": ("Verifix ma'lumotlari yuklandi", "Sana: {date}. O'zgartirishlarni kiriting (xodim almashtirish, lavozim o'zgartirish, o'chirish) va kunni yoping."),
        "uz_cyrl": ("Verifix маълумотлари юкланди", "Сана: {date}. Ўзгартиришларни киритинг (ходим алмаштириш, лавозим ўзгартириш, ўчириш) ва кунни ёпинг."),
        "ru": ("Данные Verifix загружены", "Дата: {date}. Внесите изменения (обмен сотрудниками, смена должности, удаление) и закройте день."),
        "en": ("Verifix data uploaded", "Date: {date}. Make your changes (people exchange, role change, deletion), then close the day."),
    },
    "new_role_change": {
        "uz": ("{actor_name} lavozim o'zgarishi hujjati yubordi", "{count} xodim → {new_role} | Sana: {date}"),
        "uz_cyrl": ("{actor_name} лавозим ўзгариши ҳужжати юборди", "{count} ходим → {new_role} | Сана: {date}"),
        "ru": ("Новый документ смены должности от {actor_name}", "{count} сотр. → {new_role} | Дата: {date}"),
        "en": ("New Role Change document from {actor_name}", "{count} employee(s) → {new_role} | Date: {date}"),
    },
    # An admin's own role-change filing is already posted when the parties hear
    # about it, so this says what HAPPENED — `new_role_change` announces a draft
    # somebody still has to decide, and the two must never read alike. The
    # people-exchange twin is `worker_exchange_approved`.
    "role_change_approved": {
        "uz": ("Lavozim o'zgarishi tasdiqlandi", "{count} xodim → {new_role} | Sana: {date}"),
        "uz_cyrl": ("Лавозим ўзгариши тасдиқланди", "{count} ходим → {new_role} | Сана: {date}"),
        "ru": ("Смена должности проведена", "{count} сотр. → {new_role} | Дата: {date}"),
        "en": ("Role change posted", "{count} employee(s) → {new_role} | Date: {date}"),
    },
    # A leader's late checklist day was opened by an admin: it counts again, at
    # its own score. The day stays flagged as late on the dashboard, so the text
    # says "counted", never "on time".
    "leader_late_approved": {
        "uz": ("Kechikkan hisobot qabul qilindi", "Sana: {date} | Ochdi: {decided_by} | Bu kun endi {score}% bilan hisoblanadi (kechikkan deb belgilangan)"),
        "uz_cyrl": ("Кечиккан ҳисобот қабул қилинди", "Сана: {date} | Очди: {decided_by} | Бу кун энди {score}% билан ҳисобланади (кечиккан деб белгиланган)"),
        "ru": ("Опоздавший отчёт засчитан", "Дата: {date} | Открыл(а): {decided_by} | День теперь считается с результатом {score}% (отмечен как опоздавший)"),
        "en": ("Late report accepted", "Date: {date} | Opened by: {decided_by} | The day now counts at {score}% (still flagged as late)"),
    },
    # An admin reviewed the AI's doubt about a proof photo and agreed with it.
    # `rejected` costs the task its weight for that day; `requeried` costs
    # nothing yet and asks for a replacement — so the two texts must NOT be
    # interchangeable, or a leader re-uploads thinking they still have time when
    # the day is already scored.
    "leader_proof_rejected": {
        "uz": ("Dalil rasmi qabul qilinmadi", "Sana: {date} | Vazifa: {task} | Tekshirdi: {by} | Bu vazifa shu kun uchun bajarilmagan deb hisoblanadi."),
        "uz_cyrl": ("Далил расми қабул қилинмади", "Сана: {date} | Вазифа: {task} | Текширди: {by} | Бу вазифа шу кун учун бажарилмаган деб ҳисобланади."),
        "ru": ("Фото-подтверждение не принято", "Дата: {date} | Задача: {task} | Проверил(а): {by} | Задача засчитана как невыполненная за этот день."),
        "en": ("Proof photo rejected", "Date: {date} | Task: {task} | Reviewed by: {by} | The task now counts as not done for that day."),
    },
    # ── automatic day verification (shift 1, from leader_ai.AUTO_FROM) ────────
    # Every task of a leader's day has been judged. Two audiences, two things
    # they need: the brigadir gets the unit's number, the leader gets their own
    # receipt. A clean day and a day with rejections are SEPARATE templates on
    # purpose — one sentence carrying "0 rejected" reads as an accusation the
    # reader then has to disprove, and whole-sentence templates are the house
    # rule anyway (word order differs per language).
    "leader_day_report_clean": {
        "uz": ("{leader}: kun tasdiqlandi — {score}%", "Sana: {date} | Barcha {checked} ta dalil qabul qilindi. Batafsil hisobotni quyidagi tugmadan oching."),
        "uz_cyrl": ("{leader}: кун тасдиқланди — {score}%", "Сана: {date} | Барча {checked} та далил қабул қилинди. Батафсил ҳисоботни қуйидаги тугмадан очинг."),
        "ru": ("{leader}: день подтверждён — {score}%", "Дата: {date} | Все {checked} подтверждений приняты. Подробный отчёт — по кнопке ниже."),
        "en": ("{leader}: day verified — {score}%", "Date: {date} | All {checked} proofs accepted. Open the full report with the button below."),
    },
    "leader_day_report_flagged": {
        "uz": ("{leader}: {rejected} ta vazifa o'tmadi — {score}%", "Sana: {date} | Topshirilgan: {raw}% → tasdiqlangan: {score}% | Qabul qilinmagan vazifalar: {tasks}. Sabablarini hisobotdan ko'ring."),
        "uz_cyrl": ("{leader}: {rejected} та вазифа ўтмади — {score}%", "Сана: {date} | Топширилган: {raw}% → тасдиқланган: {score}% | Қабул қилинмаган вазифалар: {tasks}. Сабабларини ҳисоботдан кўринг."),
        "ru": ("{leader}: не прошло задач — {rejected}, итог {score}%", "Дата: {date} | Сдано: {raw}% → подтверждено: {score}% | Не принято: {tasks}. Причины — в отчёте."),
        "en": ("{leader}: {rejected} task(s) failed — {score}%", "Date: {date} | Submitted: {raw}% → verified: {score}% | Not accepted: {tasks}. The reasons are in the report."),
    },
    # The score MOVED after the first report — a re-review, an admin ruling or
    # an upheld dispute. Says what it was and what it is now: a bare new number
    # is indistinguishable from the notification arriving twice.
    "leader_day_report_corrected": {
        "uz": ("{leader}: baho yangilandi — {before}% → {score}%", "Sana: {date} | Qayta ko'rib chiqilgandan so'ng kun bahosi o'zgardi. Hozir qabul qilinmagan: {rejected} ta vazifa."),
        "uz_cyrl": ("{leader}: баҳо янгиланди — {before}% → {score}%", "Сана: {date} | Қайта кўриб чиқилгандан сўнг кун баҳоси ўзгарди. Ҳозир қабул қилинмаган: {rejected} та вазифа."),
        "ru": ("{leader}: оценка изменена — {before}% → {score}%", "Дата: {date} | После пересмотра результат дня изменился. Сейчас не принято задач: {rejected}."),
        "en": ("{leader}: score updated — {before}% → {score}%", "Date: {date} | The day's result changed after review. Currently not accepted: {rejected} task(s)."),
    },
    # The leader's own copy. They are told on EVERY verified day, clean or not
    # (user, 2026-08-14): points come off automatically now, and a deduction
    # somebody discovers at the end of the month is how trust in the whole
    # system dies. The clean receipt is also what makes the flagged message
    # legible as a verdict rather than as an accusation out of nowhere.
    "leader_day_clean": {
        "uz": ("Kun hisobotingiz tasdiqlandi — {score}%", "Sana: {date} | Barcha {checked} ta dalil rasmingiz qabul qilindi. Rahmat."),
        "uz_cyrl": ("Кун ҳисоботингиз тасдиқланди — {score}%", "Сана: {date} | Барча {checked} та далил расмингиз қабул қилинди. Раҳмат."),
        "ru": ("Ваш отчёт за день подтверждён — {score}%", "Дата: {date} | Все ваши {checked} фото-подтверждений приняты. Спасибо."),
        "en": ("Your day report is verified — {score}%", "Date: {date} | All {checked} of your proof photos were accepted. Thank you."),
    },
    "leader_day_flagged": {
        "uz": ("{rejected} ta dalilingiz qabul qilinmadi — {score}%", "Sana: {date} | Topshirilgan: {raw}% → tasdiqlangan: {score}% | Vazifalar: {tasks}. Har biri uchun sababni hisobotdan ko'ring; rozi bo'lmasangiz brigadiringizga murojaat qiling."),
        "uz_cyrl": ("{rejected} та далилингиз қабул қилинмади — {score}%", "Сана: {date} | Топширилган: {raw}% → тасдиқланган: {score}% | Вазифалар: {tasks}. Ҳар бири учун сабабни ҳисоботдан кўринг; рози бўлмасангиз бригадирингизга мурожаат қилинг."),
        "ru": ("Не принято подтверждений: {rejected} — итог {score}%", "Дата: {date} | Сдано: {raw}% → подтверждено: {score}% | Задачи: {tasks}. Причина по каждой — в отчёте; если не согласны, обратитесь к своему бригадиру."),
        "en": ("{rejected} of your proofs were not accepted — {score}%", "Date: {date} | Submitted: {raw}% → verified: {score}% | Tasks: {tasks}. The reason for each is in the report; if you disagree, talk to your supervisor."),
    },
    "leader_day_corrected": {
        "uz": ("Bahoyingiz yangilandi — {before}% → {score}%", "Sana: {date} | Qayta ko'rib chiqilgandan so'ng kun bahosi o'zgardi. Hozir qabul qilinmagan: {rejected} ta vazifa."),
        "uz_cyrl": ("Баҳоингиз янгиланди — {before}% → {score}%", "Сана: {date} | Қайта кўриб чиқилгандан сўнг кун баҳоси ўзгарди. Ҳозир қабул қилинмаган: {rejected} та вазифа."),
        "ru": ("Ваша оценка изменена — {before}% → {score}%", "Дата: {date} | После пересмотра результат дня изменился. Сейчас не принято задач: {rejected}."),
        "en": ("Your score was updated — {before}% → {score}%", "Date: {date} | The day's result changed after review. Currently not accepted: {rejected} task(s)."),
    },
    # ── a day taken OUT of the results ───────────────────────────────────────
    # Not a correction: the number does not move to something else, it stops
    # existing. Both people who were told a score for this day are told it no
    # longer counts, because a figure that quietly vanishes from an average is
    # exactly the change neither of them could otherwise explain.
    "leader_day_report_excluded": {
        "uz": ("{leader}: kun hisobdan chiqarildi", "Sana: {date} | Bu kun ({score}%) endi natijalarga kirmaydi — na ortiqcha, na kamchilik. Sabab: {reason} | Kim: {by}"),
        "uz_cyrl": ("{leader}: кун ҳисобдан чиқарилди", "Сана: {date} | Бу кун ({score}%) энди натижаларга кирмайди — на ортиқча, на камчилик. Сабаб: {reason} | Ким: {by}"),
        "ru": ("{leader}: день исключён из результатов", "Дата: {date} | Этот день ({score}%) больше не входит в результаты — ни в плюс, ни в минус. Причина: {reason} | Кто: {by}"),
        "en": ("{leader}: day excluded from results", "Date: {date} | This day ({score}%) no longer counts either way. Reason: {reason} | By: {by}"),
    },
    "leader_day_excluded": {
        "uz": ("Kuningiz hisobdan chiqarildi", "Sana: {date} | Bu kun ({score}%) endi natijalaringizga kirmaydi — na ortiqcha, na kamchilik. Sabab: {reason} | Kim: {by}"),
        "uz_cyrl": ("Кунингиз ҳисобдан чиқарилди", "Сана: {date} | Бу кун ({score}%) энди натижаларингизга кирмайди — на ортиқча, на камчилик. Сабаб: {reason} | Ким: {by}"),
        "ru": ("Ваш день исключён из результатов", "Дата: {date} | Этот день ({score}%) больше не влияет на ваши результаты — ни в плюс, ни в минус. Причина: {reason} | Кто: {by}"),
        "en": ("Your day was excluded from the results", "Date: {date} | This day ({score}%) no longer affects your results either way. Reason: {reason} | By: {by}"),
    },
    # Put back: the day counts again, at the score it always had.
    "leader_day_report_restored": {
        "uz": ("{leader}: kun yana hisobga olinadi", "Sana: {date} | Bu kun ({score}%) natijalarga qaytarildi. Kim: {by}"),
        "uz_cyrl": ("{leader}: кун яна ҳисобга олинади", "Сана: {date} | Бу кун ({score}%) натижаларга қайтарилди. Ким: {by}"),
        "ru": ("{leader}: день снова учитывается", "Дата: {date} | Этот день ({score}%) возвращён в результаты. Кто: {by}"),
        "en": ("{leader}: day counts again", "Date: {date} | This day ({score}%) is back in the results. By: {by}"),
    },
    "leader_day_restored": {
        "uz": ("Kuningiz yana hisobga olinadi", "Sana: {date} | Bu kun ({score}%) natijalaringizga qaytarildi. Kim: {by}"),
        "uz_cyrl": ("Кунингиз яна ҳисобга олинади", "Сана: {date} | Бу кун ({score}%) натижаларингизга қайтарилди. Ким: {by}"),
        "ru": ("Ваш день снова учитывается", "Дата: {date} | Этот день ({score}%) возвращён в ваши результаты. Кто: {by}"),
        "en": ("Your day counts again", "Date: {date} | This day ({score}%) is back in your results. By: {by}"),
    },
    # ── a LEADER taken out of the results from a date on ─────────────────────
    # Not one day: everything from `{date}` onwards, days that do not exist yet
    # included. Sent once, at the decision, because "one message per affected
    # day" would be a message a morning forever about one fact — and both people
    # are told, since a leader who quietly stops appearing in a ranking and a
    # brigadir whose unit average changes shape are exactly the two who cannot
    # work out why on their own.
    "leader_cutoff_report_set": {
        "uz": ("{leader}: natijalar hisoblanmaydi", "Sana: {date} dan boshlab | Bu liderning kunlari endi o'rtacha natijaga umuman kirmaydi. Sabab: {reason} | Kim: {by}"),
        "uz_cyrl": ("{leader}: натижалар ҳисобланмайди", "Сана: {date} дан бошлаб | Бу лидернинг кунлари энди ўртача натижага умуман кирмайди. Сабаб: {reason} | Ким: {by}"),
        "ru": ("{leader}: результаты больше не учитываются", "С {date} | Дни этого лидера больше не входят в средний результат. Причина: {reason} | Кто: {by}"),
        "en": ("{leader}: results no longer count", "From {date} | This leader's days no longer enter the average at all. Reason: {reason} | By: {by}"),
    },
    "leader_cutoff_set": {
        "uz": ("Natijalaringiz hisoblanmaydi", "Sana: {date} dan boshlab | Kunlaringiz endi o'rtacha natijaga umuman kirmaydi — na ortiqcha, na kamchilik. Sabab: {reason} | Kim: {by}"),
        "uz_cyrl": ("Натижаларингиз ҳисобланмайди", "Сана: {date} дан бошлаб | Кунларингиз энди ўртача натижага умуман кирмайди — на ортиқча, на камчилик. Сабаб: {reason} | Ким: {by}"),
        "ru": ("Ваши результаты больше не учитываются", "С {date} | Ваши дни больше не входят в средний результат — ни в плюс, ни в минус. Причина: {reason} | Кто: {by}"),
        "en": ("Your results no longer count", "From {date} | Your days no longer enter the average either way. Reason: {reason} | By: {by}"),
    },
    "leader_cutoff_report_lifted": {
        "uz": ("{leader}: natijalar yana hisoblanadi", "Sana: {date} dan boshlangan cheklov bekor qilindi — kunlar o'z bahosi bilan qaytarildi. Kim: {by}"),
        "uz_cyrl": ("{leader}: натижалар яна ҳисобланади", "Сана: {date} дан бошланган чеклов бекор қилинди — кунлар ўз баҳоси билан қайтарилди. Ким: {by}"),
        "ru": ("{leader}: результаты снова учитываются", "Ограничение с {date} снято — дни вернулись со своими оценками. Кто: {by}"),
        "en": ("{leader}: results count again", "The cutoff from {date} was lifted — the days are back at the scores they always had. By: {by}"),
    },
    "leader_cutoff_lifted": {
        "uz": ("Natijalaringiz yana hisoblanadi", "Sana: {date} dan boshlangan cheklov bekor qilindi — kunlaringiz o'z bahosi bilan qaytarildi. Kim: {by}"),
        "uz_cyrl": ("Натижаларингиз яна ҳисобланади", "Сана: {date} дан бошланган чеклов бекор қилинди — кунларингиз ўз баҳоси билан қайтарилди. Ким: {by}"),
        "ru": ("Ваши результаты снова учитываются", "Ограничение с {date} снято — ваши дни вернулись со своими оценками. Кто: {by}"),
        "en": ("Your results count again", "The cutoff from {date} was lifted — your days are back at the scores they always had. By: {by}"),
    },
    # An admin ruled on the brigadir's objection to an automatic rejection.
    # A proof filed after the task's own deadline. The leader is told at EVERY
    # terminal stage, rejection included: somebody who explained themselves and
    # heard nothing back learns that explaining is pointless, which is the one
    # outcome that makes the whole flow worthless.
    "late_proof_uplifted": {
        "uz": ("Kechikkan isbot adminlarga yuborildi", "Sana: {date} | Vazifa: {task} | Brigadir: {by} | Izoh: {note}"),
        "uz_cyrl": ("Кечиккан исбот админларга юборилди", "Сана: {date} | Вазифа: {task} | Бригадир: {by} | Изоҳ: {note}"),
        "ru": ("Позднее подтверждение передано администраторам", "Дата: {date} | Задача: {task} | Бригадир: {by} | Комментарий: {note}"),
        "en": ("Late proof passed to the admins", "Date: {date} | Task: {task} | Brigadir: {by} | Comment: {note}"),
    },
    "late_proof_approved": {
        "uz": ("Kechikkan isbot tasdiqlandi", "Sana: {date} | Vazifa: {task} | Hal qildi: {by} | Vazifa to'liq ballini oldi."),
        "uz_cyrl": ("Кечиккан исбот тасдиқланди", "Сана: {date} | Вазифа: {task} | Ҳал қилди: {by} | Вазифа тўлиқ баллини олди."),
        "ru": ("Позднее подтверждение принято", "Дата: {date} | Задача: {task} | Решил(а): {by} | Задача получила полный балл."),
        "en": ("Late proof approved", "Date: {date} | Task: {task} | Decided by: {by} | The task got its full weight."),
    },
    "late_proof_rejected": {
        "uz": ("Kechikkan isbot rad etildi", "Sana: {date} | Vazifa: {task} | Hal qildi: {by} | Bu vazifa uchun ball berilmaydi."),
        "uz_cyrl": ("Кечиккан исбот рад этилди", "Сана: {date} | Вазифа: {task} | Ҳал қилди: {by} | Бу вазифа учун балл берилмайди."),
        "ru": ("Позднее подтверждение отклонено", "Дата: {date} | Задача: {task} | Решил(а): {by} | Балл за эту задачу не начислен."),
        "en": ("Late proof rejected", "Date: {date} | Task: {task} | Decided by: {by} | No point is given for this task."),
    },
    # ── objections to an AI rejection: the three-stage chain ────────────────
    # A leader files their account of the shift, their brigadir refuses it or
    # makes the case for it, an admin rules. Everybody is told at every stage
    # that takes the decision out of their hands — somebody who explained
    # themselves and heard nothing back learns that explaining is pointless,
    # which is the one outcome that makes the whole chain worthless.
    "leader_dispute_filed": {
        "uz": ("{leader} AI qaroriga norozilik bildirdi", "Sana: {date} | Vazifa: {task} | Izoh: {reason} | Siz ko'rib chiqasiz: rad etasiz yoki adminlarga yuborasiz."),
        "uz_cyrl": ("{leader} AI қарорига норозилик билдирди", "Сана: {date} | Вазифа: {task} | Изоҳ: {reason} | Сиз кўриб чиқасиз: рад этасиз ёки админларга юборасиз."),
        "ru": ("{leader} возражает против решения ИИ", "Дата: {date} | Задача: {task} | Комментарий: {reason} | Решение за вами: отклонить или передать администраторам."),
        "en": ("{leader} objects to an AI ruling", "Date: {date} | Task: {task} | Note: {reason} | It is yours to read: refuse it, or pass it to the admins."),
    },
    "leader_dispute_uplifted": {
        "uz": ("Norozilik adminlarga yuborildi", "Sana: {date} | Vazifa: {task} | Brigadir: {by} | Izoh: {note}"),
        "uz_cyrl": ("Норозилик админларга юборилди", "Сана: {date} | Вазифа: {task} | Бригадир: {by} | Изоҳ: {note}"),
        "ru": ("Возражение передано администраторам", "Дата: {date} | Задача: {task} | Бригадир: {by} | Комментарий: {note}"),
        "en": ("Objection passed to the admins", "Date: {date} | Task: {task} | Brigadir: {by} | Comment: {note}"),
    },
    # Refused by the BRIGADIR — it never reached an admin, and saying so is the
    # difference between "nobody agreed with you" and "nobody read it".
    "leader_dispute_sup_rejected": {
        "uz": ("Norozilikni brigadir rad etdi", "Sana: {date} | Vazifa: {task} | Rad etdi: {by} | Izoh: {note} | AI qarori kuchida qoladi."),
        "uz_cyrl": ("Норозиликни бригадир рад этди", "Сана: {date} | Вазифа: {task} | Рад этди: {by} | Изоҳ: {note} | AI қарори кучида қолади."),
        "ru": ("Возражение отклонил бригадир", "Дата: {date} | Задача: {task} | Отклонил(а): {by} | Комментарий: {note} | Решение ИИ остаётся в силе."),
        "en": ("The brigadir refused the objection", "Date: {date} | Task: {task} | Refused by: {by} | Comment: {note} | The AI ruling stands."),
    },
    "leader_dispute_approved": {
        "uz": ("Norozilik qabul qilindi", "Sana: {date} | Vazifa: {task} | Hal qildi: {by} | Izoh: {note} | Vazifa yana bajarilgan deb hisoblanadi."),
        "uz_cyrl": ("Норозилик қабул қилинди", "Сана: {date} | Вазифа: {task} | Ҳал қилди: {by} | Изоҳ: {note} | Вазифа яна бажарилган деб ҳисобланади."),
        "ru": ("Возражение принято", "Дата: {date} | Задача: {task} | Решил(а): {by} | Комментарий: {note} | Задача снова засчитана как выполненная."),
        "en": ("Objection upheld", "Date: {date} | Task: {task} | Decided by: {by} | Comment: {note} | The task counts as done again."),
    },
    "leader_dispute_rejected": {
        "uz": ("Norozilik rad etildi", "Sana: {date} | Vazifa: {task} | Hal qildi: {by} | Izoh: {note} | Vazifa bajarilmagan bo'lib qoladi."),
        "uz_cyrl": ("Норозилик рад этилди", "Сана: {date} | Вазифа: {task} | Ҳал қилди: {by} | Изоҳ: {note} | Вазифа бажарилмаган бўлиб қолади."),
        "ru": ("Возражение отклонено", "Дата: {date} | Задача: {task} | Решил(а): {by} | Комментарий: {note} | Задача остаётся незачтённой."),
        "en": ("Objection refused", "Date: {date} | Task: {task} | Decided by: {by} | Comment: {note} | The task stays not done."),
    },
    # The ruling above taken back — the task returns to the AI's verdict.
    "leader_dispute_undone": {
        "uz": ("Norozilik bo'yicha qaror bekor qilindi", "Sana: {date} | Vazifa: {task} | Bekor qildi: {by} | Vazifa yana AI xulosasi bo'yicha hisoblanadi."),
        "uz_cyrl": ("Норозилик бўйича қарор бекор қилинди", "Сана: {date} | Вазифа: {task} | Бекор қилди: {by} | Вазифа яна AI хулосаси бўйича ҳисобланади."),
        "ru": ("Решение по возражению отменено", "Дата: {date} | Задача: {task} | Отменил(а): {by} | Задача снова считается по решению ИИ."),
        "en": ("The ruling on the objection was undone", "Date: {date} | Task: {task} | Undone by: {by} | The task counts by the AI verdict again."),
    },
    "leader_proof_requeried": {
        "uz": ("Dalil rasmini qayta yuboring", "Sana: {date} | Vazifa: {task} | So'radi: {by} | Hozircha baho o'zgargani yo'q — yangi rasm yuklang."),
        "uz_cyrl": ("Далил расмини қайта юборинг", "Сана: {date} | Вазифа: {task} | Сўради: {by} | Ҳозирча баҳо ўзгаргани йўқ — янги расм юкланг."),
        "ru": ("Нужно переснять подтверждение", "Дата: {date} | Задача: {task} | Запросил(а): {by} | Оценка пока не изменена — загрузите новое фото."),
        "en": ("Please re-file your proof", "Date: {date} | Task: {task} | Requested by: {by} | Nothing has been deducted yet — upload a new photo."),
    },
    "new_edit_request": {
        "uz": ("{supervisor_name} tahrirlash so'rovi yubordi", "Xodim: {worker_name} | Sana: {date}"),
        "uz_cyrl": ("{supervisor_name} таҳрирлаш сўрови юборди", "Ходим: {worker_name} | Сана: {date}"),
        "ru": ("Запрос на редактирование от {supervisor_name}", "Сотрудник: {worker_name} | Дата: {date}"),
        "en": ("New edit request from {supervisor_name}", "Worker: {worker_name} | Date: {date}"),
    },
    "new_delete_request": {
        "uz": ("{supervisor_name} o'chirish so'rovi yubordi", "Xodim: {worker_name} | Sana: {date}"),
        "uz_cyrl": ("{supervisor_name} ўчириш сўрови юборди", "Ходим: {worker_name} | Сана: {date}"),
        "ru": ("Запрос на удаление от {supervisor_name}", "Сотрудник: {worker_name} | Дата: {date}"),
        "en": ("New delete request from {supervisor_name}", "Worker: {worker_name} | Date: {date}"),
    },
    "bulk_delete_request": {
        "uz": ("{supervisor_name} ommaviy o'chirish so'rovi yubordi", "{count} xodim | Sana: {date}"),
        "uz_cyrl": ("{supervisor_name} оммавий ўчириш сўрови юборди", "{count} ходим | Сана: {date}"),
        "ru": ("Массовый запрос на удаление от {supervisor_name}", "{count} сотр. | Дата: {date}"),
        "en": ("Bulk delete request from {supervisor_name}", "{count} worker(s) | Date: {date}"),
    },
    "request_approved_supervisor": {
        "uz": ("So'rovingiz tasdiqlandi", "Xodim: {worker_name} | Sana: {date} | Tasdiqladi: {processor_name}"),
        "uz_cyrl": ("Сўровингиз тасдиқланди", "Ходим: {worker_name} | Сана: {date} | Тасдиқлади: {processor_name}"),
        "ru": ("Ваш запрос одобрен", "Сотрудник: {worker_name} | Дата: {date} | Одобрил(а): {processor_name}"),
        "en": ("Your request was approved", "Worker: {worker_name} | Date: {date} | By: {processor_name}"),
    },
    "request_rejected_supervisor": {
        "uz": ("So'rovingiz rad etildi", "Xodim: {worker_name} | Sana: {date} | Rad etdi: {processor_name}"),
        "uz_cyrl": ("Сўровингиз рад этилди", "Ходим: {worker_name} | Сана: {date} | Рад этди: {processor_name}"),
        "ru": ("Ваш запрос отклонён", "Сотрудник: {worker_name} | Дата: {date} | Отклонил(а): {processor_name}"),
        "en": ("Your request was rejected", "Worker: {worker_name} | Date: {date} | By: {processor_name}"),
    },
    "request_approved_others": {
        "uz": ("{processor_name} so'rovni tasdiqladi", "Brigadir: {supervisor_name} | Xodim: {worker_name} | Sana: {date}"),
        "uz_cyrl": ("{processor_name} сўровни тасдиқлади", "Бригадир: {supervisor_name} | Ходим: {worker_name} | Сана: {date}"),
        "ru": ("Запрос одобрен: {processor_name}", "Бригадир: {supervisor_name} | Сотрудник: {worker_name} | Дата: {date}"),
        "en": ("Request approved by {processor_name}", "Supervisor: {supervisor_name} | Worker: {worker_name} | Date: {date}"),
    },
    "request_rejected_others": {
        "uz": ("{processor_name} so'rovni rad etdi", "Brigadir: {supervisor_name} | Xodim: {worker_name} | Sana: {date}"),
        "uz_cyrl": ("{processor_name} сўровни рад этди", "Бригадир: {supervisor_name} | Ходим: {worker_name} | Сана: {date}"),
        "ru": ("Запрос отклонён: {processor_name}", "Бригадир: {supervisor_name} | Сотрудник: {worker_name} | Дата: {date}"),
        "en": ("Request rejected by {processor_name}", "Supervisor: {supervisor_name} | Worker: {worker_name} | Date: {date}"),
    },
    "request_undone": {
        "uz": ("So'rov bekor qilindi", "Xodim: {worker_name} | Sana: {date} | Bekor qildi: {undoer}"),
        "uz_cyrl": ("Сўров бекор қилинди", "Ходим: {worker_name} | Сана: {date} | Бекор қилди: {undoer}"),
        "ru": ("Запрос отменён", "Сотрудник: {worker_name} | Дата: {date} | Отменил(а): {undoer}"),
        "en": ("A request was undone", "Worker: {worker_name} | Date: {date} | By: {undoer}"),
    },
    "admin_record_edited": {
        "uz": ("Admin xodim yozuvini tahrirladi", "Xodim: {worker_name} | Sana: {date} | Kim: {admin_name}"),
        "uz_cyrl": ("Админ ходим ёзувини таҳрирлади", "Ходим: {worker_name} | Сана: {date} | Ким: {admin_name}"),
        "ru": ("Администратор отредактировал запись", "Сотрудник: {worker_name} | Дата: {date} | Кто: {admin_name}"),
        "en": ("Admin edited a worker record", "Worker: {worker_name} | Date: {date} | By: {admin_name}"),
    },
    "admin_record_deleted": {
        "uz": ("Admin xodim yozuvini o'chirdi", "Xodim: {worker_name} | Sana: {date} | Kim: {admin_name}"),
        "uz_cyrl": ("Админ ходим ёзувини ўчирди", "Ходим: {worker_name} | Сана: {date} | Ким: {admin_name}"),
        "ru": ("Администратор удалил запись", "Сотрудник: {worker_name} | Дата: {date} | Кто: {admin_name}"),
        "en": ("Admin deleted a worker record", "Worker: {worker_name} | Date: {date} | By: {admin_name}"),
    },
    "worker_exchange_created": {
        "uz": ("{actor_name} xodim almashinuvi yaratdi", "{count} xodim → {target} | Sana: {date}"),
        "uz_cyrl": ("{actor_name} ходим алмашинуви яратди", "{count} ходим → {target} | Сана: {date}"),
        "ru": ("Новый обмен сотрудниками от {actor_name}", "{count} сотр. → {target} | Дата: {date}"),
        "en": ("New worker exchange from {actor_name}", "{count} worker(s) → {target} | Date: {date}"),
    },
    "worker_exchange_approved": {
        "uz": ("Xodim almashinuvi tasdiqlandi", "{count} xodim → {target} | Sana: {date}"),
        "uz_cyrl": ("Ходим алмашинуви тасдиқланди", "{count} ходим → {target} | Сана: {date}"),
        "ru": ("Обмен сотрудниками одобрен", "{count} сотр. → {target} | Дата: {date}"),
        "en": ("Worker exchange approved", "{count} worker(s) → {target} | Date: {date}"),
    },
    "worker_exchange_cancelled": {
        "uz": ("Xodim almashinuvi bekor qilindi", "{count} xodim → {target} | Sana: {date}"),
        "uz_cyrl": ("Ходим алмашинуви бекор қилинди", "{count} ходим → {target} | Сана: {date}"),
        "ru": ("Обмен сотрудниками отменён", "{count} сотр. → {target} | Дата: {date}"),
        "en": ("Worker exchange cancelled", "{count} worker(s) → {target} | Date: {date}"),
    },
    "document_rejected": {
        "uz": ("{actor_name} hujjatingizni rad etdi", "{doc_label} | Sana: {date}"),
        "uz_cyrl": ("{actor_name} ҳужжатингизни рад этди", "{doc_label} | Сана: {date}"),
        "ru": ("{actor_name} отклонил(а) ваш документ", "{doc_label} | Дата: {date}"),
        "en": ("{actor_name} rejected your document", "{doc_label} | Date: {date}"),
    },
    # The concern family shares ONE body shape, and every row is written the same
    # way: «quoted content» on its own line, then a blank line, then one
    #   {emoji} {Label}: {placeholder}
    # fact per line. The old single-line «Label: v | Label: v | Label: v» run-on
    # wrapped into three lines of names on a phone before the reader reached the
    # thing that had happened.
    #
    # That row shape is a contract, not decoration: _render_body reads it to
    # build the rich Telegram DM (<blockquote> for the quoted words, <b> for the
    # label — see _NOTIF_TG_ICON), so the DM and the bell come from THIS string
    # and cannot drift the way two hand-kept tables would. A row whose value is
    # empty (no leader snapshot, no reason) drops out whole, so every label here
    # is optional.
    "concern_created": {
        "uz": ("Yangi xavotir qo'shildi", "«{concern}»\n\n👤 Lider: {leader_name}\n🙋 Xavotir egasi: {owner}\n📅 Sana: {date}"),
        "uz_cyrl": ("Янги хавотир қўшилди", "«{concern}»\n\n👤 Лидер: {leader_name}\n🙋 Хавотир эгаси: {owner}\n📅 Сана: {date}"),
        "ru": ("Добавлено новое опасение", "«{concern}»\n\n👤 Лидер: {leader_name}\n🙋 Владелец: {owner}\n📅 Дата: {date}"),
        "en": ("New concern added", "“{concern}”\n\n👤 Leader: {leader_name}\n🙋 Owner: {owner}\n📅 Date: {date}"),
    },
    "concern_assigned": {
        "uz": ("{actor_name} nomingizga xavotir qo'shdi", "«{concern}»\n\n🙋 Xavotir egasi: {owner}\n📅 Sana: {date}"),
        "uz_cyrl": ("{actor_name} номингизга хавотир қўшди", "«{concern}»\n\n🙋 Хавотир эгаси: {owner}\n📅 Сана: {date}"),
        "ru": ("{actor_name} добавил(а) опасение на ваше имя", "«{concern}»\n\n🙋 Владелец: {owner}\n📅 Дата: {date}"),
        "en": ("{actor_name} added a concern for you", "“{concern}”\n\n🙋 Owner: {owner}\n📅 Date: {date}"),
    },
    # A concern's level changed and the reader is NOT the receiving handler —
    # the brigadir of the unit it is about, and the cell's leader, stay in the
    # loop instead of watching it vanish upwards. The receiving handler gets
    # concern_escalated / concern_returned ("… to you") instead.
    "concern_moved": {
        "uz": ("Xavotir {level_label} darajasiga o'tdi", "«{concern}»\n\n📝 Sabab: {reason}\n✍️ Kim: {actor_name}\n👤 Lider: {leader_name}\n📅 Sana: {date}"),
        "uz_cyrl": ("Хавотир {level_label} даражасига ўтди", "«{concern}»\n\n📝 Сабаб: {reason}\n✍️ Ким: {actor_name}\n👤 Лидер: {leader_name}\n📅 Сана: {date}"),
        "ru": ("Опасение передано на уровень «{level_label}»", "«{concern}»\n\n📝 Причина: {reason}\n✍️ Кто: {actor_name}\n👤 Лидер: {leader_name}\n📅 Дата: {date}"),
        "en": ("A concern moved to {level_label}", "“{concern}”\n\n📝 Reason: {reason}\n✍️ By: {actor_name}\n👤 Leader: {leader_name}\n📅 Date: {date}"),
    },
    # The resolution note leads the body for the same reason the move reason does
    # on concern_moved: "X closed it" without HOW is a line nobody can act on,
    # and the note is mandatory at every door that flips a concern to done.
    # Legacy notices carry no solution param, so the row simply drops out.
    "concern_resolved": {
        "uz": ("{actor_name} xavotirni hal qildi", "«{concern}»\n\n📝 Yechim: {solution}\n👤 Lider: {leader_name}\n📅 Sana: {date}"),
        "uz_cyrl": ("{actor_name} хавотирни ҳал қилди", "«{concern}»\n\n📝 Ечим: {solution}\n👤 Лидер: {leader_name}\n📅 Сана: {date}"),
        "ru": ("{actor_name} закрыл(а) опасение", "«{concern}»\n\n📝 Решение: {solution}\n👤 Лидер: {leader_name}\n📅 Дата: {date}"),
        "en": ("{actor_name} resolved a concern", "“{concern}”\n\n📝 Solution: {solution}\n👤 Leader: {leader_name}\n📅 Date: {date}"),
    },
    "concern_reopened": {
        "uz": ("{actor_name} xavotirni qayta ochdi", "«{concern}»\n\n👤 Lider: {leader_name}\n📅 Sana: {date}"),
        "uz_cyrl": ("{actor_name} хавотирни қайта очди", "«{concern}»\n\n👤 Лидер: {leader_name}\n📅 Сана: {date}"),
        "ru": ("{actor_name} переоткрыл(а) опасение", "«{concern}»\n\n👤 Лидер: {leader_name}\n📅 Дата: {date}"),
        "en": ("{actor_name} reopened a concern", "“{concern}”\n\n👤 Leader: {leader_name}\n📅 Date: {date}"),
    },
    "concern_edited": {
        "uz": ("{actor_name} xavotirni tahrirladi", "«{concern}»\n\n👤 Lider: {leader_name}\n📅 Sana: {date}"),
        "uz_cyrl": ("{actor_name} хавотирни таҳрирлади", "«{concern}»\n\n👤 Лидер: {leader_name}\n📅 Сана: {date}"),
        "ru": ("{actor_name} изменил(а) опасение", "«{concern}»\n\n👤 Лидер: {leader_name}\n📅 Дата: {date}"),
        "en": ("{actor_name} edited a concern", "“{concern}”\n\n👤 Leader: {leader_name}\n📅 Date: {date}"),
    },
    "concern_escalated": {
        "uz": ("{actor_name} xavotirni sizga yo'naltirdi", "«{concern}»\n\n📝 Sabab: {reason}\n📍 Daraja: {level_label}\n👤 Lider: {leader_name}\n📅 Sana: {date}"),
        "uz_cyrl": ("{actor_name} хавотирни сизга йўналтирди", "«{concern}»\n\n📝 Сабаб: {reason}\n📍 Даража: {level_label}\n👤 Лидер: {leader_name}\n📅 Сана: {date}"),
        "ru": ("{actor_name} передал(а) вам опасение", "«{concern}»\n\n📝 Причина: {reason}\n📍 Уровень: {level_label}\n👤 Лидер: {leader_name}\n📅 Дата: {date}"),
        "en": ("{actor_name} escalated a concern to you", "“{concern}”\n\n📝 Reason: {reason}\n📍 Level: {level_label}\n👤 Leader: {leader_name}\n📅 Date: {date}"),
    },
    "concern_returned": {
        "uz": ("{actor_name} xavotirni sizga qaytardi", "«{concern}»\n\n📝 Sabab: {reason}\n📍 Daraja: {level_label}\n👤 Lider: {leader_name}\n📅 Sana: {date}"),
        "uz_cyrl": ("{actor_name} хавотирни сизга қайтарди", "«{concern}»\n\n📝 Сабаб: {reason}\n📍 Даража: {level_label}\n👤 Лидер: {leader_name}\n📅 Сана: {date}"),
        "ru": ("{actor_name} вернул(а) вам опасение", "«{concern}»\n\n📝 Причина: {reason}\n📍 Уровень: {level_label}\n👤 Лидер: {leader_name}\n📅 Дата: {date}"),
        "en": ("{actor_name} returned a concern to you", "“{concern}”\n\n📝 Reason: {reason}\n📍 Level: {level_label}\n👤 Leader: {leader_name}\n📅 Date: {date}"),
    },
    "concern_comment": {
        "uz": ("{author_name} xavotirga izoh qoldirdi", "«{comment}»\n\n📄 Xavotir: {concern}"),
        "uz_cyrl": ("{author_name} хавотирга изоҳ қолдирди", "«{comment}»\n\n📄 Хавотир: {concern}"),
        "ru": ("{author_name} оставил(а) комментарий к опасению", "«{comment}»\n\n📄 Опасение: {concern}"),
        "en": ("{author_name} commented on a concern", "“{comment}”\n\n📄 Concern: {concern}"),
    },
    "task_created": {
        "uz": ("Yangi vazifa: {creator_name}", "Muddat: {date}\n{task}"),
        "uz_cyrl": ("Янги вазифа: {creator_name}", "Муддат: {date}\n{task}"),
        "ru": ("Новая задача от {creator_name}", "Срок: {date}\n{task}"),
        "en": ("New task from {creator_name}", "Due: {date}\n{task}"),
    },
    "task_status_changed": {
        "uz": ("{actor_name} vazifa holatini o'zgartirdi", "Yangi holat: {status_label}\n{task}"),
        "uz_cyrl": ("{actor_name} вазифа ҳолатини ўзгартирди", "Янги ҳолат: {status_label}\n{task}"),
        "ru": ("{actor_name} изменил(а) статус задачи", "Новый статус: {status_label}\n{task}"),
        "en": ("{actor_name} changed a task status", "New status: {status_label}\n{task}"),
    },
    "task_comment": {
        "uz": ("{author_name} vazifaga izoh qoldirdi", "{comment}\nVazifa: {task}"),
        "uz_cyrl": ("{author_name} вазифага изоҳ қолдирди", "{comment}\nВазифа: {task}"),
        "ru": ("{author_name} оставил(а) комментарий к задаче", "{comment}\nЗадача: {task}"),
        "en": ("{author_name} commented on a task", "{comment}\nTask: {task}"),
    },
    # Trudoyomkost call-tomorrow modal → per-brigadir "invite N workers" notice.
    # Body carries a date row, the recommended count, the upper-band maximum, and
    # an inexactness disclaimer (by design). The bell renders this plain text;
    # the Telegram DM gets an HTML variant with a real blockquote (see
    # ``_mk_notif_tg``). Keep the two structurally in sync.
    "call_forecast": {
        "uz": ("Xodim chaqirish uchun prognozlar:",
               "👤 Brigadir: {name}\n📅 Sana: {date}\n📊 Zagruzka foizi: {eff}%\n🧑‍🍳 Chaqirish tavsiya qilinadi: {count} nafar\n⚠️ Maksimum: {max} nafar\n\n"
               "Ko'rsatilgan sonlar faqat zagruzka hisobiga kiritiladigan ishchilar soni — bu barcha ishchilarning umumiy soni emas.\n\n"
               "Ushbu prognoz faqat oldingi kunlarning tarixiy ma'lumotlari asosida shakllantirilgan. "
               "Shu sababli, amaldagi vaziyat va boshqa omillarni ham inobatga olib, kerakli miqdorda odam chaqirishingizni so'raymiz!"),
        "uz_cyrl": ("Ходим чақириш учун прогнозлар:",
                    "👤 Бригадир: {name}\n📅 Сана: {date}\n📊 Загрузка фоизи: {eff}%\n🧑‍🍳 Чақириш тавсия қилинади: {count} нафар\n⚠️ Максимум: {max} нафар\n\n"
                    "Кўрсатилган сонлар фақат загрузка ҳисобига киритиладиган ишчилар сони — бу барча ишчиларнинг умумий сони эмас.\n\n"
                    "Ушбу прогноз фақат олдинги кунларнинг тарихий маълумотлари асосида шакллантирилган. "
                    "Шу сабабли, амалдаги вазият ва бошқа омилларни ҳам инобатга олиб, керакли миқдорда одам чақиришингизни сўраймиз!"),
        "ru": ("Прогноз по вызову сотрудников:",
               "👤 Бригадир: {name}\n📅 Дата: {date}\n📊 Процент загрузки: {eff}%\n🧑‍🍳 Рекомендуется вызвать: {count} чел.\n⚠️ Максимум: {max} чел.\n\n"
               "Указанные числа — это количество работников, учитываемых в загрузке, а не общее число всех работников.\n\n"
               "Этот прогноз сформирован только на основе исторических данных за предыдущие дни. "
               "Поэтому, учитывая текущую ситуацию и другие факторы, просим вызвать необходимое количество людей!"),
        "en": ("Staff call forecast:",
               "👤 Supervisor: {name}\n📅 Date: {date}\n📊 Load percentage: {eff}%\n🧑‍🍳 Recommended to call: {count} workers\n⚠️ Maximum: {max} workers\n\n"
               "The numbers shown are the count of workers included in the Workload calculation — not the total number of all workers.\n\n"
               "This forecast is based only on historical data from previous days. "
               "Therefore, taking the actual situation and other factors into account, please call the number of people you need!"),
    },
}

# Leader-task status labels for notification text — resolved from the raw
# ``task_status`` param at view time so the label follows the *viewer's*
# language (same mechanism as doc_type → doc_label). Wording matches the
# tasks page's status pills.
_TASK_STATUS_LABELS = {
    "todo":  {"uz": "Bajarilishi kerak", "uz_cyrl": "Бажарилиши керак", "ru": "К выполнению", "en": "To do"},
    "doing": {"uz": "Jarayonda",         "uz_cyrl": "Жараёнда",         "ru": "В процессе",   "en": "Doing"},
    "done":  {"uz": "Bajarildi",         "uz_cyrl": "Бажарилди",        "ru": "Выполнено",    "en": "Done"},
}

# Concern escalation-level labels — resolved from the raw ``concern_level``
# param at view time (same mechanism as task_status). Wording matches the
# concerns page's level pills.
_CONCERN_LEVEL_LABELS = {
    "leader":        {"uz": "Lider",           "uz_cyrl": "Лидер",           "ru": "Лидер",            "en": "Leader"},
    "supervisor":    {"uz": "Brigadir",        "uz_cyrl": "Бригадир",        "ru": "Бригадир",         "en": "Supervisor"},
    "shift-manager": {"uz": "Smena menejeri",  "uz_cyrl": "Смена менежери",  "ru": "Сменный менеджер", "en": "Shift manager"},
    "top-manager":   {"uz": "Top-menejment",   "uz_cyrl": "Топ-менежмент",   "ru": "Топ-менеджмент",   "en": "Top management"},
}


def _fmt_date(d, lang: str) -> str:
    if isinstance(d, str):
        try:
            d = date.fromisoformat(d)
        except ValueError:
            return str(d)
    months = _MONTHS.get(lang, _MONTHS["en"])
    month_name = months[d.month - 1]
    if lang == "en":
        return f"{month_name} {d.day}, {d.year}"
    return f"{d.day} {month_name} {d.year}"


def _get_user_lang(db: Session, telegram_id: int) -> str:
    """The recipient's saved language, used to render their Telegram DM. Seeded
    admins have no telegram_users row, so fall back to the admins table."""
    user = db.query(TelegramUser).filter_by(telegram_id=telegram_id).first()
    if user and user.language:
        return user.language
    admin = db.query(Admin).filter_by(telegram_id=telegram_id).first()
    if admin and admin.language:
        return admin.language
    return "uz"


def _same_person(a, b) -> bool:
    """Do two name params denote the same person? Compared through the English
    transliteration, so a Cyrillic snapshot matches its Latin twin and the
    tutuq-belgi variants ("O'g'li" / "Oʻgʻli" / "O‘g‘li") fold together."""
    if not isinstance(a, str) or not isinstance(b, str) or not a.strip() or not b.strip():
        return False
    fold = lambda s: " ".join(transliterate(s, "en").split()).casefold()
    return fold(a) == fold(b)


def _html_row(line: str) -> str:
    """One body TEMPLATE line → its Telegram HTML form, markup only.

    «quoted» / “quoted” (the words a person typed) becomes a real
    <blockquote>; an ``{emoji} {Label}: {value}`` row keeps its emoji outside a
    <b>bold label</b>, matching the leader day-report DMs. Anything else is
    passed through untouched. Applied to the TEMPLATE, so the label split can
    never be thrown off by a colon inside somebody's own text."""
    body = line.strip()
    if body[:1] in ("«", "“") and body[-1:] in ("»", "”"):
        return f"<blockquote>{body[1:-1]}</blockquote>"
    icon, _, rest = line.partition(" ")
    label, sep, value = rest.partition(":")
    if not sep or "{" in label:
        return line
    return f"{icon} <b>{label}:</b>{value}"


def _render_body(tmpl: str, values: dict, *, html: bool = False) -> str:
    """Format a notification body, dropping whole rows whose one value is empty.

    Notification params are optional by nature: a legacy concern carries no
    leader snapshot, a level change may have no reason, and _notif_values blanks
    a leader who IS the owner. A dangling «Лидер:» with nothing after it reads as
    missing data; no row at all reads as "not applicable", which is what it is.

    The decision is made on the TEMPLATE, one line at a time — never by pattern-
    matching the rendered text, which would let a concern the user happened to
    type as "Muammo:\\njuva ishlamayapti" lose its own first line. Only rows
    carrying EXACTLY ONE placeholder can be dropped, so the older single-line
    «Label: v | Label: v» bodies keep every field they ever showed. A missing
    key drops its row instead of raising, the same back-compat courtesy the
    ``setdefault`` fallbacks in _notif_values extend to params added after the
    fact.

    ``html=True`` returns the same rows as Telegram HTML (see _html_row); pass
    HTML-escaped values with it."""
    out: list[str] = []
    for line in tmpl.split("\n"):
        fields = [f for _, f, _, _ in string.Formatter().parse(line) if f]
        if len(fields) == 1:
            v = values.get(fields[0])
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
        rendered = (_html_row(line) if html else line).format(**values)
        # Collapse the separator a dropped row leaves stranded.
        if not rendered.strip() and (not out or not out[-1].strip()):
            continue
        out.append(rendered)
    return "\n".join(out).strip("\n")


def _notif_values(params: dict, lang: str, *, escape: bool = False) -> dict:
    """The interpolation values behind BOTH notification renderers — the plain
    bell/DM text and the HTML Telegram body. One prep step, so an HTML DM can
    never quietly lose a language-derived label the bell resolves.

    ``escape`` HTML-escapes the interpolated free text (names, a concern's own
    words) so a stray & or < in a DB value can't break the markup; ints pass
    through untouched."""
    params = params or {}
    # Latinise embedded DB values (names, job titles) for uz/en so notifications
    # match the dashboard; ru/uz_cyrl keep the original Cyrillic. No-op on the
    # already-Latin/non-string params (count, etc.).
    values = {k: transliterate(v, lang) for k, v in params.items()}
    if "date" in params:
        values["date"] = _fmt_date(params["date"], lang)
    # Back-compat: call_forecast gained ``eff`` (Zagruzka %) then ``name``
    # (supervisor) fields after some notices were already stored; fall back so
    # those old rows still render at view time.
    values.setdefault("eff", "—")
    values.setdefault("name", "—")
    # Language-derived params: resolve from the raw value so the label localises
    # to the *viewer's* language, not the creator's (doc_type → doc_label).
    if "doc_type" in params:
        values["doc_label"] = _doc_label(params["doc_type"], lang)
    if "task_status" in params:
        by_lang = _TASK_STATUS_LABELS.get(params["task_status"], {})
        values["status_label"] = by_lang.get(lang) or by_lang.get("en") or params["task_status"]
    if "concern_level" in params:
        by_lang = _CONCERN_LEVEL_LABELS.get(params["concern_level"], {})
        values["level_label"] = by_lang.get(lang) or by_lang.get("en") or params["concern_level"]
    # A concern raised by the cell's own leader printed the one name twice, under
    # two labels ("Лидер: X … Хавотир эгаси: X"), which told the reader nothing
    # the first row hadn't. Blank the duplicate and let _render_body take the
    # line out. Done here, at RENDER time, so notifications already in the bell
    # lose the repetition too.
    if _same_person(params.get("leader_name"), params.get("owner")):
        values["leader_name"] = ""
    if escape:
        values = {k: (_html_escape(v) if isinstance(v, str) else v)
                  for k, v in values.items()}
    return values


def _mk_notif(nkey: str, params: dict, lang: str) -> tuple[str, str]:
    """Render a notification template (title, body) in ``lang``. Pure — given the
    same key + raw params it produces the same output, so the bell can call it at
    *view time* in each viewer's current language (see routers/notifications.py)."""
    strings = _NOTIF_STRINGS.get(nkey, {})
    title_tmpl, body_tmpl = strings.get(lang) or strings.get("en") or (nkey, "")
    values = _notif_values(params, lang)
    return title_tmpl.format(**values), _render_body(body_tmpl, values)


# The second, cheaper door to a rich DM: a key listed here is rendered in HTML
# straight from its _NOTIF_STRINGS row shape — «quoted» content becomes a
# <blockquote>, each "{emoji} Label: value" row gets a <b>bold label</b>, and
# this emoji leads the bold title. Same look as the hand-written templates
# below, but ONE string per language stays the source for both the bell and the
# DM, which is what keeps a typo fixed in one of them from surviving in the
# other. Reach for _NOTIF_TG_HTML instead only when the DM must say something
# the bell does not (a greeting, an advisory paragraph, a premium emoji).
_NOTIF_TG_ICON = {
    "concern_created": "🔔",
    "concern_assigned": "📌",
    "concern_moved": "🔀",
    "concern_resolved": "✅",
    "concern_reopened": "🔄",
    "concern_edited": "✏️",
    "concern_escalated": "⬆️",
    "concern_returned": "↩️",
    "concern_comment": "💬",
}

# HTML-formatted Telegram bodies for notifications whose DM should render richer
# than the plain bell text — bold labels and a real <blockquote> the legacy
# Markdown parse mode can't produce. Only keys present here (or in
# _NOTIF_TG_ICON above) send in HTML mode; every other notification keeps the
# plain Markdown DM. Params are ints + a pre-formatted date (no user free-text),
# so no HTML escaping is needed.
_NOTIF_TG_HTML = {
    # Verifix upload reminder → a warm, personal DM: greeting by the brigadir's
    # name, the day-close ask, and a premium (custom) emoji sign-off. The
    # <tg-emoji> renders animated for Telegram Premium recipients and degrades
    # to the plain 🌤 for everyone else. The in-app bell keeps the terse
    # ``verifix_uploaded`` text in _NOTIF_STRINGS — only the DM is enriched.
    "verifix_uploaded": {
        "uz": ("🔔 <b>Verifix ma'lumotlari yuklandi</b> — 📅 <b>{date}</b>\n\n"
               "Assalomu alaykum, hurmatli <b>{name}</b>!\n\n"
               "Iltimos, bugungi o'zgarishlarni (xodim almashinuvi, lavozim o'zgarishi, o'chirish) "
               "kiritib, kunni yopishingizni so'raymiz ✅\n\n"
               "Samarali ish kuni tilaymiz! <tg-emoji emoji-id=\"5402477260982731644\">🌤</tg-emoji>"),
        "uz_cyrl": ("🔔 <b>Verifix маълумотлари юкланди</b> — 📅 <b>{date}</b>\n\n"
                    "Ассалому алайкум, ҳурматли <b>{name}</b>!\n\n"
                    "Илтимос, бугунги ўзгаришларни (ходим алмашинуви, лавозим ўзгариши, ўчириш) "
                    "киритиб, кунни ёпишингизни сўраймиз ✅\n\n"
                    "Самарали иш куни тилаймиз! <tg-emoji emoji-id=\"5402477260982731644\">🌤</tg-emoji>"),
        "ru": ("🔔 <b>Данные Verifix загружены</b> — 📅 <b>{date}</b>\n\n"
               "Здравствуйте, уважаемый <b>{name}</b>!\n\n"
               "Пожалуйста, внесите сегодняшние изменения (обмен сотрудниками, смена должности, удаление) "
               "и закройте день ✅\n\n"
               "Продуктивного рабочего дня! <tg-emoji emoji-id=\"5402477260982731644\">🌤</tg-emoji>"),
        "en": ("🔔 <b>Verifix data uploaded</b> — 📅 <b>{date}</b>\n\n"
               "Hello, dear <b>{name}</b>!\n\n"
               "Please make today's changes (people exchange, role change, deletion) "
               "and close the day ✅\n\n"
               "Have a productive day! <tg-emoji emoji-id=\"5402477260982731644\">🌤</tg-emoji>"),
    },
    "call_forecast": {
        "uz": ("<b>Xodim chaqirish uchun prognozlar:</b>\n\n"
               "👤 <b>Brigadir:</b> {name}\n📅 <b>Sana:</b> {date}\n📊 <b>Zagruzka foizi:</b> {eff}%\n🧑‍🍳 <b>Chaqirish tavsiya qilinadi:</b> {count} nafar\n⚠️ <b>Maksimum:</b> {max} nafar\n\n"
               "<blockquote>Ko'rsatilgan sonlar faqat zagruzka hisobiga kiritiladigan ishchilar soni — bu barcha ishchilarning umumiy soni emas.\n\n"
               "Ushbu prognoz faqat oldingi kunlarning tarixiy ma'lumotlari asosida shakllantirilgan. "
               "Shu sababli, amaldagi vaziyat va boshqa omillarni ham inobatga olib, kerakli miqdorda odam chaqirishingizni so'raymiz!</blockquote>"),
        "uz_cyrl": ("<b>Ходим чақириш учун прогнозлар:</b>\n\n"
                    "👤 <b>Бригадир:</b> {name}\n📅 <b>Сана:</b> {date}\n📊 <b>Загрузка фоизи:</b> {eff}%\n🧑‍🍳 <b>Чақириш тавсия қилинади:</b> {count} нафар\n⚠️ <b>Максимум:</b> {max} нафар\n\n"
                    "<blockquote>Кўрсатилган сонлар фақат загрузка ҳисобига киритиладиган ишчилар сони — бу барча ишчиларнинг умумий сони эмас.\n\n"
                    "Ушбу прогноз фақат олдинги кунларнинг тарихий маълумотлари асосида шакллантирилган. "
                    "Шу сабабли, амалдаги вазият ва бошқа омилларни ҳам инобатга олиб, керакли миқдорда одам чақиришингизни сўраймиз!</blockquote>"),
        "ru": ("<b>Прогноз по вызову сотрудников:</b>\n\n"
               "👤 <b>Бригадир:</b> {name}\n📅 <b>Дата:</b> {date}\n📊 <b>Процент загрузки:</b> {eff}%\n🧑‍🍳 <b>Рекомендуется вызвать:</b> {count} чел.\n⚠️ <b>Максимум:</b> {max} чел.\n\n"
               "<blockquote>Указанные числа — это количество работников, учитываемых в загрузке, а не общее число всех работников.\n\n"
               "Этот прогноз сформирован только на основе исторических данных за предыдущие дни. "
               "Поэтому, учитывая текущую ситуацию и другие факторы, просим вызвать необходимое количество людей!</blockquote>"),
        "en": ("<b>Staff call forecast:</b>\n\n"
               "👤 <b>Supervisor:</b> {name}\n📅 <b>Date:</b> {date}\n📊 <b>Load percentage:</b> {eff}%\n🧑‍🍳 <b>Recommended to call:</b> {count} workers\n⚠️ <b>Maximum:</b> {max} workers\n\n"
               "<blockquote>The numbers shown are the count of workers included in the Workload calculation — not the total number of all workers.\n\n"
               "This forecast is based only on historical data from previous days. "
               "Therefore, taking the actual situation and other factors into account, please call the number of people you need!</blockquote>"),
    },
    # ── automatic day verification: the DM a person is JUDGED by ──────────────
    # These six arrive daily and carry a number that costs points, so the plain
    # one-line DM was the wrong shape for them: «Сдано: 100% → подтверждено:
    # 55% | Не принято: №2, №4, №7, №8, №11, №13. Причины — в отчёте.» is one
    # pipe-separated wall in which the verdict, its arithmetic and the way back
    # all have the same weight. Here each fact gets its own labelled line, the
    # score leads (it is what the phone's preview shows), and the recourse sits
    # apart in a blockquote. The in-app bell keeps the terse _NOTIF_STRINGS
    # text — a bell row is a one-liner by design.
    # Two audiences, two voices: the *_report_* keys go to the unit's brigadir
    # and name the leader; the bare keys are the leader's own copy, in the
    # second person, and end on who to talk to rather than on how to object.
    "leader_day_report_clean": {
        "uz": ("✅ <b>Kun tasdiqlandi — {score}%</b>\n\n"
               "👤 <b>Lider:</b> {leader}\n"
               "📅 <b>Sana:</b> {date}\n"
               "📸 <b>Tekshirildi:</b> {total} tadan {checked} ta — barchasi qabul qilindi\n\n"
               "<blockquote>Batafsil hisobot — quyidagi tugmada.</blockquote>"),
        "uz_cyrl": ("✅ <b>Кун тасдиқланди — {score}%</b>\n\n"
                    "👤 <b>Лидер:</b> {leader}\n"
                    "📅 <b>Сана:</b> {date}\n"
                    "📸 <b>Текширилди:</b> {total} тадан {checked} та — барчаси қабул қилинди\n\n"
                    "<blockquote>Батафсил ҳисобот — қуйидаги тугмада.</blockquote>"),
        "ru": ("✅ <b>День подтверждён — {score}%</b>\n\n"
               "👤 <b>Лидер:</b> {leader}\n"
               "📅 <b>Дата:</b> {date}\n"
               "📸 <b>Проверено:</b> {checked} из {total} — все подтверждения приняты\n\n"
               "<blockquote>Подробный отчёт — по кнопке ниже.</blockquote>"),
        "en": ("✅ <b>Day verified — {score}%</b>\n\n"
               "👤 <b>Leader:</b> {leader}\n"
               "📅 <b>Date:</b> {date}\n"
               "📸 <b>Checked:</b> {checked} of {total} — all proofs accepted\n\n"
               "<blockquote>The full report is behind the button below.</blockquote>"),
    },
    "leader_day_report_flagged": {
        "uz": ("⚠️ <b>Kun tekshirildi — {score}%</b>\n\n"
               "👤 <b>Lider:</b> {leader}\n"
               "📅 <b>Sana:</b> {date}\n\n"
               "📤 <b>Topshirilgan:</b> {raw}%\n"
               "✅ <b>Tasdiqlangan:</b> {score}%\n"
               "❌ <b>Qabul qilinmadi:</b> {total} tadan {rejected} ta\n\n"
               "🔻 <b>Vazifalar:</b> {tasks}\n\n"
               "<blockquote>Har bir vazifa bo'yicha sabab — quyidagi tugmadagi hisobotda. "
               "Qaror noto'g'ri deb hisoblasangiz, o'sha yerdan e'tiroz yuborishingiz mumkin.</blockquote>"),
        "uz_cyrl": ("⚠️ <b>Кун текширилди — {score}%</b>\n\n"
                    "👤 <b>Лидер:</b> {leader}\n"
                    "📅 <b>Сана:</b> {date}\n\n"
                    "📤 <b>Топширилган:</b> {raw}%\n"
                    "✅ <b>Тасдиқланган:</b> {score}%\n"
                    "❌ <b>Қабул қилинмади:</b> {total} тадан {rejected} та\n\n"
                    "🔻 <b>Вазифалар:</b> {tasks}\n\n"
                    "<blockquote>Ҳар бир вазифа бўйича сабаб — қуйидаги тугмадаги ҳисоботда. "
                    "Қарор нотўғри деб ҳисобласангиз, ўша ердан эътироз юборишингиз мумкин.</blockquote>"),
        "ru": ("⚠️ <b>День проверен — {score}%</b>\n\n"
               "👤 <b>Лидер:</b> {leader}\n"
               "📅 <b>Дата:</b> {date}\n\n"
               "📤 <b>Сдано:</b> {raw}%\n"
               "✅ <b>Подтверждено:</b> {score}%\n"
               "❌ <b>Не принято:</b> {rejected} из {total}\n\n"
               "🔻 <b>Задачи:</b> {tasks}\n\n"
               "<blockquote>Причина по каждой задаче — в отчёте по кнопке ниже. "
               "Если решение кажется ошибочным, оттуда же можно отправить возражение.</blockquote>"),
        "en": ("⚠️ <b>Day verified — {score}%</b>\n\n"
               "👤 <b>Leader:</b> {leader}\n"
               "📅 <b>Date:</b> {date}\n\n"
               "📤 <b>Submitted:</b> {raw}%\n"
               "✅ <b>Verified:</b> {score}%\n"
               "❌ <b>Not accepted:</b> {rejected} of {total}\n\n"
               "🔻 <b>Tasks:</b> {tasks}\n\n"
               "<blockquote>The reason for each task is in the report behind the button below. "
               "If a decision looks wrong, you can file an objection from there.</blockquote>"),
    },
    "leader_day_report_corrected": {
        "uz": ("🔄 <b>Baho qayta ko'rib chiqildi — {score}%</b>\n\n"
               "👤 <b>Lider:</b> {leader}\n"
               "📅 <b>Sana:</b> {date}\n\n"
               "📊 <b>Avval:</b> {before}% → <b>hozir:</b> {score}%\n"
               "❌ <b>Hozir qabul qilinmagan:</b> {rejected} ta vazifa\n\n"
               "<blockquote>Nima o'zgargani — quyidagi tugmadagi hisobotda.</blockquote>"),
        "uz_cyrl": ("🔄 <b>Баҳо қайта кўриб чиқилди — {score}%</b>\n\n"
                    "👤 <b>Лидер:</b> {leader}\n"
                    "📅 <b>Сана:</b> {date}\n\n"
                    "📊 <b>Аввал:</b> {before}% → <b>ҳозир:</b> {score}%\n"
                    "❌ <b>Ҳозир қабул қилинмаган:</b> {rejected} та вазифа\n\n"
                    "<blockquote>Нима ўзгаргани — қуйидаги тугмадаги ҳисоботда.</blockquote>"),
        "ru": ("🔄 <b>Оценка пересмотрена — {score}%</b>\n\n"
               "👤 <b>Лидер:</b> {leader}\n"
               "📅 <b>Дата:</b> {date}\n\n"
               "📊 <b>Было:</b> {before}% → <b>стало:</b> {score}%\n"
               "❌ <b>Сейчас не принято:</b> {rejected} задач(и)\n\n"
               "<blockquote>Что именно изменилось — в отчёте по кнопке ниже.</blockquote>"),
        "en": ("🔄 <b>Score reviewed — {score}%</b>\n\n"
               "👤 <b>Leader:</b> {leader}\n"
               "📅 <b>Date:</b> {date}\n\n"
               "📊 <b>Was:</b> {before}% → <b>now:</b> {score}%\n"
               "❌ <b>Currently not accepted:</b> {rejected} task(s)\n\n"
               "<blockquote>What changed is in the report behind the button below.</blockquote>"),
    },
    "leader_day_clean": {
        "uz": ("✅ <b>Kun hisobotingiz tasdiqlandi — {score}%</b>\n\n"
               "📅 <b>Sana:</b> {date}\n"
               "📸 <b>Tekshirildi:</b> {total} tadan {checked} ta — barcha rasmlaringiz qabul qilindi\n\n"
               "<blockquote>Rahmat! Batafsil hisobot — quyidagi tugmada.</blockquote>"),
        "uz_cyrl": ("✅ <b>Кун ҳисоботингиз тасдиқланди — {score}%</b>\n\n"
                    "📅 <b>Сана:</b> {date}\n"
                    "📸 <b>Текширилди:</b> {total} тадан {checked} та — барча расмларингиз қабул қилинди\n\n"
                    "<blockquote>Раҳмат! Батафсил ҳисобот — қуйидаги тугмада.</blockquote>"),
        "ru": ("✅ <b>Ваш отчёт за день подтверждён — {score}%</b>\n\n"
               "📅 <b>Дата:</b> {date}\n"
               "📸 <b>Проверено:</b> {checked} из {total} — все ваши фото приняты\n\n"
               "<blockquote>Спасибо! Подробный отчёт — по кнопке ниже.</blockquote>"),
        "en": ("✅ <b>Your day report is verified — {score}%</b>\n\n"
               "📅 <b>Date:</b> {date}\n"
               "📸 <b>Checked:</b> {checked} of {total} — all your photos were accepted\n\n"
               "<blockquote>Thank you! The full report is behind the button below.</blockquote>"),
    },
    "leader_day_flagged": {
        "uz": ("⚠️ <b>Kuningiz tekshirildi — {score}%</b>\n\n"
               "📅 <b>Sana:</b> {date}\n\n"
               "📤 <b>Topshirilgan:</b> {raw}%\n"
               "✅ <b>Tasdiqlangan:</b> {score}%\n"
               "❌ <b>Qabul qilinmadi:</b> {total} tadan {rejected} ta\n\n"
               "🔻 <b>Vazifalar:</b> {tasks}\n\n"
               "<blockquote>Har bir vazifa bo'yicha sabab — quyidagi tugmadagi hisobotda. "
               "Rozi bo'lmasangiz, brigadiringizga murojaat qiling.</blockquote>"),
        "uz_cyrl": ("⚠️ <b>Кунингиз текширилди — {score}%</b>\n\n"
                    "📅 <b>Сана:</b> {date}\n\n"
                    "📤 <b>Топширилган:</b> {raw}%\n"
                    "✅ <b>Тасдиқланган:</b> {score}%\n"
                    "❌ <b>Қабул қилинмади:</b> {total} тадан {rejected} та\n\n"
                    "🔻 <b>Вазифалар:</b> {tasks}\n\n"
                    "<blockquote>Ҳар бир вазифа бўйича сабаб — қуйидаги тугмадаги ҳисоботда. "
                    "Рози бўлмасангиз, бригадирингизга мурожаат қилинг.</blockquote>"),
        "ru": ("⚠️ <b>Ваш день проверен — {score}%</b>\n\n"
               "📅 <b>Дата:</b> {date}\n\n"
               "📤 <b>Сдано:</b> {raw}%\n"
               "✅ <b>Подтверждено:</b> {score}%\n"
               "❌ <b>Не принято:</b> {rejected} из {total}\n\n"
               "🔻 <b>Задачи:</b> {tasks}\n\n"
               "<blockquote>Причина по каждой задаче — в отчёте по кнопке ниже. "
               "Если вы не согласны, обратитесь к своему бригадиру.</blockquote>"),
        "en": ("⚠️ <b>Your day was verified — {score}%</b>\n\n"
               "📅 <b>Date:</b> {date}\n\n"
               "📤 <b>Submitted:</b> {raw}%\n"
               "✅ <b>Verified:</b> {score}%\n"
               "❌ <b>Not accepted:</b> {rejected} of {total}\n\n"
               "🔻 <b>Tasks:</b> {tasks}\n\n"
               "<blockquote>The reason for each task is in the report behind the button below. "
               "If you disagree, talk to your supervisor.</blockquote>"),
    },
    "leader_day_report_excluded": {
        "uz": ("⊘ <b>Kun hisobdan chiqarildi</b>\n\n"
               "👤 <b>Lider:</b> {leader}\n"
               "📅 <b>Sana:</b> {date}\n"
               "📊 <b>Kun bahosi edi:</b> {score}%\n"
               "✍️ <b>Kim:</b> {by}\n\n"
               "💬 <b>Sabab:</b> {reason}\n\n"
               "<blockquote>Bu kun endi o'rtacha natijaga umuman kirmaydi — na ortiqcha, na kamchilik.</blockquote>"),
        "uz_cyrl": ("⊘ <b>Кун ҳисобдан чиқарилди</b>\n\n"
                    "👤 <b>Лидер:</b> {leader}\n"
                    "📅 <b>Сана:</b> {date}\n"
                    "📊 <b>Кун баҳоси эди:</b> {score}%\n"
                    "✍️ <b>Ким:</b> {by}\n\n"
                    "💬 <b>Сабаб:</b> {reason}\n\n"
                    "<blockquote>Бу кун энди ўртача натижага умуман кирмайди — на ортиқча, на камчилик.</blockquote>"),
        "ru": ("⊘ <b>День исключён из результатов</b>\n\n"
               "👤 <b>Лидер:</b> {leader}\n"
               "📅 <b>Дата:</b> {date}\n"
               "📊 <b>Результат дня был:</b> {score}%\n"
               "✍️ <b>Кто:</b> {by}\n\n"
               "💬 <b>Причина:</b> {reason}\n\n"
               "<blockquote>Этот день больше не входит в средний результат — ни в плюс, ни в минус.</blockquote>"),
        "en": ("⊘ <b>Day excluded from the results</b>\n\n"
               "👤 <b>Leader:</b> {leader}\n"
               "📅 <b>Date:</b> {date}\n"
               "📊 <b>The day scored:</b> {score}%\n"
               "✍️ <b>By:</b> {by}\n\n"
               "💬 <b>Reason:</b> {reason}\n\n"
               "<blockquote>This day is now out of the average entirely — neither a plus nor a minus.</blockquote>"),
    },
    "leader_day_excluded": {
        "uz": ("⊘ <b>Kuningiz hisobdan chiqarildi</b>\n\n"
               "📅 <b>Sana:</b> {date}\n"
               "📊 <b>Kun bahosi edi:</b> {score}%\n"
               "✍️ <b>Kim:</b> {by}\n\n"
               "💬 <b>Sabab:</b> {reason}\n\n"
               "<blockquote>Bu kun natijalaringizga umuman ta'sir qilmaydi — na ortiqcha, na kamchilik. "
               "Ballaringiz qolgan kunlar bo'yicha hisoblanadi.</blockquote>"),
        "uz_cyrl": ("⊘ <b>Кунингиз ҳисобдан чиқарилди</b>\n\n"
                    "📅 <b>Сана:</b> {date}\n"
                    "📊 <b>Кун баҳоси эди:</b> {score}%\n"
                    "✍️ <b>Ким:</b> {by}\n\n"
                    "💬 <b>Сабаб:</b> {reason}\n\n"
                    "<blockquote>Бу кун натижаларингизга умуман таъсир қилмайди — на ортиқча, на камчилик. "
                    "Балларингиз қолган кунлар бўйича ҳисобланади.</blockquote>"),
        "ru": ("⊘ <b>Ваш день исключён из результатов</b>\n\n"
               "📅 <b>Дата:</b> {date}\n"
               "📊 <b>Результат дня был:</b> {score}%\n"
               "✍️ <b>Кто:</b> {by}\n\n"
               "💬 <b>Причина:</b> {reason}\n\n"
               "<blockquote>Этот день никак не влияет на ваши результаты — ни в плюс, ни в минус. "
               "Средний балл считается по остальным дням.</blockquote>"),
        "en": ("⊘ <b>Your day was excluded from the results</b>\n\n"
               "📅 <b>Date:</b> {date}\n"
               "📊 <b>The day scored:</b> {score}%\n"
               "✍️ <b>By:</b> {by}\n\n"
               "💬 <b>Reason:</b> {reason}\n\n"
               "<blockquote>This day does not affect your results either way. "
               "Your average is taken over the remaining days.</blockquote>"),
    },
    "leader_day_report_restored": {
        "uz": ("↩️ <b>Kun yana hisobga olinadi</b>\n\n"
               "👤 <b>Lider:</b> {leader}\n"
               "📅 <b>Sana:</b> {date}\n"
               "📊 <b>Kun bahosi:</b> {score}%\n"
               "✍️ <b>Kim:</b> {by}"),
        "uz_cyrl": ("↩️ <b>Кун яна ҳисобга олинади</b>\n\n"
                    "👤 <b>Лидер:</b> {leader}\n"
                    "📅 <b>Сана:</b> {date}\n"
                    "📊 <b>Кун баҳоси:</b> {score}%\n"
                    "✍️ <b>Ким:</b> {by}"),
        "ru": ("↩️ <b>День снова учитывается</b>\n\n"
               "👤 <b>Лидер:</b> {leader}\n"
               "📅 <b>Дата:</b> {date}\n"
               "📊 <b>Результат дня:</b> {score}%\n"
               "✍️ <b>Кто:</b> {by}"),
        "en": ("↩️ <b>Day counts again</b>\n\n"
               "👤 <b>Leader:</b> {leader}\n"
               "📅 <b>Date:</b> {date}\n"
               "📊 <b>The day scores:</b> {score}%\n"
               "✍️ <b>By:</b> {by}"),
    },
    "leader_day_restored": {
        "uz": ("↩️ <b>Kuningiz yana hisobga olinadi</b>\n\n"
               "📅 <b>Sana:</b> {date}\n"
               "📊 <b>Kun bahosi:</b> {score}%\n"
               "✍️ <b>Kim:</b> {by}"),
        "uz_cyrl": ("↩️ <b>Кунингиз яна ҳисобга олинади</b>\n\n"
                    "📅 <b>Сана:</b> {date}\n"
                    "📊 <b>Кун баҳоси:</b> {score}%\n"
                    "✍️ <b>Ким:</b> {by}"),
        "ru": ("↩️ <b>Ваш день снова учитывается</b>\n\n"
               "📅 <b>Дата:</b> {date}\n"
               "📊 <b>Результат дня:</b> {score}%\n"
               "✍️ <b>Кто:</b> {by}"),
        "en": ("↩️ <b>Your day counts again</b>\n\n"
               "📅 <b>Date:</b> {date}\n"
               "📊 <b>The day scores:</b> {score}%\n"
               "✍️ <b>By:</b> {by}"),
    },
    "leader_cutoff_report_set": {
        "uz": ("⊘ <b>Lider natijalari hisoblanmaydi</b>\n\n"
               "👤 <b>Lider:</b> {leader}\n"
               "📅 <b>Qachondan:</b> {date}\n"
               "✍️ <b>Kim:</b> {by}\n\n"
               "💬 <b>Sabab:</b> {reason}\n\n"
               "<blockquote>Shu kundan boshlab bu liderning kunlari o'rtacha natijaga umuman kirmaydi — "
               "na ortiqcha, na kamchilik. Undan oldingi kunlar o'z bahosi bilan qoladi.</blockquote>"),
        "uz_cyrl": ("⊘ <b>Лидер натижалари ҳисобланмайди</b>\n\n"
                    "👤 <b>Лидер:</b> {leader}\n"
                    "📅 <b>Қачондан:</b> {date}\n"
                    "✍️ <b>Ким:</b> {by}\n\n"
                    "💬 <b>Сабаб:</b> {reason}\n\n"
                    "<blockquote>Шу кундан бошлаб бу лидернинг кунлари ўртача натижага умуман кирмайди — "
                    "на ортиқча, на камчилик. Ундан олдинги кунлар ўз баҳоси билан қолади.</blockquote>"),
        "ru": ("⊘ <b>Результаты лидера не учитываются</b>\n\n"
               "👤 <b>Лидер:</b> {leader}\n"
               "📅 <b>С какого дня:</b> {date}\n"
               "✍️ <b>Кто:</b> {by}\n\n"
               "💬 <b>Причина:</b> {reason}\n\n"
               "<blockquote>С этого дня дни этого лидера не входят в средний результат — ни в плюс, "
               "ни в минус. Более ранние дни остаются со своими оценками.</blockquote>"),
        "en": ("⊘ <b>This leader's results no longer count</b>\n\n"
               "👤 <b>Leader:</b> {leader}\n"
               "📅 <b>From:</b> {date}\n"
               "✍️ <b>By:</b> {by}\n\n"
               "💬 <b>Reason:</b> {reason}\n\n"
               "<blockquote>From this day on their days are out of the average entirely — neither a "
               "plus nor a minus. Everything before it keeps the score it always had.</blockquote>"),
    },
    "leader_cutoff_set": {
        "uz": ("⊘ <b>Natijalaringiz hisoblanmaydi</b>\n\n"
               "📅 <b>Qachondan:</b> {date}\n"
               "✍️ <b>Kim:</b> {by}\n\n"
               "💬 <b>Sabab:</b> {reason}\n\n"
               "<blockquote>Shu kundan boshlab kunlaringiz o'rtacha natijaga umuman kirmaydi — "
               "na ortiqcha, na kamchilik. Undan oldingi kunlaringiz o'z bahosi bilan qoladi.</blockquote>"),
        "uz_cyrl": ("⊘ <b>Натижаларингиз ҳисобланмайди</b>\n\n"
                    "📅 <b>Қачондан:</b> {date}\n"
                    "✍️ <b>Ким:</b> {by}\n\n"
                    "💬 <b>Сабаб:</b> {reason}\n\n"
                    "<blockquote>Шу кундан бошлаб кунларингиз ўртача натижага умуман кирмайди — "
                    "на ортиқча, на камчилик. Ундан олдинги кунларингиз ўз баҳоси билан қолади.</blockquote>"),
        "ru": ("⊘ <b>Ваши результаты не учитываются</b>\n\n"
               "📅 <b>С какого дня:</b> {date}\n"
               "✍️ <b>Кто:</b> {by}\n\n"
               "💬 <b>Причина:</b> {reason}\n\n"
               "<blockquote>С этого дня ваши дни не входят в средний результат — ни в плюс, ни в "
               "минус. Более ранние дни остаются со своими оценками.</blockquote>"),
        "en": ("⊘ <b>Your results no longer count</b>\n\n"
               "📅 <b>From:</b> {date}\n"
               "✍️ <b>By:</b> {by}\n\n"
               "💬 <b>Reason:</b> {reason}\n\n"
               "<blockquote>From this day on your days are out of the average entirely — neither a "
               "plus nor a minus. Everything before it keeps the score it always had.</blockquote>"),
    },
    "leader_cutoff_report_lifted": {
        "uz": ("↩️ <b>Lider natijalari yana hisoblanadi</b>\n\n"
               "👤 <b>Lider:</b> {leader}\n"
               "📅 <b>Cheklov qachondan edi:</b> {date}\n"
               "✍️ <b>Kim:</b> {by}"),
        "uz_cyrl": ("↩️ <b>Лидер натижалари яна ҳисобланади</b>\n\n"
                    "👤 <b>Лидер:</b> {leader}\n"
                    "📅 <b>Чеклов қачондан эди:</b> {date}\n"
                    "✍️ <b>Ким:</b> {by}"),
        "ru": ("↩️ <b>Результаты лидера снова учитываются</b>\n\n"
               "👤 <b>Лидер:</b> {leader}\n"
               "📅 <b>Ограничение было с:</b> {date}\n"
               "✍️ <b>Кто:</b> {by}"),
        "en": ("↩️ <b>This leader's results count again</b>\n\n"
               "👤 <b>Leader:</b> {leader}\n"
               "📅 <b>The cutoff was from:</b> {date}\n"
               "✍️ <b>By:</b> {by}"),
    },
    "leader_cutoff_lifted": {
        "uz": ("↩️ <b>Natijalaringiz yana hisoblanadi</b>\n\n"
               "📅 <b>Cheklov qachondan edi:</b> {date}\n"
               "✍️ <b>Kim:</b> {by}"),
        "uz_cyrl": ("↩️ <b>Натижаларингиз яна ҳисобланади</b>\n\n"
                    "📅 <b>Чеклов қачондан эди:</b> {date}\n"
                    "✍️ <b>Ким:</b> {by}"),
        "ru": ("↩️ <b>Ваши результаты снова учитываются</b>\n\n"
               "📅 <b>Ограничение было с:</b> {date}\n"
               "✍️ <b>Кто:</b> {by}"),
        "en": ("↩️ <b>Your results count again</b>\n\n"
               "📅 <b>The cutoff was from:</b> {date}\n"
               "✍️ <b>By:</b> {by}"),
    },
    "leader_day_corrected": {
        "uz": ("🔄 <b>Bahoyingiz yangilandi — {score}%</b>\n\n"
               "📅 <b>Sana:</b> {date}\n\n"
               "📊 <b>Avval:</b> {before}% → <b>hozir:</b> {score}%\n"
               "❌ <b>Hozir qabul qilinmagan:</b> {rejected} ta vazifa\n\n"
               "<blockquote>Nima o'zgargani — quyidagi tugmadagi hisobotda.</blockquote>"),
        "uz_cyrl": ("🔄 <b>Баҳоингиз янгиланди — {score}%</b>\n\n"
                    "📅 <b>Сана:</b> {date}\n\n"
                    "📊 <b>Аввал:</b> {before}% → <b>ҳозир:</b> {score}%\n"
                    "❌ <b>Ҳозир қабул қилинмаган:</b> {rejected} та вазифа\n\n"
                    "<blockquote>Нима ўзгаргани — қуйидаги тугмадаги ҳисоботда.</blockquote>"),
        "ru": ("🔄 <b>Ваша оценка изменена — {score}%</b>\n\n"
               "📅 <b>Дата:</b> {date}\n\n"
               "📊 <b>Было:</b> {before}% → <b>стало:</b> {score}%\n"
               "❌ <b>Сейчас не принято:</b> {rejected} задач(и)\n\n"
               "<blockquote>Что именно изменилось — в отчёте по кнопке ниже.</blockquote>"),
        "en": ("🔄 <b>Your score was updated — {score}%</b>\n\n"
               "📅 <b>Date:</b> {date}\n\n"
               "📊 <b>Was:</b> {before}% → <b>now:</b> {score}%\n"
               "❌ <b>Currently not accepted:</b> {rejected} task(s)\n\n"
               "<blockquote>What changed is in the report behind the button below.</blockquote>"),
    },
}


def _mk_notif_tg(nkey: str, params: dict, lang: str) -> str | None:
    """The HTML-formatted Telegram body for a notification, or None when the key
    has neither kind of rich variant and the caller should send the plain DM.

    Two sources, in order: a hand-written _NOTIF_TG_HTML template (the DM says
    more than the bell), else an icon in _NOTIF_TG_ICON, which promotes the key's
    own _NOTIF_STRINGS row shape to HTML. The result is self-contained — it
    carries its own title, because the HTML send path has no separate title
    line."""
    tmpls = _NOTIF_TG_HTML.get(nkey)
    icon = _NOTIF_TG_ICON.get(nkey)
    if not tmpls and not icon:
        return None                      # plain-DM key: don't prep values for nothing
    values = _notif_values(params, lang, escape=True)
    if tmpls:
        return (tmpls.get(lang) or tmpls.get("en")).format(**values)

    strings = _NOTIF_STRINGS.get(nkey, {})
    title_tmpl, body_tmpl = strings.get(lang) or strings.get("en") or ("", "")
    if not title_tmpl:
        return None
    title = title_tmpl.format(**values)
    body = _render_body(body_tmpl, values, html=True)
    return f"{icon} <b>{title}</b>\n\n{body}" if body else f"{icon} <b>{title}</b>"


def _jsonify_params(params: dict) -> dict:
    """Make a template params dict JSON-storable: dates → ISO strings; everything
    else (names, counts, slugs) is already JSON-safe."""
    return {
        k: (v.isoformat() if isinstance(v, (date, datetime)) else v)
        for k, v in (params or {}).items()
    }


def _notify(
    db: Session, telegram_id: int | None, title: str | None = None, body: str | None = None,
    type: str = "info", dm: bool = True, *,
    nkey: str | None = None, params: dict | None = None, lang: str | None = None,
    profile: str | None = None,
):
    # Ghost Mode (admin header toggle): the change still applies and is recorded
    # in the audit trail, but no bell/Telegram notification is pushed to anyone.
    if notifications_suppressed():
        return
    # telegram_id None = the addressee profile is UNCLAIMED: the bell row queues
    # on the profile (whoever claims it inherits the history) and no DM goes out.
    if telegram_id is None:
        if profile is None:
            return
        dm = False
    if nkey is not None:
        # Template row: store the key + raw params so the bell renders it in each
        # viewer's current language. title/body are also stored, rendered in the
        # recipient's language, for the Telegram DM and as a legacy fallback.
        if lang is None:
            lang = _get_user_lang(db, telegram_id)
        title, body = _mk_notif(nkey, params or {}, lang)
        db.add(Notification(
            recipient_telegram_id=telegram_id, recipient_profile=profile, nkey=nkey,
            params=_jsonify_params(params or {}), title=title, body=body, type=type,
        ))
    else:
        title, body = title or "", body or ""
        db.add(Notification(recipient_telegram_id=telegram_id, recipient_profile=profile,
                            title=title, body=body, type=type))
    if dm:
        try:
            from app.telegram_bot import send_tg_notification
            html = _mk_notif_tg(nkey, params, lang) if nkey else None
            send_tg_notification(telegram_id, title, body, html=html)
        except Exception:
            pass


def flush_queued_supervisor_dms(db: Session, telegram_id: int, manager_id: int) -> None:
    """A brigadir just got approved for a unit. Bell rows that were queued to the
    supervisor PROFILE while it was still unclaimed (e.g. a call-to-shift notice)
    were written with recipient_telegram_id NULL and never sent as a Telegram DM.
    Deliver them now and stamp them with the new holder's id so they are sent
    once and never replayed on a later re-claim. Bounded to the last 2 days so a
    fresh claim never dumps stale history into the chat. Called from the approval
    paths (decide_registration / admin add-role); guards its own errors there."""
    if notifications_suppressed():
        return
    key = _profile_key("supervisor", manager_id)
    if not key:
        return
    cutoff = datetime.now(timezone.utc) - timedelta(days=2)
    rows = db.query(Notification).filter(
        Notification.recipient_profile == key,
        Notification.recipient_telegram_id.is_(None),
        Notification.created_at >= cutoff,
    ).order_by(Notification.created_at).all()
    if not rows:
        return
    lang = _get_user_lang(db, telegram_id)
    from app.telegram_bot import send_tg_notification
    for r in rows:
        html = None
        if r.nkey:
            try:
                title, body = _mk_notif(r.nkey, r.params or {}, lang)
                html = _mk_notif_tg(r.nkey, r.params or {}, lang)
            except Exception:
                title, body = r.title or r.nkey, r.body or ""
        else:
            title, body = r.title or "", r.body or ""
        send_tg_notification(telegram_id, title, body, html=html)
        r.recipient_telegram_id = telegram_id   # delivered — don't replay
    db.commit()


def notify_profile(db: Session, profile: str | None, nkey: str, params: dict,
                   type: str = "info", exclude_account: int | None = None,
                   skip_accounts: set[int] | None = None,
                   markup_fn=None) -> set[int]:
    """Notify a PROFILE — the person — wherever they are.

    Writes ONE bell row addressed to the profile (so every account holding it
    sees exactly one copy, and a future claimer inherits it) plus a Telegram DM
    to EVERY approved holder, each rendered in that account's own language.
    A profile nobody has claimed yet queues the bell row with no DM; it is
    delivered on claim (see flush_queued_supervisor_dms).

    This is the ONLY correct way to notify a person. A single-recipient _notify
    picks one registration, which means: co-holders never hear about work
    addressed to them, and after a handover the DM goes to whoever registered
    first — typically the person who left the post.

    ``exclude_account`` suppresses the DM to the account that triggered the
    event (no "you did this" buzz) while STILL writing the profile's bell row,
    so the person's colleagues are not silenced by the actor's own action.

    ``skip_accounts`` suppresses the DM to accounts already DMed about the SAME
    event through another profile — one person may hold two of the profiles an
    event addresses (a leader who also stands in as their unit's brigadir). The
    return value is the set of accounts this call DMed, so a caller notifying
    several profiles about one event accumulates it and passes it back in.

    ``markup_fn(lang)`` attaches an inline keyboard to the DM, built per
    recipient language — a notification whose whole point is "go and look at
    this" is a dead end without the button, and the bell row (which carries no
    keyboard) is not a substitute on a phone.
    """
    if notifications_suppressed() or not profile:
        return set()
    from app.identity import profile_holders

    holder_ids = profile_holders(db, profile)
    if not holder_ids:
        # unclaimed profile → queue the bell row only; no account to DM yet
        _notify(db, None, nkey=nkey, params=params, type=type, profile=profile)
        return set()
    # one profile-addressed bell row (DMs handled per-holder below) …
    _notify(db, holder_ids[0], nkey=nkey, params=params, type=type, dm=False,
            profile=profile)
    # … then a DM to each holder in their own language (HTML variant when the
    # notification defines one, e.g. the call-forecast blockquote)
    from app.telegram_bot import send_tg_notification
    skip = set(skip_accounts or ())
    if exclude_account is not None:
        skip.add(exclude_account)
    dmed: set[int] = set()
    for tid in holder_ids:
        if tid in skip:
            continue
        lang = _get_user_lang(db, tid)
        title, body = _mk_notif(nkey, params, lang)
        html = _mk_notif_tg(nkey, params, lang)
        markup = None
        if markup_fn is not None:
            try:
                markup = markup_fn(lang)
            except Exception:
                logger.exception("notify_profile: markup build failed for %s", nkey)
        try:
            send_tg_notification(tid, title, body, html=html, markup=markup)
        except Exception:
            pass
        dmed.add(tid)
    return dmed


def _notify_supervisor_all(db: Session, manager_id: int, nkey: str,
                           params: dict, type: str = "info") -> None:
    """Notify the supervisor profile of a unit (managers.id IS its profile id)."""
    notify_profile(db, _profile_key("supervisor", manager_id), nkey, params, type)


# ── profile addressing ────────────────────────────────────────────────────────
# Bell rows are addressed to PROFILES (notifications.recipient_profile), not
# telegram accounts: one account can hold several profiles via role switching,
# and an unclaimed profile inherits its rows when claimed. The canonical key is
# "role:id" over the stable profile namespaces — role_profiles.id for
# admin / top-manager / shift-manager / leader / guest, managers.id for
# supervisor — never telegram_user_roles.id (role rows churn on re-claim).

# The implementations live in app/identity.py — THE single answer to "who is
# this person" for the whole app. These names are kept as the historical import
# surface (tasks.py, concerns.py and the bot import them from here).
_profile_key = identity.profile_key
_role_row_profile_key = identity.role_row_profile_key
_viewer_profile_key = identity.viewer_profile_key


def _get_shift_for_manager(db: Session, manager_id: int) -> int:
    mgr = db.query(Manager).filter_by(id=manager_id).first()
    return mgr.shift if mgr else 1


def _assert_day_open(db: Session, manager_id: int, d: date):
    """Supervisors may not submit changes once they have closed the day."""
    if db.query(DayApproval).filter_by(manager_id=manager_id, date=d).first():
        raise HTTPException(
            status_code=409,
            detail="Day is closed — changes can no longer be submitted for this date",
        )


def _unit_has_attendance(db: Session, manager_id: int, d: date) -> bool:
    """True once the unit's verifix data for the date has landed (≥1 named row).
    Nameless hours-only leftovers created by split exchanges don't count."""
    return db.query(Attendance.id).filter(
        Attendance.manager_id == manager_id,
        Attendance.date == d,
        Attendance.worker_name.isnot(None),
        Attendance.worker_name.notin_(["", "nan", "NaN"]),
    ).first() is not None


class ExchangeTargetNoData(HTTPException):
    """A people-exchange may only target a unit whose verifix attendance for the
    date is already uploaded: upload_verifix wipes (manager, date) wholesale, so
    rows transferred into a data-less unit would be destroyed by its eventual
    upload. approvals.py re-raises this instead of folding it into the generic
    409 → "already handled" toast."""
    def __init__(self):
        super().__init__(
            status_code=409,
            detail="Target unit has no attendance data for this date yet — upload its verifix file first",
        )


def _find_supervisor(db: Session, manager_id: int) -> Optional[TelegramUserRole]:
    """The approved supervisor role instance for a unit. Role instances live in
    telegram_user_roles (a person may hold several roles); the returned row
    carries the telegram_id and the role-scoped full_name.

    Turnover leaves the former holder's approved row in place — nothing revokes
    it — so a unit can carry several approved supervisor rows. Pick the MOST
    RECENTLY approved one so notifications and their Telegram DMs reach the
    current brigadir; a stale row would send the DM to an old/blocked account
    while the bell (addressed to the stable profile) still shows correctly."""
    return db.query(TelegramUserRole).filter(
        TelegramUserRole.role == "supervisor",
        TelegramUserRole.role_id == manager_id,
        TelegramUserRole.status == "approved",
    ).order_by(
        TelegramUserRole.approved_at.desc().nullslast(),
        TelegramUserRole.id.desc(),
    ).first()


def _notify_all_parties(
    db: Session,
    manager_id: int,
    nkey: str,
    params: dict,
    ntype: str = "info",
    actor_tg_id: int = None,
    include_supervisor: bool = True,
    admin_dm: bool = True,
):
    """Notify admins + relevant shift-managers + optionally supervisor, excluding
    the actor. Each recipient receives the notification in their own language.

    When ``admin_dm`` is False, admins still get the in-app (bell) notification
    but NOT the plain Telegram DM — used on request-creation events where admins
    instead receive the rich approve/reject button-message (see app.approvals)."""
    # ONE bell row per addressed PROFILE — per person. Building the recipient
    # set out of registrations instead gave a profile held by two accounts two
    # identical bell rows (three holders → three), because each holder
    # contributed its own row for the same person.
    admin_rows = db.query(Admin).all()
    admin_ids: set[int] = {a.telegram_id for a in admin_rows}
    profiles: set[str] = {
        _profile_key("admin", a.profile_id) for a in admin_rows if a.profile_id
    }

    # Shift-managers for this manager's shift — the profiles themselves, so a
    # profile is addressed once whether it is held by nobody, one person or three.
    shift    = _get_shift_for_manager(db, manager_id)
    role_ids = _sm_role_ids_for_shift(db, shift)
    profiles.update(_profile_key("shift-manager", rid) for rid in role_ids)

    # Supervisor
    if include_supervisor:
        profiles.add(_profile_key("supervisor", manager_id))
    profiles.discard(None)

    # The actor's own profile is skipped (no "you did this" notice), but their
    # COLLEAGUES on that same profile are not: excluding by account used to
    # silence the whole profile whenever one of its holders acted.
    actor_profiles = {
        r.profile_key or _role_row_profile_key(db, r)
        for r in db.query(TelegramUserRole).filter(
            TelegramUserRole.telegram_id == actor_tg_id,
            TelegramUserRole.status == "approved",
        ).all()
    } if actor_tg_id else set()

    dmed: set[int] = set()
    for prof in sorted(profiles):
        holders = identity.profile_holders(db, prof)
        if not holders:
            # unclaimed profile → queue the bell row, nobody to DM yet
            _notify(db, None, type=ntype, nkey=nkey, params=params, profile=prof)
            continue
        # One bell row for the profile …
        _notify(db, holders[0], type=ntype, dm=False, nkey=nkey, params=params,
                profile=prof)
        # … then at most one DM per ACCOUNT, however many profiles it holds.
        if prof in actor_profiles:
            continue
        for tg_id in holders:
            if tg_id == actor_tg_id or tg_id in dmed:
                continue
            if not admin_dm and tg_id in admin_ids:
                continue
            dmed.add(tg_id)
            lang = _get_user_lang(db, tg_id)
            title, body = _mk_notif(nkey, params, lang)
            html = _mk_notif_tg(nkey, params, lang)
            try:
                from app.telegram_bot import send_tg_notification
                send_tg_notification(tg_id, title, body, html=html)
            except Exception:
                pass


def notify_supervisor_verifix_upload(db: Session, manager_id: int, d: date):
    """Tell a unit's supervisor that fresh verifix attendance data was uploaded
    for ``d``, so they can make their changes (people exchange, role change,
    deletion) and close the day. Called by the /admin/upload handler after each
    file is inserted — ONLY the supervisor is notified (no admins/shift-managers),
    and the day's close-state is left untouched.

    Verifix uploads are admin-only (/admin/upload is verify_admin-gated), so the
    supervisor is never the person who uploaded — the notification always fires
    for the unit's current brigadir. Addressed to the PROFILE, so EVERY account
    working as that brigadir is told the data landed; picking one registration
    meant only the earliest-approved holder heard about it, which after a
    handover is the person who left the post. An unclaimed supervisor profile
    still gets the bell queued — delivered as a DM once it is claimed.
    The caller must commit; the bell row is added to ``db``, DMs sent inline."""
    prof = _profile_key("supervisor", manager_id)
    if not identity.profile_holders(db, prof):
        logger.warning(
            "verifix upload for manager %s on %s: no approved supervisor to notify",
            manager_id, d,
        )
    # The DM greets the brigadir by name (see the verifix_uploaded HTML variant);
    # the plain bell text ignores the extra param.
    mgr = db.query(Manager).filter_by(id=manager_id).first()
    params = {"date": d}
    if mgr and mgr.name:
        params["name"] = mgr.name
    notify_profile(db, prof, nkey="verifix_uploaded", params=params)


def _log_admin_action(
    db: Session,
    manager_id: int,
    attend_date: date,
    worker_name: str,
    action: str,          # "edit" | "delete"
    changes: dict,
    original: dict,
    admin_tg_id: int,
    admin_name: str,
    batch_id: Optional[str] = None,
):
    """Create a pre-approved EditRequest to log an admin's direct action."""
    supervisor = _find_supervisor(db, manager_id)
    sup_tg_id  = supervisor.telegram_id if supervisor else 0
    sup_name   = supervisor.full_name   if supervisor else ""

    logged_changes = {"_initiated_by": "admin", **changes}
    if action == "delete":
        logged_changes["_action"] = "delete"

    now = datetime.now(timezone.utc)
    req = EditRequest(
        manager_id=manager_id,
        supervisor_telegram_id=sup_tg_id,
        supervisor_name=sup_name,
        date=attend_date,
        worker_name=worker_name,
        changes=logged_changes,
        original=original,
        status="approved",
        processed_by_telegram_id=admin_tg_id,
        processed_by_name=admin_name,
        processed_at=now,
        batch_id=batch_id,
    )
    db.add(req)

    nkey = "admin_record_deleted" if action == "delete" else "admin_record_edited"
    _notify_all_parties(
        db, manager_id,
        nkey,
        {"worker_name": worker_name, "date": attend_date, "admin_name": admin_name},
        ntype="info",
        actor_tg_id=admin_tg_id,
        include_supervisor=True,
    )


# ── Field options ──────────────────────────────────────────────────────────────

@router.get("/field-options")
def field_options(db: Session = Depends(get_db), caller=Depends(_require_staff)):
    job_titles = [
        r[0] for r in db.query(distinct(Attendance.job_title))
        .filter(Attendance.job_title.isnot(None), Attendance.job_title != "nan", Attendance.job_title != "")
        .order_by(Attendance.job_title).all()
    ]
    schedules = [
        r[0] for r in db.query(distinct(Attendance.schedule))
        .filter(Attendance.schedule.isnot(None), Attendance.schedule != "nan", Attendance.schedule != "")
        .order_by(Attendance.schedule).all()
    ]
    return {
        "job_titles": job_titles,
        "schedules": schedules,
        # Subset selectable as a Role Change target (verifix-only roles removed).
        "assignable_job_titles": [j for j in job_titles if is_assignable_target_role(j)],
    }


# ── Supervisors list (admin picker) ───────────────────────────────────────────

@router.get("/supervisors")
def list_supervisors(caller=Depends(_get_caller), db: Session = Depends(get_db)):
    """Unit reference list — names and shifts only, no attendance data.

    Deliberately NOT behind `_require_staff`: it is also the unit picker the
    admin panel's Cleanup tab needs, and a cleanup grantee has no business
    holding the whole /staff page just to read unit names. That grant is the
    ONLY new way in — every other caller goes through the original two checks
    (staff/daily page access, then admin-or-shift-manager) unchanged."""
    cleanup_scope = cap_scope(db, caller, CAP_CLEANUP)
    # A staff/daily page grant at "all" is the unit picker's whole point: the
    # person may read every unit's day, so they must be able to list the units.
    sees_all_units = _staff_sees_all(db, caller)
    if cleanup_scope is None and not sees_all_units:
        if not role_can_access(caller.get("role"), ["staff", "daily"], get_page_access(db),
                               capability_pages(db, caller), caller_denied_pages(db, caller)):
            raise HTTPException(status_code=403, detail="Access denied")
        if caller.get("role") not in ("admin", "shift-manager"):
            raise HTTPException(status_code=403, detail="Admin or shift-manager only")
    q = db.query(Manager).filter(Manager.archived.is_(False)).order_by(Manager.shift, Manager.name)
    vis = _visible_manager_ids(db, caller)  # None = all (admin); shift-managers see their shift only
    # A cleanup grant at "all" is admin reach by definition — the tab exists to
    # undo an upload that landed on the wrong unit, which needs every unit.
    if cleanup_scope == "all" or sees_all_units:
        vis = None
    if vis is not None:
        q = q.filter(Manager.id.in_(vis))
    return [{"manager_id": m.id, "full_name": m.name, "shift": m.shift} for m in q.all()]


# ── Attendance fetch ───────────────────────────────────────────────────────────

@router.get("/attendance")
def get_attendance(
    attend_date: str,
    manager_id: Optional[int] = None,
    caller=Depends(_require_staff),
    db: Session = Depends(get_db),
):
    role    = caller.get("role")
    role_id = caller.get("role_id")
    # A personal staff/daily page grant at "all" reads like admin: any unit, with
    # a supervisor's own still the default when the page sends no manager_id.
    sees_all = _staff_sees_all(db, caller)

    if role == "supervisor" and not (sees_all and manager_id):
        manager_id = role_id
        if not manager_id:
            raise HTTPException(status_code=400, detail="Supervisor has no linked manager")
    elif role == "admin" or sees_all:
        if not manager_id:
            raise HTTPException(status_code=400, detail="manager_id required for admin")
    elif role == "shift-manager":
        if not manager_id:
            raise HTTPException(status_code=400, detail="manager_id required")
        if not _can_touch_manager(db, caller, manager_id):
            raise HTTPException(status_code=403, detail="Not allowed for this manager")
    else:
        raise HTTPException(status_code=403, detail="Not allowed")

    try:
        d = date.fromisoformat(attend_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    rows = db.query(Attendance).filter(
        Attendance.manager_id == manager_id,
        Attendance.date == d,
        Attendance.worker_name.isnot(None),
        Attendance.worker_name.notin_(["", "nan", "NaN"]),
    ).order_by(Attendance.worker_name).all()

    pending = db.query(EditRequest).filter(
        EditRequest.manager_id == manager_id,
        EditRequest.date == d,
        EditRequest.status == "pending",
    ).all()
    pending_map = {r.worker_name: r for r in pending}

    # Workers moved onto a task (approved people-exchange) stay on this page but
    # show "not came" + a task pill. Map worker_name → task_name.
    #
    # below_min_eff: → task workers whose split cleared 2h on NEITHER side (e.g. sent
    # home ~1h in via the "Uyga qaytarilgan" task) don't count as anyone's worker.
    # Newly-applied exchanges already blank them off the roster (_apply_split_exchange),
    # but a doc applied under the old rule — or restored by a verifix re-upload — leaves
    # the worker named with her before-T hours, so drop her here too and fold her
    # effective before-T hours into extra_hours (mirrors the blanking, at read time).
    task_map: dict[str, str] = {}
    below_min_eff: dict[str, float] = {}
    for ex in db.query(HrDocument).filter(
        HrDocument.doc_type   == "people_exchange",
        HrDocument.manager_id == manager_id,
        HrDocument.date       == d,
        HrDocument.status     == "approved",
    ).all():
        pl = ex.payload or {}
        if pl.get("target_type") != "task":
            continue
        ttime = pl.get("transfer_time")
        for emp in pl.get("employees", []):
            wn = emp.get("worker_name")
            if not wn:
                continue
            task_map[wn] = pl.get("task_name")
            if ttime:
                plan = _compute_split(emp.get("snapshot") or {}, ttime, pl.get("return_time"))
                if plan and max(plan["part1"], plan["part2"]) < MIN_MOVED_ZAGRUZKA_HOURS:
                    below_min_eff[wn] = plan["part1_eff"]

    def _serialize(row: Attendance):
        pr = pending_map.get(row.worker_name)
        return {
            "id":                row.id,
            "worker_name":       row.worker_name,
            "job_title":         row.job_title,
            "schedule":          row.schedule,
            "clock_in_out":      row.clock_in_out,
            "hours_worked":      float(row.hours_worked)      if row.hours_worked      is not None else None,
            "early_arrival_min": float(row.early_arrival_min) if row.early_arrival_min is not None else None,
            "effective_hours":   float(row.effective_hours)   if row.effective_hours   is not None else None,
            # Which cell the single-file «Davomat» upload filed the row under.
            # NULL on days that came in through the older per-supervisor files.
            "verifix_code":      row.verifix_code,
            # A worker split across two of this unit's cells («Yacheykalar»):
            # both halves keep the NAME — the worker is on this roster either
            # way — and `hc_weight` is what stops the per-cell headcount
            # counting them twice. NULL stays NULL on the wire rather than
            # being normalised to 1.0: only the null tells an unsplit row from
            # a half that happens to be worth a whole person.
            "hc_weight":         float(row.hc_weight) if row.hc_weight is not None else None,
            "split_of":          row.split_of,
            # The unit's own brigadir — on the roster, never in the load.
            "is_supervisor":     bool(row.is_supervisor),
            "on_task":           task_map.get(row.worker_name),
            "pending_request":   {"id": pr.id, "changes": pr.changes, "original": pr.original} if pr else None,
        }

    # Sum of hours for rows that have no worker name (hidden from table but counted in totals)
    extra_hours = db.query(func.sum(Attendance.hours_worked)).filter(
        Attendance.manager_id == manager_id,
        Attendance.date == d,
        or_(
            Attendance.worker_name.is_(None),
            Attendance.worker_name.in_(["", "nan", "NaN"]),
        ),
        Attendance.hours_worked.isnot(None),
        Attendance.hours_worked > 0,
    ).scalar() or 0.0
    extra_hours = float(extra_hours)   # func.sum on a Numeric col returns Decimal

    mgr = db.query(Manager).filter_by(id=manager_id).first()

    # Drop below-min → task workers that are still named (old-rule doc or post
    # re-upload) and credit their effective before-T hours to the unit's extra_hours,
    # so the count/table/totals match a freshly-blanked worker without a re-approval.
    workers = []
    for r in rows:
        if r.worker_name in below_min_eff:
            extra_hours += below_min_eff[r.worker_name]
            continue
        workers.append(_serialize(r))

    # Name catalog for the codes present today, so the page can label the cell
    # column in the viewer's language. Codes with no Cell record (e.g. a brand
    # new «Код подразделения») still render — as the bare code.
    codes = sorted({r.verifix_code for r in rows if r.verifix_code})
    cells = []
    if codes:
        by_code = {c.verifix_code: c for c in db.query(Cell).filter(Cell.verifix_code.in_(codes)).all()}
        # The cell's LEADER — what the Yacheyka column and its filter print
        # beside the code, since a cell is never written out by its workshop
        # name (frontend `utils/cellName.js`). One query for the whole day.
        lids = {c.leader_id for c in by_code.values() if c.leader_id}
        lead_names = {
            p.id: p.name
            for p in db.query(RoleProfile).filter(RoleProfile.id.in_(lids)).all()
        } if lids else {}
        for code in codes:
            c = by_code.get(code)
            cells.append({
                # cell_id lets the Yacheyka column link to /cells/:id; None for
                # codes the registry doesn't know (they render as inert text).
                "cell_id":      c.id if c else None,
                "verifix_code": code,
                "leader_name":  lead_names.get(c.leader_id) if c else None,
                "name_uz":      c.name_workshop_uz      if c else None,
                "name_uz_cyrl": c.name_workshop_uz_cyrl if c else None,
                "name_ru":      c.name_workshop_ru      if c else None,
                "name_en":      c.name_workshop_en      if c else None,
            })

    return {
        "manager_id":   manager_id,
        "manager_name": mgr.name if mgr else None,
        "date":         attend_date,
        "workers":      workers,
        "cells":        cells,
        "extra_hours":  round(float(extra_hours), 2),
    }


# ── Admin direct update ────────────────────────────────────────────────────────

class DirectUpdateBody(BaseModel):
    manager_id:   int
    attend_date:  str
    worker_name:  str
    job_title:    Optional[str]   = None
    schedule:     Optional[str]   = None
    hours_worked: Optional[float] = None


@router.post("/attendance/update")
def admin_update(body: DirectUpdateBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    _require_cap_over_unit(caller, db, CAP_ATTENDANCE_EDIT, body.manager_id)

    d = date.fromisoformat(body.attend_date)
    row = db.query(Attendance).filter(
        Attendance.manager_id == body.manager_id,
        Attendance.date == d,
        Attendance.worker_name == body.worker_name,
        Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Attendance record not found")

    row_id = row.id      # snapshot: the commit below expires the instance
    original = {
        "job_title":    row.job_title   or "",
        "schedule":     row.schedule    or "",
        "hours_worked": float(row.hours_worked) if row.hours_worked is not None else None,
    }
    changes = {}
    if body.job_title    is not None: changes["job_title"]    = body.job_title
    if body.schedule     is not None: changes["schedule"]     = body.schedule
    if body.hours_worked is not None: changes["hours_worked"] = body.hours_worked

    if body.job_title    is not None: row.job_title    = body.job_title
    if body.schedule     is not None: row.schedule     = body.schedule
    if body.hours_worked is not None: row.hours_worked = body.hours_worked

    if changes:
        _log_admin_action(
            db, body.manager_id, d, body.worker_name,
            "edit", changes, original,
            int(caller["sub"]), caller.get("full_name", "Admin"),
        )
    db.commit()
    if changes:
        unit = unit_name(db, body.manager_id)
        diff = [(f, original.get(f), v) for f, v in changes.items()]
        alert_grant_use(
            db, caller, CAP_ATTENDANCE_EDIT, "attendance.edit",
            details=[("unit", unit),
                     ("worker", body.worker_name),
                     ("date", body.attend_date)],
            changes=diff,
        )
        action_log.enrich(
            target_kind="worker", target_id=row_id, target_name=body.worker_name,
            unit_id=body.manager_id, unit_name=unit, day=d,
            details=[("unit", unit), ("worker", body.worker_name),
                     ("date", body.attend_date)],
            changes=diff,
        )
    return {"ok": True}


# ── Admin direct delete ────────────────────────────────────────────────────────

class AdminDeleteBody(BaseModel):
    manager_id:  int
    attend_date: str
    worker_name: str


@router.post("/attendance/delete")
def admin_delete(body: AdminDeleteBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    _require_cap_over_unit(caller, db, CAP_ATTENDANCE_DELETE, body.manager_id)

    d = date.fromisoformat(body.attend_date)
    row = db.query(Attendance).filter(
        Attendance.manager_id == body.manager_id,
        Attendance.date == d,
        Attendance.worker_name == body.worker_name,
        Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Attendance record not found")

    row_id = row.id      # snapshot: the row is gone after the delete below
    original = {
        "job_title":    row.job_title   or "",
        "schedule":     row.schedule    or "",
        "hours_worked": float(row.hours_worked) if row.hours_worked is not None else None,
    }

    _log_admin_action(
        db, body.manager_id, d, body.worker_name,
        "delete", {}, original,
        int(caller["sub"]), caller.get("full_name", "Admin"),
    )
    db.delete(row)
    db.commit()
    unit = unit_name(db, body.manager_id)
    diff = [(f, v, None) for f, v in original.items()]
    alert_grant_use(
        db, caller, CAP_ATTENDANCE_DELETE, "attendance.delete",
        details=[("unit", unit),
                 ("worker", body.worker_name),
                 ("date", body.attend_date)],
        changes=diff,
    )
    action_log.enrich(
        target_kind="worker", target_id=row_id, target_name=body.worker_name,
        unit_id=body.manager_id, unit_name=unit, day=d,
        details=[("unit", unit), ("worker", body.worker_name),
                 ("date", body.attend_date)],
        changes=diff,
    )
    return {"ok": True}


# ── Admin / Supervisor bulk delete from attendance ────────────────────────────

class BulkDeleteBody(BaseModel):
    manager_id:       Optional[int]  = None   # required for admin
    attend_date:      str
    worker_names:     List[str]
    replace_batch_id: Optional[str]  = None   # supervisor: withdraw old batch before creating new


@router.post("/attendance/bulk-delete")
def bulk_delete_attendance(
    body: BulkDeleteBody,
    caller=Depends(_require_staff),
    db: Session = Depends(get_db),
):
    role = caller.get("role")
    # `direct` = delete the rows outright instead of filing pending requests.
    # Admins always did; a staff.attendance.delete grant is precisely the power
    # to do the same, so a granted supervisor stops queueing their own requests
    # and deletes — which is what "fix attendance directly" was granted for.
    granted = has_cap(db, caller, CAP_ATTENDANCE_DELETE)
    if role not in ("admin", "supervisor") and not granted:
        raise HTTPException(status_code=403, detail="Not allowed")
    if not body.worker_names:
        raise HTTPException(status_code=400, detail="No workers specified")

    direct = role == "admin" or granted
    if direct:
        # Grantees act from their own unit's page, which sends no manager_id —
        # fall back to their unit before demanding one.
        manager_id = body.manager_id or caller.get("role_id")
        if not manager_id:
            raise HTTPException(status_code=400, detail="manager_id required for admin")
        _require_cap_over_unit(caller, db, CAP_ATTENDANCE_DELETE, manager_id)
    else:
        manager_id = caller.get("role_id")
        if not manager_id:
            raise HTTPException(status_code=400, detail="Supervisor has no linked manager")

    try:
        d = date.fromisoformat(body.attend_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    affected = 0
    pending_admin_batch = None   # (batch_id, manager_id, date, supervisor_name, names) for supervisor batches

    if direct:
        # One batch_id per bulk action so the logged requests appear as a
        # single grouped row in the Requests tab.
        admin_batch_id = str(uuid4())
        deleted_rows: list[tuple] = []   # (worker, job_title, None) for the grant alert
        for worker_name in body.worker_names:
            row = db.query(Attendance).filter(
                Attendance.manager_id  == manager_id,
                Attendance.date        == d,
                Attendance.worker_name == worker_name,
                Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
            ).first()
            if not row:
                continue
            original = {
                "job_title":    row.job_title   or "",
                "schedule":     row.schedule    or "",
                "hours_worked": float(row.hours_worked) if row.hours_worked is not None else None,
            }
            _log_admin_action(
                db, manager_id, d, worker_name,
                "delete", {}, original,
                int(caller["sub"]), caller.get("full_name", "Admin"),
                batch_id=admin_batch_id,
            )
            deleted_rows.append((worker_name, original["job_title"], None))
            db.delete(row)
            affected += 1
    else:
        # Supervisor → create pending delete requests
        _assert_day_open(db, manager_id, d)
        supervisor_tg_id = int(caller["sub"])
        supervisor_name  = caller.get("full_name", "")

        # If editing an existing batch: withdraw all its pending requests first
        if body.replace_batch_id:
            old_reqs = db.query(EditRequest).filter(
                EditRequest.batch_id == body.replace_batch_id,
                EditRequest.status   == "pending",
            ).all()
            for req in old_reqs:
                req.status = "rejected"

        # Generate one batch_id for all requests created in this call
        new_batch_id = str(uuid4())
        created_names: list[str] = []

        for worker_name in body.worker_names:
            row = db.query(Attendance).filter(
                Attendance.manager_id  == manager_id,
                Attendance.date        == d,
                Attendance.worker_name == worker_name,
                Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
            ).first()
            if not row:
                continue
            # Skip if there's already an active pending request for this worker
            # (outside of the batch being replaced)
            existing = db.query(EditRequest).filter(
                EditRequest.manager_id  == manager_id,
                EditRequest.date        == d,
                EditRequest.worker_name == worker_name,
                EditRequest.status      == "pending",
            ).first()
            if existing:
                continue
            original = {
                "job_title":    row.job_title   or "",
                "schedule":     row.schedule    or "",
                "hours_worked": float(row.hours_worked) if row.hours_worked is not None else None,
            }
            db.add(EditRequest(
                manager_id=manager_id,
                supervisor_telegram_id=supervisor_tg_id,
                supervisor_name=supervisor_name,
                date=d,
                worker_name=worker_name,
                changes={"_action": "delete"},
                original=original,
                status="pending",
                batch_id=new_batch_id,
            ))
            created_names.append(worker_name)
            affected += 1

        if affected > 0:
            _notify_all_parties(
                db, manager_id,
                "bulk_delete_request",
                {"supervisor_name": supervisor_name, "count": affected, "date": body.attend_date},
                ntype="info",
                actor_tg_id=supervisor_tg_id,
                include_supervisor=False,
                admin_dm=False,        # admins get the rich approve/reject message instead
            )
            pending_admin_batch = (new_batch_id, manager_id, d, supervisor_name, created_names)

    print(f"[bulk-delete] role={role} manager_id={manager_id} date={body.attend_date} requested={len(body.worker_names)} affected={affected} — committing")
    db.commit()
    print(f"[bulk-delete] commit OK")
    if direct and affected:
        unit = unit_name(db, manager_id)
        alert_grant_use(
            db, caller, CAP_ATTENDANCE_DELETE, "attendance.bulk_delete",
            details=[("unit", unit),
                     ("date", body.attend_date),
                     ("count", affected)],
            changes=deleted_rows,
        )
        action_log.enrich(
            target_kind="batch", target_id=admin_batch_id,
            unit_id=manager_id, unit_name=unit, day=d,
            details=[("unit", unit), ("date", body.attend_date),
                     ("count", affected)],
            changes=deleted_rows,
        )
    elif pending_admin_batch:
        # The supervisor branch deletes nothing — it FILES a batch of delete
        # requests, so this row must not read as a deletion that happened.
        _b_id, _b_mid, _b_date, _b_sup, _b_names = pending_admin_batch
        action_log.enrich(
            action="attendance.request_filed",
            target_kind="batch", target_id=_b_id,
            unit_id=manager_id, day=d,
            details=[("date", body.attend_date), ("count", affected),
                     ("workers", ", ".join(_b_names[:10])
                      + (f" +{len(_b_names) - 10}" if len(_b_names) > 10 else "")),
                     ("status", "pending")],
        )
    # A replaced batch's old requests were just rejected — clear its admin message.
    if body.replace_batch_id:
        try:
            from app.approvals import edit_admin_notices
            edit_admin_notices("edit_batch", str(body.replace_batch_id), "rejected",
                               caller.get("full_name", ""))
        except Exception:
            pass
    # Supervisor batch → send one approve/reject button-message to admins.
    if pending_admin_batch:
        try:
            from app.approvals import send_edit_batch_to_admins
            send_edit_batch_to_admins(db, *pending_admin_batch)
        except Exception:
            pass
    return {"ok": True, "affected": affected}


# ── Deleted workers (restorable) ─────────────────────────────────────────────

@router.get("/attendance/deleted")
def get_deleted_workers(
    manager_id: Optional[int] = None,
    caller=Depends(_require_staff),
    db: Session = Depends(get_db),
):
    """
    Returns all approved delete-requests that have NOT been undone yet.
    Admin: optional manager_id filter; omit to see all managers' deletions.
    Supervisor: always scoped to their own manager.
    """
    role    = caller.get("role")
    role_id = caller.get("role_id")

    if role == "admin":
        pass  # manager_id is an optional filter; None → all managers
    elif role == "supervisor":
        manager_id = role_id
        if not manager_id:
            raise HTTPException(status_code=400, detail="Supervisor has no linked manager")
    else:
        raise HTTPException(status_code=403, detail="Not allowed")

    q = db.query(EditRequest).filter(
        EditRequest.status == "approved",
        EditRequest.changes["_action"].astext == "delete",
    )
    if manager_id:
        q = q.filter(EditRequest.manager_id == manager_id)

    rows = q.order_by(EditRequest.processed_at.desc().nullslast()).all()

    mgr_names = {m.id: m.name for m in db.query(Manager).all()}

    return [
        {
            "id":            r.id,
            "manager_id":    r.manager_id,
            "manager_name":  mgr_names.get(r.manager_id, "—"),
            "worker_name":   r.worker_name,
            "date":          r.date.isoformat(),
            "original":      r.original or {},
            "deleted_by":    r.processed_by_name or r.supervisor_name or "—",
            "deleted_at":    r.processed_at.isoformat() if r.processed_at else None,
        }
        for r in rows
    ]


# ── Create request (supervisor — edit or delete) ───────────────────────────────

class CreateRequestBody(BaseModel):
    attend_date: str
    worker_name: str
    action:      str  = "edit"   # "edit" | "delete"
    changes:     dict = {}
    original:    dict


@router.post("/requests", status_code=201)
def create_request(body: CreateRequestBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    if caller.get("role") != "supervisor":
        raise HTTPException(status_code=403, detail="Supervisors only")

    manager_id       = caller.get("role_id")
    supervisor_tg_id = int(caller["sub"])
    supervisor_name  = caller.get("full_name", "")

    if not manager_id:
        raise HTTPException(status_code=400, detail="Supervisor has no linked manager")

    d = date.fromisoformat(body.attend_date)
    _assert_day_open(db, manager_id, d)

    existing = db.query(EditRequest).filter(
        EditRequest.manager_id == manager_id,
        EditRequest.date == d,
        EditRequest.worker_name == body.worker_name,
        EditRequest.status == "pending",
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="A pending request already exists for this row")

    changes  = {"_action": "delete"} if body.action == "delete" else body.changes
    req = EditRequest(
        manager_id=manager_id,
        supervisor_telegram_id=supervisor_tg_id,
        supervisor_name=supervisor_name,
        date=d,
        worker_name=body.worker_name,
        changes=changes,
        original=body.original,
        status="pending",
    )
    db.add(req)
    db.flush()

    req_nkey = "new_delete_request" if body.action == "delete" else "new_edit_request"
    _notify_all_parties(
        db, manager_id,
        req_nkey,
        {"supervisor_name": supervisor_name, "worker_name": body.worker_name, "date": body.attend_date},
        ntype="info",
        actor_tg_id=supervisor_tg_id,
        include_supervisor=False,  # supervisor created it, no need to notify them
        admin_dm=False,            # admins get the rich approve/reject message instead
    )

    req_id = req.id      # snapshot: the commit below expires the instance
    db.commit()
    action_log.enrich(
        target_kind="request", target_id=req_id, target_name=body.worker_name,
        unit_id=manager_id, day=d,
        details=[("worker", body.worker_name), ("date", body.attend_date),
                 ("mode", body.action), ("status", "pending")],
        changes=[(f, (body.original or {}).get(f), v)
                 for f, v in changes.items() if not f.startswith("_")],
    )
    # Admins get an approve/reject button-message with the full request detail.
    try:
        from app.approvals import send_edit_request_to_admins
        send_edit_request_to_admins(db, req)
    except Exception:
        pass
    return {"id": req_id}


# ── Pending count ──────────────────────────────────────────────────────────────

@router.get("/requests/pending-count")
def pending_count(caller=Depends(_require_staff), db: Session = Depends(get_db)):
    role   = caller.get("role")
    tg_id  = int(caller["sub"])

    q = db.query(EditRequest).filter(EditRequest.status == "pending")

    if role == "supervisor":
        # The unit IS the supervisor profile (managers.id), so scoping by it
        # counts everything the PERSON filed — from any of their logins — and
        # nothing from the other units a multi-role user also supervises.
        # Counting by account instead gave two people running one unit two
        # different pending totals.
        if caller.get("role_id"):
            q = q.filter(EditRequest.manager_id == caller["role_id"])
        else:
            q = q.filter(EditRequest.supervisor_telegram_id == tg_id)
    elif role == "shift-manager":
        sm_slot = caller.get("role_id")
        if not sm_slot:
            return {"count": 0}
        shift   = _sm_shift(db, sm_slot)
        mgr_ids = [m.id for m in db.query(Manager).filter(Manager.shift == shift, Manager.archived.is_(False)).all()]
        q = q.filter(EditRequest.manager_id.in_(mgr_ids))

    return {"count": q.count()}


# ── List requests ──────────────────────────────────────────────────────────────

@router.get("/requests")
def list_requests(caller=Depends(_require_staff), db: Session = Depends(get_db)):
    role    = caller.get("role")
    role_id = caller.get("role_id")
    tg_id   = int(caller["sub"])

    q = db.query(EditRequest)

    if role == "supervisor":
        # Everything filed for the unit — the unit IS this person's profile, so
        # a co-holder's requests are this person's requests. (Previously each
        # account saw only what it had personally filed, so the other holder's
        # pending requests were invisible while still blocking the day.)
        # Includes admin's logged actions on the unit's workers.
        if role_id:
            q = q.filter(EditRequest.manager_id == role_id)
        else:
            q = q.filter(EditRequest.supervisor_telegram_id == tg_id)
    elif role == "shift-manager":
        if not role_id:
            return []
        # The shift comes from the PROFILE. The old id<=2 guess sent every
        # shift-1 profile created after the original four slots to shift 2.
        shift       = _sm_shift(db, role_id)
        mgr_ids     = [m.id for m in db.query(Manager).filter(Manager.shift == shift, Manager.archived.is_(False)).all()]
        q = q.filter(EditRequest.manager_id.in_(mgr_ids))
    # admin sees all

    rows = q.order_by(EditRequest.created_at.desc()).all()

    def _ser(r: EditRequest):
        return {
            "id":                       r.id,
            "manager_id":               r.manager_id,
            "supervisor_name":          r.supervisor_name,
            "date":                     r.date.isoformat(),
            "worker_name":              r.worker_name,
            "changes":                  r.changes,
            "original":                 r.original,
            "status":                   r.status,
            "processed_by_name":        r.processed_by_name,
            "processed_by_telegram_id": r.processed_by_telegram_id,
            "created_at":               r.created_at.isoformat() if r.created_at else None,
            "processed_at":             r.processed_at.isoformat() if r.processed_at else None,
        }

    return [_ser(r) for r in rows]


# ── Approve / Reject ───────────────────────────────────────────────────────────

def _process_request(req_id: int, action: str, caller: dict, db: Session):
    granted = has_cap(db, caller, CAP_REQUESTS_APPROVE)
    if caller.get("role") not in ("admin", "shift-manager") and not granted:
        raise HTTPException(status_code=403, detail="Not authorised")

    req = db.query(EditRequest).filter_by(id=req_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req.status != "pending":
        raise HTTPException(status_code=409, detail="Request already processed")

    # A grant at "own" scope adds the ACTION but not the reach: the request must
    # still sit inside the rows this profile already sees. Roles that could
    # already approve (admin, shift-manager) are unaffected — grants are
    # additive and never narrow an existing authority.
    if (granted and caller.get("role") not in ("admin", "shift-manager")
            and not scope_is_all(db, caller, CAP_REQUESTS_APPROVE)):
        units = _caller_unit_ids(caller, db)
        if units is not None and req.manager_id not in units:
            raise HTTPException(status_code=403, detail="Not authorised")

    # A native shift-manager may act only on their own shift's requests — the
    # same rule undo_request enforces. Without this, approve/reject was looser
    # than undo, letting a shift-1 manager edit shift-2 attendance.
    if caller.get("role") == "shift-manager":
        sm_slot = caller.get("role_id")
        if not sm_slot:
            raise HTTPException(status_code=403, detail="No shift assigned")
        if _sm_shift(db, sm_slot) != _get_shift_for_manager(db, req.manager_id):
            raise HTTPException(status_code=403, detail="Not responsible for this shift")

    processor_tg_id = int(caller["sub"])
    processor_name  = caller.get("full_name", "")
    now = datetime.now(timezone.utc)

    req.status                   = action
    req.processed_by_telegram_id = processor_tg_id
    req.processed_by_name        = processor_name
    req.processed_at             = now

    if action == "approved":
        changes = req.changes or {}
        att_row = db.query(Attendance).filter(
            Attendance.manager_id == req.manager_id,
            Attendance.date       == req.date,
            Attendance.worker_name == req.worker_name,
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        if att_row:
            if changes.get("_action") == "delete":
                db.delete(att_row)
            else:
                field_changes = {k: v for k, v in changes.items() if not k.startswith("_")}
                if "job_title"    in field_changes: att_row.job_title    = field_changes["job_title"]
                if "schedule"     in field_changes: att_row.schedule     = field_changes["schedule"]
                if "hours_worked" in field_changes: att_row.hours_worked = float(field_changes["hours_worked"])

    is_approved = action == "approved"
    ntype       = "success" if is_approved else "warning"
    sup_nkey    = "request_approved_supervisor" if is_approved else "request_rejected_supervisor"
    others_nkey = "request_approved_others"     if is_approved else "request_rejected_others"

    # Notify supervisor about their request result (different message). The
    # request belongs to a unit, so it addresses that unit's supervisor profile.
    if req.supervisor_telegram_id:
        _notify(
            db, req.supervisor_telegram_id,
            type="success" if is_approved else "error",
            nkey=sup_nkey,
            params={"worker_name": req.worker_name, "date": req.date, "processor_name": processor_name},
            profile=_profile_key("supervisor", req.manager_id),
        )

    # Notify admin + shift-managers (supervisor already notified above)
    _notify_all_parties(
        db, req.manager_id,
        others_nkey,
        {"processor_name": processor_name, "supervisor_name": req.supervisor_name,
         "worker_name": req.worker_name, "date": req.date},
        ntype=ntype,
        actor_tg_id=processor_tg_id,
        include_supervisor=False,
    )

    db.commit()
    # Grant-use warning (covers the web app AND the Telegram inline tap — both
    # funnel through this core). Native admins/shift-managers are filtered out
    # inside; a diff only accompanies an APPLIED change.
    req_changes = req.changes or {}
    if action == "approved":
        if req_changes.get("_action") == "delete":
            alert_changes = [(f, v, None) for f, v in (req.original or {}).items()]
        else:
            alert_changes = [(f, (req.original or {}).get(f), v)
                             for f, v in req_changes.items() if not f.startswith("_")]
    else:
        alert_changes = None
    unit = unit_name(db, req.manager_id)
    alert_details = [("unit", unit),
                     ("supervisor", req.supervisor_name),
                     ("worker", req.worker_name),
                     ("date", str(req.date))]
    if req_changes.get("_action") == "delete":
        alert_details.insert(0, ("request", tv("v.request_delete")))
    alert_grant_use(db, caller, CAP_REQUESTS_APPROVE, f"request.{action}",
                    details=alert_details, changes=alert_changes)
    action_log.enrich(
        target_kind="request", target_id=req_id, target_name=req.worker_name,
        unit_id=req.manager_id, unit_name=unit, day=req.date,
        details=[("unit", unit), ("brigadir", req.supervisor_name),
                 ("worker", req.worker_name), ("date", str(req.date)),
                 ("mode", "delete" if req_changes.get("_action") == "delete" else "edit")],
        changes=(alert_changes or []) + [("status", "pending", action)],
    )
    # Edit every admin's Telegram approve/reject message with the outcome,
    # whoever decided (this runs for both the web app and the Telegram tap).
    try:
        from app.approvals import edit_admin_notices
        edit_admin_notices("edit_request", str(req_id), action, processor_name)
    except Exception:
        pass
    return {"ok": True, "status": action}


@router.post("/requests/{req_id}/withdraw")
def withdraw_request(req_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    """Supervisor withdraws their own pending deletion request before it is confirmed."""
    role  = caller.get("role")
    tg_id = int(caller["sub"])
    if role not in ("admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Not authorised")

    req = db.query(EditRequest).filter_by(id=req_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req.status != "pending":
        raise HTTPException(status_code=409, detail="Can only withdraw pending requests")

    # A supervisor may withdraw anything filed for their own unit: the unit IS
    # their profile, so a co-holder's request is this same person's request.
    # Comparing accounts here 403'd two people who run one unit together.
    if role == "supervisor" and not (
        (caller.get("role_id") and req.manager_id == caller["role_id"])
        or req.supervisor_telegram_id == tg_id
    ):
        raise HTTPException(status_code=403, detail="Not your request")

    was_delete = (req.changes or {}).get("_action") == "delete"
    w_name, w_mid, w_date = req.worker_name, req.manager_id, req.date
    req.status = "rejected"
    db.commit()
    action_log.enrich(
        target_kind="request", target_id=req_id, target_name=w_name,
        unit_id=w_mid, day=w_date,
        details=[("worker", w_name), ("date", str(w_date)),
                 ("mode", "delete" if was_delete else "edit")],
        changes=[("status", "pending", "withdrawn")],
    )
    try:
        from app.approvals import edit_admin_notices
        edit_admin_notices("edit_request", str(req_id), "rejected", caller.get("full_name", ""))
    except Exception:
        pass
    return {"ok": True}


@router.post("/requests/{req_id}/approve")
def approve_request(req_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    return _process_request(req_id, "approved", caller, db)


@router.post("/requests/{req_id}/reject")
def reject_request(req_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    return _process_request(req_id, "rejected", caller, db)


@router.post("/requests/{req_id}/undo")
def undo_request(req_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    role   = caller.get("role")
    tg_id  = int(caller["sub"])
    undoer = caller.get("full_name", "")

    if role not in ("admin", "shift-manager"):
        raise HTTPException(status_code=403, detail="Not authorised")

    req = db.query(EditRequest).filter_by(id=req_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req.status != "approved":
        raise HTTPException(status_code=409, detail="Can only undo approved requests")

    # Shift-manager: verify they're responsible for this manager's shift
    if role == "shift-manager":
        sm_slot = caller.get("role_id")
        if not sm_slot:
            raise HTTPException(status_code=403, detail="No shift assigned")
        sm_shift  = _sm_shift(db, sm_slot)
        mgr_shift = _get_shift_for_manager(db, req.manager_id)
        if sm_shift != mgr_shift:
            raise HTTPException(status_code=403, detail="Not responsible for this shift")

    changes  = req.changes  or {}
    original = req.original or {}
    is_delete = changes.get("_action") == "delete"

    if is_delete:
        # Recreate the deleted attendance row from original data
        exists = db.query(Attendance).filter(
            Attendance.manager_id  == req.manager_id,
            Attendance.date        == req.date,
            Attendance.worker_name == req.worker_name,
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        if not exists:
            try:
                hw = float(original["hours_worked"]) if original.get("hours_worked") is not None else None
            except (TypeError, ValueError):
                hw = None
            db.add(Attendance(
                manager_id   = req.manager_id,
                date         = req.date,
                worker_name  = req.worker_name,
                job_title    = original.get("job_title")  or "",
                schedule     = original.get("schedule")   or "",
                hours_worked = hw,
            ))
    else:
        # Restore original field values
        att = db.query(Attendance).filter(
            Attendance.manager_id  == req.manager_id,
            Attendance.date        == req.date,
            Attendance.worker_name == req.worker_name,
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        if att:
            field_changes = {k: v for k, v in changes.items() if not k.startswith("_")}
            if "job_title"    in field_changes: att.job_title    = original.get("job_title",    "")
            if "schedule"     in field_changes: att.schedule     = original.get("schedule",     "")
            if "hours_worked" in field_changes:
                try:
                    att.hours_worked = float(original["hours_worked"]) if original.get("hours_worked") is not None else None
                except (TypeError, ValueError):
                    pass

    req.status = "undone"

    _notify_all_parties(
        db, req.manager_id,
        "request_undone",
        {"worker_name": req.worker_name, "date": req.date, "undoer": undoer},
        ntype="warning",
        actor_tg_id=tg_id,
        include_supervisor=True,
    )

    u_name, u_mid, u_date = req.worker_name, req.manager_id, req.date
    db.commit()
    action_log.enrich(
        target_kind="request", target_id=req_id, target_name=u_name,
        unit_id=u_mid, day=u_date,
        details=[("worker", u_name), ("date", str(u_date)),
                 ("mode", "delete" if is_delete else "edit")],
        changes=[("status", "approved", "undone")]
                + [(f, changes.get(f), original.get(f))
                   for f in changes if not f.startswith("_")],
    )
    return {"ok": True}


# ── Export attendance to Excel → send to Telegram ─────────────────────────────

class ExportRow(BaseModel):
    worker_name:       Optional[str]   = None
    job_title:         Optional[str]   = None
    cell:              Optional[str]   = None
    schedule:          Optional[str]   = None
    clock_in_out:      Optional[str]   = None
    hours_worked:      Optional[float] = None
    early_arrival_min: Optional[float] = None
    effective_hours:   Optional[float] = None


class ExportBody(BaseModel):
    manager_id:  int
    attend_date: str
    rows:        List[ExportRow]


@router.post("/attendance/export")
def export_attendance(request: Request, body: ExportBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    tg_id = int(caller["sub"])

    # Resolve manager name
    mgr = db.query(Manager).filter_by(id=body.manager_id).first()
    manager_name = mgr.name if mgr else f"Manager {body.manager_id}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Yacheyka mirrors the page: the column exists only when the day's rows
    # actually carry a cell code (legacy per-supervisor days have none).
    has_cell = any((r.cell or "").strip() for r in body.rows)
    headers = ["Date", "Manager", "Worker", "Lavozim"] \
        + (["Yacheyka"] if has_cell else []) \
        + ["Jadval", "Clock In/Out", "Soat", "Early Arrival (min)", "Eff. Hours"]
    ws.append(headers)

    # Style header row
    hdr_fill = PatternFill(fill_type="solid", fgColor="1C4ED8")
    for col_i in range(1, len(headers) + 1):
        cell = ws.cell(1, col_i)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Data rows with alternating shading
    even_fill = PatternFill(fill_type="solid", fgColor="F1F5F9")
    for row_i, r in enumerate(body.rows, 2):
        row = [
            body.attend_date,
            manager_name,
            r.worker_name        or "",
            r.job_title          or "",
        ]
        if has_cell:
            row.append(r.cell or "")
        row += [
            r.schedule           or "",
            r.clock_in_out       or "",
            r.hours_worked,
            r.early_arrival_min,
            r.effective_hours,
        ]
        ws.append(row)
        if row_i % 2 == 0:
            for col_i in range(1, len(headers) + 1):
                ws.cell(row_i, col_i).fill = even_fill

    # Column widths: Date, Manager, Worker, Lavozim, [Yacheyka], Jadval, Clock, Soat, Early, Eff
    widths = [13, 30, 42, 26] + ([12] if has_cell else []) + [18, 18, 8, 18, 12]
    for col_i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col_i).column_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"attendance_{body.attend_date}.xlsx"
    caption  = f"📊 Attendance — {body.attend_date}  •  {manager_name}  •  {len(body.rows)} workers"

    try:
        resp = deliver_xlsx(request, caller, filename, buf.read(), caption, chat_id=tg_id)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Telegram send failed: {e}")
    action_log.enrich(
        target_kind="day", target_id=f"{body.manager_id}:{body.attend_date}",
        unit_id=body.manager_id, unit_name=manager_name, day=body.attend_date,
        details=[("unit", manager_name), ("date", body.attend_date),
                 ("rows", len(body.rows)), ("file", filename)],
    )
    return resp


# ── Batch-level request endpoints ─────────────────────────────────────────────

class BatchApproveBody(BaseModel):
    ids: Optional[List[int]] = None  # subset of req_ids to approve; None/empty = all pending


def _batch_id_filter(batch_id: str):
    """Resolve a batch token to an EditRequest filter.

    Requests created via the bulk delete modal share a UUID `batch_id`.
    Single per-row requests have `batch_id = NULL` and are addressed by the
    frontend as 'solo-{request_id}'. Without this, canceling/approving a
    null-batch request hit `batch_id == 'null'` and silently matched nothing.
    """
    if batch_id.startswith("solo-"):
        try:
            return EditRequest.id == int(batch_id[len("solo-"):])
        except ValueError:
            return EditRequest.id == -1  # matches nothing → clean 404
    return EditRequest.batch_id == batch_id


def _process_batch(batch_token: str, action: str, caller: dict, db: Session, ids=None) -> int:
    """Approve/reject every pending request in a batch (or the ``ids`` subset).
    Shared by the HTTP endpoints and the Telegram callback; on approval of a
    delete-batch the attendance rows are removed. Returns the count processed
    and edits each admin's Telegram message with the outcome.

    Bulk batches are always deletions (created by bulk_delete_attendance), so
    an approval deletes the attendance row exactly like the old endpoint did."""
    if caller.get("role") not in ("admin", "shift-manager"):
        raise HTTPException(status_code=403, detail="Not authorised")

    q = db.query(EditRequest).filter(
        _batch_id_filter(batch_token),
        EditRequest.status == "pending",
    )
    if ids:
        q = q.filter(EditRequest.id.in_(ids))
    reqs = q.all()
    if not reqs:
        raise HTTPException(status_code=404, detail="No pending requests found in batch")

    processor_name  = caller.get("full_name", "")
    processor_tg_id = int(caller["sub"])
    now = datetime.now(timezone.utc)

    for req in reqs:
        req.status                   = action
        req.processed_by_telegram_id = processor_tg_id
        req.processed_by_name        = processor_name
        req.processed_at             = now
        if action == "approved":
            att_row = db.query(Attendance).filter(
                Attendance.manager_id  == req.manager_id,
                Attendance.date        == req.date,
                Attendance.worker_name == req.worker_name,
                Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
            ).first()
            if att_row:
                db.delete(att_row)

    # Snapshot before the commit expires the instances — the register must not
    # cost one re-SELECT per request in the batch.
    b_mids  = {r.manager_id for r in reqs}
    b_days  = {r.date for r in reqs}
    b_names = [r.worker_name for r in reqs]

    db.commit()
    action_log.enrich(
        target_kind="batch", target_id=batch_token,
        unit_id=next(iter(b_mids)) if len(b_mids) == 1 else None,
        day=next(iter(b_days)) if len(b_days) == 1 else None,
        details=[("count", len(reqs)),
                 ("workers", ", ".join(b_names[:10])
                  + (f" +{len(b_names) - 10}" if len(b_names) > 10 else ""))],
        changes=[("status", "pending", action)],
    )
    try:
        from app.approvals import edit_admin_notices
        edit_admin_notices("edit_batch", str(batch_token), action, processor_name)
        # A solo token addresses a single request that may instead have been
        # tracked as an "edit_request" notice (create_request path) — edit that
        # key too so the admin buttons clear whichever way the notice was filed.
        if str(batch_token).startswith("solo-"):
            edit_admin_notices("edit_request", str(batch_token)[len("solo-"):], action, processor_name)
    except Exception:
        pass
    return len(reqs)


@router.post("/requests/batch/{batch_id}/approve")
def approve_batch(
    batch_id: str,
    body: BatchApproveBody,
    caller=Depends(_require_staff),
    db: Session = Depends(get_db),
):
    """Admin/shift-manager approves selected (or all) pending requests in a batch."""
    n = _process_batch(batch_id, "approved", caller, db, ids=body.ids)
    return {"ok": True, "approved": n}


@router.post("/requests/batch/{batch_id}/reject")
def reject_batch(
    batch_id: str,
    caller=Depends(_require_staff),
    db: Session = Depends(get_db),
):
    """Admin/shift-manager rejects all pending requests in a batch."""
    n = _process_batch(batch_id, "rejected", caller, db)
    return {"ok": True, "rejected": n}


@router.post("/requests/batch/{batch_id}/withdraw")
def withdraw_batch(
    batch_id: str,
    caller=Depends(_require_staff),
    db: Session = Depends(get_db),
):
    """Supervisor withdraws all their own pending requests in a batch."""
    role    = caller.get("role")
    role_id = caller.get("role_id")
    tg_id   = int(caller["sub"])
    if role not in ("admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Not authorised")

    q = db.query(EditRequest).filter(
        _batch_id_filter(batch_id),
        EditRequest.status   == "pending",
    )
    if role == "supervisor":
        # Scope to the unit (= the profile), not the login, so a batch filed by
        # the other holder of this brigadir profile can still be withdrawn.
        q = q.filter(EditRequest.manager_id == role_id) if role_id else \
            q.filter(EditRequest.supervisor_telegram_id == tg_id)

    reqs = q.all()
    if not reqs:
        raise HTTPException(status_code=404, detail="No pending requests found in batch")

    w_mids  = {r.manager_id for r in reqs}
    w_days  = {r.date for r in reqs}
    w_names = [r.worker_name for r in reqs]
    for req in reqs:
        req.status = "rejected"

    db.commit()
    action_log.enrich(
        target_kind="batch", target_id=batch_id,
        unit_id=next(iter(w_mids)) if len(w_mids) == 1 else None,
        day=next(iter(w_days)) if len(w_days) == 1 else None,
        details=[("count", len(reqs)),
                 ("workers", ", ".join(w_names[:10])
                  + (f" +{len(w_names) - 10}" if len(w_names) > 10 else ""))],
        changes=[("status", "pending", "withdrawn")],
    )
    try:
        from app.approvals import edit_admin_notices
        name = caller.get("full_name", "")
        edit_admin_notices("edit_batch", str(batch_id), "rejected", name)
        if str(batch_id).startswith("solo-"):
            edit_admin_notices("edit_request", str(batch_id)[len("solo-"):], "rejected", name)
    except Exception:
        pass
    return {"ok": True, "withdrawn": len(reqs)}


# ════════════════════════════════════════════════════════════════════════════════
# HR Documents — document-driven change workflow (Role Change, …)
# ════════════════════════════════════════════════════════════════════════════════

DOC_TYPE_LABELS = {
    "role_change":     "Role Change",
    "people_exchange": "People Exchange",
    "graphic_change":  "Graphic Change",
}

# Per-language doc-type labels, used when the label is baked into a notification
# string sent to a specific recipient (the serialized API field above stays the
# English key — the frontend localizes it via its own t() keys).
DOC_TYPE_LABELS_I18N = {
    "role_change":     {"uz": "Lavozimni o'zgartirish", "uz_cyrl": "Лавозимни ўзгартириш", "ru": "Смена должности",   "en": "Role Change"},
    "people_exchange": {"uz": "Xodimlarni almashtirish", "uz_cyrl": "Ходимларни алмаштириш", "ru": "Обмен сотрудниками", "en": "People Exchange"},
    "graphic_change":  {"uz": "Jadvalni o'zgartirish",   "uz_cyrl": "Жадвални ўзгартириш",   "ru": "Смена графика",      "en": "Graphic Change"},
}


def _doc_label(doc_type: str, lang: str) -> str:
    """Localized doc-type label for notification text (falls back en → key)."""
    by_lang = DOC_TYPE_LABELS_I18N.get(doc_type)
    if not by_lang:
        return doc_type
    return by_lang.get(lang) or by_lang.get("en") or doc_type


def _caller_unit_ids(caller, db: Session) -> list[int] | None:
    """Unit ids the caller's OWN row scoping covers; None = no restriction.

    Delegates to the shared definition in app/capabilities so an "own"-scoped
    grant reaches exactly the rows the person already sees — and exactly the
    rows they get notified about. Mirrors ``_scope_documents`` /
    ``_scope_deletion_requests``, with one deliberate difference: a supervisor
    with no unit gets the empty list rather than a fallback to "rows I created",
    because a grant must never be broader than the scoping it claims to reuse."""
    return profile_unit_ids(db, identity.viewer_profile_key(db, caller))


def _scope_deletion_requests(caller, db: Session):
    """Return deletion EditRequests visible to the caller — the FULL history:
    pending, approved, rejected (incl. withdrawn) and undone, so processed
    requests stay visible on the Requests tab instead of disappearing."""
    role    = caller.get("role")
    role_id = caller.get("role_id")
    tg_id   = int(caller["sub"])

    q = db.query(EditRequest).filter(
        EditRequest.changes["_action"].astext == "delete",
        EditRequest.status.in_(["pending", "approved", "rejected", "undone"]),
    )
    # An "all"-scoped grant means admin reach: the queue this person was given
    # is the whole factory's, so lift the role filters entirely.
    if scope_is_all(db, caller, CAP_REQUESTS_APPROVE):
        return q.order_by(EditRequest.date.desc()).all()
    if role == "supervisor":
        if role_id:
            q = q.filter(EditRequest.manager_id == role_id)
        else:
            q = q.filter(EditRequest.supervisor_telegram_id == tg_id)
    elif role == "shift-manager":
        if not role_id:
            return []
        shift   = _sm_shift(db, role_id)   # from the profile, not an id guess
        mgr_ids = [m.id for m in db.query(Manager).filter(Manager.shift == shift, Manager.archived.is_(False)).all()]
        q = q.filter(EditRequest.manager_id.in_(mgr_ids))
    # admin → all
    return q.order_by(EditRequest.date.desc()).all()


_REAL_DOC_TYPES = ("people_exchange", "role_change")


def _real_docs(db: Session):
    """THE starting query for every /staff document read — real documents only.

    /staff is the LIVE register: approving a row here runs ``_apply_doc_effects``
    and writes attendance. The clause used to exclude the cell-level rehearsal
    page's sandbox doc types; that page and its types are gone (2026-08-30), so
    every document is real now. The WHITELIST stays, because it answers a second
    question the sandbox never owned: this register serves every type, so an
    unanticipated one would appear in the list, in the sidebar pending badge and
    on the approve door with no reader having decided what it means there.

    Every ``_scope_documents(...)`` call site starts here so the rule has ONE
    spelling and no door can be forgotten. ``services/day_state`` hand-copies
    the same tuple on purpose — see the comment there.

    TRAP: a whitelist DEFAULTS TO HIDING. That is the safe default for the
    day-state queries but the wrong one for a register, so a fifth real
    doc_type (``graphic_change`` is a placeholder today, with no writer) must be
    added here or it will silently not show on this page.
    """
    return db.query(HrDocument).filter(HrDocument.doc_type.in_(_REAL_DOC_TYPES))


def _scope_documents(q, caller, db: Session):
    """Restrict a HrDocument query to what the caller is allowed to see."""
    role    = caller.get("role")
    role_id = caller.get("role_id")
    tg_id   = int(caller["sub"])

    # An "all"-scoped grant means admin reach — the transfer queue this person
    # handles is every unit's, not just their own.
    if scope_is_all(db, caller, CAP_DOCUMENTS_APPROVE):
        return q
    if role == "supervisor":
        if not role_id:
            return q.filter(HrDocument.created_by_telegram_id == tg_id)
        # Own unit's documents + people-exchange documents addressed TO this
        # supervisor's unit, so the receiving supervisor can see and approve
        # incoming worker moves.
        incoming = and_(
            HrDocument.doc_type == "people_exchange",
            HrDocument.payload["target_manager_id"].astext == str(role_id),
        )
        return q.filter(or_(HrDocument.manager_id == role_id, incoming))
    if role == "shift-manager":
        if not role_id:
            return q.filter(HrDocument.id < 0)   # always-empty
        shift   = _sm_shift(db, role_id)   # from the profile, not an id guess
        mgr_ids = [m.id for m in db.query(Manager).filter(Manager.shift == shift, Manager.archived.is_(False)).all()]
        return q.filter(HrDocument.manager_id.in_(mgr_ids))
    # admin → everything
    return q


def _can_approve(caller) -> bool:
    return caller.get("role") in ("admin", "shift-manager")


def _granted_over_doc(doc: HrDocument, caller: dict, db: Session) -> bool:
    """True if a ``staff.documents.approve`` grant covers THIS document.

    "all" covers everything; "own" covers only documents inside the profile's
    normal unit scoping — including a transfer addressed TO their unit, which
    is the incoming leg a receiving supervisor already sees."""
    scope = cap_scope(db, caller, CAP_DOCUMENTS_APPROVE)
    if scope is None:
        return False
    if scope == "all":
        return True
    units = _caller_unit_ids(caller, db)
    if units is None:
        return True
    if doc.manager_id in units:
        return True
    payload = doc.payload or {}
    return (doc.doc_type == "people_exchange"
            and payload.get("target_type") == "supervisor"
            and payload.get("target_manager_id") in units)


def _native_can_approve_doc(doc: HrDocument, caller: dict, db: Session) -> bool:
    """Approval authority the caller's ROLE carries on its own — no grants.

    role_change     → admin or shift-manager (the existing rule).
    people_exchange → • to a supervisor: admin OR the RECEIVING supervisor.
                      • to a task:        admin OR a shift-manager of the
                        sending unit's shift.
    """
    role = caller.get("role")
    if role == "admin":
        return True
    if doc.doc_type != "people_exchange":
        return _can_approve(caller)
    payload = doc.payload or {}
    if payload.get("target_type") == "supervisor":
        return role == "supervisor" and caller.get("role_id") == payload.get("target_manager_id")
    # task target → a shift-manager of the sending unit's shift
    if role == "shift-manager":
        shift = _get_shift_for_manager(db, doc.manager_id)
        return _sm_shift(db, caller.get("role_id")) == shift
    return False


def _can_approve_doc(doc: HrDocument, caller: dict, db: Session) -> bool:
    """Approval authority, per document type. One approval is always enough.

    The role's own authority (_native_can_approve_doc), plus, additively,
    anyone granted ``staff.documents.approve``: at "all" scope over every
    document, at "own" scope only over documents inside their normal reach.
    The grant never removes an authority — it is the mechanism for "this
    person handles transfers", without making them an admin.
    """
    return (_native_can_approve_doc(doc, caller, db)
            or _granted_over_doc(doc, caller, db))


def _doc_via_grant(doc: HrDocument, caller: dict, db: Session) -> bool:
    """True when the caller's authority over THIS document exists only through
    the ``staff.documents.approve`` grant — the trigger for the admin warning
    DM (capability_alerts). Evaluate BEFORE mutating: the answer is what
    authorized the action, not what would authorize it afterwards."""
    return (not _native_can_approve_doc(doc, caller, db)
            and _granted_over_doc(doc, caller, db))


def _doc_alert_details(db: Session, doc: HrDocument) -> list:
    # An unknown type falls back to its raw column value rather than a missing
    # translation: the warning DM must still name what it is about.
    kind = (tv("doc." + doc.doc_type)
            if doc.doc_type in _REAL_DOC_TYPES else doc.doc_type)
    details = [("document", kind),
               ("unit", unit_name(db, doc.manager_id)),
               ("date", str(doc.date))]
    emps = [e.get("worker_name") or "?" for e in (doc.payload or {}).get("employees") or []]
    if emps:
        names = ", ".join(emps[:10]) + (f" +{len(emps) - 10}" if len(emps) > 10 else "")
        details.append(("workers", f"{len(emps)}: {names}"))
    return details


def _doc_log_fields(doc: HrDocument, changes: list | None = None) -> dict:
    """Action-log identification of an HR document — `action_log.enrich(**…)`.

    Costs no query: ``supervisor_name`` is the unit's name snapshotted when the
    document was filed and the employees ride in the payload. Call it BEFORE a
    delete — a removed row's attributes cannot be read back after the commit.
    A → supervisor exchange also states the move itself (old_unit → new_unit),
    which is the one fact the status transition alone never carries.
    """
    payload = doc.payload or {}
    emps = [(e.get("worker_name") or "").strip()
            for e in payload.get("employees") or []]
    emps = [e for e in emps if e]
    unit = doc.supervisor_name
    details = [("doc_type", doc.doc_type), ("unit", unit or "—"),
               ("date", str(doc.date))]
    if payload.get("target_type"):
        details.append(("target", _exchange_target_label(payload)))
    if payload.get("new_role"):
        details.append(("role", payload["new_role"]))
    if emps:
        details.append(("workers", f"{len(emps)}: " + ", ".join(emps[:10])
                                   + (f" +{len(emps) - 10}" if len(emps) > 10 else "")))
    chg = list(changes or [])
    if payload.get("target_type") == "supervisor":
        # old_unit → new_unit, as ONE old→new row: `changes` is already that table.
        chg.append(("unit", unit, payload.get("target_manager_name")))
    return {
        "target_kind": "document", "target_id": doc.id,
        "target_name": emps[0] if len(emps) == 1 else None,
        "unit_id": doc.manager_id, "unit_name": unit,
        "day": doc.date, "details": details, "changes": chg,
    }


def _record_history(db: Session, doc: HrDocument, action: str, caller: dict, detail: dict | None = None):
    db.add(HrDocumentHistory(
        document_id=doc.id,
        action=action,
        actor_telegram_id=int(caller["sub"]) if caller.get("sub") else None,
        actor_name=caller.get("full_name", ""),
        detail=detail,
    ))


def _apply_role_change(db: Session, doc: HrDocument):
    """Set job_title → new_role for every employee in the document's date/unit."""
    payload   = doc.payload or {}
    new_role  = payload.get("new_role")
    for emp in payload.get("employees", []):
        att = db.query(Attendance).filter(
            Attendance.manager_id  == doc.manager_id,
            Attendance.date        == doc.date,
            Attendance.worker_name == emp.get("worker_name"),
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        if att:
            att.job_title = new_role


def _revert_role_change(db: Session, doc: HrDocument):
    """Restore each employee's job_title back to its stored old_role."""
    for emp in (doc.payload or {}).get("employees", []):
        att = db.query(Attendance).filter(
            Attendance.manager_id  == doc.manager_id,
            Attendance.date        == doc.date,
            Attendance.worker_name == emp.get("worker_name"),
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        if att:
            att.job_title = emp.get("old_role") or ""


# ── People exchange (worker move) ────────────────────────────────────────────────

# ── Transfer-time split helpers (admin-only people-exchange feature) ─────────────
# When an admin sets a transfer time T on an exchange, each worker's day is split
# so the TOTAL worked hours are conserved:
#   part1 = (T - clock_in)/60        → before-T worked time, INCLUDES early arrival
#   part2 = total_worked - part1     → after-T remainder (the day's lunch/break
#                                       deduction therefore stays inside part2)
# The worker's NAME goes to whichever side is larger (tie → original unit); the
# smaller side keeps only a nameless "hours-only" row (folded into extra_hours).
# Early arrival belongs to the worker's REAL (named) row only:
#   → if the name STAYS (before-T side wins), the original unit keeps the early on
#     the worker's row (effective_hours nets it out, early_arrival_min preserved).
#   → if the name MOVES (after-T side wins), the before-T remainder becomes the
#     nameless leftover and is credited EFFECTIVE hours (part1_eff) — early is
#     dropped, since no real row claims it. The receiving side's early is always 0.
#   → supervisor: the receiving side is the target unit, so part2 lands there.
#   → task:       there is no receiving unit, so part2 is simply DROPPED (the worker
#                 isn't credited for on-task time); only the before-T portion
#                 survives on the sending unit. If the name stays it's her own row
#                 (clock C-T, early kept); if the name leaves she is removed from the
#                 roster and the before-T effective hours become a nameless leftover.
#                 (A → task move with NO transfer time still marks the worker X/0.)
#
# Minimum bar: a MOVED (split) worker must have cleared MIN_MOVED_ZAGRUZKA_HOURS on
# her larger/named side to count as a real worker for any unit. If she reached it on
# neither side (e.g. sent home ~1h into the shift, an hour before-T and nothing/half
# an hour after), she is credited to NO ONE: her name leaves the roster and each side
# becomes a nameless hours-only leftover — the same blanking the "worked more away"
# case already uses, just also triggered when both sides fall short.
MIN_MOVED_ZAGRUZKA_HOURS = 2.0

def _parse_hhmm(s) -> Optional[int]:
    """'08:00' / '8-00' / '08.00' / '17:04 (8.43)' → minutes from midnight, else None.

    Tolerates the verifix clock format which carries a trailing ' (8.43)' worked-
    hours suffix and spaces around the dash (e.g. '07:49 - 17:04 (8.43)')."""
    if not s:
        return None
    txt = str(s).split("(")[0].strip().replace(".", ":").replace("-", ":")
    parts = txt.split(":")
    try:
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 and parts[1] != "" else 0
        return h * 60 + m
    except (ValueError, IndexError):
        return None


def _fmt_hhmm(mins) -> str:
    # Wrap to a wall-clock time so an overnight-normalised minute (e.g. 1478 for a
    # 00:38 clock-out carried past midnight) formats as "00:38", not "24:38".
    mins = int(round(mins)) % 1440
    return f"{mins // 60:02d}:{mins % 60:02d}"


def _schedule_start_min(schedule) -> Optional[int]:
    if not schedule:
        return None
    return _parse_hhmm(str(schedule).split("до")[0])


def _clock_bounds_min(clock_in_out):
    """'08:00-19:47' → (clock_in_min, clock_out_min); (None, None) if unparseable."""
    if not clock_in_out or "-" not in str(clock_in_out):
        return None, None
    left, _, right = str(clock_in_out).strip().partition("-")
    return _parse_hhmm(left), _parse_hhmm(right)


def _normalize_transfer_time(caller: dict, ttype: Optional[str], raw) -> Optional[str]:
    """Honour a transfer time for admins and supervisors, moving to a supervisor
    OR a task. Returns a canonical 'HH:MM' string or None."""
    if not raw or ttype not in ("supervisor", "task"):
        return None
    mins = _parse_hhmm(raw)
    return _fmt_hhmm(mins) if mins is not None else None


def _normalize_return_time(ttype: Optional[str], transfer_time: Optional[str], raw) -> Optional[str]:
    """A return time R is the END of the away stint and is only meaningful when a
    transfer time T is also set (→ supervisor or task). Returns a canonical
    'HH:MM' string or None."""
    if not raw or not transfer_time or ttype not in ("supervisor", "task"):
        return None
    mins = _parse_hhmm(raw)
    return _fmt_hhmm(mins) if mins is not None else None


def _compute_split(snapshot: dict, transfer_time: str, return_time: Optional[str] = None) -> Optional[dict]:
    """Resolve how a single worker's day splits around the transfer time T, and —
    when a return time R is given — the moment they come back (the carve-out).

    TWO-WAY (no return) — the worker leaves at T and never returns:
      part1 = (T - clock_in)/60     → before-T worked time, INCLUDES early arrival
      part2 = total_worked - part1  → after-T remainder

    CARVE-OUT (return time R, C ≤ T ≤ R ≤ O) — the worker is away only for [T, R]
    and ends the day back home, so the two home slices [C,T]+[R,O] are one side:
      away = (R - T)/60             → the away stint, at clock duration
      part1 = total_worked - away   → HOME side (both slices), keeps early + break
      part2 = away                  → AWAY side
    The home named row therefore keeps its full C–O clock (two slices can't be one
    HH:MM range) while the away row, if the name moves, shows T–R.

    Either way the NAME goes to the bigger of part1/part2 (tie → original unit), and
    early arrival is only ever credited to the original unit (receiving side = 0).
    Returns None when the worker can't be split (missing/invalid times or hours) so
    the caller can fall back to a plain full move. Hours are in decimal hours.
    """
    C, O  = _clock_bounds_min(snapshot.get("clock_in_out"))
    T     = _parse_hhmm(transfer_time)
    total = snapshot.get("hours_worked")
    early = float(snapshot.get("early_arrival_min") or 0)
    if T is None or C is None or O is None or total is None:
        return None
    # Overnight shift: a clock-out at/under the clock-in crossed midnight, so carry
    # it (and a post-midnight transfer time) into the next day to keep C ≤ T ≤ O.
    if O <= C:
        O += 1440
    if T < C:
        T += 1440
    if O <= C:                                         # still degenerate → can't split
        return None
    total = float(total)
    T     = max(C, min(T, O))                          # clamp into the worked window

    R = _parse_hhmm(return_time) if return_time else None
    if R is not None:
        # ── Carve-out: the away stint is [T, R]; everything else stays home. ──
        if R < C:                                      # return crossed midnight too
            R += 1440
        R = max(T, min(R, O))                          # clamp into [T, O]
        away  = max(0.0, min((R - T) / 60.0, total))   # away stint at clock duration
        part1 = max(0.0, total - away)                 # home side (both slices), incl. break+early
        part2 = away                                   # away side
        return {
            "T":          _fmt_hhmm(T),
            "C":          _fmt_hhmm(C),
            "O":          _fmt_hhmm(O),
            "R":          _fmt_hhmm(R),
            "stay":       part1 >= part2,              # tie → stays on the original unit
            "part1":      round(part1, 4),             # home-side hours (incl. early)
            "part2":      round(part2, 4),             # away-side hours
            "part1_eff":  round(max(0.0, part1 - early / 60.0), 4),
            "home_clock": f"{_fmt_hhmm(C)}-{_fmt_hhmm(O)}",  # name stays → full C–O span
            "away_clock": f"{_fmt_hhmm(T)}-{_fmt_hhmm(R)}",  # name moves → just the [T,R] stint
            "early_min":  early,
        }

    part1 = max(0.0, min((T - C) / 60.0, total))       # before-T (incl. early), capped at total
    part2 = max(0.0, total - part1)                    # after-T remainder; total conserved
    return {
        "T":         _fmt_hhmm(T),
        "C":         _fmt_hhmm(C),
        "O":         _fmt_hhmm(O),
        "stay":      part1 >= part2,                   # tie → stays on the original unit
        "part1":     round(part1, 4),                  # original-side hours (incl. early)
        "part2":     round(part2, 4),                  # receiving-side hours (early already on orig)
        "part1_eff": round(max(0.0, part1 - early / 60.0), 4),  # original effective (early removed)
        "early_min": early,
    }


def _apply_split_exchange(db: Session, doc: HrDocument):
    """Apply an exchange that carries a transfer time, splitting each worker's day
    at T. For a → supervisor move the task/receiving side is handed to the other
    unit; for a → task move that side is simply dropped. Records what it did back
    into the payload so a later cancel/delete can revert precisely."""
    payload = doc.payload or {}
    is_task = payload.get("target_type") == "task"
    target  = payload.get("target_manager_id")
    ttime   = payload.get("transfer_time")
    rtime   = payload.get("return_time")
    for emp in payload.get("employees", []):
        att = db.query(Attendance).filter(
            Attendance.manager_id  == doc.manager_id,
            Attendance.date        == doc.date,
            Attendance.worker_name == emp.get("worker_name"),
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        if not att:
            continue
        plan = _compute_split(emp.get("snapshot") or {}, ttime, rtime)
        if not plan or (not is_task and not target):
            # Can't split → fall back to a plain full move.
            if is_task:
                att.clock_in_out      = "X"
                att.hours_worked      = 0
                att.effective_hours   = None
                att.early_arrival_min = None
            else:
                att.manager_id = target
                # The worker arrives at the SUPERVISOR, not at one of their
                # cells — the sender is not asked to guess inside somebody
                # else's unit. The receiving supervisor places them on the
                # /staff «Yacheykalar» tab, and their day will not close until
                # they have (`_unplaced_workers`).
                att.verifix_code = None
            emp["applied"] = {"side": "move", "leftover_id": None, "plain": True}
            continue

        leftover_id = None
        # Neither side cleared the minimum worked-hours bar → the worker counts as
        # nobody's worker. Her name leaves the roster and each side becomes a nameless
        # hours-only leftover: her own row is repurposed into the before-T effective
        # leftover on the sending unit (mirrors the → task blanking below, so revert
        # restores it in place by id), and a → supervisor move also drops the after-T
        # hours as a nameless leftover on the receiving unit (→ task drops them).
        if max(plan["part1"], plan["part2"]) < MIN_MOVED_ZAGRUZKA_HOURS:
            recv_leftover_id = None
            if not is_task and target and plan["part2"] > 0:
                # Cell-less like every arrival, and NAMELESS besides — so no
                # placement can ever reach it. `_unplaced_workers` deliberately
                # counts named rows only: the day-close gate must not demand a
                # cell for a row the placement tab cannot show.
                row = Attendance(manager_id=target, date=doc.date, worker_name=None,
                                 hours_worked=plan["part2"],
                                 verifix_code=None)
                db.add(row); db.flush()
                recv_leftover_id = row.id
            att.worker_name       = None
            att.job_title         = None
            att.schedule          = None
            att.clock_in_out      = None
            att.hours_worked      = plan["part1_eff"]
            att.effective_hours   = None
            att.early_arrival_min = None
            emp["applied"] = {"side": "below_min", "leftover_id": att.id,
                              "recv_leftover_id": recv_leftover_id, "task_blanked": True}
            continue
        if plan["stay"]:
            # Home side wins: worker keeps their name on the sending unit. No return
            # → clock-out trimmed to T; carve-out (return) → full C–O span, since the
            # home slices [C,T]+[R,O] can't be one HH:MM range. Early stays here.
            att.clock_in_out    = plan.get("home_clock") or f'{plan["C"]}-{plan["T"]}'
            att.hours_worked    = plan["part1"]
            att.effective_hours = plan["part1_eff"]
            # early_arrival_min unchanged — early belongs to the original unit
            if not is_task and plan["part2"] > 0:
                # → supervisor: the after-T hours land on the receiving unit,
                # cell-less and nameless (see the below-min branch above).
                # → task: dropped (no row).
                row = Attendance(manager_id=target, date=doc.date, worker_name=None,
                                 hours_worked=plan["part2"],
                                 verifix_code=None)
                db.add(row); db.flush()
                leftover_id = row.id
            emp["applied"] = {"side": "stay", "leftover_id": leftover_id}
        else:
            # After-T side wins: the worker's name leaves the sending unit's roster.
            if is_task:
                # → task: part2 is dropped and there is no receiving unit, so she is
                # REMOVED from the roster. Her own row is repurposed into the nameless
                # before-T leftover — blanking the name drops her from the table and
                # folds the hours into extra_hours. Value-only, exactly like a
                # supervisor leftover: effective before-T hours (part1_eff, early
                # stripped), no clock / title / early.
                att.worker_name       = None
                att.job_title         = None
                att.schedule          = None
                att.clock_in_out      = None
                att.hours_worked      = plan["part1_eff"]
                att.effective_hours   = None
                att.early_arrival_min = None
                emp["applied"] = {"side": "move", "leftover_id": att.id, "task_blanked": True}
            else:
                # → supervisor: the row moves to the target with the after-T hours,
                # and the before-T remainder stays as a nameless hours-only row on
                # the sending unit. Credited the EFFECTIVE hours (part1_eff = early
                # stripped): once the name has left, the original unit isn't credited
                # for the worker clocking in before their scheduled start.
                att.manager_id        = target
                att.verifix_code      = None      # placed by the receiving supervisor
                # No return → away runs T–O; carve-out → just the [T,R] stint.
                att.clock_in_out      = plan.get("away_clock") or f'{plan["T"]}-{plan["O"]}'
                att.hours_worked      = plan["part2"]
                att.early_arrival_min = 0          # early stays on the original unit
                att.effective_hours   = plan["part2"]
                if plan["part1_eff"] > 0:
                    # The before-T remainder stays credited to the worker's ORIGINAL cell.
                    row = Attendance(manager_id=doc.manager_id, date=doc.date, worker_name=None,
                                     hours_worked=plan["part1_eff"],
                                     verifix_code=emp.get("old_verifix_code"))
                    db.add(row); db.flush()
                    leftover_id = row.id
                emp["applied"] = {"side": "move", "leftover_id": leftover_id}
    flag_modified(doc, "payload")


def _revert_split_exchange(db: Session, doc: HrDocument):
    """Undo an applied transfer-time split: restore the worker's full row from the
    snapshot and delete the nameless leftover row it created. For a → task move the
    worker's own row was blanked into the leftover, so it is restored in place by id
    (re-attach name + snapshot) rather than deleted."""
    payload = doc.payload or {}
    is_task = payload.get("target_type") == "task"
    target  = payload.get("target_manager_id")
    for emp in payload.get("employees", []):
        applied = emp.get("applied") or {}
        snap    = emp.get("snapshot") or {}
        wname   = emp.get("worker_name")
        side    = applied.get("side", "move")
        if applied.get("task_blanked"):
            # → task move: her own row was blanked into a nameless leftover. Restore
            # it in place by id (re-attach the name + snapshot); never delete it.
            row = db.query(Attendance).filter(Attendance.id == applied.get("leftover_id")).first()
            if row:
                row.worker_name       = wname
                row.manager_id        = emp.get("old_manager_id") or doc.manager_id
                row.job_title         = snap.get("job_title")
                row.schedule          = snap.get("schedule")
                row.clock_in_out      = snap.get("clock_in_out")
                row.hours_worked      = snap.get("hours_worked")
                row.early_arrival_min = snap.get("early_arrival_min")
                row.effective_hours   = snap.get("effective_hours")
                if "old_verifix_code" in emp:
                    row.verifix_code  = emp.get("old_verifix_code")
            # A below-min → supervisor move also parked the after-T hours as a nameless
            # leftover on the receiving unit — drop it as the full row is restored.
            recv_lid = applied.get("recv_leftover_id")
            if recv_lid:
                recv = db.query(Attendance).filter(Attendance.id == recv_lid).first()
                if recv:
                    db.delete(recv)
            emp.pop("applied", None)
            continue
        # For a → supervisor move the full row lives on the target if it moved,
        # else on the sending unit. A → task move never relocates the row.
        cur_mgr = target if (side == "move" and not is_task) else doc.manager_id
        att = db.query(Attendance).filter(
            Attendance.manager_id  == cur_mgr,
            Attendance.date        == doc.date,
            Attendance.worker_name == wname,
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        if att:
            att.manager_id        = emp.get("old_manager_id") or doc.manager_id
            att.job_title         = snap.get("job_title")
            att.schedule          = snap.get("schedule")
            att.clock_in_out      = snap.get("clock_in_out")
            att.hours_worked      = snap.get("hours_worked")
            att.early_arrival_min = snap.get("early_arrival_min")
            att.effective_hours   = snap.get("effective_hours")
            if "old_verifix_code" in emp:
                att.verifix_code  = emp.get("old_verifix_code")
        lid = applied.get("leftover_id")
        if lid:
            row = db.query(Attendance).filter(Attendance.id == lid).first()
            if row:
                db.delete(row)
        emp.pop("applied", None)
    flag_modified(doc, "payload")


def _apply_people_exchange(db: Session, doc: HrDocument):
    """Apply an approved worker move for the document's date.

    → supervisor: reassign the attendance row to the receiving unit (the worker
                  leaves the sender's grid/KPI and is counted as the receiver's).
    → task:       keep the row on the sending supervisor's page but mark the
                  worker "not came" (clock_in_out="X", hours_worked=0). The KPI
                  filter only counts hours_worked > 0, so a marked worker drops
                  out of every calculation while staying visible on the roster.
                  job_title and schedule are preserved; a snapshot in the payload
                  lets a later cancel restore the original came-state.
    """
    payload = doc.payload or {}
    ttype   = payload.get("target_type")
    target  = payload.get("target_manager_id")
    if payload.get("transfer_time") and ((ttype == "supervisor" and target) or ttype == "task"):
        _apply_split_exchange(db, doc)
        return
    for emp in payload.get("employees", []):
        att = db.query(Attendance).filter(
            Attendance.manager_id  == doc.manager_id,
            Attendance.date        == doc.date,
            Attendance.worker_name == emp.get("worker_name"),
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        if not att:
            continue
        if ttype == "supervisor" and target:
            att.manager_id = target
            # Cell-less on arrival, unconditionally — including for a document
            # filed under the old rule with a `target_cell` still in its
            # payload. Honouring that key would drop the worker into a cell the
            # RECEIVING supervisor never chose, which is exactly the guess this
            # change removed; the sender-side `old_verifix_code` is untouched,
            # so a revert still restores the original cell.
            att.verifix_code = None
        else:
            att.clock_in_out      = "X"
            att.hours_worked      = 0
            att.effective_hours   = None
            att.early_arrival_min = None


def _revert_people_exchange(db: Session, doc: HrDocument):
    """Undo an applied worker move (cancel / delete of an approved exchange)."""
    payload = doc.payload or {}
    ttype   = payload.get("target_type")
    target  = payload.get("target_manager_id")
    if payload.get("transfer_time") and ((ttype == "supervisor" and target) or ttype == "task"):
        _revert_split_exchange(db, doc)
        return
    for emp in payload.get("employees", []):
        wname = emp.get("worker_name")
        if ttype == "supervisor" and target:
            att = db.query(Attendance).filter(
                Attendance.manager_id  == target,
                Attendance.date        == doc.date,
                Attendance.worker_name == wname,
                Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
            ).first()
            if att:
                att.manager_id = emp.get("old_manager_id") or doc.manager_id
                if "old_verifix_code" in emp:
                    att.verifix_code = emp.get("old_verifix_code")
        else:
            # Restore the worker's original came-state from the snapshot.
            snap = emp.get("snapshot") or {}
            att = db.query(Attendance).filter(
                Attendance.manager_id  == doc.manager_id,
                Attendance.date        == doc.date,
                Attendance.worker_name == wname,
                Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
            ).first()
            if att:
                att.job_title         = snap.get("job_title")
                att.schedule          = snap.get("schedule")
                att.clock_in_out      = snap.get("clock_in_out")
                att.hours_worked      = snap.get("hours_worked")
                att.early_arrival_min = snap.get("early_arrival_min")
                att.effective_hours   = snap.get("effective_hours")
            else:
                # robustness: if the row is gone, recreate it from the snapshot
                db.add(Attendance(
                    manager_id        = doc.manager_id,
                    date              = doc.date,
                    worker_name       = wname,
                    job_title         = snap.get("job_title"),
                    schedule          = snap.get("schedule"),
                    clock_in_out      = snap.get("clock_in_out"),
                    hours_worked      = snap.get("hours_worked"),
                    early_arrival_min = snap.get("early_arrival_min"),
                    effective_hours   = snap.get("effective_hours"),
                    verifix_code      = emp.get("old_verifix_code"),
                ))


def _resolve_exchange_target(db: Session, sender_id: int, d: date, ttype: Optional[str],
                             target_manager_id_in: Optional[int], task_name_in: Optional[str]):
    """Validate the move target; returns (ttype, target_manager_id,
    target_manager_name, task_name). Enforces: real target unit, not the sender,
    and the receiving unit's day must still be open.

    It no longer resolves a destination CELL (2026-08-30). The sender was being
    asked which of the RECEIVING unit's cells the worker would land in — a guess
    about somebody else's shopfloor, made before the shift had run. The worker
    now arrives cell-less and the receiving supervisor places them on the /staff
    «Yacheykalar» tab, where the day-close gate makes it unskippable."""
    if ttype not in ("supervisor", "task"):
        raise HTTPException(status_code=400, detail="target_type must be 'supervisor' or 'task'")
    if ttype == "supervisor":
        if not target_manager_id_in:
            raise HTTPException(status_code=400, detail="target_manager_id is required")
        if target_manager_id_in == sender_id:
            raise HTTPException(status_code=400, detail="Cannot exchange workers to the same unit")
        target = db.query(Manager).filter_by(id=target_manager_id_in).first()
        if not target:
            raise HTTPException(status_code=404, detail="Target supervisor not found")
        _assert_day_open(db, target.id, d)   # can't move into a closed unit
        if not _unit_has_attendance(db, target.id, d):
            raise ExchangeTargetNoData()     # can't move into a unit with no data yet
        return ttype, target.id, target.name, None
    name = (task_name_in or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="task_name is required")
    return ttype, None, None, name


def _build_exchange_payload(db: Session, manager_id: int, d: date, target_type: str,
                            target_manager_id: Optional[int], target_manager_name: Optional[str],
                            task_name: Optional[str], employees: List[str],
                            transfer_time: Optional[str] = None, return_time: Optional[str] = None):
    emp_rows = []
    for wname in employees:
        att = db.query(Attendance).filter(
            Attendance.manager_id  == manager_id,
            Attendance.date        == d,
            Attendance.worker_name == wname,
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        if not att:
            continue   # "worker must be present" — silently drop rows with no record
        row = {
            "worker_name":    wname,
            "old_manager_id": manager_id,
            "old_role":       att.job_title or "",
            # The worker's cell BEFORE the move — restored on cancel/revert.
            "old_verifix_code": att.verifix_code,
        }
        # A full snapshot lets a later cancel restore the original row. Needed for
        # task moves (which blank the row) and for transfer-time splits (which
        # mutate clock-out / hours and may relocate the row to the receiver).
        if target_type == "task" or transfer_time:
            row["snapshot"] = {
                "job_title":         att.job_title,
                "schedule":          att.schedule,
                "clock_in_out":      att.clock_in_out,
                "hours_worked":      float(att.hours_worked)      if att.hours_worked      is not None else None,
                "early_arrival_min": float(att.early_arrival_min) if att.early_arrival_min is not None else None,
                "effective_hours":   float(att.effective_hours)   if att.effective_hours   is not None else None,
            }
        emp_rows.append(row)
    return {
        "target_type":         target_type,
        "target_manager_id":   target_manager_id,
        "target_manager_name": target_manager_name,
        "task_name":           task_name,
        "transfer_time":       transfer_time,
        "return_time":         return_time,
        "employees":           emp_rows,
    }


def _exchange_target_label(payload: dict) -> str:
    """Who the workers are moving to, for a notification.

    The receiving SUPERVISOR and nothing else: a move no longer names a
    destination cell, because the receiving supervisor chooses it after the
    shift has actually run."""
    if (payload or {}).get("target_type") == "supervisor":
        return payload.get("target_manager_name") or "—"
    return (payload or {}).get("task_name") or "—"


def _notify_exchange(db: Session, doc: HrDocument, event: str, actor_tg_id: int, admin_dm: bool = True):
    """Notify the parties for a worker-exchange action. _notify_all_parties covers
    the sending unit's admins/shift-managers/supervisor; the receiving supervisor
    sits in another unit, so notify them separately."""
    payload   = doc.payload or {}
    nkey_map  = {
        "created":   "worker_exchange_created",
        "approved":  "worker_exchange_approved",
        "cancelled": "worker_exchange_cancelled",
    }
    nkey   = nkey_map.get(event, "worker_exchange_created")
    params = {
        "actor_name": doc.created_by_name or "",
        "count":      len(payload.get("employees", [])),
        "target":     _exchange_target_label(payload),
        "date":       doc.date,
    }
    _notify_all_parties(db, doc.manager_id, nkey, params, ntype="info",
                        actor_tg_id=actor_tg_id, include_supervisor=True, admin_dm=admin_dm)
    if payload.get("target_type") == "supervisor" and payload.get("target_manager_id"):
        sup = _find_supervisor(db, payload["target_manager_id"])
        if sup and sup.telegram_id != actor_tg_id:
            # On creation the receiving supervisor also gets a rich inline
            # approve/reject message (app.approvals.send_hr_document_to_admins),
            # so skip the duplicate plain DM here — keep only the in-app bell.
            # approved/cancelled events carry no inline message, so DM as usual.
            _notify(db, sup.telegram_id, type="info", dm=event != "created",
                    nkey=nkey, params=params,
                    profile=_profile_key("supervisor", sup.role_id))


def _serialize_doc(doc: HrDocument, mgr_name: str | None = None, detailed: bool = False):
    payload   = doc.payload or {}
    employees = payload.get("employees", [])
    out = {
        "id":               doc.id,
        "doc_type":         doc.doc_type,
        "doc_type_label":   DOC_TYPE_LABELS.get(doc.doc_type, doc.doc_type),
        "manager_id":       doc.manager_id,
        "supervisor_name":  doc.supervisor_name or mgr_name,
        "date":             doc.date.isoformat() if doc.date else None,
        "status":           doc.status,                       # draft | approved | rejected
        "approved":         doc.status == "approved",         # → Да / Нет
        "new_role":         payload.get("new_role"),
        "target_type":          payload.get("target_type"),
        "target_manager_id":    payload.get("target_manager_id"),
        "target_manager_name":  payload.get("target_manager_name"),
        "task_name":            payload.get("task_name"),
        "transfer_time":        payload.get("transfer_time"),
        "return_time":          payload.get("return_time"),
        "employee_count":   len(employees),
        "created_by_telegram_id": doc.created_by_telegram_id,
        "created_by_name":  doc.created_by_name,
        "approved_by_name": doc.approved_by_name,
        "created_at":       doc.created_at.isoformat() if doc.created_at else None,
        "approved_at":      doc.approved_at.isoformat() if doc.approved_at else None,
    }
    if detailed:
        out["employees"] = employees
        out["payload"]   = payload
    return out


def _resolve_manager(caller, db: Session, manager_id: Optional[int]):
    """Determine which manager (unit) a document belongs to + its display name."""
    role = caller.get("role")
    if role == "supervisor":
        mid = caller.get("role_id")
        if not mid:
            raise HTTPException(status_code=400, detail="Supervisor has no linked manager")
    else:
        if not manager_id:
            raise HTTPException(status_code=400, detail="manager_id required")
        mid = manager_id
    mgr = db.query(Manager).filter_by(id=mid).first()
    return mid, (mgr.name if mgr else None)


# ── People-exchange option sources (targets + tasks) ─────────────────────────────

@router.get("/exchange-targets")
def exchange_targets(attend_date: str, manager_id: Optional[int] = None,
                     caller=Depends(_require_staff), db: Session = Depends(get_db)):
    """Supervisors a worker exchange may move INTO for a date — every unit except
    the sender, excluding any unit that has already closed that day or has no
    attendance data uploaded for it yet (rows moved into a data-less unit would
    be destroyed by that unit's eventual verifix upload).

    A target is a SUPERVISOR and nothing more. It used to carry that unit's cell
    codes so the sender could pick a destination cell; the receiving supervisor
    places the worker themselves now (/staff «Yacheykalar»), so the sender is no
    longer asked to guess inside another unit."""
    if caller.get("role") not in ("admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Admin or supervisor only")
    d = date.fromisoformat(attend_date)
    sender_id = caller.get("role_id") if caller.get("role") == "supervisor" else manager_id
    closed = {
        c.manager_id for c in db.query(DayApproval).filter(DayApproval.date == d).all()
    }
    has_data = {
        mid for (mid,) in db.query(Attendance.manager_id).filter(
            Attendance.date == d,
            Attendance.worker_name.isnot(None),
            Attendance.worker_name.notin_(["", "nan", "NaN"]),
        ).distinct().all()
    }
    out = []
    for m in db.query(Manager).filter(Manager.archived.is_(False)).order_by(Manager.shift, Manager.name).all():
        if m.id == sender_id or m.id in closed or m.id not in has_data:
            continue
        out.append({"manager_id": m.id, "full_name": m.name, "shift": m.shift})
    return out


def _ensure_exchange_task(db: Session, name: Optional[str], caller: dict) -> None:
    """Persist a task name to the permanent shared list (create, or reactivate a
    previously removed one). Called whenever a people-exchange targets a task, so
    the '＋ Yangi vazifa' name an admin types sticks around for everyone on every
    date. No-op for blank names. Caller commits.

    Adding a task to the shared list is admin-only: supervisors may target an
    existing, active task but cannot introduce a new one (nor revive a removed
    one). Referencing an already-active task is a no-op, so it stays open to all."""
    n = (name or "").strip()
    if not n:
        return
    t = db.query(ExchangeTask).filter(ExchangeTask.name == n).first()
    if t is not None and t.active:
        return
    if caller.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only an admin can add a new task")
    if t is None:
        db.add(ExchangeTask(name=n, active=True, created_by_telegram_id=int(caller["sub"])))
    else:
        t.active = True


@router.get("/tasks")
def list_exchange_tasks(attend_date: Optional[str] = None,
                        caller=Depends(_require_staff), db: Session = Depends(get_db)):
    """Permanent, shared list of worker-exchange task names. Tasks persist across
    every date until an admin removes them; any supervisor or admin who creates
    one makes it an option for the rest. (attend_date is accepted but ignored —
    kept for backward compatibility with older clients.)"""
    names = [
        t.name for t in db.query(ExchangeTask)
        .filter(ExchangeTask.active.is_(True))
        .order_by(func.lower(ExchangeTask.name))
        .all()
    ]
    return {"tasks": names}


class TaskDeleteBody(BaseModel):
    name: str


@router.post("/tasks/delete")
def delete_exchange_task(body: TaskDeleteBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    """Admin-only soft removal of a task from the shared picker. The row is kept
    (active=False) so existing exchange documents that reference the name keep
    resolving — it simply stops being offered for new exchanges."""
    if caller.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    t = db.query(ExchangeTask).filter(ExchangeTask.name == name).first()
    if not t or not t.active:
        raise HTTPException(status_code=404, detail="Task not found")
    task_id = t.id
    t.active = False
    db.commit()
    action_log.enrich(
        target_kind="task", target_id=task_id, target_name=name,
        details=[("task", name)],
        changes=[("enabled", True, False)],
    )
    return {"ok": True}


# ── List / Get ─────────────────────────────────────────────────────────────────

@router.get("/documents")
def list_documents(caller=Depends(_require_staff), db: Session = Depends(get_db)):
    rows = _scope_documents(_real_docs(db), caller, db) \
        .order_by(HrDocument.created_at.desc()).all()
    mgr_names = {m.id: m.name for m in db.query(Manager).all()}
    docs = [_serialize_doc(d, mgr_names.get(d.manager_id)) for d in rows]

    # Deletion EditRequests — group by batch_id, appear alongside role-change documents
    del_rows = _scope_deletion_requests(caller, db)

    # Group by batch_id (or solo-{id} for legacy/null batch_id requests)
    batch_map: dict = defaultdict(list)
    for r in del_rows:
        key = r.batch_id if r.batch_id else f"solo-{r.id}"
        batch_map[key].append(r)

    del_items = []
    for batch_key, reqs in batch_map.items():
        reqs.sort(key=lambda r: r.id)
        first       = reqs[0]
        has_pending = any(r.status == "pending" for r in reqs)
        # "undone" counts as applied for history purposes — the deletion WAS
        # approved, then the worker was restored. A batch is 'approved' when
        # at least one row was applied; fully-rejected batches show 'rejected'.
        any_applied  = any(r.status in ("approved", "undone") for r in reqs)
        batch_status = "pending" if has_pending else ("approved" if any_applied else "rejected")
        processed_by = next((r.processed_by_name for r in reqs if r.processed_by_name), None)

        # Use earliest created_at as the sort key for the batch
        earliest_created = min(
            (r.created_at for r in reqs if r.created_at),
            default=None,
        )

        del_items.append({
            "id":               first.id,
            "batch_id":         first.batch_id,
            "_source":          "deletion",
            "doc_type":         "deletion",
            "doc_type_label":   "Deletion request",
            "manager_id":       first.manager_id,
            "manager_name":     mgr_names.get(first.manager_id, "—"),
            "supervisor_name":  first.supervisor_name or mgr_names.get(first.manager_id, "—"),
            "supervisor_telegram_id": first.supervisor_telegram_id,
            "date":             first.date.isoformat(),
            "status":           batch_status,
            "approved":         batch_status == "approved",
            "new_role":         None,
            "employee_count":   len(reqs),
            "created_by_name":  first.supervisor_name,
            "approved_by_name": processed_by if batch_status != "pending" else None,
            "created_at":       earliest_created.isoformat() if earliest_created else first.date.isoformat(),
            "workers": [
                {
                    "id":              r.id,
                    "worker_name":     r.worker_name,
                    "status":          r.status,
                    "approved_by_name": r.processed_by_name,
                    "original":        r.original or {},
                }
                for r in reqs
            ],
        })

    combined = docs + del_items
    combined.sort(key=lambda x: x.get("created_at") or x.get("date") or "", reverse=True)
    return combined


@router.get("/documents/pending-count")
def documents_pending_count(caller=Depends(_require_staff), db: Session = Depends(get_db)):
    # Must match the "Requests" tab badge, which counts BOTH pending role-change
    # documents AND pending deletion-request batches from /documents. Counting
    # only HrDocument drafts (the old behaviour) ignored deletion requests, so
    # the sidebar badge was smaller than the tab badge.

    # 1) Pending role-change documents (HrDocument drafts)
    doc_count = _scope_documents(_real_docs(db), caller, db) \
        .filter(HrDocument.status == "draft").count()

    # 2) Pending deletion-request batches — grouped by batch_id exactly like
    #    /documents, counted once per batch that has any pending request.
    pending_batches = set()
    for r in _scope_deletion_requests(caller, db):
        if r.status == "pending":
            pending_batches.add(r.batch_id if r.batch_id else f"solo-{r.id}")

    return {"count": doc_count + len(pending_batches)}


@router.get("/documents/{doc_id}")
def get_document(doc_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    doc = _scope_documents(_real_docs(db), caller, db).filter(HrDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    mgr = db.query(Manager).filter_by(id=doc.manager_id).first()
    return _serialize_doc(doc, mgr.name if mgr else None, detailed=True)


@router.get("/documents/{doc_id}/history")
def document_history(doc_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    doc = _scope_documents(_real_docs(db), caller, db).filter(HrDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    rows = db.query(HrDocumentHistory).filter_by(document_id=doc_id) \
        .order_by(HrDocumentHistory.created_at.asc()).all()
    return [{
        "id":         h.id,
        "action":     h.action,
        "actor_name": h.actor_name,
        "detail":     h.detail,
        "created_at": h.created_at.isoformat() if h.created_at else None,
    } for h in rows]


# ── Create / Update ─────────────────────────────────────────────────────────────

class DocCreateBody(BaseModel):
    doc_type:    str            = "role_change"
    attend_date: str
    manager_id:  Optional[int]  = None      # required for admin (sending unit)
    employees:   List[str]                  # worker_name list
    # role_change
    new_role:    Optional[str]  = None
    # people_exchange
    target_type:       Optional[str] = None   # "supervisor" | "task"
    target_manager_id: Optional[int] = None
    task_name:         Optional[str] = None
    transfer_time:     Optional[str] = None   # "HH:MM" — split (→ supervisor or task)
    return_time:       Optional[str] = None   # "HH:MM" — carve-out end (the away stint is [T,R])


def _build_role_payload(db: Session, manager_id: int, d: date, new_role: str, employees: List[str]):
    emp_rows = []
    for wname in employees:
        att = db.query(Attendance).filter(
            Attendance.manager_id  == manager_id,
            Attendance.date        == d,
            Attendance.worker_name == wname,
            Attendance.split_of.is_(None),   # the PRIMARY half of a split is canonical
        ).first()
        emp_rows.append({
            "worker_name": wname,
            "old_role":    (att.job_title if att else "") or "",
        })
    return {"new_role": new_role, "employees": emp_rows}


@router.post("/documents", status_code=201)
def create_document(body: DocCreateBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    if caller.get("role") not in ("admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Not allowed to create documents")
    if body.doc_type not in ("role_change", "people_exchange"):
        raise HTTPException(status_code=400, detail="Unsupported document type")
    if not body.employees:
        raise HTTPException(status_code=400, detail="Select at least one employee")

    d = date.fromisoformat(body.attend_date)
    manager_id, mgr_name = _resolve_manager(caller, db, body.manager_id)

    if body.doc_type == "people_exchange":
        return _create_people_exchange(db, caller, body, d, manager_id, mgr_name)

    # ── role_change ──
    if not body.new_role:
        raise HTTPException(status_code=400, detail="new_role is required")
    if not is_assignable_target_role(body.new_role):
        raise HTTPException(status_code=400, detail="This role can only be set from verifix files and cannot be chosen as a role-change target")
    if caller.get("role") == "supervisor":
        _assert_day_open(db, manager_id, d)
    payload = _build_role_payload(db, manager_id, d, body.new_role, body.employees)

    doc = HrDocument(
        doc_type="role_change",
        manager_id=manager_id,
        supervisor_name=mgr_name,
        date=d,
        payload=payload,
        status="draft",
        created_by_telegram_id=int(caller["sub"]),
        created_by_name=caller.get("full_name", ""),
        created_by_role=caller.get("role"),
    )
    db.add(doc)
    db.flush()
    _record_history(db, doc, "created", caller, {
        "new_role": body.new_role, "employee_count": len(payload["employees"]),
    })
    # An ADMIN's own filing IS the approval. There is nobody above them to ask,
    # so a draft waiting on its author's second tap is a step that decides
    # nothing — and Ghost Mode auto-approved for exactly that reason already.
    # The two branches differ ONLY in whether anyone is told: Ghost Mode is
    # silent by definition (notify + broadcast are gated), an ordinary admin
    # filing is announced as a DONE deed. No approve/reject card follows either
    # way, because there is nothing left for anyone to decide.
    ghost = notifications_suppressed()
    if ghost or caller.get("role") == "admin":
        _approve_doc(doc, caller, db)
        log = _doc_log_fields(doc, [("status", None, "approved")])
        if not ghost:
            # include_supervisor: the roles are already changed on their unit's
            # day, so the brigadir hears about it — the draft path leaves them
            # out because there is still a decision pending. admin_dm stays on:
            # the rich approve/reject message this create path normally defers
            # to is never sent on this branch.
            _notify_all_parties(
                db, manager_id,
                "role_change_approved",
                {"actor_name": caller.get("full_name", ""), "count": len(payload["employees"]),
                 "new_role": body.new_role, "date": body.attend_date},
                ntype="info",
                actor_tg_id=int(caller["sub"]),
            )
        db.commit()
        action_log.enrich(**log)
        return {"id": doc.id, "status": doc.status}
    log = _doc_log_fields(doc, [("status", None, "draft")])
    _notify_all_parties(
        db, manager_id,
        "new_role_change",
        {"actor_name": caller.get("full_name", ""), "count": len(payload["employees"]),
         "new_role": body.new_role, "date": body.attend_date},
        ntype="info",
        actor_tg_id=int(caller["sub"]),
        include_supervisor=False,
        admin_dm=False,            # admins get the rich approve/reject message instead
    )
    db.commit()
    action_log.enrich(**log)
    try:
        from app.approvals import send_hr_document_to_admins
        send_hr_document_to_admins(db, doc)
    except Exception:
        pass
    return {"id": doc.id}


def _create_people_exchange(db: Session, caller: dict, body: "DocCreateBody",
                            d: date, manager_id: int, mgr_name: Optional[str]):
    # The sending unit's day must still be open
    _assert_day_open(db, manager_id, d)
    ttype, tgt_id, tgt_name, task_name = _resolve_exchange_target(
        db, manager_id, d, body.target_type, body.target_manager_id, body.task_name,
    )
    ttime = _normalize_transfer_time(caller, ttype, body.transfer_time)
    rtime = _normalize_return_time(ttype, ttime, body.return_time)
    payload = _build_exchange_payload(db, manager_id, d, ttype, tgt_id, tgt_name, task_name,
                                      body.employees, transfer_time=ttime, return_time=rtime)
    if not payload["employees"]:
        raise HTTPException(status_code=400, detail="None of the selected workers have a record on this date")
    if ttype == "task":
        _ensure_exchange_task(db, task_name, caller)

    doc = HrDocument(
        doc_type="people_exchange",
        manager_id=manager_id,
        supervisor_name=mgr_name,
        date=d,
        payload=payload,
        status="draft",
        created_by_telegram_id=int(caller["sub"]),
        created_by_name=caller.get("full_name", ""),
        created_by_role=caller.get("role"),
    )
    db.add(doc)
    db.flush()
    _record_history(db, doc, "created", caller, {
        "target": _exchange_target_label(payload), "employee_count": len(payload["employees"]),
    })
    # An admin's own filing IS the approval, exactly as in create_document
    # above; Ghost Mode auto-approves for anyone and stays silent.
    ghost = notifications_suppressed()
    if ghost or caller.get("role") == "admin":
        _approve_doc(doc, caller, db)
        log = _doc_log_fields(doc, [("status", None, "approved")])
        if not ghost:
            # "approved", not "created": the parties are told what happened, not
            # asked to decide it. admin_dm defaults on — the rich approve/reject
            # message the draft path defers to is not sent on this branch, so
            # withholding the plain DM would leave admins with a bell row only.
            _notify_exchange(db, doc, "approved", int(caller["sub"]))
        db.commit()
        action_log.enrich(**log)
        return {"id": doc.id, "status": doc.status}
    log = _doc_log_fields(doc, [("status", None, "draft")])
    _notify_exchange(db, doc, "created", int(caller["sub"]), admin_dm=False)
    db.commit()
    action_log.enrich(**log)
    try:
        from app.approvals import send_hr_document_to_admins
        send_hr_document_to_admins(db, doc)
    except Exception:
        pass
    return {"id": doc.id}


class DocUpdateBody(BaseModel):
    employees: List[str]
    new_role:  Optional[str] = None           # role_change
    target_type:       Optional[str] = None   # people_exchange
    target_manager_id: Optional[int] = None
    task_name:         Optional[str] = None
    transfer_time:     Optional[str] = None
    return_time:       Optional[str] = None


def _is_doc_creator(doc: HrDocument, caller: dict) -> bool:
    """Was this document created by the caller's PROFILE?

    A document belongs to the unit that filed it, and the unit IS the
    supervisor's profile — so every account working as that brigadir may edit,
    withdraw or reject its own unit's draft. Comparing Telegram accounts made
    two people running one unit unable to touch each other's drafts even though
    the org model says they are the same person. Non-supervisor callers keep the
    account comparison (their profile is not the unit).
    """
    if (caller.get("role") == "supervisor" and caller.get("role_id")
            and doc.manager_id == caller["role_id"]):
        return True
    return doc.created_by_telegram_id == int(caller["sub"])


@router.put("/documents/{doc_id}")
def update_document(doc_id: int, body: DocUpdateBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    doc = _scope_documents(_real_docs(db), caller, db).filter(HrDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if doc.status != "draft":
        raise HTTPException(status_code=409, detail="Only draft (Нет) documents can be edited")
    is_creator = _is_doc_creator(doc, caller)
    if caller.get("role") not in ("admin", "shift-manager") and not is_creator:
        raise HTTPException(status_code=403, detail="Not allowed to edit this document")
    if not body.employees:
        raise HTTPException(status_code=400, detail="Select at least one employee")

    if doc.doc_type == "people_exchange":
        _assert_day_open(db, doc.manager_id, doc.date)   # sender's day must be open
        prev    = doc.payload or {}
        ttype   = body.target_type or prev.get("target_type")
        tgt_in  = body.target_manager_id if body.target_manager_id is not None else prev.get("target_manager_id")
        task_in = body.task_name if body.task_name is not None else prev.get("task_name")
        ttype, tgt_id, tgt_name, task_name = _resolve_exchange_target(
            db, doc.manager_id, doc.date, ttype, tgt_in, task_in)
        ttime_in = body.transfer_time if body.transfer_time is not None else prev.get("transfer_time")
        ttime    = _normalize_transfer_time(caller, ttype, ttime_in)
        rtime_in = body.return_time if body.return_time is not None else prev.get("return_time")
        rtime    = _normalize_return_time(ttype, ttime, rtime_in)
        payload = _build_exchange_payload(db, doc.manager_id, doc.date, ttype, tgt_id, tgt_name, task_name,
                                          body.employees, transfer_time=ttime, return_time=rtime)
        if not payload["employees"]:
            raise HTTPException(status_code=400, detail="None of the selected workers have a record on this date")
        if ttype == "task":
            _ensure_exchange_task(db, task_name, caller)
        doc.payload = payload
        _record_history(db, doc, "edited", caller, {
            "target": _exchange_target_label(payload), "employee_count": len(payload["employees"]),
        })
        log = _doc_log_fields(doc)
        db.commit()
        action_log.enrich(**log)
        return {"ok": True}

    # ── role_change ──
    if caller.get("role") == "supervisor":
        _assert_day_open(db, doc.manager_id, doc.date)
    if not body.new_role:
        raise HTTPException(status_code=400, detail="new_role is required")
    if not is_assignable_target_role(body.new_role):
        raise HTTPException(status_code=400, detail="This role can only be set from verifix files and cannot be chosen as a role-change target")
    doc.payload = _build_role_payload(db, doc.manager_id, doc.date, body.new_role, body.employees)
    _record_history(db, doc, "edited", caller, {
        "new_role": body.new_role, "employee_count": len(body.employees),
    })
    log = _doc_log_fields(doc)
    db.commit()
    action_log.enrich(**log)
    return {"ok": True}


# ── Approve (Провести) / Cancel (Отменить) / Delete (Удалить) ────────────────────

def _apply_doc_effects(db: Session, doc: HrDocument):
    if doc.doc_type == "role_change":
        _apply_role_change(db, doc)
    elif doc.doc_type == "people_exchange":
        _apply_people_exchange(db, doc)


def _revert_doc_effects(db: Session, doc: HrDocument):
    if doc.doc_type == "role_change":
        _revert_role_change(db, doc)
    elif doc.doc_type == "people_exchange":
        _revert_people_exchange(db, doc)


def reapply_task_exchanges(db: Session, manager_id: int, d: date) -> int:
    """Re-apply approved → task people-exchange effects after a verifix re-upload
    (admin.upload_verifix) wiped and re-inserted a unit's attendance for the date.

    A re-upload restores every worker's full row while the approved exchange docs
    stay in place, so a worker sent to a task would otherwise reappear with full
    hours + the task pill and get re-counted (zagruzka, «came», etc.) — exactly the
    thing that move was meant to remove. A → task effect lives entirely inside the
    sending unit (part2 is dropped, no receiving row), so re-applying it over the
    fresh rows is idempotent and safe: a supervisor-majority split is re-trimmed to
    its before-T hours, a task-majority worker is re-blanked off the roster, and a
    plain full move is re-marked X/0. → supervisor moves are intentionally skipped —
    they touch a second unit that this upload did not wipe, so re-applying them would
    duplicate the receiving-side rows; the receiver's own re-upload handles its side.
    Returns the number of docs re-applied."""
    docs = db.query(HrDocument).filter(
        HrDocument.doc_type   == "people_exchange",
        HrDocument.manager_id == manager_id,
        HrDocument.date       == d,
        HrDocument.status     == "approved",
    ).all()
    applied = 0
    for doc in docs:
        if (doc.payload or {}).get("target_type") != "task":
            continue
        _apply_people_exchange(db, doc)
        applied += 1
    return applied


# How old a document's DATE may be and still be posted. Approving a stale draft
# is not a paperwork detail: it APPLIES the document, rewriting a day whose
# attendance was uploaded, confirmed and counted weeks ago — and every
# lost-worker incident traced on 2026-08-22 came from exactly that shape, an
# exchange landing on a unit-day long after the upload that built it.
#
# On 2026-08-22 a backlog of June drafts was posted in one burst, silently
# moving workers across June days. This is the stop: one bound, checked in
# `_approve_doc`, so EVERY door inherits it — the API, the bulk action and the
# Telegram ✅ button — instead of three places that have to agree.
#
# A draft older than this is not deleted and not rejected; it simply cannot be
# posted any more. Raise the number here if a genuinely old correction has to
# go through.
STALE_APPROVE_DAYS = 14


class StaleDocument(HTTPException):
    """A document whose date is too far in the past to post."""

    def __init__(self, doc_date, age_days: int):
        super().__init__(status_code=409, detail={
            "code": "doc_too_old",
            "date": doc_date.isoformat() if doc_date else None,
            "age_days": age_days,
            "max_age_days": STALE_APPROVE_DAYS,
            "message": (f"Hujjat sanasi {age_days} kun oldin "
                        f"({doc_date}). {STALE_APPROVE_DAYS} kundan eski hujjatni "
                        f"tasdiqlab bo'lmaydi — u o'sha kungi davomatni qayta yozadi."),
        })


def _approve_doc(doc: HrDocument, caller: dict, db: Session):
    if doc.status == "approved":
        return
    if doc.status == "rejected":
        raise HTTPException(status_code=409, detail="Rejected documents cannot be posted")
    # Stale drafts cannot be posted — see STALE_APPROVE_DAYS. Checked before any
    # effect is applied, so a refused document is left exactly as it was.
    if doc.date:
        age = (date.today() - doc.date).days
        if age > STALE_APPROVE_DAYS:
            logger.warning("DOC-STALE refused approve of #%s (%s, %s days old) by %s",
                           doc.id, doc.date, age, caller.get("full_name") or caller.get("sub"))
            raise StaleDocument(doc.date, age)
    # A → supervisor exchange must not land in a unit whose verifix data for the
    # date isn't uploaded yet — that unit's eventual upload would wipe the
    # transferred rows. Creation already enforces this; re-check here for drafts
    # created before the guard existed (or if the target's rows were deleted).
    payload = doc.payload or {}
    if (doc.doc_type == "people_exchange"
            and payload.get("target_type") == "supervisor"
            and payload.get("target_manager_id")
            and not _unit_has_attendance(db, payload["target_manager_id"], doc.date)):
        raise ExchangeTargetNoData()
    _apply_doc_effects(db, doc)
    doc.status                  = "approved"
    doc.approved_by_telegram_id = int(caller["sub"])
    doc.approved_by_name        = caller.get("full_name", "")
    doc.approved_at             = datetime.now(timezone.utc)
    _record_history(db, doc, "approved", caller)


def _cancel_doc(doc: HrDocument, caller: dict, db: Session):
    if doc.status != "approved":
        return
    _revert_doc_effects(db, doc)
    doc.status                  = "draft"
    doc.approved_by_telegram_id = None
    doc.approved_by_name        = None
    doc.approved_at             = None
    _record_history(db, doc, "cancelled", caller)


def _reject_document(doc: HrDocument, caller: dict, db: Session):
    """Reject a *draft* HR document — keep the row as a rejected record (the
    processor lands in the approved_by_* columns, exactly like deletion batches
    surface their processor) and notify its creator. This is the Telegram/app
    counterpart of approving (posting) a draft. Approved documents cannot be
    rejected; cancel or delete them instead."""
    if doc.status != "draft":
        raise HTTPException(status_code=409, detail="Only draft documents can be rejected")
    doc.status                  = "rejected"
    doc.approved_by_telegram_id = int(caller["sub"])
    doc.approved_by_name        = caller.get("full_name", "")
    doc.approved_at             = datetime.now(timezone.utc)
    _record_history(db, doc, "rejected", caller)
    if doc.created_by_telegram_id and doc.created_by_telegram_id != int(caller["sub"]):
        _notify(db, doc.created_by_telegram_id, type="error",
                nkey="document_rejected",
                params={
                    "actor_name": caller.get("full_name", ""),
                    "doc_type":   doc.doc_type,
                    "date":       doc.date,
                })


def _may_reject_doc(doc: HrDocument, caller: dict, db: Session) -> bool:
    """Who may reject a draft: anyone with approval authority for it, an
    admin/shift-manager (who could always delete the draft), or its creator
    (rejecting your own draft = withdrawing it)."""
    return (_can_approve_doc(doc, caller, db)
            or caller.get("role") in ("admin", "shift-manager")
            or _is_doc_creator(doc, caller))


def _doc_reject_via_grant(doc: HrDocument, caller: dict, db: Session) -> bool:
    """_doc_via_grant for the wider reject/withdraw authority: rejecting your
    own draft or acting as admin/shift-manager is native, so only a pure
    grantee trips the warning."""
    return (_granted_over_doc(doc, caller, db)
            and not _native_can_approve_doc(doc, caller, db)
            and caller.get("role") not in ("admin", "shift-manager")
            and not _is_doc_creator(doc, caller))


@router.post("/documents/{doc_id}/approve")
def approve_document(doc_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    doc = _scope_documents(_real_docs(db), caller, db).filter(HrDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if not _can_approve_doc(doc, caller, db):
        raise HTTPException(status_code=403, detail="Not authorised to post this document")
    if doc.status == "approved":
        # Idempotent, and deliberately silent: _approve_doc would return without
        # touching anything, so a second press must not re-DM the parties, warn
        # about a grant use that changed nothing, or re-stamp the inline admin
        # card with the name of someone who did not approve it.
        return {"ok": True, "status": doc.status}
    via_grant = _doc_via_grant(doc, caller, db)
    _approve_doc(doc, caller, db)
    if doc.doc_type == "people_exchange":
        _notify_exchange(db, doc, "approved", int(caller["sub"]))
    log = _doc_log_fields(doc, [("status", "draft", "approved")])
    db.commit()
    action_log.enrich(**log)
    alert_grant_use(db, caller, CAP_DOCUMENTS_APPROVE, "document.approved",
                    details=_doc_alert_details(db, doc),
                    changes=[("status", tv("v.draft"), tv("v.approved"))],
                    native=not via_grant)
    try:
        from app.approvals import edit_admin_notices
        edit_admin_notices("hr_document", str(doc_id), "approved", caller.get("full_name", ""))
    except Exception:
        pass
    return {"ok": True, "status": doc.status}


@router.post("/documents/{doc_id}/reject")
def reject_document(doc_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    """Reject a pending (draft) document — the webapp counterpart of the
    Telegram ❌ button. The record stays visible with a rejected status."""
    doc = _scope_documents(_real_docs(db), caller, db).filter(HrDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if not _may_reject_doc(doc, caller, db):
        raise HTTPException(status_code=403, detail="Not authorised to reject this document")
    via_grant = _doc_reject_via_grant(doc, caller, db)
    _reject_document(doc, caller, db)
    log = _doc_log_fields(doc, [("status", "draft", "rejected")])
    db.commit()
    action_log.enrich(**log)
    alert_grant_use(db, caller, CAP_DOCUMENTS_APPROVE, "document.rejected",
                    details=_doc_alert_details(db, doc),
                    changes=[("status", tv("v.draft"), tv("v.rejected"))],
                    native=not via_grant)
    try:
        from app.approvals import edit_admin_notices
        edit_admin_notices("hr_document", str(doc_id), "rejected", caller.get("full_name", ""))
    except Exception:
        pass
    return {"ok": True, "status": doc.status}


@router.post("/documents/{doc_id}/cancel")
def cancel_document(doc_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    doc = _scope_documents(_real_docs(db), caller, db).filter(HrDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if not _can_approve_doc(doc, caller, db):
        raise HTTPException(status_code=403, detail="Not authorised to un-post this document")
    if doc.status != "approved":
        # Nothing to un-post — idempotent and silent, as on the approve side.
        return {"ok": True, "status": doc.status}
    via_grant = _doc_via_grant(doc, caller, db)
    _cancel_doc(doc, caller, db)
    if doc.doc_type == "people_exchange":
        _notify_exchange(db, doc, "cancelled", int(caller["sub"]))
    log = _doc_log_fields(doc, [("status", "approved", "draft")])
    db.commit()
    action_log.enrich(**log)
    alert_grant_use(db, caller, CAP_DOCUMENTS_APPROVE, "document.cancelled",
                    details=_doc_alert_details(db, doc),
                    changes=[("status", tv("v.approved"), tv("v.draft"))],
                    native=not via_grant)
    return {"ok": True, "status": doc.status}


@router.post("/documents/{doc_id}/delete")
def delete_document(doc_id: int, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    doc = _scope_documents(_real_docs(db), caller, db).filter(HrDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    is_creator = _is_doc_creator(doc, caller)
    # Details snapshot BEFORE the delete below: expired attributes of a deleted
    # row can't be read back after commit.
    grant_alert = None   # (details, changes) when a grant authorized this
    # Approved docs may only be removed by an approver (reverts effects first).
    if doc.status == "approved":
        if not _can_approve_doc(doc, caller, db):
            raise HTTPException(status_code=403, detail="Approved documents can only be deleted by an approver")
        if _doc_via_grant(doc, caller, db):
            grant_alert = (_doc_alert_details(db, doc),
                           [("status", tv("v.approved"), None)])
        _revert_doc_effects(db, doc)
    elif doc.status == "draft":
        # Deleting a pending draft IS its rejection — keep the record instead of
        # erasing it, same as the Telegram ❌ button. Clears the admins' buttons.
        if not _may_reject_doc(doc, caller, db):
            raise HTTPException(status_code=403, detail="Not allowed to reject this document")
        via_grant = _doc_reject_via_grant(doc, caller, db)
        _reject_document(doc, caller, db)
        log = _doc_log_fields(doc, [("status", "draft", "rejected")])
        db.commit()
        # Deleting a DRAFT rejects it — the register must say what happened,
        # not what the button was called.
        action_log.enrich(action="document.rejected", **log)
        alert_grant_use(db, caller, CAP_DOCUMENTS_APPROVE, "document.rejected",
                        details=_doc_alert_details(db, doc),
                        changes=[("status", tv("v.draft"), tv("v.rejected"))],
                        native=not via_grant)
        try:
            from app.approvals import edit_admin_notices
            edit_admin_notices("hr_document", str(doc_id), "rejected", caller.get("full_name", ""))
        except Exception:
            pass
        return {"ok": True, "status": doc.status}
    elif caller.get("role") not in ("admin", "shift-manager") and not is_creator:
        raise HTTPException(status_code=403, detail="Not allowed to delete this document")

    log = _doc_log_fields(doc, [("status", doc.status, None)])
    db.delete(doc)
    db.commit()
    action_log.enrich(**log)
    if grant_alert:
        alert_grant_use(db, caller, CAP_DOCUMENTS_APPROVE, "document.deleted",
                        details=grant_alert[0], changes=grant_alert[1],
                        native=False)
    return {"ok": True}


# ── Bulk toolbar (Провести / Отменить / Удалить on many) ─────────────────────────

class DocBulkBody(BaseModel):
    ids:    List[int]
    action: str           # approve | cancel | delete


@router.post("/documents/bulk")
def bulk_documents(body: DocBulkBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    docs = _scope_documents(_real_docs(db), caller, db) \
        .filter(HrDocument.id.in_(body.ids)).all()

    done = 0
    resolved: list[tuple[int, str]] = []   # (doc_id, outcome) for admin-message cross-edit
    # (label, old, new) per grant-authorized doc — ONE aggregated warning DM
    # after the commit instead of one per document. Built before mutation, so
    # deleted rows are still readable.
    grant_rows: list[tuple] = []
    # Exchange notifications are fired AFTER the commit, never inside the loop.
    # A Telegram DM is not transactional: sent mid-transaction it survives a
    # rollback, so a batch that fails half-way notifies people about approvals
    # that never happened — and does it again on every retry.
    pending_notify: list[tuple] = []
    refused: list[dict] = []

    def _grant_row(doc, old_key, new_key):
        if _doc_via_grant(doc, caller, db):
            grant_rows.append((f"#{doc.id} · {unit_name(db, doc.manager_id)} · {doc.date}",
                               tv(old_key), tv(new_key) if new_key else None))

    for doc in docs:
        # Approval authority is per-document (e.g. a receiving supervisor may
        # post their own incoming exchange but not someone else's role change).
        if body.action == "approve":
            # Only a DRAFT can be posted. An already-approved document is a
            # silent no-op inside _approve_doc, so everything below it would
            # record an approval that never happened — the exchange DM most of
            # all, which is what turned one «Провести» over a select-all (the
            # header checkbox takes every visible row, posted ones included)
            # into a fresh announcement of the whole day's exchanges to every
            # admin. Rejected documents cannot be posted at all.
            if doc.status != "draft" or not _can_approve_doc(doc, caller, db):
                continue
            try:
                _approve_doc(doc, caller, db)
            except (StaleDocument, ExchangeTargetNoData) as exc:
                # One un-postable document must not abort the batch. Raising here
                # rolled the WHOLE transaction back — so nothing was saved, while
                # every Telegram DM already sent for the documents processed
                # before it had gone out for real, and each retry re-sent them.
                refused.append({
                    "doc_id": doc.id,
                    "date":   doc.date.isoformat() if doc.date else None,
                    "reason": exc.detail.get("code") if isinstance(exc.detail, dict) else "refused",
                })
                continue
            # Recorded only once the approval actually stuck.
            _grant_row(doc, "v.draft", "v.approved")
            pending_notify.append((doc, "approved"))
            resolved.append((doc.id, "approved"))
        elif body.action == "cancel":
            # Mirror of the approve branch: only an APPROVED document can be
            # un-posted. _cancel_doc no-ops on a draft, so without this a
            # select-all «Отменить» announces the cancellation of exchanges
            # that were never posted.
            if doc.status != "approved" or not _can_approve_doc(doc, caller, db):
                continue
            _grant_row(doc, "v.approved", "v.draft")
            _cancel_doc(doc, caller, db)
            if doc.doc_type == "people_exchange":
                pending_notify.append((doc, "cancelled"))
        elif body.action == "delete":
            is_creator = _is_doc_creator(doc, caller)
            if doc.status == "approved":
                if not _can_approve_doc(doc, caller, db):
                    continue
                _grant_row(doc, "v.approved", None)
                _revert_doc_effects(db, doc)
                db.delete(doc)
            elif doc.status == "draft":
                # Deleting a pending draft IS its rejection — keep the record
                # instead of erasing it, same as the Telegram ❌ button.
                if not _may_reject_doc(doc, caller, db):
                    continue
                if _doc_reject_via_grant(doc, caller, db):
                    grant_rows.append((f"#{doc.id} · {unit_name(db, doc.manager_id)} · {doc.date}",
                                       tv("v.draft"), tv("v.rejected")))
                _reject_document(doc, caller, db)
                resolved.append((doc.id, "rejected"))
            else:  # rejected — permanent cleanup of the record
                if caller.get("role") not in ("admin", "shift-manager") and not is_creator:
                    continue
                db.delete(doc)
        else:
            raise HTTPException(status_code=400, detail="Unknown action")
        done += 1

    db.commit()

    # Now that the batch is durable, tell people about it. Best-effort: a
    # Telegram hiccup must never undo approvals that are already saved.
    for doc, event in pending_notify:
        if doc.doc_type != "people_exchange":
            continue
        try:
            _notify_exchange(db, doc, event, int(caller["sub"]))
        except Exception:
            logger.exception("bulk %s: notification failed for doc %s", body.action, doc.id)
    if pending_notify:
        try:
            db.commit()
        except Exception:
            db.rollback()
            logger.exception("bulk %s: could not persist notifications", body.action)

    if grant_rows:
        alert_grant_use(db, caller, CAP_DOCUMENTS_APPROVE, "document.bulk",
                        details=[("count", len(grant_rows))],
                        changes=grant_rows, native=False)
    # ONE row for the whole toolbar press, naming the operation and the count —
    # never one per document.
    action_log.enrich(
        target_kind="batch",
        target_id=",".join(str(i) for i in body.ids[:20]),
        details=[("mode", body.action), ("count", done),
                 ("skipped", max(len(body.ids) - done, 0))],
    )
    if resolved:
        try:
            from app.approvals import edit_admin_notices
            name = caller.get("full_name", "")
            for doc_id, outcome in resolved:
                edit_admin_notices("hr_document", str(doc_id), outcome, name)
        except Exception:
            pass
    return {"ok": True, "affected": done, "refused": refused}


# ══════════════════════════════════════════════════════════════════════════════
#  DAY CLOSE — supervisors close their own day (no admin approval needed).
#  Existence of a DayApproval row = day CLOSED. The day becomes CONFIRMED once
#  every request for that date is processed (approved or rejected). Only
#  confirmed (manager, date) pairs are calculated/shown anywhere on dashboards.
#  Only an admin can re-open a closed day (deletes the row → back to OPEN).
# ══════════════════════════════════════════════════════════════════════════════

def _staff_sees_all(db: Session, caller) -> bool:
    """True when a personal page grant lifts this caller's unit pin on the
    staff/daily reads.

    Both pages share these endpoints, so either grant at "all" counts — the
    scope means "reads every unit's day", and which of the two pages the person
    was handed doesn't change what that sentence promises."""
    return (page_scope_is_all(db, caller, "staff")
            or page_scope_is_all(db, caller, "daily"))


def _visible_manager_ids(db: Session, caller) -> Optional[List[int]]:
    """Manager ids a caller may see/approve. None = all (admin).

    A personal staff/daily page grant at "all" is admin reach by definition —
    the whole point of that scope is that this person reads every unit's day."""
    role    = caller.get("role")
    role_id = caller.get("role_id")
    if role == "admin" or _staff_sees_all(db, caller):
        return None
    if role == "supervisor":
        return [role_id] if role_id else []
    if role == "shift-manager":
        shift = _sm_shift(db, role_id)
        return [m.id for m in db.query(Manager)
                .filter(Manager.shift == shift, Manager.archived.is_(False)).all()]
    return []


def _can_touch_manager(db: Session, caller, manager_id: int) -> bool:
    vis = _visible_manager_ids(db, caller)
    return vis is None or manager_id in vis


def _staff_target_manager(db: Session, caller, manager_id: Optional[int]) -> Optional[int]:
    """The unit a staff/daily READ targets.

    A supervisor is pinned to their own unit, so their pages never need to send
    manager_id — except when a staff/daily page grant at "all" widens them, in
    which case an explicitly requested unit wins and their own stays the
    default. ``_can_touch_manager`` still has the last word on the result."""
    if caller.get("role") == "supervisor" and not (
            manager_id and _staff_sees_all(db, caller)):
        return caller.get("role_id")
    return manager_id


@router.get("/approvals/calendar")
def approvals_calendar(
    year: int,
    month: int,
    manager_id: Optional[int] = None,
    caller=Depends(_require_staff),
    db: Session = Depends(get_db),
):
    """
    Per-day status for one manager across a calendar month:
      confirmed → day closed and all its requests processed
      closed    → day closed but requests still await review
      open      → the day has worker data but is not closed yet
      (absent)  → no worker data that day
    """
    manager_id = _staff_target_manager(db, caller, manager_id)
    if not manager_id:
        raise HTTPException(status_code=400, detail="manager_id required")
    if not _can_touch_manager(db, caller, manager_id):
        raise HTTPException(status_code=403, detail="Not allowed for this manager")

    start = date(year, month, 1)
    end   = date(year + (month == 12), (month % 12) + 1, 1)  # first day of next month

    # Dates with real worker data
    data_dates = {
        d for (d,) in db.query(distinct(Attendance.date)).filter(
            Attendance.manager_id == manager_id,
            Attendance.date >= start, Attendance.date < end,
            Attendance.worker_name.isnot(None),
            Attendance.worker_name.notin_(["", "nan", "NaN"]),
        ).all()
    }

    closures = db.query(DayApproval).filter(
        DayApproval.manager_id == manager_id,
        DayApproval.date >= start, DayApproval.date < end,
    ).all()
    closed_map = {a.date: a for a in closures}

    # Dates still blocked by unprocessed requests / draft documents
    pending_dates = {
        d for (d,) in db.query(distinct(EditRequest.date)).filter(
            EditRequest.manager_id == manager_id,
            EditRequest.date >= start, EditRequest.date < end,
            EditRequest.status == "pending",
        ).all()
    } | {
        # Only a REAL document blocks a day here — the twin of the same clause in
        # `services/day_state.pending_counts`. An unrecognised type must not hold
        # a day at «closed» on this calendar on the strength of a reader nobody
        # has written yet.
        d for (d,) in db.query(distinct(HrDocument.date)).filter(
            HrDocument.manager_id == manager_id,
            HrDocument.date >= start, HrDocument.date < end,
            HrDocument.status == "draft",
            HrDocument.doc_type.in_(_REAL_DOC_TYPES),
        ).all()
    }

    all_dates = data_dates | set(closed_map.keys())
    days = {}
    for d in sorted(all_dates):
        iso = d.isoformat()
        if d in closed_map:
            a = closed_map[d]
            days[iso] = {
                "status":    "closed" if d in pending_dates else "confirmed",
                "closed_by": a.approved_by_name,
                "closed_at": a.approved_at.isoformat() if a.approved_at else None,
            }
        else:
            days[iso] = {"status": "open"}

    mgr = db.query(Manager).filter_by(id=manager_id).first()
    return {
        "manager_id": manager_id,
        "manager_name": mgr.name if mgr else None,
        "year": year, "month": month,
        "days": days,
    }


@router.get("/approvals/day")
def approval_day(
    attend_date: str,
    manager_id: Optional[int] = None,
    caller=Depends(_require_staff),
    db: Session = Depends(get_db),
):
    """Day-close state for a single (manager, date) — used by the Daily/Staff pages."""
    manager_id = _staff_target_manager(db, caller, manager_id)
    if not manager_id:
        raise HTTPException(status_code=400, detail="manager_id required")
    if not _can_touch_manager(db, caller, manager_id):
        raise HTTPException(status_code=403, detail="Not allowed for this manager")
    try:
        d = date.fromisoformat(attend_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    state, closure, counts = day_state(db, manager_id, d)
    # What the close will REFUSE on, published before the press — the same
    # `_unplaced_workers` the endpoint enforces with, so the warning and the
    # refusal can never name different people.
    unplaced = [] if closure is not None else _unplaced_workers(db, manager_id, d)
    return {
        "manager_id":       manager_id,
        "date":             attend_date,
        "state":            state,   # open | closed | confirmed
        "closed":           closure is not None,
        "closed_by":        closure.approved_by_name if closure else None,
        "closed_at":        closure.approved_at.isoformat() if closure and closure.approved_at else None,
        "pending_requests": counts["pending_requests"] + counts["draft_docs"],
        "can_reopen":       _cap_covers_unit(caller, db, CAP_DAY_REOPEN, manager_id),
        "needs_cell":       len(unplaced),
        "needs_cell_names": [r.worker_name for r in unplaced[:20]],
    }


class ApprovalBody(BaseModel):
    manager_id: Optional[int] = None
    date: str


# ── Cell placement («Yacheykalar» tab on /staff) ──────────────────────────────
#
# Where a supervisor says WHICH CELL each of their people actually worked in.
# It exists because an accepted people-exchange no longer names a destination
# cell (2026-08-30): the sender was guessing about somebody else's shopfloor
# before the shift had run, so the worker now arrives on the SUPERVISOR and the
# supervisor places them here. `_unplaced_workers` is the gate that makes it
# unskippable, and it is the same predicate this tab clears.
#
# Everything is STAGED on the client and written by one PUT, the admin «Davomat»
# tab's model: a supervisor rearranging people should be able to change their
# mind before anything is real, and one write means one action-log row.


def _split_hours(att: Attendance, hhmm: str):
    """Split one attendance row's hours at a wall-clock time → (h1, h2).

    Deliberately NOT `_compute_split`: that one answers a different question —
    how a day divides between two UNITS, with an early-arrival rule, a
    return/carve-out window and a minimum-hours test that can strip the worker's
    name off the smaller side. None of that applies inside one unit, where both
    halves stay named and stay on the same roster.

    The two halves are scaled to sum to EXACTLY `hours_worked`, so no unit total
    moves: the clock is only used for the RATIO. Returns None when the row
    carries no usable clock or the time falls outside it — the caller refuses
    rather than inventing a division.
    """
    total = float(att.hours_worked or 0)
    if total <= 0:
        return None
    # `_clock_bounds_min` / `_parse_hhmm` are the file's ONE clock parser and
    # already tolerate the verifix formats («07:49 - 17:04 (8.43)», '8-00',
    # '08.00'). A second spelling here is how the split and the transfer-time
    # arithmetic would start disagreeing about the same string.
    t = _parse_hhmm(hhmm)
    c, o = _clock_bounds_min(att.clock_in_out)
    if t is None or c is None or o is None:
        return None
    if o <= c:
        o += 1440                       # crossed midnight — the house rule
    if t <= c:
        t += 1440
    if not (c < t < o):
        return None
    frac = (t - c) / (o - c)
    h1 = round(total * frac, 4)
    h2 = round(total - h1, 4)
    if h1 <= 0 or h2 <= 0:
        return None
    return h1, h2


def _placement_cells(db: Session, manager_id: int, codes_in_use: set) -> list:
    """The cells this supervisor may place somebody INTO.

    The unit's REGISTRY cells (cells.manager_id) unioned with whatever codes the
    day's rows already carry. The registry half is what lets an EMPTY cell be a
    destination — /api/staff/attendance derives its catalog from the codes
    present in today's rows, so on its own it can only ever offer a cell that
    already has somebody standing in it. The in-use half is what keeps a code
    the registry has never heard of visible instead of silently unplaceable:
    the two registers are allowed to disagree, in public.
    """
    clauses = [Cell.manager_id == manager_id]
    if codes_in_use:
        clauses.append(Cell.verifix_code.in_(list(codes_in_use)))
    return db.query(Cell).filter(or_(*clauses)).all()


def _lender_names(db: Session, manager_id: int, d: date) -> dict:
    """worker_name → the unit that LENT them here, off the approved exchange.

    Only for display: it answers "why is this person in my list at all", which
    is the first question a supervisor asks of a name they do not recognise.
    """
    out: dict = {}
    docs = db.query(HrDocument).filter(
        HrDocument.date == d,
        HrDocument.doc_type == "people_exchange",
        HrDocument.status == "approved",
    ).all()
    for doc in docs:
        payload = doc.payload or {}
        if payload.get("target_manager_id") != manager_id:
            continue
        sender = doc.supervisor_name or unit_name(db, doc.manager_id)
        for emp in payload.get("employees") or []:
            wn = (emp.get("worker_name") or "").strip()
            if wn:
                out[wn] = sender
    return out


@router.get("/cell-placement")
def cell_placement(attend_date: str, manager_id: Optional[int] = None,
                   caller=Depends(_require_staff), db: Session = Depends(get_db)):
    """The unit's day, grouped by cell, with the cell-less workers first."""
    manager_id = _staff_target_manager(db, caller, manager_id)
    if not manager_id:
        raise HTTPException(status_code=400, detail="manager_id required")
    if not _can_touch_manager(db, caller, manager_id):
        raise HTTPException(status_code=403, detail="Not allowed for this manager")
    try:
        d = date.fromisoformat(attend_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    from app.services.kpi_calculator import is_direct_role
    rows = db.query(Attendance).filter(
        Attendance.manager_id == manager_id,
        Attendance.date == d,
        Attendance.worker_name.isnot(None),
        Attendance.worker_name.notin_(["", "nan", "NaN"]),
    ).order_by(Attendance.worker_name).all()

    closed = db.query(DayApproval).filter_by(manager_id=manager_id, date=d).first()
    lenders = _lender_names(db, manager_id, d)

    def _hc(r) -> float:
        """This row's contribution to its cell's headcount N.

        The SAME rule `idle_source` and `kpi_calculator` now weigh a cell by:
        a split worker is named in both of their cells and is worth their share
        of a person in each, so the two halves add back to exactly 1.0 and no
        UNIT total moves. Printing a plain row count here instead would put a
        number on screen that no KPI agrees with.
        """
        if not is_direct_role(r.job_title, r.hours_worked, bool(r.is_supervisor)):
            return 0.0
        return 1.0 if r.hc_weight is None else float(r.hc_weight)

    def w_json(r):
        return {
            "id":            r.id,
            "worker_name":   r.worker_name,
            "job_title":     r.job_title or "",
            "hours":         float(r.hours_worked) if r.hours_worked is not None else 0.0,
            "clock_in_out":  r.clock_in_out or "",
            "counted":       is_direct_role(r.job_title, r.hours_worked, bool(r.is_supervisor)),
            "hc_weight":     float(r.hc_weight) if getattr(r, "hc_weight", None) is not None else None,
            "split_of":      getattr(r, "split_of", None),
            "from_unit":     lenders.get(r.worker_name),
        }

    by_code: dict = {}
    unplaced = []
    for r in rows:
        code = (r.verifix_code or "").strip()
        # The unit's own brigadir is cell-less by construction and no placement
        # can ever give them one — they are neither unplaced nor in a cell.
        if not code:
            if r.is_supervisor:
                continue
            unplaced.append(w_json(r))
        else:
            by_code.setdefault(code, []).append(r)

    cells_out = []
    for c in _placement_cells(db, manager_id, set(by_code)):
        crows = by_code.pop(c.verifix_code, [])
        cells_out.append({
            "verifix_code": c.verifix_code,
            "cell_id":      c.id,
            "leader_name":  _cell_leader_name(db, c),
            "workers":      [w_json(r) for r in crows],
            "counted":      round(sum(_hc(r) for r in crows), 2),
            "total":        len(crows),
            "hours":        round(sum(float(r.hours_worked or 0) for r in crows), 1),
        })
    # A code on a row that the registry does not know — shown, never dropped.
    for code, crows in by_code.items():
        cells_out.append({
            "verifix_code": code, "cell_id": None, "leader_name": None,
            "workers":      [w_json(r) for r in crows],
            "counted":      round(sum(_hc(r) for r in crows), 2),
            "total":        len(crows),
            "hours":        round(sum(float(r.hours_worked or 0) for r in crows), 1),
        })
    cells_out.sort(key=lambda c: c["verifix_code"])

    mgr = db.query(Manager).filter_by(id=manager_id).first()
    return {
        "manager_id":   manager_id,
        "manager_name": mgr.name if mgr else "",
        "shift":        mgr.shift if mgr else None,
        "date":         attend_date,
        "day_closed":   closed is not None,
        "can_edit":     closed is None and _can_edit_placement(caller, manager_id),
        "unplaced":     unplaced,
        "cells":        cells_out,
        "totals": {
            "cells":    len(cells_out),
            "workers":  len(rows),
            "counted":  sum(c["counted"] for c in cells_out),
            "hours":    round(sum(c["hours"] for c in cells_out), 1),
            "unplaced": len(unplaced),
        },
    }


def _cell_leader_name(db: Session, c: Cell) -> Optional[str]:
    """The cell's leader — the second fact a code is allowed to carry.
    A cell is its CODE; the workshop name is never printed (utils/cellName.js)."""
    if not c.leader_id:
        return None
    p = db.query(RoleProfile).filter_by(id=c.leader_id).first()
    return p.full_name if p else None


def _can_edit_placement(caller: dict, manager_id: int) -> bool:
    """Who may WRITE a placement. Reading follows the page's ordinary scope, but
    writing is the unit's own business: a supervisor widened by a page grant can
    BROWSE another unit and must not rearrange its people — the same rule
    `canCreateHere` already applies to documents on this page."""
    role = caller.get("role")
    if role == "admin":
        return True
    if role == "supervisor":
        return caller.get("role_id") == manager_id
    return False


class PlacementMove(BaseModel):
    attendance_id: int
    verifix_code:  str


class PlacementSplit(BaseModel):
    attendance_id: int
    verifix_code:  str          # the cell the FIRST half stays in
    second_code:   str          # the cell the second half moves to
    split_at:      str          # "HH:MM"


class PlacementBody(BaseModel):
    manager_id: Optional[int] = None
    date:       str
    moves:      List[PlacementMove]  = []
    splits:     List[PlacementSplit] = []
    unsplits:   List[int]            = []   # the SECONDARY row's id


@router.put("/cell-placement")
def save_cell_placement(body: PlacementBody, caller=Depends(_require_staff),
                        db: Session = Depends(get_db)):
    """Write a batch of placements, splits and un-splits for one unit-day."""
    manager_id = _staff_target_manager(db, caller, body.manager_id)
    if not manager_id:
        raise HTTPException(status_code=400, detail="manager_id required")
    if not _can_touch_manager(db, caller, manager_id):
        raise HTTPException(status_code=403, detail="Not allowed for this manager")
    if not _can_edit_placement(caller, manager_id):
        raise HTTPException(status_code=403, detail="Only this unit's supervisor or an admin may place workers")
    try:
        d = date.fromisoformat(body.date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    # A closed day is frozen like everything else on this page. 409 so the
    # client can tell "somebody closed it under me" from a validation refusal.
    if db.query(DayApproval).filter_by(manager_id=manager_id, date=d).first():
        raise HTTPException(status_code=409, detail="Day is already closed")

    valid_codes = {c.verifix_code for c in _placement_cells(db, manager_id, set())}

    def _own_row(att_id: int) -> Attendance:
        r = db.query(Attendance).filter_by(id=att_id).first()
        # Re-checked per row: an id is typeable, and the caller's authority is
        # over a UNIT-DAY, not over an arbitrary attendance row.
        if not r or r.manager_id != manager_id or r.date != d:
            raise HTTPException(status_code=404, detail="Worker row not found for this unit and date")
        return r

    def _check_code(code: str) -> str:
        code = (code or "").strip()
        if code not in valid_codes:
            raise HTTPException(status_code=400,
                                detail=f"«{code}» — bu brigadaning yacheykasi emas")
        return code

    moved = split_n = unsplit_n = 0

    for u in body.unsplits:
        sec = _own_row(u)
        if sec.split_of is None:
            raise HTTPException(status_code=400, detail="That row is not the second half of a split")
        pri = db.query(Attendance).filter_by(id=sec.split_of).first()
        if pri is not None:
            # Put the day back together: the halves were scaled to sum to the
            # original, and the clocks were cut "C-T" / "T-O", so both restore
            # exactly without having stored a copy of either.
            pri.hours_worked = float(pri.hours_worked or 0) + float(sec.hours_worked or 0)
            pri.hc_weight    = None
            pri_start = (pri.clock_in_out or "").partition("-")[0]
            sec_end   = (sec.clock_in_out or "").rpartition("-")[2]
            if pri_start and sec_end:
                pri.clock_in_out = f"{pri_start}-{sec_end}"
        db.delete(sec)
        unsplit_n += 1

    for m in body.moves:
        r = _own_row(m.attendance_id)
        r.verifix_code = _check_code(m.verifix_code)
        moved += 1

    for sp in body.splits:
        r = _own_row(sp.attendance_id)
        if r.split_of is not None:
            raise HTTPException(status_code=400, detail="Cannot split a half of an existing split")
        if db.query(Attendance).filter_by(split_of=r.id).first():
            raise HTTPException(status_code=400, detail="That worker is already split")
        first  = _check_code(sp.verifix_code)
        second = _check_code(sp.second_code)
        if first == second:
            raise HTTPException(status_code=400, detail="Ikkala yacheyka bir xil bo'lishi mumkin emas")
        parts = _split_hours(r, sp.split_at)
        if not parts:
            raise HTTPException(
                status_code=400,
                detail=f"{r.worker_name}: «{sp.split_at}» bu xodimning ish vaqtidan tashqarida")
        h1, h2 = parts
        total = h1 + h2
        c_o = (r.clock_in_out or "").strip()
        start_s, _, end_s = c_o.partition("-")
        # The SECOND half is a new row carrying the same name — the worker is on
        # this supervisor's roster either way, so stripping the name off the
        # smaller side (what a cross-unit split does) would be a lie here. What
        # keeps the arithmetic honest instead is `hc_weight`: the two halves sum
        # to 1.0, so every per-cell N sees a fraction and every UNIT total is
        # unmoved.
        second_row = Attendance(
            manager_id        = manager_id,
            date              = d,
            worker_name       = r.worker_name,
            job_title         = r.job_title,
            schedule          = r.schedule,
            clock_in_out      = f"{sp.split_at}-{end_s}" if end_s else None,
            hours_worked      = h2,
            early_arrival_min = 0,          # early belongs to the first half
            effective_hours   = None,
            verifix_code      = second,
            is_supervisor     = bool(r.is_supervisor),
            hc_weight         = round(h2 / total, 6),
            split_of          = r.id,
        )
        r.verifix_code    = first
        r.hours_worked    = h1
        r.effective_hours = None
        r.hc_weight       = round(h1 / total, 6)
        if start_s:
            r.clock_in_out = f"{start_s}-{sp.split_at}"
        db.add(second_row)
        split_n += 1

    db.commit()

    left = _unplaced_workers(db, manager_id, d)
    action_log.enrich(
        target_kind="day", target_id=f"{manager_id}:{body.date}",
        unit_id=manager_id, unit_name=unit_name(db, manager_id), day=d,
        details=[("date", body.date), ("moved", moved), ("split", split_n),
                 ("unsplit", unsplit_n), ("still_unplaced", len(left))],
    )
    return {"ok": True, "moved": moved, "split": split_n,
            "unsplit": unsplit_n, "unplaced": len(left)}



# The day a worker's row first carried a cell («Код подразделения», the
# single-file «Davomat» upload — models.Attendance.verifix_code). EVERY row
# before it is cell-less by construction, and an admin can reopen any historical
# day from four surfaces, so without this floor re-opening one would make it
# permanently unclosable — on a platform with no shell.
CELLS_REQUIRED_FROM = date(2026, 8, 1)


def _unplaced_workers(db: Session, manager_id: int, d: date) -> list:
    """Named, counted workers in this unit-day who still have no cell.

    THE predicate behind the day-close gate and behind the «Yacheykalar» tab —
    one spelling, so the button that refuses and the page that clears the
    refusal can never disagree about who is missing.

    Three exclusions, each of which makes the gate unclearable if dropped:
      · `is_supervisor` — the unit's OWN brigadir is written cell-less on every
        re-projection (attendance_batch._cellless_by_manager) and no placement
        can ever give them one. Counting them locks every unit out forever.
      · nameless rows — the hours-only leftovers a transfer-time split writes
        (`worker_name IS NULL`). /api/staff/attendance folds them into
        `extra_hours` and the tab cannot show them, so nobody could name one
        into a cell.
      · rows that did not come — `CALC_ROWS_FILTER` already demands hours > 0.
        An absent worker belongs to no cell that day.

    An empty-string code counts as no cell: the parser writes `code or None`,
    but nothing forces that shape on every writer, and a stray "" would
    otherwise pass the gate while still counting towards no cell anywhere.
    """
    if d < CELLS_REQUIRED_FROM:
        return []
    from app.routers.workers import CALC_ROWS_FILTER
    rows = db.query(Attendance).filter(
        Attendance.manager_id == manager_id,
        Attendance.date == d,
        Attendance.worker_name.isnot(None),
        Attendance.worker_name.notin_(["", "nan", "NaN"]),
        CALC_ROWS_FILTER,
        or_(Attendance.verifix_code.is_(None), Attendance.verifix_code == ""),
    ).order_by(Attendance.worker_name).all()
    return rows


def _unplaced_detail(rows: list) -> str:
    """The 409 body, as a PLAIN STRING.

    `utils/api.js` rewrites any non-string `detail` (a dict without
    msg/message/detail becomes JSON.stringify), and both close dialogs render
    `detail` only when `typeof d === "string"` — so a structured refusal arrives
    as the generic "save failed" with the reason stripped off, which is the one
    outcome a hard gate must never produce. Capped at five names because
    ConfirmDialog's card has no max-height and no scroll: an uncapped list
    pushes the buttons off-screen and the operator can neither read the reason
    nor dismiss the dialog.
    """
    names = [r.worker_name for r in rows[:5]]
    more  = len(rows) - len(names)
    tail  = f" +{more}" if more > 0 else ""
    return (f"{len(rows)} ta xodim yacheykaga biriktirilmagan: "
            f"{', '.join(names)}{tail}. "
            f"«Yacheykalar» bo'limida ularni joylashtiring.")


@router.post("/daily/close")
def close_day(body: ApprovalBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    """
    Supervisor closes their own day (final — no approval needed), or an admin
    closes a day on behalf of a supervisor (manager_id required in the body;
    the supervisor gets notified). After closing, the supervisor can no longer
    submit changes for this date. Data appears on dashboards once every request
    for the date is processed (the day becomes 'confirmed').
    """
    role = caller.get("role")
    if role == "supervisor":
        manager_id = caller.get("role_id")
        if not manager_id:
            raise HTTPException(status_code=400, detail="Supervisor has no assigned manager")
    elif role == "admin":
        manager_id = body.manager_id
        if not manager_id:
            raise HTTPException(status_code=400, detail="manager_id required")
    else:
        raise HTTPException(status_code=403, detail="Only supervisors or admins can close a day")
    try:
        d = date.fromisoformat(body.date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    if d > date.today():
        raise HTTPException(status_code=400, detail="Cannot close a future date")

    if db.query(DayApproval).filter_by(manager_id=manager_id, date=d).first():
        raise HTTPException(status_code=409, detail="Day is already closed")

    # A worker with no cell is a worker whose hours belong to no cell's load and
    # to no cell's headcount — Σ(Nᵢ·Tᵢ)/ΣNᵢ simply never sees them. Since
    # 2026-08-30 an accepted people-exchange lands its workers on the SUPERVISOR
    # rather than on a cell, so this is the step that makes the receiving
    # supervisor say where they actually worked, while the day is still fresh in
    # their head. HARD, and for an admin closing on behalf too: an override
    # would be used, and the number it protects is one nobody can reconstruct
    # afterwards.
    unplaced = _unplaced_workers(db, manager_id, d)
    if unplaced:
        # enrich() BEFORE the raise — the middleware has already opened an
        # `attendance.day_closed` row for this request and will stamp it
        # 'refused'; without this it names no reason at all.
        action_log.enrich(
            target_kind="day", target_id=f"{manager_id}:{body.date}",
            unit_id=manager_id, unit_name=unit_name(db, manager_id), day=d,
            details=[("date", body.date), ("blocked", "cells_missing"),
                     ("workers", len(unplaced))],
        )
        raise HTTPException(status_code=409, detail=_unplaced_detail(unplaced))

    # Per-cell ojidaniya never blocks a close (from 2026-08-22): a leader's entry
    # counts the moment it is saved, so nothing on /idle-cell can be pending.
    # The close dialog asks the brigadir whether they have looked at what their
    # leaders entered (`GET /api/idle-cell/day-summary`) — a question, not a
    # gate — and the day's lock then freezes those rows with everything else.

    closer_name = caller.get("full_name", "")
    db.add(DayApproval(
        manager_id=manager_id,
        date=d,
        approved_by_telegram_id=int(caller["sub"]),
        approved_by_name=closer_name,
        approved_at=datetime.now(timezone.utc),
    ))
    _notify_all_parties(
        db, manager_id,
        "day_closed",
        {"closer_name": closer_name, "date": body.date},
        ntype="info",
        actor_tg_id=int(caller["sub"]),
        include_supervisor=(role == "admin"),
    )
    db.commit()

    state, _, counts = day_state(db, manager_id, d)
    left = counts["pending_requests"] + counts["draft_docs"]
    action_log.enrich(
        target_kind="day", target_id=f"{manager_id}:{body.date}",
        unit_id=manager_id, unit_name=unit_name(db, manager_id), day=d,
        details=[("date", body.date), ("state", state)]
                + ([("pending", left)] if left else []),
        changes=[("status", "open", "closed")],
    )
    return {
        "ok": True, "state": state, "manager_id": manager_id, "date": body.date,
        "pending_requests": counts["pending_requests"] + counts["draft_docs"],
    }


@router.post("/approvals/reopen")
def reopen_day(body: ApprovalBody, caller=Depends(_require_staff), db: Session = Depends(get_db)):
    if not body.manager_id:
        raise HTTPException(status_code=400, detail="manager_id required")
    if not _cap_covers_unit(caller, db, CAP_DAY_REOPEN, body.manager_id):
        raise HTTPException(status_code=403, detail="Only an admin can re-open a closed day")
    try:
        d = date.fromisoformat(body.date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    existing = db.query(DayApproval).filter_by(manager_id=body.manager_id, date=d).first()
    if existing:
        # Snapshot before the delete: the row that says who closed this day is
        # about to be gone, and the register becomes its only remaining trace.
        closer = existing.approved_by_name
        db.delete(existing)
        _notify_all_parties(
            db, body.manager_id,
            "day_reopened",
            {"reopener_name": caller.get("full_name", "admin"), "date": body.date},
            ntype="warning",
            actor_tg_id=int(caller["sub"]),
            include_supervisor=True,
        )
        db.commit()
        unit = unit_name(db, body.manager_id)
        alert_grant_use(db, caller, CAP_DAY_REOPEN, "day.reopened",
                        details=[("unit", unit), ("date", body.date)])
        action_log.enrich(
            target_kind="day", target_id=f"{body.manager_id}:{body.date}",
            unit_id=body.manager_id, unit_name=unit, day=d,
            details=[("unit", unit), ("date", body.date),
                     ("closed_by", closer or "—")],
            changes=[("status", "closed", "open")],
        )
    return {"ok": True, "state": "open", "manager_id": body.manager_id, "date": body.date}


@router.get("/approvals/cells")
def approved_cells(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_page("zagruzka", "staff", "daily")),
):
    """
    Open endpoint (mirrors /api/heatmap): the CONFIRMED (manager_id, date)
    pairs in a range — closed days with no unprocessed requests left. The
    dashboard treats everything else as having no data. Returns ISO dates.
    """
    def _parse(s, fallback):
        try:
            return date.fromisoformat(s) if s else fallback
        except ValueError:
            return fallback

    d_from = _parse(date_from, date(2000, 1, 1))
    d_to   = _parse(date_to,   date(2100, 1, 1))
    pairs  = confirmed_pairs(db, d_from, d_to)
    return {"cells": [{"manager_id": mid, "date": d.isoformat()} for mid, d in sorted(pairs)]}
