"""
Production-planning API (ABC form).

Brigadir-facing (supervisor scoped to their own Manager via the JWT role_id;
admin may pass ?manager_id=):
    GET  /api/production/dashboard?date=YYYY-MM-DD
    GET  /api/production/dates
    POST /api/production/override          {date, sap_code, work_center, field, value}
    POST /api/production/reconciliation     {date, data}

Admin-only:
    POST /api/production/wc-override        {date, work_center, people, shtatka}
    POST /admin/production/upload           file(s) + manager_id + date + mode
    GET  /admin/production/work-centers?manager_id=
    PUT  /admin/production/work-centers/{id}    {shtatka, capacity}
    GET    /admin/production/catalog?manager_id=
    POST   /admin/production/catalog            {manager_id, sap_code, name, work_center, labor_time}
    PUT    /admin/production/catalog/{id}       {labor_time, name, sap_code, work_center, active}
    DELETE /admin/production/catalog/{id}
"""
from __future__ import annotations

import statistics
from datetime import date, datetime, timedelta
from functools import lru_cache
from io import BytesIO
from typing import Annotated, Optional

import jwt
from jwt import PyJWTError as JWTError
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.config import settings
from app.database import get_db
from app.models import (
    Manager, AppSetting, ProductionData, PPProduct, PPWorkCenter, PPWorkCenterDaily,
    PPDaily, PPReconciliation, PPUpload, ForecastCallNotice, TelegramUser,
    TelegramUserRole, RoleProfile,
)
from app.permissions import require_page
from app.services.pp_parser import read_workbook_slices, parse_catalog_workbook, FAZA_COLUMNS
from app.services.pp_calc import compute_dashboard, DEFAULT_SHIFT_MIN, DEFAULT_PRODUCTIVE_MIN
from app.services.name_map import sheet_alias_map

router = APIRouter(tags=["production"])
_oauth2 = OAuth2PasswordBearer(tokenUrl="/api/auth/webapp")

PAGE = "production"

POSITIONS_TITLE = {"uz": "Pozitsiyalar", "uz_cyrl": "Позициялар", "ru": "Позиции", "en": "Positions"}

# «ABC форма» workbook layout — the Excel export reproduces the brigadirs' manual
# ABC form («Форма ABC … 8 соатлик») sheet-for-sheet: shift totals in row 1,
# headers in row 2, position rows from row 3, the per-team block (M:W), the
# indicator block (X:Y) and the staffing block (Z:AA).
#
# Only the four true inputs are written as values — Трудоемкость (C), Команда
# (D), Факт (G), ПЛАН (H) — plus Штатка (W) and the reconciliation counts.
# EVERYTHING else is a live formula, so editing any of those recalculates the
# whole sheet exactly as the manual form does. Labels stay in the template's
# original mixed ru/uz wording regardless of UI language.
ABC_HEADERS = ["Сап код", "SKU", "Трудоемкость", "Команда", "ЛЮДИ", "вып %",
               "Факт План", "ПЛАН", "Общ.трудаёмкост", "Минут", "Парето"]
ABC_WIDTHS = {"A": 12.5, "B": 42, "C": 12.5, "D": 10.5, "E": 8, "F": 8, "G": 9.5,
              "H": 9, "I": 12.5, "J": 8.5, "K": 8.5, "L": 4.5, "M": 10, "N": 8.5,
              "O": 15, "P": 10, "Q": 9, "R": 8.5, "S": 9.5, "T": 8.5, "U": 7.5,
              "V": 3, "W": 8.5, "X": 34, "Y": 11, "Z": 24, "AA": 9}
ABC_SPARE_ROWS = 15   # bordered formula rows under the data for hand-added SKUs
ABC_DATA_START = 3    # first position row (row 1 = totals, row 2 = headers)
ABC_TEAM_START = 6    # first row of the M:W per-team block, as in the form

# Per-команда identity colour — MUST stay in sync with `WC_PALETTE` / `wcColor`
# in frontend/src/pages/Production.jsx so a team wears the same colour in the
# export as it does in the «Позиции» table (hash of the code → palette slot).
WC_PALETTE = ["6366F1", "0EA5E9", "10B981", "F59E0B", "EF4444", "EC4899",
              "8B5CF6", "14B8A6", "F97316", "84CC16", "06B6D4", "A855F7"]
WC_TINT = 0.16       # chip background = colour blended onto white, as in the app


@lru_cache(maxsize=256)
def _wc_style(code: str) -> tuple:
    """(fill, font colour) for a work-center chip — pale tint + the colour itself."""
    h = 0
    for ch in code:
        h = (h * 31 + ord(ch)) & 0xFFFFFFFF          # JS: (h*31 + charCodeAt) >>> 0
    hexc = WC_PALETTE[h % len(WC_PALETTE)]
    rgb = (int(hexc[0:2], 16), int(hexc[2:4], 16), int(hexc[4:6], 16))
    tint = "".join(f"{round(c * WC_TINT + 255 * (1 - WC_TINT)):02X}" for c in rgb)
    return PatternFill("solid", fgColor=tint), hexc


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def _verify_admin(token: Annotated[str, Depends(_oauth2)]) -> dict:
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=[settings.algorithm])
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    if payload.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return payload


def _shift_manager_shift(payload: dict, db: Session) -> int:
    """The shift (1|2) a shift-manager profile covers. Their JWT role_id points
    at role_profiles.id, and the shift lives there — the JWT itself has no shift
    field."""
    rp = db.query(RoleProfile).filter(
        RoleProfile.id == payload.get("role_id"),
        RoleProfile.role == "shift-manager",
    ).first()
    if not rp or rp.shift is None:
        raise HTTPException(status_code=403, detail="No shift assigned to this shift-manager profile")
    return int(rp.shift)


def _configured_manager_ids(db: Session) -> set[int]:
    """Managers whose ABC production is set up — they have a catalog (pp_products)
    or at least one uploaded daily snapshot (pp_daily). Only these have a
    meaningful dashboard, so they are the only units offered in the picker."""
    ids = {m for (m,) in db.query(PPProduct.manager_id).distinct()}
    ids |= {m for (m,) in db.query(PPDaily.manager_id).distinct()}
    return ids


def _resolve_manager_id(payload: dict, requested: Optional[int], db: Session) -> int:
    """Resolve the single brigadir unit a request targets, enforcing role scope:

        supervisor    → pinned to their own unit (JWT role_id); ?manager_id= ignored.
        shift-manager → any unit *in their own shift* (?manager_id= required).
        top-manager   → any unit (?manager_id= required).
        admin         → any unit (?manager_id= required).

    Everyone else is refused. Shift scope is enforced here (not only in the
    picker) so a shift-manager can't reach another shift by forging manager_id.
    """
    role = payload.get("role")
    if role == "supervisor":
        mid = payload.get("role_id")
        if not mid:
            raise HTTPException(status_code=403, detail="No unit assigned to this supervisor")
        return int(mid)
    if role in ("admin", "top-manager", "shift-manager"):
        if not requested:
            raise HTTPException(status_code=400, detail="manager_id is required")
        mid = int(requested)
        mgr = db.query(Manager).filter(Manager.id == mid, Manager.archived.is_(False)).first()
        if not mgr:
            raise HTTPException(status_code=404, detail=f"Manager {mid} not found")
        if role == "shift-manager" and mgr.shift != _shift_manager_shift(payload, db):
            raise HTTPException(status_code=403, detail="This unit is not in your shift")
        return mid
    raise HTTPException(status_code=403, detail="Not allowed to view production data")


def _constants(db: Session) -> tuple[float, float]:
    rows = {r.key: r.value for r in db.query(AppSetting).filter(
        AppSetting.key.in_(["pp_shift_min", "pp_productive_min"])).all()}
    def num(k, default):
        try:
            return float(rows[k])
        except (KeyError, ValueError, TypeError):
            return default
    return num("pp_shift_min", DEFAULT_SHIFT_MIN), num("pp_productive_min", DEFAULT_PRODUCTIVE_MIN)


def _parse_date(s: Optional[str]) -> date:
    if not s:
        return date.today()
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")


def _build_dashboard(db: Session, manager_id: int, day: date) -> dict:
    products = (
        db.query(PPProduct)
        .filter(PPProduct.manager_id == manager_id, PPProduct.active.is_(True))
        .order_by(PPProduct.sort_order, PPProduct.id)
        .all()
    )
    wcs = (
        db.query(PPWorkCenter)
        .filter(PPWorkCenter.manager_id == manager_id, PPWorkCenter.active.is_(True))
        .order_by(PPWorkCenter.sort_order, PPWorkCenter.id)
        .all()
    )
    daily = db.query(PPDaily).filter(PPDaily.manager_id == manager_id, PPDaily.date == day).all()
    wc_daily = db.query(PPWorkCenterDaily).filter(
        PPWorkCenterDaily.manager_id == manager_id, PPWorkCenterDaily.date == day).all()
    wc_overrides = {o.work_center: {"people": o.people, "shtatka": o.shtatka} for o in wc_daily}

    quantities: dict[tuple[str, str], dict] = {}
    for d in daily:
        plan_eff = d.plan_override if d.plan_override is not None else d.plan_qty
        actual_eff = d.actual_override if d.actual_override is not None else d.actual_qty
        quantities[(d.sap_code, d.work_center)] = {
            "plan_qty": float(plan_eff or 0),
            "actual_qty": float(actual_eff or 0),
            "plan_overridden": d.plan_override is not None,
            "actual_overridden": d.actual_override is not None,
        }

    shift_min, productive_min = _constants(db)
    result = compute_dashboard(
        products=[{
            "id": p.id,
            "sap_code": p.sap_code, "name": p.name, "work_center": p.work_center,
            "labor_time": (float(p.labor_time) if p.labor_time is not None else None),
            "sort_order": p.sort_order,
        } for p in products],
        quantities=quantities,
        work_centers=[{
            "code": w.code, "shtatka": w.shtatka,
            "capacity": (float(w.capacity) if w.capacity is not None else None),
            "sort_order": w.sort_order,
        } for w in wcs],
        shift_min=shift_min,
        productive_min=productive_min,
        wc_overrides=wc_overrides,
    )

    # SKUs present in the SAP snapshot but absent from the catalog
    catalog_keys = {(p.sap_code, p.work_center) for p in products}
    unknown = sorted({k for k in quantities if k not in catalog_keys})

    # Operation / phase code («Опер.») per (SKU, work center), surfaced from the
    # day's GLOBAL фаза upload (manager_id NULL) so the Positions table can show
    # it. The op lives only in the raw фаза rows ([order, op, wc, sku, …]); the
    # daily snapshot drops it. Distinct ops for one pair are joined with " / ".
    faza_up = db.query(PPUpload).filter(
        PPUpload.manager_id.is_(None), PPUpload.date == day,
        PPUpload.file_type == "faza").first()
    op_sets: dict[tuple[str, str], list[str]] = {}
    for fr in (faza_up.rows if faza_up else []):
        if len(fr) < 4:
            continue
        op, wc, sku = str(fr[1] or "").strip(), str(fr[2] or ""), str(fr[3] or "")
        if not op or not sku or sku == "—":
            continue
        ops = op_sets.setdefault((sku, wc), [])
        if op not in ops:
            ops.append(op)
    op_by_key = {k: " / ".join(sorted(v)) for k, v in op_sets.items()}
    for prow in result["rows"]:
        prow["op"] = op_by_key.get((prow["sap_code"], prow["work_center"]))

    recon = db.query(PPReconciliation).filter(
        PPReconciliation.manager_id == manager_id, PPReconciliation.date == day).first()

    mgr = db.query(Manager).filter(Manager.id == manager_id).first()
    result.update({
        "manager_id": manager_id,
        "manager_name": mgr.name if mgr else None,
        "date": day.isoformat(),
        "reconciliation": (recon.data if recon else {}),
        "unknown_skus": [{"sap_code": s, "work_center": w} for s, w in unknown],
        "missing_labor_count": sum(1 for r in result["rows"] if not r["has_labor"]),
    })
    return result


# --------------------------------------------------------------------------- #
# brigadir-facing
# --------------------------------------------------------------------------- #
@router.get("/api/production/dashboard")
def get_dashboard(
    date: Optional[str] = Query(None),
    manager_id: Optional[int] = Query(None),
    payload: dict = Depends(require_page(PAGE)),
    db: Session = Depends(get_db),
):
    mid = _resolve_manager_id(payload, manager_id, db)
    return _build_dashboard(db, mid, _parse_date(date))


class PositionsExportBody(BaseModel):
    date: Optional[str] = None
    manager_id: Optional[int] = None
    lang: str = "ru"
    # PPProduct ids in the EXACT on-screen order (post search/team-filter/sort)
    # at the moment the button was pressed. Empty → full default SAP order.
    order: list[int] = []
    # Kept for wire compatibility — the export is the fixed «загрузка» template
    # now, so the on-screen column picker no longer shapes the file.
    columns: list[str] = []


@router.post("/api/production/export.xlsx")
def export_positions(
    body: PositionsExportBody,
    payload: dict = Depends(require_page(PAGE)),
    db: Session = Depends(get_db),
):
    """Excel export = the brigadirs' manual «ABC форма» sheet, pre-filled from the
    day's dashboard and delivered to the caller's private Telegram chat.

    The layout mirrors that form cell-for-cell — totals in row 1, headers in row
    2, positions from row 3, the per-team block in M:W, indicators in X:Y and the
    staffing block in Z:AA. Only the true inputs are values (Трудоемкость, Команда,
    Факт, ПЛАН, Штатка and the reconciliation counts); every derived cell is a live
    formula — ЛЮДИ via VLOOKUP into the M:N team table, Общ.трудаёмкост =C*H/60,
    Парето against the I1 shift sum, per-team SUMIFS loading — so the file keeps
    recalculating as the brigadir edits it during the shift.

    Two deliberate departures from the manual form: the indicator block's three
    headcount figures are independent instead of all pointing at one cell (see the
    X:Y section), and division-prone cells are wrapped in IFERROR so the spare rows
    stay clean. Rows render in the exact order the client sends (`body.order`)."""
    lang = body.lang
    mid = _resolve_manager_id(payload, body.manager_id, db)
    day = _parse_date(body.date)
    dash = _build_dashboard(db, mid, day)
    rows = dash["rows"]
    if body.order:
        by_id = {r.get("id"): r for r in rows}
        rows = [by_id[i] for i in body.order if i in by_id]
    wcs = dash.get("work_centers") or []
    recon = dash.get("reconciliation") or {}
    sm, pm = _constants(db)
    sm, pm = int(sm), float(pm or DEFAULT_PRODUCTIVE_MIN)

    title_word = POSITIONS_TITLE.get(lang, POSITIONS_TITLE["ru"])
    mgr_name = dash.get("manager_name") or ""
    day_h = day.strftime("%d.%m.%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = title_word[:31]

    green_head = PatternFill("solid", fgColor="E2EFDA")   # header band / labels
    green_cell = PatternFill("solid", fgColor="C6E0B4")   # editable data area
    yellow = PatternFill("solid", fgColor="FFFF00")       # key inputs (date, Штатка)
    bold = Font(bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    head_al = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ds = ABC_DATA_START
    data_end = ds + len(rows) + ABC_SPARE_ROWS - 1        # incl. spare formula rows
    t0 = ABC_TEAM_START                                   # first team row
    t1 = t0 + max(len(wcs), 1) - 1                        # last team row
    ttot = t1 + 1                                         # SUM row under the block

    # --- row 1: shift totals (вып % = actual/plan minutes, and the two sums) --
    ws["E1"] = "=IFERROR(F1/I1,0)"
    ws["F1"] = f"=+SUM(F{ds}:F{data_end})"
    ws["H1"] = day                                        # the form's date cell
    ws["I1"] = f"=SUM(I{ds}:I{data_end})"
    for coord, nf in (("E1", "0%"), ("F1", "0.00"), ("H1", "DD.MMM"), ("I1", "#\\ ##0.00")):
        c = ws[coord]
        c.border, c.alignment, c.font, c.number_format = border, center, bold, nf
    ws["E1"].fill, ws["H1"].fill = green_cell, yellow

    # --- row 2: headers -------------------------------------------------------
    for i, h in enumerate(ABC_HEADERS, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font, c.alignment, c.border = bold, head_al, border
    for col, h in ((13, "Команда"), (14, "O. SONI"), (15, "Загруженность"),
                   (23, "Штатка"), (24, "Показатель"), (25, "Кол-во"),
                   (26, "Сколько должна на штатке")):
        c = ws.cell(row=2, column=col, value=h)
        c.font, c.alignment, c.border = bold, head_al, border
    ws.row_dimensions[2].height = 25.35

    # --- position rows (row 3+), then spare formula rows for hand-added SKUs --
    for idx, r in enumerate(rows):
        rn = ds + idx
        ws.cell(row=rn, column=1, value=r.get("sap_code") or "")
        ws.cell(row=rn, column=2, value=r.get("name") or "")
        ws.cell(row=rn, column=3, value=r.get("labor_time") if r.get("has_labor") else None)
        ws.cell(row=rn, column=4, value=r.get("work_center") or "")
        act, plan = r.get("actual_qty"), r.get("plan_qty")
        ws.cell(row=rn, column=7, value=act if act else None)
        ws.cell(row=rn, column=8, value=plan if plan else None)
    for rn in range(ds, data_end + 1):
        ws.cell(row=rn, column=5, value=f"=+IFERROR(VLOOKUP(D{rn},$M:$N,2,0),0)")   # ЛЮДИ
        ws.cell(row=rn, column=6, value=f"=C{rn}*G{rn}/60")                          # вып %
        ws.cell(row=rn, column=9, value=f"=C{rn}*H{rn}/60")                          # Общ.трудаёмкост
        ws.cell(row=rn, column=10, value=f"=IFERROR(I{rn}/E{rn},0)")                 # Минут
        ws.cell(row=rn, column=11, value=f"=+IFERROR(I{rn}/$I$1,0)")                 # Парето
        for cn in range(1, 12):
            c = ws.cell(row=rn, column=cn)
            c.border = border
            c.alignment = left if cn in (1, 2) else center
            if cn in (3, 4, 7, 8):        # the editable inputs
                c.fill = green_cell
        # Команда wears its identity colour instead of the plain editable green,
        # so teams read at a glance exactly as the chips do in the app's table.
        code = ws.cell(row=rn, column=4).value
        if code:
            fill, colour = _wc_style(str(code))
            dc = ws.cell(row=rn, column=4)
            dc.fill, dc.font = fill, Font(color=colour, bold=True)
        for cn, nf in ((3, "0.0"), (5, "0.0"), (6, "0.0"), (7, "0.0"), (8, "0"),
                       (9, "0.0"), (10, "0.0"), (11, "0%")):
            ws.cell(row=rn, column=cn).number_format = nf

    # --- M:W — per-team block (feeds the ЛЮДИ VLOOKUP) ------------------------
    # S (capacity) is emitted as =W*productive_min so that editing Штатка cascades
    # into load and headcount; a hand-tuned PPWorkCenter.capacity stays a literal.
    ws.cell(row=t0 - 1, column=19, value=f"Для {pm / sm:.0%} труд").alignment = center
    ws.cell(row=t0 - 1, column=23, value=f"=SUM(W{t0}:W{t1})").number_format = "0"
    ws.cell(row=t0 - 1, column=23).font = bold
    for idx in range(t1 - t0 + 1):
        rn = t0 + idx
        w = wcs[idx] if idx < len(wcs) else {}
        code = w.get("work_center") or ""
        shtatka = int(w.get("shtatka") or 0)
        cap = float(w.get("capacity") or 0)
        ws.cell(row=rn, column=13, value=code)                                        # M Команда
        ws.cell(row=rn, column=14, value=f"=ROUND(U{rn},0)")                          # N O. SONI
        ws.cell(row=rn, column=15, value=(                                            # O Загруженность
            f"=+IFERROR(SUMIFS($I:$I,$D:$D,$M{rn})/({sm}*VLOOKUP($M{rn},$D:$E,2,0)),0)"))
        ws.cell(row=rn, column=16, value=code)                                        # P Команда
        ws.cell(row=rn, column=17, value=f"=SUMIFS(I:I,D:D,P{rn})")                   # Q минут
        ws.cell(row=rn, column=18, value=f"=IFERROR(Q{rn}/S{rn},0)")                  # R real load
        hand_tuned = shtatka > 0 and abs(cap - shtatka * pm) > 0.01
        ws.cell(row=rn, column=19,                                                    # S capacity
                value=(round(cap, 2) if hand_tuned else f"=W{rn}*{pm:g}"))
        ws.cell(row=rn, column=20, value=f"=IFERROR(SUMIFS(S:S,P:P,M{rn})/({sm}*W{rn}),0)")  # T
        ws.cell(row=rn, column=21, value=f"=W{rn}*R{rn}")                             # U kerak (fract.)
        ws.cell(row=rn, column=23, value=shtatka or None)                             # W Штатка
        for cn in (13, 14, 15, 16, 17, 18, 19, 20, 21, 23):
            c = ws.cell(row=rn, column=cn)
            c.border, c.alignment = border, center
        ws.cell(row=rn, column=23).fill = yellow
        for cn, nf in ((14, "0.0"), (15, "0%"), (17, "0"), (18, "0.0%"),
                       (19, "0"), (20, "0.0%"), (21, "0.0"), (23, "0")):
            ws.cell(row=rn, column=cn).number_format = nf
    for cn, nf in ((14, "0.0"), (21, "0.0")):
        c = ws.cell(row=ttot, column=cn,
                    value=f"=SUM({get_column_letter(cn)}{t0}:{get_column_letter(cn)}{t1})")
        c.border, c.alignment, c.font, c.number_format = border, center, bold, nf

    # --- X:Y — indicator block ------------------------------------------------
    # The manual form points «keldi», «kelishi kerak edi» and «kerak» at the same
    # cell, which pins spare-people to 0, обеспеч to 100% and абсетеизм to 0%.
    # Here they are three independent figures:
    #   keldi           = AA8 «Сравнение» — roster minus brigadir/лидер/мицу/отдихает
    #   kelishi kerak   = AA3 «По штатке Факт» — the roster the day was planned on
    #   kerak           = N<ttot> — headcount the workload actually requires
    indicators = [
        ("Nechta odam keldi",                   "=AA8",                          "0"),
        ("Nechta odam kelishi kerak edi",       "=AA3",                          "0"),
        ("Nechta odam kerak",                   f"=N{ttot}",                     "0"),
        ("Bo`sh odam/kerakli odam",             "=+Y3-Y5",                       "0"),
        ("Kerakli odam bilan o`rtacha bandlik", f"=IFERROR(I1/(Y5*{sm}),0)",     "0%"),
        ("Hozirgi odam bilan o`rtacha bandlik", f"=IFERROR(I1/(Y3*{sm}),0)",     "0%"),
        ("% обеспеч",                           "=IFERROR(Y3/Y5,0)",             "0%"),
        ("% абсетеизм",                         '=IFERROR(1-(Y3/Y4),"")',        "0%"),
        ("Общ.трудаёмкост",                     "=I1",                           "#\\ ##0.00"),
    ]
    for idx, (label, formula, nf) in enumerate(indicators):
        rn = ds + idx
        xc = ws.cell(row=rn, column=24, value=label)
        yc = ws.cell(row=rn, column=25, value=formula)
        xc.font, xc.alignment, xc.border = bold, left, border
        yc.alignment, yc.border, yc.number_format = center, border, nf
        xc.fill = yc.fill = green_head

    # --- Z:AA — staffing (штатка) block, pre-filled from the reconciliation ---
    staffing = [
        ("По штатке Факт", recon.get("po_shtatke_fact")),
        ("Бригадир",       recon.get("brigadir")),
        ("Лидер",          recon.get("lider")),
        ("Мицу",           recon.get("mitsu")),
        ("Отдихает",       recon.get("otdihaet")),
        ("Сравнение",      "=AA3-AA4-AA5-AA6-AA7"),
        ("Верифекс",       "=AA8+AA6+AA5+AA4"),
    ]
    for idx, (label, val) in enumerate(staffing):
        rn = ds + idx
        zc = ws.cell(row=rn, column=26, value=label)
        ac = ws.cell(row=rn, column=27, value=val)
        zc.font, zc.alignment, zc.border = bold, center, border
        ac.alignment, ac.border, ac.number_format = center, border, "0"
        zc.fill = green_head
        if not (isinstance(val, str) and val.startswith("=")):
            ac.fill = yellow          # hand-entered counts

    ws.freeze_panes = "A3"
    for col, w in ABC_WIDTHS.items():
        ws.column_dimensions[col].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"{day_h} ABC форма {mgr_name}.xlsx" if mgr_name else f"{day_h} ABC форма.xlsx"
    from app.telegram_bot import bot
    caption = f"📊 {title_word}" + (f" — {mgr_name}" if mgr_name else "") + f"  •  {day_h}"
    try:
        bot.send_document(chat_id=int(payload["sub"]), document=(fname, bio.read()), caption=caption)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Telegram send failed: {e}")
    return {"ok": True}


@router.get("/api/production/dates")
def get_dates(
    manager_id: Optional[int] = Query(None),
    payload: dict = Depends(require_page(PAGE)),
    db: Session = Depends(get_db),
):
    mid = _resolve_manager_id(payload, manager_id, db)
    own = db.query(PPDaily.date).filter(PPDaily.manager_id == mid).distinct().all()
    # Dates with a stored GLOBAL SAP file count for every brigadir — the raw
    # фаза/заголовок views have content even when this unit got no PPDaily
    # rows out of that day's join.
    glob = db.query(PPUpload.date).filter(PPUpload.manager_id.is_(None)).distinct().all()
    dates = sorted({r[0] for r in own} | {r[0] for r in glob}, reverse=True)
    return {"dates": [d.isoformat() for d in dates]}


@router.get("/api/production/managers")
def list_production_managers(
    payload: dict = Depends(require_page(PAGE)),
    db: Session = Depends(get_db),
):
    """Brigadir units to offer in the dashboard picker, scoped to the caller:
    supervisor → only their own unit; shift-manager → configured units in their
    shift; top-manager / admin → every configured unit. 'Configured' = has an ABC
    catalog or an uploaded daily snapshot (see _configured_manager_ids)."""
    role = payload.get("role")
    q = db.query(Manager).filter(Manager.archived.is_(False))
    if role == "supervisor":
        mgrs = q.filter(Manager.id == payload.get("role_id")).all()
    elif role in ("admin", "top-manager", "shift-manager"):
        if role == "shift-manager":
            q = q.filter(Manager.shift == _shift_manager_shift(payload, db))
        configured = _configured_manager_ids(db)
        mgrs = [m for m in q.order_by(Manager.name).all() if m.id in configured]
    else:
        raise HTTPException(status_code=403, detail="Not allowed to view production data")
    return {"managers": [{"manager_id": m.id, "name": m.name, "shift": m.shift} for m in mgrs]}


class OverrideBody(BaseModel):
    date: str
    sap_code: str
    work_center: str
    field: str            # 'plan' | 'actual'
    value: Optional[float]  # null clears the override


@router.post("/api/production/override")
def set_override(
    body: OverrideBody,
    manager_id: Optional[int] = Query(None),
    payload: dict = Depends(require_page(PAGE)),
    db: Session = Depends(get_db),
):
    if body.field not in ("plan", "actual"):
        raise HTTPException(status_code=400, detail="field must be 'plan' or 'actual'")
    mid = _resolve_manager_id(payload, manager_id, db)
    day = _parse_date(body.date)

    row = db.query(PPDaily).filter(
        PPDaily.manager_id == mid, PPDaily.date == day,
        PPDaily.sap_code == body.sap_code, PPDaily.work_center == body.work_center,
    ).first()
    if not row:
        # allow overriding a row that has no SAP snapshot yet
        row = PPDaily(manager_id=mid, date=day, sap_code=body.sap_code,
                      work_center=body.work_center, plan_qty=0, actual_qty=0)
        db.add(row)

    if body.field == "plan":
        row.plan_override = body.value
    else:
        row.actual_override = body.value
    db.commit()
    return _build_dashboard(db, mid, day)


class WcOverrideBody(BaseModel):
    date: str
    work_center: str
    # BOTH fields are authoritative on every call — the staffing-card modal always
    # submits its two inputs together. null = clear the pin and fall back to the
    # computed N / configured штатка.
    people: Optional[int] = None
    shtatka: Optional[int] = None


@router.post("/api/production/wc-override")
def set_wc_override(
    body: WcOverrideBody,
    manager_id: Optional[int] = Query(None),
    payload: dict = Depends(_verify_admin),
    db: Session = Depends(get_db),
):
    """Pin «O. SONI» and/or «штатка» for one work center on ONE date.

    Admin-only: everyone else sees the staffing cards read-only. The pin lives in
    pp_work_center_daily, so the master pp_work_centers config and every other
    date keep their values — clearing both fields drops the row entirely."""
    mid = _resolve_manager_id(payload, manager_id, db)
    day = _parse_date(body.date)
    code = (body.work_center or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="work_center is required")
    for name, val in (("people", body.people), ("shtatka", body.shtatka)):
        if val is not None and not (0 <= val <= 9999):
            raise HTTPException(status_code=400, detail=f"{name} must be between 0 and 9999")

    row = db.query(PPWorkCenterDaily).filter(
        PPWorkCenterDaily.manager_id == mid, PPWorkCenterDaily.date == day,
        PPWorkCenterDaily.work_center == code,
    ).first()

    if body.people is None and body.shtatka is None:
        if row:
            db.delete(row)
    elif row:
        row.people, row.shtatka = body.people, body.shtatka
    else:
        db.add(PPWorkCenterDaily(manager_id=mid, date=day, work_center=code,
                                 people=body.people, shtatka=body.shtatka))
    db.commit()
    return _build_dashboard(db, mid, day)


class ReconciliationBody(BaseModel):
    date: str
    data: dict


@router.post("/api/production/reconciliation")
def save_reconciliation(
    body: ReconciliationBody,
    manager_id: Optional[int] = Query(None),
    payload: dict = Depends(require_page(PAGE)),
    db: Session = Depends(get_db),
):
    mid = _resolve_manager_id(payload, manager_id, db)
    day = _parse_date(body.date)
    row = db.query(PPReconciliation).filter(
        PPReconciliation.manager_id == mid, PPReconciliation.date == day).first()
    if not row:
        row = PPReconciliation(manager_id=mid, date=day, data=body.data or {})
        db.add(row)
    else:
        row.data = body.data or {}
    db.commit()
    return {"ok": True, "data": row.data}


# --------------------------------------------------------------------------- #
# admin
# --------------------------------------------------------------------------- #
def _upsert_upload(db, manager_id, day, file_type, columns, rows, filename):
    up = db.query(PPUpload).filter(
        PPUpload.manager_id == manager_id, PPUpload.date == day,
        PPUpload.file_type == file_type).first()
    if not up:
        up = PPUpload(manager_id=manager_id, date=day, file_type=file_type)
        db.add(up)
    up.columns = columns
    up.rows = rows
    up.row_count = len(rows)
    up.filename = filename


def _ingest_for_manager(db, manager_id: int, day: date, mode: str, *,
                        faza_ops: list[dict], order_sku: dict, order_deliv: dict) -> int:
    """Write ONE brigadir's PPDaily snapshot from the globally-parsed slices of
    an upload: filter фаза ops to their work centers, join → SKU
    (catalog-filtered), aggregate ПЛАН/ФАКТ. Returns rows written. The SAP
    export is one plant-wide file — upload_phase parses it once, stores the raw
    slices globally, and calls this per configured brigadir."""
    # Scope to this brigadir: own work centers (config ∪ catalog) and catalog SKUs.
    products = db.query(PPProduct).filter(PPProduct.manager_id == manager_id).all()
    own_wcs = {w.code for w in db.query(PPWorkCenter).filter(
        PPWorkCenter.manager_id == manager_id).all()} | {p.work_center for p in products}
    catalog_skus = {p.sap_code for p in products}

    # Join фаза operations → SKU, aggregate plan/actual by (SKU, work center).
    #   ПЛАН  = Σ «Кол-во операции» over the matching operations          (Excel col F)
    #   ФАКТ  = Σ order «Поставлено» over the matching operations          (Excel «План пост», col M)
    # «Поставлено» is order-level and repeats per operation, exactly like the
    # Excel SUMIFS over «План пост» — so we add it once per matching фаза row.
    faza_agg: dict[tuple[str, str], dict] = {}
    for op in faza_ops:
        if own_wcs and op["wc"] not in own_wcs:   # not this brigadir's work center
            continue
        sku = order_sku.get(op["order"])
        if sku and (not catalog_skus or sku in catalog_skus):
            a = faza_agg.setdefault((sku, op["wc"]), {"plan_qty": 0.0, "actual_qty": 0.0})
            a["plan_qty"] += op["plan"]
            a["actual_qty"] += order_deliv.get(op["order"], 0.0)

    updated = 0
    if faza_agg:
        # mode 'both' = fresh daily snapshot → replace the date (also clears overrides).
        if mode == "both":
            db.query(PPDaily).filter(PPDaily.manager_id == manager_id, PPDaily.date == day).delete()
            db.flush()
        for (sap, wc), agg in faza_agg.items():
            row = db.query(PPDaily).filter(
                PPDaily.manager_id == manager_id, PPDaily.date == day,
                PPDaily.sap_code == sap, PPDaily.work_center == wc).first()
            if not row:
                row = PPDaily(manager_id=manager_id, date=day, sap_code=sap, work_center=wc,
                              plan_qty=0, actual_qty=0)
                db.add(row)
            if mode in ("plan", "both"):
                row.plan_qty = agg["plan_qty"]
                row.plan_override = None      # SAP upload resets the manual override
            if mode in ("actual", "both"):
                row.actual_qty = agg["actual_qty"]
                row.actual_override = None
            updated += 1
    return updated


def _num(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _stored_slices(db, day: date) -> tuple[list[dict], dict, dict]:
    """Rebuild (faza_ops, order_sku, order_deliv) for a date from the PPUpload
    rows kept at upload time, so a brigadir configured AFTER the SAP files
    landed can be ingested without re-uploading them."""
    ups = db.query(PPUpload).filter(PPUpload.date == day).all()

    # The global (manager_id NULL) row is the whole plant file; legacy
    # per-brigadir slices are disjoint crops, so fall back to all of them.
    faza_ups = [u for u in ups if u.file_type == "faza"]
    faza_ups = [u for u in faza_ups if u.manager_id is None] or faza_ups
    faza_ops = [
        {"order": str(r[0]), "wc": str(r[2] or ""), "plan": _num(r[5])}
        for up in faza_ups for r in (up.rows or [])   # [order, op, wc, sku, name, plan, …]
        if len(r) >= 6 and r[0]
    ]

    order_sku: dict[str, str] = {}
    order_deliv: dict[str, float] = {}
    for up in sorted((u for u in ups if u.file_type == "zaga"),
                     key=lambda u: u.manager_id is not None):   # global first
        for r in (up.rows or []):     # [order, sku, plant, ordqty, deliv, conf, …]
            if len(r) >= 2 and r[0] and r[1]:
                order_sku.setdefault(str(r[0]), str(r[1]))
                if len(r) > 4:
                    order_deliv.setdefault(str(r[0]), _num(r[4]))
    return faza_ops, order_sku, order_deliv


def _backfill_manager(db, manager_id: int) -> dict:
    """Ingest EVERY date whose raw SAP slices are already stored, for a brigadir
    who has just been given a catalog. Without this, a фаза/заголовок upload
    that predates the catalog import leaves the unit with no pp_daily rows and
    the files would have to be uploaded again."""
    days = [d for (d,) in db.query(PPUpload.date).filter(
        PPUpload.file_type == "faza").distinct().order_by(PPUpload.date).all()]
    filled_days = filled_rows = 0
    for day in days:
        faza_ops, order_sku, order_deliv = _stored_slices(db, day)
        if not faza_ops:
            continue
        n = _ingest_for_manager(db, manager_id, day, "both", faza_ops=faza_ops,
                                order_sku=order_sku, order_deliv=order_deliv)
        if n:
            filled_days += 1
            filled_rows += n
    return {"days": filled_days, "rows": filled_rows}


@router.post("/admin/production/upload")
async def upload_phase(
    files: list[UploadFile] = File(...),
    manager_id: Optional[int] = Form(None),  # None → fan out to every configured brigadir
    date: str = Form(...),
    mode: str = Form("both"),           # 'plan' | 'actual' | 'both'
    file_type: Optional[str] = Form(None),  # 'faza' | 'zaga' | None (auto-detect)
    _: dict = Depends(_verify_admin),
    db: Session = Depends(get_db),
):
    if mode not in ("plan", "actual", "both"):
        raise HTTPException(status_code=400, detail="mode must be plan|actual|both")
    force_type = file_type if file_type in ("faza", "zaga") else None
    day = _parse_date(date)

    # The SAP фаза/заголовок export is ONE plant-wide file, so parse it once
    # (unfiltered — empty scope sets skip the per-brigadir filter) into global
    # slices, then fan out to each configured brigadir below.
    blobs = [(f.filename, await f.read()) for f in files]
    faza_ops: list[dict] = []          # raw operation dicts (no SKU yet)
    faza_dates: set = set()
    order_sku: dict[str, str] = {}     # order → SKU, from заголовок (global)
    order_deliv: dict[str, float] = {} # order → «Поставлено» (= Excel «План пост»), drives «Факт»
    zaga_rows_all: list[list] = []
    zaga_cols = None
    faza_present = zaga_present = False
    faza_file = zaga_file = None
    file_reports = []

    for name, content in blobs:
        slices = read_workbook_slices(content, day, set(), set(), force_type=force_type)
        rep = {"file": name, "faza": None, "zaga": None}
        fz = slices.get("faza")
        if fz is not None:
            faza_present = True
            faza_ops += fz["raw"]
            faza_dates.update(fz["dates"])
            faza_file = name
            rep["faza"] = {"operations": len(fz["raw"]), "dates": [d.isoformat() for d in fz["dates"]]}
        zg = slices.get("zaga")
        if zg is not None:
            zaga_present = True
            order_sku.update(zg["order_sku"])
            order_deliv.update(zg.get("order_deliv", {}))
            zaga_rows_all += zg["rows"]
            zaga_cols = zg["columns"]
            zaga_file = name
            rep["zaga"] = {"orders": len(zg["order_sku"]), "rows": len(zg["rows"])}
        file_reports.append(rep)

    if not faza_present and not zaga_present:
        raise HTTPException(
            status_code=400,
            detail="Не удалось распознать тип файла автоматически. Выберите «Тип файла» (Фаза или Заголовок) и загрузите снова.",
        )

    # A фаза-only upload still needs order→SKU: fall back to the заголовок
    # already stored for this date — the global row first, then any legacy
    # per-brigadir slices (order→SKU/«Поставлено» pairs are global truth).
    if faza_present and not zaga_present:
        stored = db.query(PPUpload).filter(
            PPUpload.date == day, PPUpload.file_type == "zaga",
        ).order_by(PPUpload.manager_id.isnot(None)).all()
        for up in stored:
            # stored zaga row: [order, sku, plant, ordqty, deliv, conf, date, name, status]
            for r in (up.rows or []):
                if len(r) >= 2 and r[0] and r[1]:
                    order_sku.setdefault(str(r[0]), str(r[1]))
                    if len(r) > 4:
                        order_deliv.setdefault(str(r[0]), float(r[4] or 0))

    # Store the raw slices ONCE, globally (manager_id NULL) — the file is
    # plant-wide; the raw views scope it to a brigadir at read time.
    if faza_present:
        faza_rows = [[op["order"], op["op"], op["wc"], order_sku.get(op["order"]) or "—",
                      op["name"], op["plan"], op["status"], op["date"], op["conf"]]
                     for op in faza_ops]
        _upsert_upload(db, None, day, "faza", FAZA_COLUMNS, faza_rows, faza_file)
    if zaga_present:
        _upsert_upload(db, None, day, "zaga", zaga_cols, zaga_rows_all, zaga_file)

    # Target: a specific brigadir if requested, else every configured one.
    if manager_id is not None:
        if not db.query(Manager).filter(Manager.id == manager_id).first():
            raise HTTPException(status_code=404, detail=f"Manager {manager_id} not found")
        targets = [manager_id]
    else:
        targets = sorted(_configured_manager_ids(db))
        if not targets:
            raise HTTPException(
                status_code=400,
                detail="Нет настроенных бригадиров — сначала импортируйте каталог хотя бы одному.",
            )

    total_rows = 0
    for mid in targets:
        total_rows += _ingest_for_manager(
            db, mid, day, mode, faza_ops=faza_ops, order_sku=order_sku,
            order_deliv=order_deliv)
    db.commit()
    return {
        "status": "ok", "date": day.isoformat(), "mode": mode,
        "brigadirs": len(targets), "rows_written": total_rows,
        "faza_operations": len(faza_ops) if faza_present else 0,
        "zaga_orders": len(order_sku),
        "files": file_reports,
    }


@router.get("/api/production/raw")
def get_raw(
    file_type: str = Query(...),       # 'faza' | 'zaga'
    date: Optional[str] = Query(None),
    manager_id: Optional[int] = Query(None),
    payload: dict = Depends(require_page(PAGE)),
    db: Session = Depends(get_db),
):
    if file_type not in ("faza", "zaga"):
        raise HTTPException(status_code=400, detail="file_type must be faza|zaga")
    mid = _resolve_manager_id(payload, manager_id, db)
    day = _parse_date(date)
    # Global row first (manager_id NULL = the plant-wide file), else the
    # brigadir's own legacy slice from before global storage.
    up = db.query(PPUpload).filter(
        PPUpload.manager_id.is_(None), PPUpload.date == day,
        PPUpload.file_type == file_type).first()
    is_global = up is not None
    if not up:
        up = db.query(PPUpload).filter(
            PPUpload.manager_id == mid, PPUpload.date == day,
            PPUpload.file_type == file_type).first()
    if not up:
        return {"present": False, "columns": [], "rows": [], "file_type": file_type, "date": day.isoformat()}
    rows = up.rows or []
    if is_global:
        # Scope the plant-wide file to this brigadir at read time — the same
        # filters legacy slices had baked in at upload time.
        products = db.query(PPProduct).filter(PPProduct.manager_id == mid).all()
        if file_type == "faza":
            # faza row: [order, op, wc, sku, name, plan, status, date, conf]
            own_wcs = {w.code for w in db.query(PPWorkCenter).filter(
                PPWorkCenter.manager_id == mid).all()} | {p.work_center for p in products}
            if own_wcs:
                rows = [r for r in rows if len(r) > 2 and r[2] in own_wcs]
        else:
            # zaga row: [order, sku, plant, ordqty, deliv, conf, date, name, status]
            catalog_skus = {p.sap_code for p in products}
            if catalog_skus:
                rows = [r for r in rows if len(r) > 1 and r[1] in catalog_skus]
    return {
        "present": True, "file_type": file_type, "date": day.isoformat(),
        "columns": up.columns, "rows": rows, "row_count": len(rows),
        "filename": up.filename,
        "uploaded_at": up.uploaded_at.isoformat() if up.uploaded_at else None,
    }


@router.post("/admin/production/catalog/import")
async def import_catalog(
    file: UploadFile = File(...),
    manager_id: int = Form(...),
    sheet_name: Optional[str] = Form(None),
    _: dict = Depends(_verify_admin),
    db: Session = Depends(get_db),
):
    """Replace a brigadir's catalog from an uploaded 'Sheet1 …' sheet: products
    (SKU, name, labor, work center) + work-center штатка/capacity. Junk '0' rows
    are dropped.

    Then re-derive pp_daily for every date whose raw SAP slices are stored, so
    a catalog imported AFTER the фаза/заголовок upload still produces numbers.
    That rewrites this brigadir's snapshots (and clears their manual overrides)
    on the backfilled dates — other brigadirs are untouched."""
    if not db.query(Manager).filter(Manager.id == manager_id).first():
        raise HTTPException(status_code=404, detail=f"Manager {manager_id} not found")
    parsed = parse_catalog_workbook(await file.read(), sheet_name)
    if not parsed["products"]:
        raise HTTPException(
            status_code=400,
            detail="Каталог не найден. Укажите имя листа (напр. «Sheet1 Торт») с колонками Трудоёмкость/Команда.",
        )

    db.query(PPProduct).filter(PPProduct.manager_id == manager_id).delete()
    for i, p in enumerate(parsed["products"]):
        db.add(PPProduct(
            manager_id=manager_id, sap_code=p["sap_code"], name=p.get("name") or "",
            work_center=p.get("work_center") or "", labor_time=p.get("labor_time"),
            sort_order=i,
        ))

    existing = {w.code: w for w in db.query(PPWorkCenter).filter(
        PPWorkCenter.manager_id == manager_id).all()}
    wc_added = wc_updated = 0
    for w in parsed["work_centers"]:
        wc = existing.get(w["code"])
        if wc:
            wc.shtatka = w.get("shtatka") or 0
            if w.get("capacity") is not None:
                wc.capacity = w["capacity"]
            wc_updated += 1
        else:
            db.add(PPWorkCenter(
                manager_id=manager_id, code=w["code"], shtatka=w.get("shtatka") or 0,
                capacity=w.get("capacity"), sort_order=w.get("sort_order", 0)))
            wc_added += 1
    db.commit()   # catalog must be visible to the scope queries in the backfill

    filled = _backfill_manager(db, manager_id)
    db.commit()
    return {
        "status": "ok", "manager_id": manager_id, "sheet": parsed["sheet"],
        "products": len(parsed["products"]),
        "work_centers_added": wc_added, "work_centers_updated": wc_updated,
        "backfilled_days": filled["days"], "backfilled_rows": filled["rows"],
    }


@router.get("/admin/production/work-centers")
def admin_work_centers(manager_id: int = Query(...), _: dict = Depends(_verify_admin),
                       db: Session = Depends(get_db)):
    wcs = db.query(PPWorkCenter).filter(PPWorkCenter.manager_id == manager_id).order_by(
        PPWorkCenter.sort_order, PPWorkCenter.id).all()
    return [{"id": w.id, "code": w.code, "shtatka": w.shtatka,
             "capacity": (float(w.capacity) if w.capacity is not None else None),
             "active": w.active} for w in wcs]


class WorkCenterBody(BaseModel):
    shtatka: Optional[int] = None
    capacity: Optional[float] = None


@router.put("/admin/production/work-centers/{wc_id}")
def admin_update_work_center(wc_id: int, body: WorkCenterBody,
                             _: dict = Depends(_verify_admin), db: Session = Depends(get_db)):
    w = db.query(PPWorkCenter).filter(PPWorkCenter.id == wc_id).first()
    if not w:
        raise HTTPException(status_code=404, detail="work center not found")
    if body.shtatka is not None:
        w.shtatka = body.shtatka
    if body.capacity is not None:
        w.capacity = body.capacity
    db.commit()
    return {"ok": True}


@router.get("/admin/production/catalog")
def admin_catalog(manager_id: int = Query(...), _: dict = Depends(_verify_admin),
                  db: Session = Depends(get_db)):
    rows = db.query(PPProduct).filter(PPProduct.manager_id == manager_id).order_by(
        PPProduct.sort_order, PPProduct.id).all()
    return [{"id": p.id, "sap_code": p.sap_code, "name": p.name,
             "work_center": p.work_center,
             "labor_time": (float(p.labor_time) if p.labor_time is not None else None),
             "active": p.active} for p in rows]


class CatalogCreateBody(BaseModel):
    manager_id: int
    sap_code: str
    name: Optional[str] = ""
    work_center: str
    labor_time: Optional[float] = None


@router.post("/admin/production/catalog")
def admin_create_catalog(body: CatalogCreateBody,
                         _: dict = Depends(_verify_admin), db: Session = Depends(get_db)):
    """Add a single catalog line (SKU) for a brigadir. sap_code + work_center are
    the (NOT NULL) join key onto the daily SAP snapshot, so both are required; the
    daily plan/fact rows join on that key at read time (no migration needed)."""
    if not db.query(Manager).filter(Manager.id == body.manager_id).first():
        raise HTTPException(status_code=404, detail=f"Manager {body.manager_id} not found")
    sap = (body.sap_code or "").strip()
    wc = (body.work_center or "").strip()
    if not sap:
        raise HTTPException(status_code=400, detail="sap_code cannot be empty")
    if not wc:
        raise HTTPException(status_code=400, detail="work_center cannot be empty")
    max_sort = db.query(func.max(PPProduct.sort_order)).filter(
        PPProduct.manager_id == body.manager_id).scalar() or 0
    p = PPProduct(
        manager_id=body.manager_id, sap_code=sap, name=(body.name or "").strip(),
        work_center=wc, labor_time=body.labor_time, sort_order=max_sort + 1,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return {"ok": True, "id": p.id}


class CatalogBody(BaseModel):
    labor_time: Optional[float] = None
    name: Optional[str] = None
    sap_code: Optional[str] = None
    work_center: Optional[str] = None
    active: Optional[bool] = None


@router.put("/admin/production/catalog/{prod_id}")
def admin_update_catalog(prod_id: int, body: CatalogBody,
                         _: dict = Depends(_verify_admin), db: Session = Depends(get_db)):
    p = db.query(PPProduct).filter(PPProduct.id == prod_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="product not found")
    if body.labor_time is not None:
        p.labor_time = body.labor_time
    if body.name is not None:
        p.name = body.name
    # sap_code + work_center are the (NOT NULL) join key onto the daily SAP
    # snapshot, so reject blanks. Renaming them re-points which SKU/unit this
    # catalog line tracks; the daily plan/fact rows join on the new key at read
    # time (they are keyed by the SAP upload, not by this row), so no migration.
    if body.sap_code is not None:
        sap = body.sap_code.strip()
        if not sap:
            raise HTTPException(status_code=400, detail="sap_code cannot be empty")
        p.sap_code = sap
    if body.work_center is not None:
        wc = body.work_center.strip()
        if not wc:
            raise HTTPException(status_code=400, detail="work_center cannot be empty")
        p.work_center = wc
    if body.active is not None:
        p.active = body.active
    db.commit()
    return {"ok": True}


@router.delete("/admin/production/catalog/{prod_id}")
def admin_delete_catalog(prod_id: int,
                         _: dict = Depends(_verify_admin), db: Session = Depends(get_db)):
    """Remove a single catalog line (SKU). The daily plan/fact rows join on the
    SAP snapshot key (sap_code + work_center), not on this row's id, so deleting a
    catalog line only drops it from the dashboard's SKU list — no daily data is
    destroyed and nothing else references it (no FK), so a hard delete is safe."""
    p = db.query(PPProduct).filter(PPProduct.id == prod_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="product not found")
    db.delete(p)
    db.commit()
    return {"ok": True}


# --------------------------------------------------------------------------- #
# Trudoyomkost analysis — cross-brigadir, by-weekday view + trend + Excel.
#
# Planned trudoyomkost is read straight from the synced *source* Google Sheet
# (admin → "Manba"): production_data.prod_plan holds planned production minutes
# per brigadir per day, for every brigadir in the sheet — not the SAP/ABC pilot.
# We fold each date onto its weekday and aggregate. Returns minutes; the client
# converts to norm-hours on the unit toggle.
# --------------------------------------------------------------------------- #
ANALYSIS_PAGE = "trudoyomkost"

WEEKDAY_LABELS = {
    "uz":      ["Du", "Se", "Cho", "Pay", "Ju", "Sha", "Yak"],
    "uz_cyrl": ["Ду", "Се", "Чо", "Пай", "Жу", "Ша", "Як"],
    "ru":      ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"],
    "en":      ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
}


def _date_strings(d_from: date, d_to: date) -> list[str]:
    """Inclusive list of 'DD.MM.YYYY' keys — production_data.date is stored as text."""
    out, cur = [], d_from
    while cur <= d_to:
        out.append(cur.strftime("%d.%m.%Y"))
        cur += timedelta(days=1)
    return out


def _load_plan_by_manager(db, manager_ids, shift, d_from, d_to) -> dict:
    """Planned trudoyomkost from the synced *source* sheet (admin → "Manba").

    production_data.prod_plan = planned production minutes per brigadir per day,
    for every brigadir in the sheet. Rows are keyed back to Manager.id by an exact
    name match, so non-brigadir rows (totals/categories) are dropped. An optional
    shift / manager_ids filter narrows the brigadir set (same as other endpoints).

    Returns {manager_id: {"name": str, "days": {date: {"plan": m, "actual": m}}}}.
    """
    managers = db.query(Manager).filter(Manager.archived.is_(False))
    if shift:
        managers = managers.filter(Manager.shift == shift)
    if manager_ids:
        managers = managers.filter(Manager.id.in_([int(x) for x in manager_ids]))
    by_name = {m.name: m for m in managers.all()}
    if not by_name:
        return {}

    # production_data spells brigadirs in either alphabet; accept every known
    # spelling and resolve each row back to its canonical Manager.
    alias = sheet_alias_map(db, by_name.keys())

    rows = db.query(ProductionData).filter(
        ProductionData.manager_name.in_(list(alias.keys())),
        ProductionData.date.in_(_date_strings(d_from, d_to)),
    ).all()

    out: dict = {}
    for r in rows:
        mgr = by_name.get(alias.get(r.manager_name))
        if not mgr:
            continue
        try:
            day = datetime.strptime(r.date, "%d.%m.%Y").date()
        except ValueError:
            continue
        e = out.setdefault(mgr.id, {"name": mgr.name, "days": {}})
        e["days"][day] = {"plan": float(r.prod_plan or 0), "actual": float(r.prod_actual or 0)}
    return out


def _trudoyomkost_payload(db, manager_ids, d_from, d_to, shift=None) -> dict:
    # Load the current window plus the preceding equal-length window in one pass,
    # so the Δ KPI reuses the same data.
    span = (d_to - d_from).days + 1
    prev_to = d_from - timedelta(days=1)
    prev_from = prev_to - timedelta(days=span - 1)
    loaded = _load_plan_by_manager(db, manager_ids, shift, prev_from, d_to)

    matrix: list[dict] = []
    profile_avgs = [[] for _ in range(7)]   # per weekday: each brigadir's weekday-avg
    profile_tot = [0.0] * 7
    daily_out: list[dict] = []
    period_total = 0.0
    prev_total = 0.0
    distinct_dates: set = set()

    for mid, entry in loaded.items():
        wd_plan = [[] for _ in range(7)]
        in_window = False
        for day, v in entry["days"].items():
            if prev_from <= day <= prev_to:
                prev_total += v["plan"]
                continue
            if not (d_from <= day <= d_to):
                continue
            in_window = True
            wd = day.weekday()
            wd_plan[wd].append(v["plan"])
            period_total += v["plan"]
            distinct_dates.add(day)
            daily_out.append({"manager_id": mid, "date": day.isoformat(),
                              "weekday": wd, "plan": v["plan"], "actual": v["actual"]})
        if not in_window:
            continue
        by_weekday, row_vals = [], []
        for wd in range(7):
            vals = wd_plan[wd]
            avg = (sum(vals) / len(vals)) if vals else 0.0
            by_weekday.append({"avg": avg, "total": sum(vals), "count": len(vals)})
            if vals:
                profile_avgs[wd].append(avg)
                profile_tot[wd] += sum(vals)
                row_vals.extend(vals)
        matrix.append({
            "manager_id": mid, "name": entry["name"],
            "by_weekday": by_weekday,
            "row_avg": (sum(row_vals) / len(row_vals)) if row_vals else 0.0,
            "row_total": sum(row_vals),
        })

    matrix.sort(key=lambda r: r["name"].lower())

    weekday_profile = [{
        "weekday": wd,
        "avg": (sum(profile_avgs[wd]) / len(profile_avgs[wd])) if profile_avgs[wd] else 0.0,
        "total": profile_tot[wd],
    } for wd in range(7)]

    n_dates = len(distinct_dates) or 1
    nonzero = [(wd, profile_tot[wd]) for wd in range(7) if profile_tot[wd] > 0]
    busiest = max(nonzero, key=lambda x: x[1]) if nonzero else (None, 0.0)
    lightest = min(nonzero, key=lambda x: x[1]) if nonzero else (None, 0.0)
    delta_pct = ((period_total - prev_total) / prev_total * 100.0) if prev_total > 0 else None

    return {
        "range": {"from": d_from.isoformat(), "to": d_to.isoformat(), "days": span},
        "supervisors": [{"id": m["manager_id"], "name": m["name"]} for m in matrix],
        "matrix": matrix,
        "weekday_profile": weekday_profile,
        "daily": daily_out,
        "kpis": {
            "period_total": period_total,
            "daily_avg": period_total / n_dates,
            "busiest_weekday": busiest[0], "busiest_value": busiest[1],
            "lightest_weekday": lightest[0], "lightest_value": lightest[1],
            "prev_total": prev_total, "delta_pct": delta_pct,
        },
        "unit": "min",
    }


def _parse_range(date_from: Optional[str], date_to: Optional[str]) -> tuple[date, date]:
    d_from, d_to = _parse_date(date_from), _parse_date(date_to)
    if d_to < d_from:
        raise HTTPException(status_code=400, detail="date_to must be on or after date_from")
    if (d_to - d_from).days > 370:
        raise HTTPException(status_code=400, detail="Range too large (max ~1 year)")
    return d_from, d_to


@router.get("/api/production/trudoyomkost")
def trudoyomkost_analysis(
    date_from: str = Query(...),
    date_to: str = Query(...),
    manager_id: list[int] = Query(default=[]),
    shift: Optional[int] = Query(None),
    payload: dict = Depends(require_page(ANALYSIS_PAGE, PAGE)),
    db: Session = Depends(get_db),
):
    d_from, d_to = _parse_range(date_from, date_to)
    return _trudoyomkost_payload(db, manager_id, d_from, d_to, shift)


@router.get("/api/production/trudoyomkost/export.xlsx")
def trudoyomkost_export(
    date_from: str = Query(...),
    date_to: str = Query(...),
    manager_id: list[int] = Query(default=[]),
    mode: str = Query("avg"),       # 'avg' | 'total'
    unit: str = Query("min"),       # 'min' | 'hrs'
    lang: str = Query("uz"),
    shift: Optional[int] = Query(None),
    send: int = Query(0),           # 1 = send to caller's Telegram chat instead of streaming
    payload: dict = Depends(require_page(ANALYSIS_PAGE, PAGE)),
    db: Session = Depends(get_db),
):
    d_from, d_to = _parse_range(date_from, date_to)
    data = _trudoyomkost_payload(db, manager_id, d_from, d_to, shift)

    labels = WEEKDAY_LABELS.get(lang, WEEKDAY_LABELS["uz"])
    div = 60.0 if unit == "hrs" else 1.0
    key = "total" if mode == "total" else "avg"
    rkey = "row_total" if mode == "total" else "row_avg"
    summary_label = "Jami" if mode == "total" else "O'rtacha"
    unit_label = "norm-soat" if unit == "hrs" else "min"

    wb = Workbook()
    ws = wb.active
    ws.title = "Trudoyomkost"

    gold = PatternFill("solid", fgColor="C8973F")
    head_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.append([f"Trudoyomkost — {summary_label} ({unit_label})  ·  {d_from.isoformat()} → {d_to.isoformat()}"])
    ws.append([f"Brigadir"] + labels + [summary_label])
    for c in ws[2]:
        c.fill, c.font, c.alignment, c.border = gold, head_font, center, border

    for row in data["matrix"]:
        vals = [(round(row["by_weekday"][wd][key] / div, 1) if row["by_weekday"][wd]["count"] else "")
                for wd in range(7)]
        ws.append([row["name"]] + vals + [round(row[rkey] / div, 1)])

    prof = data["weekday_profile"]
    foot_vals = [(round(prof[wd][key] / div, 1) if prof[wd]["total"] > 0 else "") for wd in range(7)]
    present = [prof[wd][key] for wd in range(7) if prof[wd]["total"] > 0]
    foot_summary = round(((sum(present) / len(present)) if mode != "total" else sum(present)) / div, 1) if present else 0
    ws.append([summary_label] + foot_vals + [foot_summary])

    for r in range(3, ws.max_row + 1):
        for c in ws[r]:
            c.border = border
            if c.column > 1:
                c.alignment = center
    ws.column_dimensions["A"].width = 26
    for col in range(2, 10):
        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 9
    ws.freeze_panes = "B3"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"trudoyomkost_{d_from.isoformat()}_{d_to.isoformat()}.xlsx"
    if send:
        from app.telegram_bot import bot
        caption = f"📊 Trudoyomkost — {summary_label} ({unit_label})  •  {d_from.isoformat()} → {d_to.isoformat()}"
        try:
            bot.send_document(chat_id=int(payload["sub"]), document=(fname, bio.read()), caption=caption)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Telegram send failed: {e}")
        return {"ok": True}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# --------------------------------------------------------------------------- #
# Trudoyomkost — worker prediction & statistics.
#
# Derives a *required worker count per brigadir per day* from the planned
# trudoyomkost (production_data.prod_plan, minutes) and runs the full statistical
# battery on it, folded onto weekday and month-phase, so a supervisor can predict
# how many workers to call for an upcoming shift and how confident that prediction
# is.
#
# Workers N = ROUND(prod_plan_min / 480) — total planned trudoyomkost divided by
# one worker's full standard shift (480 min). Because N is a constant × prod_plan,
# the relative-dispersion stats (CV, and hence the confidence rating) measure
# *plan stability*; layering in actual attendance is a later phase.
# --------------------------------------------------------------------------- #
SHIFT_STD_MIN = 480.0
CAPACITY_PER_WORKER_MIN = SHIFT_STD_MIN  # 480 — one worker = one full standard shift
MIN_SAMPLE = 3                       # below this a cell is "insufficient data"
FULL_SAMPLE = 6                      # below this, "high" is capped to "medium"
CV_HIGH = 0.10                       # CV < 0.10 → high · ≤ 0.20 → medium · else low
CV_MED = 0.20
MONTH_PHASES = (("early", 1, 10), ("mid", 11, 20), ("late", 21, 31))


def _capacity_min(capacity_pct: float | None) -> float:
    """Productive minutes one worker covers = capacity_pct% of the 480-min shift.
    100% → 480, 85% → 408. Falls back to the full shift for bad/empty input."""
    try:
        pct = float(capacity_pct)
    except (TypeError, ValueError):
        pct = 100.0
    pct = max(1.0, min(100.0, pct))
    return SHIFT_STD_MIN * pct / 100.0


def _workers_from_plan(plan_min: float, cap_min: float = CAPACITY_PER_WORKER_MIN) -> int:
    """Required workers for one shift from planned trudoyomkost minutes."""
    cap = cap_min if (cap_min and cap_min > 0) else CAPACITY_PER_WORKER_MIN
    return int(round((plan_min or 0.0) / cap))


def _phase_of(day: date) -> str:
    for name, lo, hi in MONTH_PHASES:
        if lo <= day.day <= hi:
            return name
    return "late"


def _confidence(n: int, cv: Optional[float]) -> str:
    """High/medium/low from coefficient of variation, gated by sample size."""
    if n < MIN_SAMPLE or cv is None:
        return "insufficient"
    if cv < CV_HIGH:
        base = "high"
    elif cv <= CV_MED:
        base = "medium"
    else:
        base = "low"
    if n < FULL_SAMPLE and base == "high":
        base = "medium"              # small-sample penalty
    return base


def _cell_stats(values: list[int]) -> dict:
    """Full stat battery for one (brigadir × weekday) or month-phase sample.

    recommend = round(median) (robust 'typical day'); band = mean ± σ.
    """
    n = len(values)
    if n == 0:
        return {"n": 0, "min": None, "max": None, "range": None, "mean": None,
                "median": None, "mode": None, "variance": None, "std": None,
                "cv": None, "confidence": "insufficient",
                "recommend": None, "band_lo": None, "band_hi": None}
    mn, mx = min(values), max(values)
    mean = statistics.mean(values)
    median = statistics.median(values)
    mode = statistics.multimode(values)[0]      # integer counts → mode is meaningful
    variance = statistics.variance(values) if n >= 2 else 0.0   # sample (n−1)
    std = statistics.stdev(values) if n >= 2 else 0.0
    cv = (std / mean) if mean > 0 else None
    return {
        "n": n, "min": mn, "max": mx, "range": mx - mn,
        "mean": round(mean, 2), "median": median, "mode": mode,
        "variance": round(variance, 2), "std": round(std, 2),
        "cv": (round(cv, 4) if cv is not None else None),
        "confidence": _confidence(n, cv),
        "recommend": int(round(median)),
        "band_lo": max(0, int(round(mean - std))),
        "band_hi": int(round(mean + std)),
    }


def _explained_fraction(all_values: list[int], groups: list[list[int]]) -> Optional[float]:
    """η² — fraction of total variance explained by a grouping (between-group SS
    ÷ total SS), using population variances. Higher = the grouping is the better
    predictor of daily worker count."""
    n = len(all_values)
    if n < 2:
        return None
    total_var = statistics.pvariance(all_values)
    if total_var == 0:
        return None
    within = sum(len(g) * (statistics.pvariance(g) if len(g) >= 2 else 0.0) for g in groups) / n
    return round(max(0.0, min(1.0, 1 - within / total_var)), 3)


def _worker_stats_payload(db, manager_ids, d_from, d_to, shift=None, capacity_pct=100.0) -> dict:
    span = (d_to - d_from).days + 1
    cap_min = _capacity_min(capacity_pct)
    loaded = _load_plan_by_manager(db, manager_ids, shift, d_from, d_to)

    supervisors: list[dict] = []
    cells: list[dict] = []
    by_sup: list[dict] = []
    wd_pool = [[] for _ in range(7)]     # per weekday: every day's worker count (all brigadirs)
    wd_cvs = [[] for _ in range(7)]      # per weekday: each brigadir's cell CV
    wd_predictable = [0] * 7
    wd_total_sup = [0] * 7
    daily_total: dict[date, int] = {}    # date → workers summed over brigadirs

    for mid, entry in sorted(loaded.items(), key=lambda kv: kv[1]["name"].lower()):
        name = entry["name"]
        wd_vals = [[] for _ in range(7)]     # (date, workers) pairs per weekday
        all_vals: list[int] = []
        for day, v in entry["days"].items():
            if not (d_from <= day <= d_to):
                continue
            w = _workers_from_plan(v["plan"], cap_min)
            wd = day.weekday()
            wd_vals[wd].append((day, w))
            all_vals.append(w)
            daily_total[day] = daily_total.get(day, 0) + w
        if not all_vals:
            continue
        supervisors.append({"id": mid, "name": name})

        sup_cvs, predictable_wds, rated = [], [], []
        for wd in range(7):
            vals = [w for _, w in wd_vals[wd]]
            st = _cell_stats(vals)
            # Чақириш mirrors the forecast table: moving average (mean ± σ) over
            # the FORECAST_WEEKS most recent same-weekday values — i.e. the call
            # for the *next* occurrence of this weekday — not the full-range median.
            ma = _cell_stats([w for _, w in sorted(wd_vals[wd])[-FORECAST_WEEKS:]])
            st["recommend"] = int(round(ma["mean"])) if ma["mean"] is not None else None
            st["band_lo"], st["band_hi"] = ma["band_lo"], ma["band_hi"]
            cells.append({"manager_id": mid, "name": name, "weekday": wd, **st})
            if st["n"] > 0:
                wd_pool[wd].extend(vals)
                wd_total_sup[wd] += 1
            if st["cv"] is not None:
                sup_cvs.append(st["cv"])
                wd_cvs[wd].append(st["cv"])
                rated.append((wd, st["cv"]))
            if st["confidence"] in ("high", "medium"):
                predictable_wds.append(wd)
                wd_predictable[wd] += 1
        mean_cv = (sum(sup_cvs) / len(sup_cvs)) if sup_cvs else None
        by_sup.append({
            "manager_id": mid, "name": name,
            "n_total": len(all_vals),
            "mean_workers": round(statistics.mean(all_vals), 1),
            "mean_cv": (round(mean_cv, 4) if mean_cv is not None else None),
            "confidence": _confidence(len(all_vals), mean_cv),
            "predictable_weekdays": predictable_wds,
            "best_weekday": (min(rated, key=lambda x: x[1])[0] if rated else None),
            "worst_weekday": (max(rated, key=lambda x: x[1])[0] if rated else None),
        })

    by_weekday = []
    for wd in range(7):
        vals, cvs = wd_pool[wd], wd_cvs[wd]
        mcv = (sum(cvs) / len(cvs)) if cvs else None
        by_weekday.append({
            "weekday": wd, "n": len(vals),
            "mean_workers": (round(statistics.mean(vals), 1) if vals else None),
            "mean_cv": (round(mcv, 4) if mcv is not None else None),
            "confidence": _confidence(len(vals), mcv),
            "predictable_supervisors": wd_predictable[wd],
            "total_supervisors": wd_total_sup[wd],
        })

    # month-phase + which grouping explains daily worker count better
    dates_sorted = sorted(daily_total)
    totals = [daily_total[d] for d in dates_sorted]
    phase_groups = {name: [] for name, _, _ in MONTH_PHASES}
    wd_groups: dict[int, list[int]] = {wd: [] for wd in range(7)}
    for d in dates_sorted:
        phase_groups[_phase_of(d)].append(daily_total[d])
        wd_groups[d.weekday()].append(daily_total[d])
    phases = [{"phase": name, **{k: _cell_stats(phase_groups[name])[k]
                                 for k in ("n", "min", "max", "mean", "median", "std", "cv")}}
              for name, _, _ in MONTH_PHASES]
    exp_wd = _explained_fraction(totals, list(wd_groups.values()))
    exp_ph = _explained_fraction(totals, list(phase_groups.values()))

    rated_sup = [s for s in by_sup if s["mean_cv"] is not None]
    rated_wd = [w for w in by_weekday if w["mean_cv"] is not None]
    overall = {
        "mean_daily_total_workers": (round(statistics.mean(totals), 1) if totals else None),
        "total_supervisors": len(supervisors),
        "distinct_days": len(dates_sorted),
        "most_predictable_supervisor": (min(rated_sup, key=lambda s: s["mean_cv"])["name"] if rated_sup else None),
        "least_predictable_supervisor": (max(rated_sup, key=lambda s: s["mean_cv"])["name"] if rated_sup else None),
        "most_predictable_weekday": (min(rated_wd, key=lambda w: w["mean_cv"])["weekday"] if rated_wd else None),
        "least_predictable_weekday": (max(rated_wd, key=lambda w: w["mean_cv"])["weekday"] if rated_wd else None),
    }

    return {
        "range": {"from": d_from.isoformat(), "to": d_to.isoformat(), "days": span},
        "capacity_per_worker_min": cap_min,
        "capacity_pct": round(_capacity_min(capacity_pct) / SHIFT_STD_MIN * 100.0, 1),
        "supervisors": supervisors,
        "cells": cells,
        "by_supervisor": by_sup,
        "by_weekday": by_weekday,
        "month_phase": {
            "phases": phases,
            "explained": {"weekday": exp_wd, "month_phase": exp_ph,
                          "winner": ("weekday" if (exp_wd or 0) >= (exp_ph or 0) else "month_phase")},
        },
        "overall": overall,
        "unit": "workers",
    }


@router.get("/api/production/trudoyomkost/worker-stats")
def trudoyomkost_worker_stats(
    date_from: str = Query(...),
    date_to: str = Query(...),
    manager_id: list[int] = Query(default=[]),
    shift: Optional[int] = Query(None),
    capacity_pct: float = Query(100.0, ge=1, le=100, description="Productive % of the 480-min shift one worker covers"),
    payload: dict = Depends(require_page(ANALYSIS_PAGE, PAGE)),
    db: Session = Depends(get_db),
):
    d_from, d_to = _parse_range(date_from, date_to)
    return _worker_stats_payload(db, manager_id, d_from, d_to, shift, capacity_pct)


# --------------------------------------------------------------------------- #
# Workers-to-call forecast — per brigadir × weekday, for one chosen week.
#
# For each (brigadir, weekday) of the selected week we forecast how many workers
# to call via a moving average over the SAME weekday in the FORECAST_WEEKS
# immediately-preceding weeks (default 3). The band is mean ± σ of those samples
# and the confidence reuses _confidence's CV rule (so a 3-sample MA tops out at
# "medium"). When the shown week's day already has loaded plan data we also
# return the actual worker count, letting the client compare forecast vs actual.
# --------------------------------------------------------------------------- #
FORECAST_WEEKS = 3   # moving-average window: same weekday over the last N weeks


def _monday_of(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _forecast_payload(db, manager_ids, week_start, weeks=FORECAST_WEEKS,
                      shift=None, capacity_pct=100.0) -> dict:
    week_start = _monday_of(week_start)
    week_end = week_start + timedelta(days=6)
    cap_min = _capacity_min(capacity_pct)
    # Pull the same-weekday history from the `weeks` preceding weeks together with
    # the shown week's own actuals, in one query.
    hist_start = week_start - timedelta(days=7 * weeks)
    loaded = _load_plan_by_manager(db, manager_ids, shift, hist_start, week_end)

    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    supervisors: list[dict] = []
    cells: list[dict] = []

    for mid, entry in sorted(loaded.items(), key=lambda kv: kv[1]["name"].lower()):
        name = entry["name"]
        days = entry["days"]               # {date: {"plan", "actual"}}
        supervisors.append({"id": mid, "name": name})
        for day in week_dates:
            wd = day.weekday()
            # same-weekday samples from the `weeks` preceding weeks (oldest→newest)
            samples = []
            for k in range(weeks, 0, -1):
                sd = day - timedelta(days=7 * k)
                v = days.get(sd)
                if v is not None:
                    samples.append({"date": sd.isoformat(),
                                    "workers": _workers_from_plan(v["plan"], cap_min)})
            st = _cell_stats([s["workers"] for s in samples])
            forecast = int(round(st["mean"])) if st["mean"] is not None else None
            av = days.get(day)
            actual = _workers_from_plan(av["plan"], cap_min) if av is not None else None
            cells.append({
                "manager_id": mid, "weekday": wd, "date": day.isoformat(),
                "forecast": forecast,
                "band_lo": st["band_lo"], "band_hi": st["band_hi"],
                "confidence": st["confidence"], "n": st["n"],
                "mean": st["mean"], "std": st["std"], "cv": st["cv"],
                "samples": samples, "actual": actual,
            })

    return {
        "week": {"start": week_start.isoformat(), "end": week_end.isoformat(),
                 "dates": [d.isoformat() for d in week_dates]},
        "weeks": weeks,
        "capacity_per_worker_min": cap_min,
        "supervisors": supervisors,
        "cells": cells,
        "unit": "workers",
    }


@router.get("/api/production/trudoyomkost/forecast")
def trudoyomkost_forecast(
    week_start: str = Query(..., description="Any date in the target week (ISO); snapped to Monday"),
    weeks: int = Query(FORECAST_WEEKS, ge=1, le=12, description="Moving-average window in weeks"),
    manager_id: list[int] = Query(default=[]),
    shift: Optional[int] = Query(None),
    capacity_pct: float = Query(100.0, ge=1, le=100, description="Productive % of the 480-min shift one worker covers"),
    payload: dict = Depends(require_page(ANALYSIS_PAGE, PAGE)),
    db: Session = Depends(get_db),
):
    return _forecast_payload(db, manager_id, _parse_date(week_start), weeks, shift, capacity_pct)


# --------------------------------------------------------------------------- #
# Call modal — one row per brigadir with the forecast for a target date (default
# tomorrow, any date pickable per shift section), whether the supervisor profile
# is claimed (can receive a Telegram DM), and the latest notice already sent for
# that date (resend guard). Deliberately UNfiltered on the page's brigadir
# filters so nobody is silently left out of the call; the ``shift`` param only
# serves the modal's per-shift sections, which together still cover everyone.
# The forecast counts DO follow the page's "Smena unumi" (shift-efficiency)
# setting via capacity_pct, so the numbers sent to supervisors match what the
# forecast/stats tables show.
# --------------------------------------------------------------------------- #
def _tomorrow() -> date:
    return date.today() + timedelta(days=1)


@router.get("/api/production/trudoyomkost/call-tomorrow")
def trudoyomkost_call_tomorrow(
    capacity_pct: float = Query(100.0, ge=1, le=100, description="Productive % of the 480-min shift one worker covers"),
    for_date: Optional[str] = Query(None, description="ISO target date; defaults to tomorrow"),
    shift: Optional[int] = Query(None, ge=1, le=2, description="Limit rows to one shift (the modal fetches per-shift sections)"),
    payload: dict = Depends(require_page(ANALYSIS_PAGE, PAGE)),
    db: Session = Depends(get_db),
):
    if for_date:
        try:
            target = date.fromisoformat(for_date)
        except ValueError:
            raise HTTPException(400, "Bad for_date")
    else:
        target = _tomorrow()
    cap_min = _capacity_min(capacity_pct)
    hist_start = target - timedelta(days=7 * FORECAST_WEEKS)
    loaded = _load_plan_by_manager(db, [], shift, hist_start, target)

    mq = db.query(Manager).filter(Manager.archived.is_(False))
    if shift is not None:
        mq = mq.filter(Manager.shift == shift)
    managers = mq.all()
    ids = [m.id for m in managers]
    # claimed supervisor profiles → a Telegram DM can actually reach someone
    claimed = {
        r.role_id for r in db.query(TelegramUserRole).filter(
            TelegramUserRole.role == "supervisor",
            TelegramUserRole.role_id.in_(ids),
            TelegramUserRole.status == "approved",
        )
    }
    # latest notice per unit for the target date (ascending order → last wins)
    last_notice: dict[int, ForecastCallNotice] = {}
    for n in db.query(ForecastCallNotice).filter(
        ForecastCallNotice.for_date == target,
        ForecastCallNotice.manager_id.in_(ids),
    ).order_by(ForecastCallNotice.sent_at):
        last_notice[n.manager_id] = n
    sender_names = {}
    sender_ids = {n.sent_by for n in last_notice.values()}
    if sender_ids:
        sender_names = {
            u.telegram_id: (u.tg_name or u.full_name)
            for u in db.query(TelegramUser).filter(TelegramUser.telegram_id.in_(sender_ids))
        }

    rows = []
    for m in sorted(managers, key=lambda m: m.name.lower()):
        days = loaded.get(m.id, {}).get("days", {})
        samples = [
            _workers_from_plan(days[sd]["plan"], cap_min)
            for k in range(FORECAST_WEEKS, 0, -1)
            if (sd := target - timedelta(days=7 * k)) in days
        ]
        st = _cell_stats(samples)
        ln = last_notice.get(m.id)
        rows.append({
            "manager_id": m.id, "name": m.name, "shift": m.shift,
            "forecast": int(round(st["mean"])) if st["mean"] is not None else None,
            "band_lo": st["band_lo"], "band_hi": st["band_hi"],
            "confidence": st["confidence"], "n": st["n"],
            "registered": m.id in claimed,
            "last_notice": {
                "workers": ln.workers,
                "sent_at": ln.sent_at.isoformat() if ln.sent_at else None,
                "by": sender_names.get(ln.sent_by),
            } if ln else None,
        })
    return {"date": target.isoformat(), "weeks": FORECAST_WEEKS, "rows": rows}


class CallNotifyItem(BaseModel):
    manager_id: int
    workers: int                   # recommended count (editable in the modal)
    max_workers: int | None = None # upper band (band_hi) shown as "Maksimum"
    date: str | None = None        # per-item target date (shift sections pick their own); falls back to request date


class CallNotifyRequest(BaseModel):
    date: str | None = None        # fallback target date for items without their own
    capacity_pct: float = 100.0    # page "Smena unumi" — shown in the DM as "Zagruzka foizi"
    items: list[CallNotifyItem]


@router.post("/api/production/trudoyomkost/call-notify")
def trudoyomkost_call_notify(
    req: CallNotifyRequest,
    payload: dict = Depends(require_page(ANALYSIS_PAGE, PAGE)),
    db: Session = Depends(get_db),
):
    # function-level import: staff.py is heavy and imports would be circular-prone
    from app.routers.staff import _notify_supervisor_all

    if not req.items:
        raise HTTPException(400, "No supervisors selected")
    # each item's date comes from its shift section's picker (explicit, so no
    # midnight-rollover guard needed); any calendar date is allowed by design
    targets: dict[int, date] = {}
    for i, item in enumerate(req.items):
        raw = item.date or req.date
        if not raw:
            raise HTTPException(400, "Missing date")
        try:
            targets[i] = date.fromisoformat(raw)
        except ValueError:
            raise HTTPException(400, "Bad date")

    actor = int(payload["sub"])
    eff = int(round(req.capacity_pct))   # Zagruzka % the counts were computed at
    by_id = {m.id: m for m in db.query(Manager).filter(
        Manager.id.in_([i.manager_id for i in req.items]),
        Manager.archived.is_(False),
    )}
    sent = []
    for i, item in enumerate(req.items):
        mgr = by_id.get(item.manager_id)
        if mgr is None or item.workers < 0:
            continue
        target = targets[i]
        # bell row keyed to the supervisor PROFILE (seen by every account holding
        # it) + a Telegram DM to EVERY holder of that profile, each in their own
        # language; an unclaimed profile queues the bell row for whoever claims it
        # later. Maksimum = the upper band; fall back to the recommended count when
        # the client didn't send one (older client / insufficient-data row).
        max_workers = item.max_workers if item.max_workers is not None else item.workers
        _notify_supervisor_all(
            db, mgr.id,
            nkey="call_forecast",
            params={"name": mgr.name, "date": target, "eff": eff, "count": item.workers, "max": max_workers},
        )
        db.add(ForecastCallNotice(manager_id=mgr.id, for_date=target,
                                  workers=item.workers, sent_by=actor))
        sent.append(mgr.id)
    db.commit()
    return {"sent": len(sent), "manager_ids": sent}
