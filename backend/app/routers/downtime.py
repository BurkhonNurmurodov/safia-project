from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.permissions import require_page
from app.models import (Cell, CellOjidaniyaInterval, DowntimeData, Factory, Manager,
                        RoleProfile)
from app.services.day_state import confirmed_pairs
from app.services.factory_scope import empty_scope, scoped_manager_ids
from app.services import (action_log, deck_narrative, idle_intervals, idle_source,
                          ojidaniya_deck, report_week)
from app.xlsx_delivery import PPTX_MIME, deliver_file, deliver_xlsx
from app.services.ojidaniya_export import build_ojidaniya_workbook
from app.services.name_map import sheet_alias_map
from app.services.sheets_reader import OJIDANIYA_ONLY_CATS

router = APIRouter(prefix="/api", tags=["downtime"])


@router.get("/downtime")
def get_downtime(
    date_from: date = Query(default=None),
    date_to: date = Query(default=None),
    shift: Optional[int] = Query(default=None),
    manager_id: List[int] = Query(default=[]),
    # kpi_only=1: strip the Ojidaniya-page-only categories (OJIDANIYA_ONLY_CATS)
    # from totals, breakdowns and flags. Every consumer of this endpoint OTHER
    # than the Ojidaniya page itself (today: the Daily performance block) must
    # pass it, so those categories exist nowhere outside /downtime and the
    # totals shown match the equip_downtime used by the загрузка KPIs.
    kpi_only: bool = Query(default=False),
    # Which plant. Omitted / null = «All factories». Ojidaniya rows key off the
    # supervisor NAME, so the factory narrows the manager set first and the
    # alias map then resolves only those names (services/factory_scope).
    factory: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("downtime", "daily")),
):
    if not date_to:
        date_to = date.today()
    if not date_from:
        date_from = date_to - timedelta(days=13)
    return _downtime(db, payload, date_from, date_to, shift, manager_id, kpi_only, factory)


def _downtime(db: Session, payload: dict, date_from: date, date_to: date,
              shift: Optional[int], manager_id: List[int], kpi_only: bool,
              factory: Optional[int]) -> dict:
    """The page's own numbers — ONE computation, read by the endpoint above and
    by the workbook export at the bottom of this file, so the file can never
    state a figure the screen it was pressed on does not."""
    scoped = scoped_manager_ids(db, payload, factory, manager_id)
    managers = db.query(Manager).filter(Manager.archived.is_(False))
    if shift:
        managers = managers.filter(Manager.shift == shift)
    if scoped is not None:
        managers = managers.filter(Manager.id.in_(scoped))
    managers = [] if empty_scope(scoped) else managers.all()
    # DowntimeData spells brigadirs in either alphabet; accept every known
    # spelling and resolve each row back to the canonical Manager.name.
    alias = sheet_alias_map(db, (m.name for m in managers))
    manager_names = set(alias.keys())

    cur = date_from
    dates = []
    while cur <= date_to:
        dates.append(cur.strftime("%d.%m.%Y"))
        cur += timedelta(days=1)

    dt_rows = db.query(DowntimeData).filter(
        DowntimeData.manager_name.in_(manager_names),
        DowntimeData.date.in_(dates),
    ).all()

    # Both halves of every category pair travel together: `total`/`by_category`
    # are the «тўхтаганда» numbers (the wait stopped the cell) and the `_ns`
    # fields the «тўхтамаганда» ones. The page tabs between them client-side, so
    # one fetch serves both and switching tabs costs no round-trip.
    dt_total: dict[str, dict[str, float]] = {}
    dt_by_cat: dict[str, dict[str, dict]] = {}
    dt_total_ns: dict[str, dict[str, float]] = {}
    dt_by_cat_ns: dict[str, dict[str, dict]] = {}
    cat_names_set: set[str] = set()
    for r in dt_rows:
        canon = alias.get(r.manager_name, r.manager_name)
        dt_total.setdefault(canon, {})[r.date] = float(r.total_minutes or 0)
        dt_by_cat.setdefault(canon, {})[r.date] = r.by_category or {}
        dt_total_ns.setdefault(canon, {})[r.date] = float(r.total_minutes_ns or 0)
        dt_by_cat_ns.setdefault(canon, {})[r.date] = r.by_category_ns or {}
        cat_names_set.update((r.by_category or {}).keys())
        cat_names_set.update((r.by_category_ns or {}).keys())

    # Every unit answers with its CELLS from `idle_source.CELLS_FROM` on —
    # earlier where the register switched one by hand — the headcount-weighted
    # mean of its cells' interval unions, and the sheet row for such a day is
    # dropped first: present or not, the sheet is no longer a source. The day-close gate
    # below is left exactly as it is: a switched day with nothing filed yet is
    # simply "not reported", so a confirmed one reads as a real zero and an open
    # one stays hidden, the same two answers the sheet model gives. A day with
    # at least one approved range on some cell is the unit's report for it.
    #
    # `kpi_only` is honoured HERE for the derived figures (the
    # Ojidaniya-only categories dropped before the union for `total`, stripped
    # from the breakdowns), so the subtract-after step in the loop below finds
    # nothing left to subtract.
    units = idle_source.cell_units(db)
    switched, lo = idle_source.switched_in_range(
        units, (m.id for m in managers), date_from, date_to)
    derived_days: list[tuple[str, str]] = []      # (manager name, d_str) overridden
    if switched:
        derived = idle_source.unit_downtime(db, switched, lo, date_to)
        by_id = {m.id: m for m in managers}
        for mid in switched:
            name = by_id[mid].name
            for d_str in dates:
                d_obj = datetime.strptime(d_str, "%d.%m.%Y").date()
                if not idle_source.uses_cells(units, mid, d_obj):
                    continue
                for bucket in (dt_total, dt_by_cat, dt_total_ns, dt_by_cat_ns):
                    bucket.get(name, {}).pop(d_str, None)
                row = derived.get((mid, d_obj.isoformat()))
                if not row or not row["cells_with_idle"]:
                    continue
                cats = dict(row["by_category"])
                cats_ns = dict(row["by_category_ns"])
                total = float(row["total"] if kpi_only else row["total_all"])
                total_ns = float(row["total_ns"])
                if kpi_only:
                    cats = {k: v for k, v in cats.items() if k not in OJIDANIYA_ONLY_CATS}
                    cats_ns = {k: v for k, v in cats_ns.items() if k not in OJIDANIYA_ONLY_CATS}
                    # A weighted plain sum is linear, so the not-stopped total
                    # without the Ojidaniya-only categories is exactly the sum
                    # of the remaining ones.
                    total_ns = float(sum(cats_ns.values()))
                dt_total.setdefault(name, {})[d_str] = total
                dt_by_cat.setdefault(name, {})[d_str] = cats
                dt_total_ns.setdefault(name, {})[d_str] = total_ns
                dt_by_cat_ns.setdefault(name, {})[d_str] = cats_ns
                cat_names_set.update(cats.keys())
                cat_names_set.update(cats_ns.keys())
                derived_days.append((name, d_str))

    cat_names = sorted(cat_names_set)
    if kpi_only:
        cat_names = [c for c in cat_names if c not in OJIDANIYA_ONLY_CATS]
    # A derived breakdown names only the categories that were filed; pad it to
    # the page's column set so every row carries every column, as sheet rows do.
    for name, d_str in derived_days:
        for c in cat_names:
            dt_by_cat[name][d_str].setdefault(c, 0.0)
            dt_by_cat_ns[name][d_str].setdefault(c, 0.0)

    # Day-close state — here it decides only whether an unreported day counts as
    # a reported zero (see the loop below), not whether reported data is shown.
    confirmed = confirmed_pairs(db, date_from, date_to, [m.id for m in managers])

    rows = []
    for mgr in sorted(managers, key=lambda m: m.name or ""):
        for d_str in dates:
            d_obj = datetime.strptime(d_str, "%d.%m.%Y").date()
            # The shift report is a source of its own: the brigadir submits it
            # once at end of shift and it carries no attendance, so a submitted
            # report shows as soon as it syncs — open day or not. The day-close
            # gate still governs the silent case: only on a confirmed day does
            # "no report" mean a real zero rather than "not reported yet".
            reported = d_str in dt_total.get(mgr.name, {})
            if not reported and (mgr.id, d_obj) not in confirmed:
                continue
            total = dt_total.get(mgr.name, {}).get(d_str, 0.0)
            cats = dt_by_cat.get(mgr.name, {}).get(d_str, {c: 0.0 for c in cat_names})
            total_ns = dt_total_ns.get(mgr.name, {}).get(d_str, 0.0)
            cats_ns = dt_by_cat_ns.get(mgr.name, {}).get(d_str, {c: 0.0 for c in cat_names})
            if kpi_only:
                total = max(total - sum(float(cats.get(c) or 0) for c in OJIDANIYA_ONLY_CATS), 0.0)
                total_ns = max(total_ns - sum(float(cats_ns.get(c) or 0) for c in OJIDANIYA_ONLY_CATS), 0.0)
                cats = {k: v for k, v in cats.items() if k not in OJIDANIYA_ONLY_CATS}
                cats_ns = {k: v for k, v in cats_ns.items() if k not in OJIDANIYA_ONLY_CATS}
            rows.append({
                # The unit's id rides beside its name so a reader can address
                # ONE unit — the detail endpoint below is keyed by id, and the
                # sheet spells brigadirs in two alphabets, so a name is not an
                # address. Additive: every earlier consumer reads the name.
                "manager_id": mgr.id,
                "manager_name": mgr.name,
                "shift": mgr.shift,
                "date": d_str,
                # Which source answered this day — the cells' interval model or
                # the «Смена отчёт» row. Additive; the export prints it.
                "source": "cells" if idle_source.uses_cells(units, mgr.id, d_obj) else "sheet",
                "total": total,
                "flagged": total > 50,
                "by_category": cats,
                "total_ns": total_ns,
                "flagged_ns": total_ns > 50,
                "by_category_ns": cats_ns,
            })

    summary: dict[str, dict] = {}
    for r in rows:
        n = r["manager_name"]
        if n not in summary:
            summary[n] = {
                "manager_id": r["manager_id"], "manager_name": n, "shift": r["shift"],
                "total": 0.0, "flagged_days": 0,
                "total_ns": 0.0, "flagged_days_ns": 0,
            }
        summary[n]["total"] += r["total"]
        summary[n]["total_ns"] += r["total_ns"]
        if r["flagged"]:
            summary[n]["flagged_days"] += 1
        if r["flagged_ns"]:
            summary[n]["flagged_days_ns"] += 1

    return {
        "dates": dates,
        "cat_names": cat_names,
        "rows": rows,
        "summary": sorted(summary.values(), key=lambda x: x["total"], reverse=True),
    }


@router.get("/downtime/seasonality")
def get_downtime_seasonality(
    year: Optional[int] = Query(default=None),
    shift: Optional[int] = Query(default=None),
    manager_id: List[int] = Query(default=[]),
    # Same meaning as on /downtime: drop the Ojidaniya-only categories so the
    # grid shows only the waiting that the загрузка KPIs count.
    kpi_only: bool = Query(default=False),
    factory: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("downtime", "daily")),
):
    """Category × calendar-month waiting minutes for one year (the Ojidaniya
    seasonality grid). Its own time axis — deliberately independent of the page
    date range, which still narrows nothing here beyond shift/supervisor — so a
    whole year is one small aggregate instead of ~11k daily rows on the wire.

    `col_totals` is the SUM OF THE CATEGORIES of that month, not the reported
    `total_minutes`: it is the denominator the grid's percentages divide by, so
    a column always adds up to 100%.
    """
    scoped = scoped_manager_ids(db, payload, factory, manager_id)
    managers = db.query(Manager).filter(Manager.archived.is_(False))
    if shift:
        managers = managers.filter(Manager.shift == shift)
    if scoped is not None:
        managers = managers.filter(Manager.id.in_(scoped))
    managers = [] if empty_scope(scoped) else managers.all()
    alias = sheet_alias_map(db, (m.name for m in managers))
    manager_names = set(alias.keys())

    if not manager_names:
        return {"years": [], "year": year, "cat_names": [],
                "col_totals": [0.0] * 12, "col_totals_ns": [0.0] * 12,
                "by_category": {}, "by_category_ns": {}}

    # Units read their CELLS from `idle_source.CELLS_FROM` on (earlier where
    # the register switched one by hand): their sheet rows on and after that
    # day are dropped and the derived per-day breakdown is added in their
    # place. The year list also carries the year each unit's switch starts in,
    # or a unit switched before its first sheet row would have no year to pick.
    units = idle_source.cell_units(db)
    mgr_ids = [m.id for m in managers]
    name_to_id = {m.name: m.id for m in managers}
    switch_years = {idle_source.start_day(units, mid).year for mid in mgr_ids}

    # Years that actually hold shift reports for the scoped supervisors, newest
    # first — the card's year selector.
    years = sorted({
        int(d[-4:]) for (d,) in db.query(DowntimeData.date)
        .filter(DowntimeData.manager_name.in_(manager_names)).distinct().all()
        if d and len(d) >= 4 and d[-4:].isdigit()
    } | switch_years, reverse=True)
    if year is None:
        today_year = date.today().year
        year = today_year if today_year in years else (years[0] if years else today_year)

    dt_rows = db.query(DowntimeData).filter(
        DowntimeData.manager_name.in_(manager_names),
        DowntimeData.date.like(f"%.{year}"),
    ).all()

    col_totals = [0.0] * 12
    col_totals_ns = [0.0] * 12
    by_cat: dict[str, list[float]] = {}
    by_cat_ns: dict[str, list[float]] = {}

    def _add(m: int, cats_stopped: dict, cats_ns: dict) -> None:
        for cats, cols, dest in (
            (cats_stopped or {}, col_totals, by_cat),
            (cats_ns or {}, col_totals_ns, by_cat_ns),
        ):
            for cat, val in cats.items():
                if kpi_only and cat in OJIDANIYA_ONLY_CATS:
                    continue
                v = float(val or 0)
                dest.setdefault(cat, [0.0] * 12)[m] += v
                cols[m] += v

    for r in dt_rows:
        try:
            d_obj = datetime.strptime(r.date or "", "%d.%m.%Y").date()
        except ValueError:
            continue
        canon = alias.get(r.manager_name, r.manager_name)
        mid = name_to_id.get(canon)
        if mid is not None and idle_source.uses_cells(units, mid, d_obj):
            continue                      # the cells answer this day, not the sheet
        _add(d_obj.month - 1, r.by_category, r.by_category_ns)

    y_lo, y_hi = date(year, 1, 1), date(year, 12, 31)
    switched, lo = idle_source.switched_in_range(units, mgr_ids, y_lo, y_hi)
    if switched:
        derived = idle_source.unit_downtime(db, switched, lo, y_hi)
        for (mid, iso), row in derived.items():
            d_obj = date.fromisoformat(iso)
            if not idle_source.uses_cells(units, mid, d_obj):
                continue
            _add(d_obj.month - 1, row["by_category"], row["by_category_ns"])

    return {
        "years": years,
        "year": year,
        "cat_names": sorted(set(by_cat) | set(by_cat_ns)),
        "col_totals": [round(v, 2) for v in col_totals],
        "col_totals_ns": [round(v, 2) for v in col_totals_ns],
        "by_category": {k: [round(v, 2) for v in vals] for k, vals in by_cat.items()},
        "by_category_ns": {k: [round(v, 2) for v in vals] for k, vals in by_cat_ns.items()},
    }


# ── What the unit's bar is MADE OF: the per-cell ojidaniya behind one supervisor ──
#
# The Ojidaniya page's supervisor bar is a number with no way in: it says a unit
# waited 464 minutes over a fortnight and nothing about which cell stopped, when,
# or why. From 2026-08-30 pressing the bar opens the detail — date by date, cell
# by cell, each cell's day drawn to scale over the table of its own events.
#
# Two rules this endpoint exists to keep:
#
# * **It serves the EVENTS, never a second answer to "how much".** The figure the
#   chart counted for a date stays the page's own row (`/api/downtime`'s `rows`,
#   which is where the headcount-weighted unit mean and the sheet row both come
#   from). This endpoint answers only "what did the cells file", so nothing here
#   can drift from the bar it was opened out of.
# * **A day the unit does NOT read from its cells has no per-cell answer at all.**
#   Before `idle_source.CELLS_FROM` (earlier where the register moved a unit) the
#   number came from the «Смена отчёт» row, which carries category minutes and no
#   endpoints — so such a day is named as a sheet day and its dates are listed,
#   rather than rendering as a cells day that nobody filed anything on. The two
#   are indistinguishable from emptiness alone, and reading one as the other is
#   how a reported day looks like a silent one.
@router.get("/downtime/cell-detail")
def get_downtime_cell_detail(
    manager_id: int = Query(...),
    date_from: date = Query(...),
    date_to: date = Query(...),
    # The page's own narrowings, mirrored: which half of every category pair is
    # on screen, whether the загрузка-only category set is in force, and the
    # doughnut's category picks. The modal is a zoom-in on one bar, so it must
    # be narrowed by exactly what narrowed that bar.
    stopped: bool = Query(default=True),
    kpi_only: bool = Query(default=False),
    cats: List[str] = Query(default=[]),
    factory: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("downtime", "daily")),
):
    """One supervisor's ojidaniya EVENTS for a period, grouped date → cell.

    Scoped exactly as the page is: a viewer who cannot see the unit on the chart
    cannot read its cells here either — `manager_id` is a query parameter, so the
    check is on the server and not on which bars were drawn."""
    scoped = scoped_manager_ids(db, payload, factory, [manager_id])
    if empty_scope(scoped) or (scoped is not None and manager_id not in scoped):
        raise HTTPException(status_code=403, detail="Out of scope")
    mgr = db.query(Manager).filter(Manager.id == manager_id).first()
    if not mgr:
        raise HTTPException(status_code=404, detail="Supervisor not found")

    detail = _cell_detail(db, manager_id, date_from, date_to, stopped, kpi_only, cats)
    return {
        "manager_id": manager_id,
        "manager_name": mgr.name,
        "shift": mgr.shift,
        **detail,
    }


def _cell_detail(db: Session, manager_id: int, date_from: date, date_to: date,
                 stopped: bool, kpi_only: bool, cats: List[str]) -> dict:
    """The unit's filed events per date, and WHICH dates its cells answer for.

    `cells_days` is not decorative: a cells day with nothing filed and a sheet
    day both arrive with no intervals, and only this list tells them apart."""
    units = idle_source.cell_units(db)
    cells_days = [
        d.isoformat()
        for d in (date_from + timedelta(days=i)
                  for i in range((date_to - date_from).days + 1))
        if idle_source.uses_cells(units, manager_id, d)
    ]
    cells = db.query(Cell).filter(Cell.manager_id == manager_id).all()
    if not cells or not cells_days:
        return {"cells_days": cells_days, "days": {}}

    wanted = set(cats) if cats else None
    rows = db.query(CellOjidaniyaInterval).filter(
        CellOjidaniyaInterval.cell_id.in_([c.id for c in cells]),
        CellOjidaniyaInterval.date.in_(cells_days),
        CellOjidaniyaInterval.status == "approved",
    ).all()

    lids = {c.leader_id for c in cells if c.leader_id}
    leaders = {p.id: p.name for p in db.query(RoleProfile).filter(
        RoleProfile.id.in_(lids)).all()} if lids else {}
    by_cell = {c.id: c for c in cells}

    # date -> cell_id -> [row]. Only the half on screen, only the categories the
    # page is counting: an event the bar did not count has no business being
    # totalled underneath it.
    grouped: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for e in rows:
        if bool(e.stopped) is not stopped:
            continue
        if kpi_only and e.category in OJIDANIYA_ONLY_CATS:
            continue
        if wanted is not None and e.category not in wanted:
            continue
        grouped[e.date][e.cell_id].append({
            "id": e.id,
            "category": e.category,
            "start": e.start,
            "end": e.end,
            "minutes": idle_intervals.duration(e.start, e.end),
            "stopped": bool(e.stopped),
            "note": e.note or "",
        })

    days: dict[str, list] = {}
    for d, per_cell in grouped.items():
        out = []
        for cid, ivs in per_cell.items():
            c = by_cell.get(cid)
            if not c:
                continue
            ivs.sort(key=lambda r: (idle_intervals.to_min(r["start"]) or 0, r["end"]))
            # The union of exactly the rows above — on the To'xtaganda half that
            # is the cell's downtime, the same figure the загрузка reads; on the
            # other it is how long the cell carried a wait it kept working
            # through. `sum_min` rides along so the header can say what the
            # overlap cost when the two differ.
            merged = idle_intervals.merged_spans(ivs, stopped_only=False)
            out.append({
                "cell_id": c.id,
                "code": c.verifix_code,
                "leader": leaders.get(c.leader_id),
                "total": sum(s["minutes"] for s in merged),
                "sum_min": sum(r["minutes"] for r in ivs),
                "intervals": ivs,
                "summary": {"merged": merged},
            })
        # Code order — the cell IS its code, so that is the order a reader scans.
        out.sort(key=lambda r: (r["code"] or "").lower())
        days[d] = out
    return {"cells_days": cells_days, "days": days}


# ── the same detail as a workbook ────────────────────────────────────────────
_DT_HEAD_FONT = Font(bold=True, size=10, color="FFFFFF")
_DT_HEAD_FILL = PatternFill("solid", fgColor="C8973F")
_DT_BODY_FONT = Font(size=10)
_DT_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_DT_RIGHT = Alignment(horizontal="right", vertical="center")
_DT_CENTER = Alignment(horizontal="center", vertical="center")

# key, fallback header, width, kind
_DT_COLS = [
    ("date",     "Sana",        12, "text"),
    ("cell",     "Yacheyka",    11, "text"),
    ("leader",   "Lider",       26, "text"),
    ("category", "Kategoriya",  14, "text"),
    ("start",    "Boshlandi",   11, "text"),
    ("end",      "Tugadi",      11, "text"),
    ("minutes",  "Daqiqa",      10, "num"),
    ("stopped",  "To'xtadi",    14, "text"),
    ("note",     "Izoh",        50, "text"),
    ("source",   "Manba",       18, "text"),
]


def _dt_xl(v):
    """A leader's name and a leader's free-text reason go into a cell verbatim,
    so two guards: a control character makes openpyxl raise (the whole export
    500s), and a leading = + - @ turns the text into a formula when it opens."""
    if v is None or not isinstance(v, str):
        return v
    v = ILLEGAL_CHARACTERS_RE.sub("", v)
    if v[:1] in ("=", "+", "-", "@"):
        v = "'" + v
    return v


class CellDetailExportBody(BaseModel):
    """The modal's own state, so the file is what the reader is looking at.
    `labels` carries the headers in the viewer's language — the sheet is read by
    the same person who read the screen, in the same words."""
    manager_id: int
    date_from: str
    date_to: str
    stopped: bool = True
    kpi_only: bool = False
    cats: List[str] = []
    factory: Optional[int] = None
    labels: dict[str, str] = {}
    caption: Optional[str] = None


@router.post("/downtime/cell-detail/export.xlsx")
def export_downtime_cell_detail(
    request: Request,
    body: CellDetailExportBody,
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("downtime", "daily")),
):
    """Excel of one supervisor's ojidaniya detail, exactly as the modal shows it
    — one row per EVENT on a cells day, and one row per category on a day that
    still came from the shift report, marked as such. A browser session
    downloads it; inside Telegram it lands in the caller's chat."""
    scoped = scoped_manager_ids(db, payload, body.factory, [body.manager_id])
    if empty_scope(scoped) or (scoped is not None and body.manager_id not in scoped):
        raise HTTPException(status_code=403, detail="Out of scope")
    mgr = db.query(Manager).filter(Manager.id == body.manager_id).first()
    if not mgr:
        raise HTTPException(status_code=404, detail="Supervisor not found")
    try:
        d_from = date.fromisoformat(body.date_from)
        d_to = date.fromisoformat(body.date_to)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date")

    L = body.labels or {}
    detail = _cell_detail(db, body.manager_id, d_from, d_to,
                          body.stopped, body.kpi_only, body.cats)
    cells_days = set(detail["cells_days"])
    yes = L.get("yes") or "Ha"
    no = L.get("no") or "Yo'q"
    src_cells = L.get("srcCells") or "Yacheykalar"
    src_sheet = L.get("srcSheet") or "Smena hisoboti"

    rows: list[dict] = []
    for d in sorted(detail["days"], reverse=True):
        for c in detail["days"][d]:
            for iv in c["intervals"]:
                rows.append({
                    "date": d, "cell": c["code"], "leader": c["leader"] or "",
                    "category": iv["category"], "start": iv["start"], "end": iv["end"],
                    "minutes": iv["minutes"], "stopped": yes if iv["stopped"] else no,
                    "note": iv["note"], "source": src_cells,
                })

    # The sheet days — no endpoints exist for them, so they carry the category
    # and its minutes and say plainly where the number came from. Leaving them
    # out would make a file shorter than the screen it claims to be.
    sheet_days = [d for d in (d_from + timedelta(days=i)
                              for i in range((d_to - d_from).days + 1))
                  if d.isoformat() not in cells_days]
    if sheet_days:
        alias = sheet_alias_map(db, [mgr.name])
        stamps = {d.strftime("%d.%m.%Y"): d.isoformat() for d in sheet_days}
        for r in db.query(DowntimeData).filter(
            DowntimeData.manager_name.in_(set(alias.keys())),
            DowntimeData.date.in_(list(stamps.keys())),
        ).all():
            cats = (r.by_category if body.stopped else r.by_category_ns) or {}
            for cat, val in sorted(cats.items()):
                if not float(val or 0):
                    continue
                if body.kpi_only and cat in OJIDANIYA_ONLY_CATS:
                    continue
                if body.cats and cat not in body.cats:
                    continue
                rows.append({
                    "date": stamps[r.date], "cell": "", "leader": "", "category": cat,
                    "start": "", "end": "", "minutes": round(float(val), 1),
                    "stopped": yes if body.stopped else no, "note": "", "source": src_sheet,
                })
    rows.sort(key=lambda r: (r["date"], r["cell"], r["start"]), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ojidaniya"
    for i, (key, fallback, width, _kind) in enumerate(_DT_COLS, 1):
        c = ws.cell(row=1, column=i, value=_dt_xl(L.get(key) or fallback))
        c.font, c.fill, c.alignment = _DT_HEAD_FONT, _DT_HEAD_FILL, _DT_CENTER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    for ri, r in enumerate(rows, 2):
        for ci, (key, _f, _w, kind) in enumerate(_DT_COLS, 1):
            v = r.get(key)
            c = ws.cell(row=ri, column=ci, value=_dt_xl(v) if kind == "text" else v)
            c.font = _DT_BODY_FONT
            if kind == "num":
                c.number_format = "0.#"
                c.alignment = _DT_RIGHT
            else:
                c.alignment = _DT_LEFT
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_DT_COLS))}{len(rows) + 1}"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"ojidaniya_{(mgr.name or str(mgr.id)).replace(' ', '_')}_{body.date_from}_{body.date_to}.xlsx"
    caption = body.caption or f"📄 {mgr.name} · {body.date_from} — {body.date_to}"
    blob = bio.read()
    resp = deliver_xlsx(request, payload, fname, blob, caption)
    action_log.enrich(
        target_kind="report", target_id=fname, target_name=mgr.name,
        details=[("file", fname), ("rows", len(rows)), ("size", len(blob)),
                 ("from_date", body.date_from), ("to_date", body.date_to),
                 ("half", "stopped" if body.stopped else "not_stopped"),
                 ("scope", "kpi_only" if body.kpi_only else "all")],
    )
    return resp


# ── the WHOLE page as a workbook ─────────────────────────────────────────────
#
# The toolbar's «Excel» button. Everything the page shows — the KPI strip, the
# per-brigadir bars, the category doughnut, the trend, the daily table and, one
# level down, every event the cells filed — under exactly the filters on screen.
#
# The client sends the SCOPE (its filter state) and the WORDS (names in the
# viewer's alphabet, labels, what each category means and which colour it
# wears). Every number is computed here, through `_downtime` — the same function
# the page reads — and `_cell_detail`, the same one the bar's modal reads. Then
# `services/ojidaniya_export.py` lays it out as a report: the formatting a person
# would otherwise do by hand before the file was fit to forward is done once,
# there.
class OjidaniyaExportBody(BaseModel):
    date_from: str
    date_to: str
    shift: Optional[int] = None
    manager_id: List[int] = []
    factory: Optional[int] = None
    stopped: bool = True
    kpi_only: bool = True
    cats: List[str] = []
    names: dict[str, str] = {}          # manager_id → display name (viewer's alphabet)
    cat_meta: dict[str, dict] = {}      # category → {label, note, color}
    cat_order: List[str] = []           # canonical A→Z order for the legend tab
    labels: dict[str, str] = {}
    sheets: dict[str, str] = {}
    meta: List[dict] = []               # [{label, value}] — the scope, in the viewer's words
    title: Optional[str] = None
    subtitle: Optional[str] = None
    caption: Optional[str] = None
    filename: Optional[str] = None


_EXPORT_MAX_DAYS = 400


@router.post("/downtime/export.xlsx")
def export_downtime(
    request: Request,
    body: OjidaniyaExportBody,
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("downtime", "daily")),
):
    """Excel report of the Ojidaniya page for the period and filters on screen.
    A browser session downloads it; inside Telegram it lands in the caller's chat."""
    try:
        d_from = date.fromisoformat(body.date_from)
        d_to = date.fromisoformat(body.date_to)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date")
    if d_from > d_to:
        raise HTTPException(status_code=400, detail="date_from is after date_to")
    if (d_to - d_from).days > _EXPORT_MAX_DAYS:
        raise HTTPException(status_code=400, detail=f"Period longer than {_EXPORT_MAX_DAYS} days")

    data = _downtime(db, payload, d_from, d_to, body.shift, body.manager_id,
                     body.kpi_only, body.factory)
    L = body.labels or {}
    tkey = "total" if body.stopped else "total_ns"
    ckey = "by_category" if body.stopped else "by_category_ns"
    fkey = "flagged" if body.stopped else "flagged_ns"
    cats = list(data["cat_names"])
    # The doughnut's picks narrow the bars to those categories' sum, exactly as
    # the chart does; the 50-min flag stays a fact about the unit's WHOLE day.
    wanted = [c for c in cats if c in body.cats] if body.cats else cats

    def disp(mid: int, name: str) -> str:
        return body.names.get(str(mid)) or name or ""

    def narrowed(row: dict) -> dict[str, float]:
        by = row.get(ckey) or {}
        return {c: float(by.get(c) or 0) for c in wanted}

    per: dict[int, dict] = {}
    matrix: dict[int, dict[str, float]] = defaultdict(dict)
    fleet: dict[str, float] = defaultdict(float)
    cat_tot: dict[str, float] = defaultdict(float)
    daily_rows: list[dict] = []
    for r in data["rows"]:
        iso = datetime.strptime(r["date"], "%d.%m.%Y").date().isoformat()
        by = narrowed(r)
        tot = float(sum(by.values())) if body.cats else float(r.get(tkey) or 0)
        mid = r["manager_id"]
        s = per.get(mid)
        if s is None:
            s = per[mid] = {
                "key": mid, "name": disp(mid, r["manager_name"]), "shift": r.get("shift"),
                "total": 0.0, "days": 0, "flagged_days": 0, "by_cat": defaultdict(float),
            }
        s["total"] += tot
        s["days"] += 1
        s["flagged_days"] += 1 if r.get(fkey) else 0
        for c, v in by.items():
            s["by_cat"][c] += v
            cat_tot[c] += v
        matrix[mid][iso] = round(tot, 1)
        fleet[iso] += tot
        daily_rows.append({
            "date": iso, "name": s["name"], "shift": r.get("shift"), "total": round(tot, 1),
            "flagged": bool(r.get(fkey)), "by_cat": {c: round(v, 1) for c, v in by.items()},
            "source": r.get("source"),
        })

    summary = sorted(per.values(), key=lambda x: -x["total"])
    for s in summary:
        s["total"] = round(s["total"], 1)
        s["avg"] = round(s["total"] / s["days"], 1) if s["days"] else 0
        top = max(s["by_cat"].items(), key=lambda kv: kv[1], default=(None, 0.0))
        s["top_cat"] = top[0] if top[1] > 0 else ""
        s["by_cat"] = dict(s["by_cat"])
    daily_rows.sort(key=lambda x: (-x["total"], x["name"]))
    daily_rows.sort(key=lambda x: x["date"], reverse=True)

    grand = round(sum(s["total"] for s in summary), 1)
    days_n = sum(s["days"] for s in summary)
    flagged_sups = sum(1 for s in summary if s["flagged_days"])
    top_all = max(cat_tot.items(), key=lambda kv: kv[1], default=(None, 0.0))
    worst = top_all[0] if top_all[1] > 0 else ""
    totals = {
        "total": grand, "days": days_n,
        "flagged_days": sum(s["flagged_days"] for s in summary),
        "avg": round(grand / days_n, 1) if days_n else 0, "top_cat": worst,
    }
    cat_share = sorted(
        ({"name": c, "label": (body.cat_meta.get(c) or {}).get("label", ""),
          "minutes": round(cat_tot.get(c, 0.0), 1), "counted": c not in OJIDANIYA_ONLY_CATS}
         for c in wanted),
        key=lambda x: -x["minutes"],
    )
    h, m = divmod(int(round(grand)), 60)
    hours_txt = f"{h} {L.get('unitHour', 'h')} {m} {L.get('unitMin', 'min')}"
    worst_meta = body.cat_meta.get(worst) or {}
    kpis = [
        {"value": grand, "label": L.get("kpiTotal", ""), "hint": hours_txt,
         "color": "#C8973F", "fmt": "#,##0.#"},
        {"value": flagged_sups, "label": L.get("kpiFlagged", ""), "hint": L.get("hintFlagged", ""),
         "color": "#ef4444" if flagged_sups else "#94a3b8"},
        {"value": worst or "—", "label": L.get("kpiWorst", ""),
         "hint": worst_meta.get("label") or L.get("hintWorst", ""),
         "color": worst_meta.get("color") or "#94a3b8"},
        {"value": len(summary), "label": L.get("kpiSups", ""), "hint": L.get("hintSups", ""),
         "color": "#6366f1"},
        {"value": days_n, "label": L.get("kpiDays", ""), "hint": L.get("hintDays", ""),
         "color": "#22c55e"},
        {"value": totals["avg"], "label": L.get("kpiAvg", ""), "hint": L.get("hintAvg", ""),
         "color": "#eab308", "fmt": "#,##0.#"},
    ]

    # One level down: every event the cells filed, for every unit on the chart —
    # the bar's own modal, for all bars at once. A day still read off the shift
    # report has no events, so it is one row per category, marked as such.
    events: list[dict] = []
    for s in summary:
        det = _cell_detail(db, s["key"], d_from, d_to, body.stopped, body.kpi_only, body.cats)
        for d, cells in det["days"].items():
            for c in cells:
                for iv in c["intervals"]:
                    events.append({
                        "date": d, "name": s["name"], "cell": c["code"] or "",
                        "leader": c["leader"] or "", "category": iv["category"],
                        "start": iv["start"], "end": iv["end"], "minutes": iv["minutes"],
                        "stopped": bool(iv["stopped"]), "note": iv["note"], "source": "cells",
                    })
    for r in daily_rows:
        if r["source"] != "sheet":
            continue
        for c, v in sorted(r["by_cat"].items()):
            if v:
                events.append({
                    "date": r["date"], "name": r["name"], "cell": "", "leader": "",
                    "category": c, "start": "", "end": "", "minutes": v,
                    "stopped": body.stopped, "note": "", "source": "sheet",
                })
    events.sort(key=lambda e: (e["name"], e["cell"], e["start"]))
    events.sort(key=lambda e: e["date"], reverse=True)

    dates_iso = [datetime.strptime(d, "%d.%m.%Y").date().isoformat() for d in data["dates"]]
    p = {
        "title": body.title or "Ojidaniya", "subtitle": body.subtitle or "",
        "sheets": body.sheets, "labels": L, "meta": body.meta, "kpis": kpis,
        "cats": wanted, "cat_meta": body.cat_meta,
        "cat_order": body.cat_order or list(body.cat_meta.keys()),
        "summary": summary, "totals": totals, "cat_share": cat_share,
        "dates": dates_iso, "matrix": {k: v for k, v in matrix.items()},
        "fleet_by_day": {k: round(v, 1) for k, v in fleet.items()},
        "daily_rows": daily_rows, "events": events,
    }
    bio = build_ojidaniya_workbook(p)
    blob = bio.read()

    fname = (body.filename or f"ojidaniya_{body.date_from}_{body.date_to}.xlsx").replace("/", "-").replace("\\", "-")
    if not fname.lower().endswith(".xlsx"):
        fname += ".xlsx"
    caption = body.caption or f"📊 Ojidaniya · {body.date_from} — {body.date_to}"
    resp = deliver_xlsx(request, payload, fname, blob, caption)
    action_log.enrich(
        target_kind="report", target_id=fname,
        details=[("file", fname), ("rows", len(daily_rows)), ("events", len(events)),
                 ("size", len(blob)), ("from_date", body.date_from), ("to_date", body.date_to),
                 ("half", "stopped" if body.stopped else "not_stopped"),
                 ("scope", "kpi_only" if body.kpi_only else "all"),
                 ("supervisors", len(summary))],
    )
    return resp


# ── the weekly deck ──────────────────────────────────────────────────────────
# One plant, by the operator's ruling (2026-09-03). Resolved by NAME rather
# than by id: ids differ between this checkout and production, and a seeded id
# goes stale the moment somebody rebuilds the register — the lesson
# `startup.seed_pp_autofill_default` already records about `startup.MANAGERS`.
DECK_FACTORY = "Uchtepa"


def _deck_factory(db: Session) -> Factory:
    """The plant the weekly deck is about, or a 500 that says what is wrong.

    A silent fallback to «all factories» would be the worst outcome available:
    the file would look right, carry another plant's units, and nothing on it
    would say so.
    """
    wanted = DECK_FACTORY.strip().casefold()
    rows = db.query(Factory).filter(Factory.archived.is_(False)).all()
    for f in rows:
        names = (f.code, f.name_uz, f.name_uz_cyrl, f.name_ru, f.name_en)
        if any((n or "").strip().casefold() == wanted for n in names):
            return f
    raise HTTPException(
        status_code=500,
        detail=(f"«{DECK_FACTORY}» zavodi topilmadi. Mavjud: "
                + ", ".join(f.code or f.name_uz or str(f.id) for f in rows)))


def _deck_events(db: Session, managers: list[Manager],
                 d_from: date, d_to: date) -> tuple[list[dict], list[dict]]:
    """Every event the plant's cells filed in the window, both halves, flat.

    `_cell_detail` answers for ONE unit and ONE half at a time — it is the bar
    modal's reader — so the deck asks it once per unit per half and flattens
    the result. The per-cell DAY unions come back alongside, because a cell's
    week is the sum of its daily unions and never the sum of its events: two
    causes overlapping on one clock would otherwise be counted twice.
    """
    events: list[dict] = []
    cell_days: list[dict] = []
    for m in managers:
        for stopped in (True, False):
            detail = _cell_detail(db, m.id, d_from, d_to, stopped, False, [])
            for iso, cells in detail["days"].items():
                d_obj = date.fromisoformat(iso)
                for c in cells:
                    cell_days.append({
                        "date": iso, "date_obj": d_obj, "cell": c["code"],
                        "leader": c.get("leader"), "supervisor": m.name,
                        "union_minutes": float(c["total"] or 0),
                        "sum_minutes": float(c["sum_min"] or 0),
                        "stopped": stopped,
                    })
                    for iv in c["intervals"]:
                        events.append({
                            "date": iso, "date_obj": d_obj,
                            "cell": c["code"], "leader": c.get("leader"),
                            "supervisor": m.name, "supervisor_id": m.id,
                            "shift": m.shift, "category": iv["category"],
                            "start": iv["start"], "end": iv["end"],
                            "minutes": float(iv["minutes"] or 0),
                            "stopped": bool(iv["stopped"]), "note": iv["note"] or "",
                        })
    return events, cell_days


@router.get("/downtime/deck-window")
def get_deck_window(payload: dict = Depends(require_page("downtime", "daily"))):
    """Which week the deck button is about to build, and for which plant.

    A label, but served rather than computed in the browser: the window is a
    rule (`services/report_week`), and a JavaScript copy of it would be a
    second spelling that drifts — the confirm would then name one week while
    the file carried another. Admin-only like the export it describes, so the
    button cannot even learn the scope it is not allowed to produce.
    """
    if payload.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Faqat administratorlar uchun")
    win = report_week.window()
    return {
        "date_from": win[0].isoformat(),
        "date_to": win[1].isoformat(),
        "label": report_week.label(win),
        "days": (win[1] - win[0]).days + 1,
        "factory": DECK_FACTORY,
    }


class DeckExportBody(BaseModel):
    """The deck takes no filters — it is a fixed weekly report, not a view of
    the page. The two dates exist only so an operator can re-run an EARLIER
    week; left out, `report_week` answers for the week that just closed."""
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    # Off only for a re-run when Gemini is down or its quota is spent and the
    # operator wants the numbers now. The deck already survives a failure on
    # its own; this skips the wait.
    narrative: bool = True


@router.post("/downtime/export.pptx")
def export_downtime_deck(
    request: Request,
    body: DeckExportBody,
    db: Session = Depends(get_db),
    payload: dict = Depends(require_page("downtime", "daily")),
):
    """The weekly Ojidaniya report as a PowerPoint deck.

    **Admin only**, checked here and not merely by hiding the button: the deck
    covers the whole plant — every unit's minutes, cells and note text — which
    /downtime deliberately withholds from a supervisor, and this endpoint is
    reachable without the UI.

    Scope is fixed and ignores whatever the page is filtered to: one plant,
    both shifts, every supervisor, all categories, the stopped half as the
    headline with the not-stopped half named beside it.
    """
    if payload.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Faqat administratorlar uchun")

    if body.date_from and body.date_to:
        try:
            win = (date.fromisoformat(body.date_from), date.fromisoformat(body.date_to))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date")
        if win[0] > win[1]:
            raise HTTPException(status_code=400, detail="date_from is after date_to")
        if (win[1] - win[0]).days > 31:
            raise HTTPException(status_code=400, detail="Period longer than 31 days")
    else:
        win = report_week.window()
    prev_win = report_week.previous(win)

    factory = _deck_factory(db)
    managers = (db.query(Manager)
                .filter(Manager.archived.is_(False), Manager.factory_id == factory.id)
                .order_by(Manager.name).all())

    cur = _downtime(db, payload, win[0], win[1], None, [], False, factory.id)
    prev = _downtime(db, payload, prev_win[0], prev_win[1], None, [], False, factory.id)
    events, cell_days = _deck_events(db, managers, win[0], win[1])

    data = ojidaniya_deck.collect(
        cur=cur, prev=prev, events=events, cell_days=cell_days,
        win=win, prev_win=prev_win,
        factory_name=(factory.name_uz or factory.code or DECK_FACTORY),
        supervisors=[{"id": m.id, "name": m.name, "shift": m.shift} for m in managers],
    )

    narrative = deck_narrative.write(data) if body.narrative else None
    deck = ojidaniya_deck.build(data, narrative)
    name = ojidaniya_deck.filename(data)

    action_log.enrich(
        request,
        target=f"{factory.name_uz or factory.code} · {report_week.label(win)}",
        detail={"events": data["events"], "minutes": round(data["total"]),
                "supervisors": data["sup_count"], "cells": data["cell_count"],
                "narrative": bool(narrative), "bytes": len(deck)},
    )
    return deliver_file(
        request, payload, name, deck, PPTX_MIME,
        caption=(f"📊 Yacheykalardagi kutish vaqtlari · {data['period']}\n"
                 f"{data['factory']} · {ojidaniya_deck.num(data['total'])} daqiqa · "
                 f"{data['events']} hodisa"))
