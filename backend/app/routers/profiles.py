"""
Pre-created profiles.

Admins create every identity here (Profiles tab); registration only binds one
of them to a Telegram account — nobody types a name at registration anymore.

Profile storage:
  supervisor              → the `managers` row itself (id = Verifix file id)
  top-manager / shift-manager / leader / admin / guest → `role_profiles`

Guest is the one exception to "admins create every identity": a guest profile
is auto-created (or re-claimed) during bot registration and only managed here
(rename / delete / unassign) — there is no admin "create guest" path.

Binding resolution (who holds a profile):
  supervisor      telegram_user_roles: role='supervisor',   role_id = manager.id
  shift-manager   telegram_user_roles: role='shift-manager',role_id = profile.id
  top-manager     telegram_user_roles: role='top-manager',  role_id = profile.id
  guest           telegram_user_roles: role='guest',        role_id = profile.id
  leader          telegram_user_roles: role='leader',
                  role_id = profile.manager_id AND full_name = profile.name
                  (leader role rows keep pointing at the unit — JWT/Concerns contract)
  admin           admins.profile_id = profile.id

Multilingual names: one canonical (Uzbek Latin) name per profile; per-language
display variants are `name.<canonical>` override keys in `translations`
(rendered by the frontend tl() helper, auto-transliterated when absent).
"""
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Annotated, Optional

from fastapi import (
    APIRouter, Depends, File, Form, HTTPException, Request, Response, UploadFile,
)
from fastapi.security import OAuth2PasswordBearer
import jwt
from jwt import PyJWTError as JWTError
from PIL import Image, ImageOps
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.orm import Session

from app import web_auth
from app.capability_alerts import alert_grant_use, tv, unit_name
from app.capabilities import CAP_CELLS_MANAGE, CAP_PROFILES_MANAGE, require_cap
from app.config import settings
from app.database import get_db
from app.identity import (
    find_role_row, parse_profile_key, profile_display_name, profile_holders,
    viewer_profile_key,
)
from app.permissions import require_page
from app.models import (
    Admin, Cell, Factory, LeaderConcern, LeaderTask, Manager, ProfilePhoto,
    RoleProfile, TelegramUser, TelegramUserRole, Translation, WebCredential,
)
from app.upload_guard import validate_avatar
from app.reg_token import validate_reg_token
from app.routers.auth import _validate_init_data
from app.xlsx_delivery import deliver_xlsx

router = APIRouter(prefix="/api/profiles", tags=["profiles"])
_oauth2 = OAuth2PasswordBearer(tokenUrl="/api/auth/webapp")

PROFILE_TYPES = {"top-manager", "shift-manager", "supervisor", "leader", "admin", "guest"}

# Every relational column that keys on managers.id — used when an admin
# re-keys a supervisor's Verifix ID. (JSONB payloads inside hr_documents may
# embed manager ids too; those are historical snapshots and stay untouched.)
_MANAGER_ID_REFS = [
    ("attendance", "manager_id"),
    ("comments", "manager_id"),
    ("edit_requests", "manager_id"),
    ("hr_documents", "manager_id"),
    ("day_approvals", "manager_id"),
    ("daily_submissions", "manager_id"),
    ("pp_products", "manager_id"),
    ("pp_work_centers", "manager_id"),
    ("pp_daily", "manager_id"),
    ("pp_reconciliation", "manager_id"),
    ("pp_uploads", "manager_id"),
    ("leader_concerns", "brigadir_manager_id"),
    ("leader_tasks", "supervisor_manager_id"),
    ("role_profiles", "manager_id"),
]


def _caller(token: Annotated[str, Depends(_oauth2)]) -> dict:
    try:
        return jwt.decode(token, settings.secret_key, algorithms=[settings.algorithm])
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")


def _deny_admin_profile(caller: dict, *ptypes: str | None) -> None:
    """The escalation stop: someone holding `admin.profiles.manage` may run the
    Profiles tab, but ADMIN identities are off-limits to them.

    Without this, a grantee could create an admin profile (or switch their own
    to admin), claim it, and become a full admin — turning one delegated job
    into total control. Real admins pass through untouched."""
    if caller.get("role") == "admin":
        return
    if any(p == "admin" for p in ptypes):
        raise HTTPException(status_code=403, detail="Only an admin can manage admin profiles")


# ── Shared helpers ────────────────────────────────────────────────────────────

def _rekey_name_overrides(db: Session, old_name: str, new_name: str) -> None:
    """Move `name.<old>` translation overrides to `name.<new>` so per-language
    display variants survive a canonical rename."""
    if not old_name or old_name == new_name:
        return
    old_key, new_key = f"name.{old_name}", f"name.{new_name}"
    taken = {(t.lang, t.key) for t in db.query(Translation).filter(Translation.key == new_key).all()}
    for t in db.query(Translation).filter(Translation.key == old_key).all():
        if (t.lang, new_key) in taken:
            db.delete(t)  # target already customised — keep it, drop the stale source
        else:
            t.key = new_key


def _remove_role_row(db: Session, role_row: TelegramUserRole) -> None:
    """Delete one role binding, mirroring admin.delete_user_role semantics:
    the user's last role takes the whole account with it, and a deleted active
    role falls back to the first remaining approved one."""
    tid = role_row.telegram_id
    ref = role_row.id
    db.delete(role_row)
    remaining = (
        db.query(TelegramUserRole)
        .filter(TelegramUserRole.telegram_id == tid, TelegramUserRole.id != ref)
        .all()
    )
    user = db.query(TelegramUser).filter_by(telegram_id=tid).first()
    if not remaining:
        if user:
            db.delete(user)
    elif user and user.active_role_ref == ref:
        approved = [r for r in remaining if r.status == "approved"]
        user.active_role_ref = approved[0].id if approved else None


def _bound_role_rows(db: Session, ptype: str, pid: int) -> list[TelegramUserRole]:
    """Every registration that holds a profile — the accounts working as this
    person. Drives the Holders column, unassign, and the switch-role impacts.

    Leaders match on the stamped profile_key first: matching on the name alone
    meant a holder whose name drifted from the profile silently vanished from
    this list while keeping full access, so an admin could neither see nor
    unassign them. Unstamped legacy rows still fall back to the name."""
    if ptype in ("supervisor", "shift-manager", "top-manager", "guest"):
        return db.query(TelegramUserRole).filter_by(role=ptype, role_id=pid).all()
    if ptype == "leader":
        p = db.query(RoleProfile).filter_by(id=pid, role="leader").first()
        if not p:
            return []
        key = f"leader:{pid}"
        return [
            r for r in db.query(TelegramUserRole).filter_by(
                role="leader", role_id=p.manager_id).all()
            if r.profile_key == key or (not r.profile_key and r.full_name == p.name)
        ]
    return []


def _rename_profile(db: Session, ptype: str, pid: int, new_name: str) -> str:
    """Canonical rename + full cascade (role rows, translation overrides).
    Returns the old name. Caller commits."""
    new_name = (new_name or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Name is required")

    if ptype == "supervisor":
        mgr = db.query(Manager).filter_by(id=pid).first()
        if not mgr:
            raise HTTPException(status_code=404, detail="Unit not found")
        old = mgr.name
        if old == new_name:
            return old
        mgr.name = new_name
        for r in db.query(TelegramUserRole).filter_by(role="supervisor", role_id=pid).all():
            r.full_name = new_name
    else:
        p = db.query(RoleProfile).filter_by(id=pid).first()
        if not p or p.role != ptype:
            raise HTTPException(status_code=404, detail="Profile not found")
        old = p.name
        if old == new_name:
            return old
        # Registration resolves profiles by name — block ambiguous duplicates
        # (leaders are scoped per unit, the rest per role). Guests are exempt:
        # their profiles resolve by id only, and two real people may share a
        # name.
        if ptype != "guest":
            dup = db.query(RoleProfile).filter(
                RoleProfile.role == ptype, RoleProfile.name == new_name, RoleProfile.id != pid,
            )
            if ptype == "leader":
                dup = dup.filter(RoleProfile.manager_id == p.manager_id)
            if dup.first():
                raise HTTPException(status_code=409, detail="Profile with this name already exists")
        for r in _bound_role_rows(db, ptype, pid):
            r.full_name = new_name
        p.name = new_name

    _rekey_name_overrides(db, old, new_name)
    return old


def _user_info(db: Session) -> dict[int, dict]:
    return {
        u.telegram_id: {"full_name": u.full_name, "username": u.username,
                        "phone": u.phone, "tg_name": u.tg_name,
                        # Can the bot DM this account? Stamped by the notifier
                        # when Telegram refuses permanently and cleared by the
                        # next delivered message (see telegram_bot
                        # _record_dm_outcome). An approved holder the bot cannot
                        # reach receives NOTHING while the app looks healthy —
                        # the Profiles tab is where that becomes visible.
                        "dm_ok": u.dm_failed_at is None,
                        "dm_error": u.dm_error,
                        "dm_failed_at": u.dm_failed_at.isoformat() if u.dm_failed_at else None}
        for u in db.query(TelegramUser).all()
    }


def _set_leader_cells(db: Session, leader_id: int, codes: list[str]) -> None:
    """Reconcile a leader's owned cells to exactly `codes`: release rows no
    longer listed (leader_id → NULL — cells are first-class rows now, their
    sap_code/workshop names survive reassignment), claim or create the rest.
    A code owned by ANOTHER leader is a 409 — cells are unique, reassign it
    from its current owner first. A claimed/created cell inherits the leader's
    supervisor unit; released rows keep their supervisor (a cell belongs to a
    supervisor with or without a leader)."""
    leader = db.query(RoleProfile).filter_by(id=leader_id).first()
    mgr_id = leader.manager_id if leader else None
    want: list[str] = []
    for c in codes or []:
        c = " ".join((c or "").split())
        if c and c not in want:
            want.append(c)
    existing = db.query(Cell).filter_by(leader_id=leader_id).all()
    for row in existing:
        if row.verifix_code not in want:
            row.leader_id = None
    have = {row.verifix_code for row in existing}
    for code in want:
        if code in have:
            continue
        row = db.query(Cell).filter_by(verifix_code=code).first()
        if row and row.leader_id not in (None, leader_id):
            owner = db.query(RoleProfile).filter_by(id=row.leader_id).first()
            raise HTTPException(
                status_code=409,
                detail=f"Cell {code} is already assigned to "
                       f"{owner.name if owner else 'another leader'}")
        if row:
            row.leader_id = leader_id
            if mgr_id:
                row.manager_id = mgr_id
        else:
            db.add(Cell(verifix_code=code, leader_id=leader_id, manager_id=mgr_id))


def _release_leader_cells(db: Session, leader_id: int) -> None:
    """Unassign every cell owned by the profile (delete + role switch away from
    leader). The rows stay — cell metadata outlives its owner."""
    db.query(Cell).filter_by(leader_id=leader_id).update({"leader_id": None})


def _manager_has_data(db: Session, manager_id: int) -> bool:
    for table, col in _MANAGER_ID_REFS:
        if table == "role_profiles":
            continue  # leader profiles are config, not history — deleted along
        row = db.execute(
            text(f"SELECT 1 FROM {table} WHERE {col} = :mid LIMIT 1"), {"mid": manager_id}
        ).first()
        if row:
            return True
    return False


# ── Admin: list ───────────────────────────────────────────────────────────────

@router.get("/admin/list")
def admin_list_profiles(db: Session = Depends(get_db),
                        _: dict = Depends(require_cap(CAP_PROFILES_MANAGE, CAP_CELLS_MANAGE))):
    users = _user_info(db)

    def binding(r: TelegramUserRole) -> dict:
        info = users.get(r.telegram_id, {})
        return {
            "role_ref":    r.id,
            "telegram_id": r.telegram_id,
            "status":      r.status,
            "user_name":   info.get("full_name"),
            "username":    info.get("username"),
            "tg_name":     info.get("tg_name"),
            "dm_ok":       info.get("dm_ok", True),
            "dm_error":    info.get("dm_error"),
            "dm_failed_at": info.get("dm_failed_at"),
        }

    role_rows = db.query(TelegramUserRole).order_by(TelegramUserRole.id).all()
    by_key: dict[tuple, list] = {}
    for r in role_rows:
        if r.status == "rejected":
            continue
        by_key.setdefault((r.role, r.role_id), []).append(r)

    # Browser logins, attached to the profile they belong to. `deliverable` is
    # computed from the bindings already in hand rather than re-querying per
    # row: a profile with no approved holder has nowhere to receive a password.
    creds = {c.profile_key: c for c in db.query(WebCredential).all()}

    # Avatar versions per profile — the frontend fetches bytes lazily and only
    # for rows that actually have a photo, so the list stays light.
    photo_vers = {
        key: int(ts.timestamp()) if ts else 1
        for key, ts in db.query(ProfilePhoto.profile_key, ProfilePhoto.updated_at).all()
    }

    def web_state(key: str, bindings: list[dict]) -> Optional[dict]:
        cred = creds.get(key)
        if not cred:
            return None
        return {
            "username":      cred.username,
            "enabled":       bool(cred.enabled),
            "last_login_at": cred.last_login_at.isoformat() if cred.last_login_at else None,
            "locked":        web_auth.lock_seconds_left(cred) > 0,
            "deliverable":   any(b.get("status") == "approved" for b in bindings),
            # Whether the password can be READ back at all. The value itself is
            # never in this list — it is fetched one profile at a time from the
            # reveal endpoint below, which is admin-only and audited.
            "has_password":  bool(cred.password_enc),
        }

    supervisors = []
    for m in db.query(Manager).order_by(Manager.id).all():
        sup_bindings = [binding(r) for r in by_key.get(("supervisor", m.id), [])]
        supervisors.append({
            "id": m.id, "name": m.name, "shift": m.shift, "archived": bool(m.archived),
            "has_data": _manager_has_data(db, m.id),
            "bindings": sup_bindings,
            "profile_key": f"supervisor:{m.id}",
            "web": web_state(f"supervisor:{m.id}", sup_bindings),
            "photo_ver": photo_vers.get(f"supervisor:{m.id}"),
        })

    profiles = db.query(RoleProfile).order_by(RoleProfile.id).all()
    mgr_names = {m.id: m.name for m in db.query(Manager).all()}
    admin_rows = db.query(Admin).all()
    admins_by_profile = {a.profile_id: a for a in admin_rows if a.profile_id}
    cell_rows = db.query(Cell).order_by(Cell.verifix_code).all()
    cells_by_leader: dict[int, list[str]] = {}
    for c in cell_rows:
        if c.leader_id:
            cells_by_leader.setdefault(c.leader_id, []).append(c.verifix_code)

    out = {"supervisors": supervisors, "top_managers": [], "shift_managers": [],
           "leaders": [], "admins": [], "guests": []}
    for p in profiles:
        item = {"id": p.id, "name": p.name, "name_uz_cyrl": p.name_uz_cyrl,
                "name_ru": p.name_ru, "name_en": p.name_en}
        if p.role == "top-manager":
            item["bindings"] = [binding(r) for r in by_key.get(("top-manager", p.id), [])]
            out["top_managers"].append(item)
        elif p.role == "guest":
            item["bindings"] = [binding(r) for r in by_key.get(("guest", p.id), [])]
            out["guests"].append(item)
        elif p.role == "shift-manager":
            item["shift"] = p.shift
            item["bindings"] = [binding(r) for r in by_key.get(("shift-manager", p.id), [])]
            out["shift_managers"].append(item)
        elif p.role == "leader":
            item["manager_id"] = p.manager_id
            item["supervisor"] = mgr_names.get(p.manager_id)
            item["cells"] = cells_by_leader.get(p.id, [])
            item["bindings"] = [
                binding(r) for r in by_key.get(("leader", p.manager_id), [])
                if r.full_name == p.name
            ]
            out["leaders"].append(item)
        elif p.role == "admin":
            a = admins_by_profile.get(p.id)
            info = users.get(a.telegram_id, {}) if a else {}
            item["bindings"] = [{
                "role_ref": None, "telegram_id": a.telegram_id, "status": "approved",
                "user_name": info.get("full_name"), "username": info.get("username"),
                "tg_name": info.get("tg_name"), "dm_ok": info.get("dm_ok", True),
                "dm_error": info.get("dm_error"),
                "dm_failed_at": info.get("dm_failed_at"),
            }] if a else []
            # pending /adminreg requests for this profile
            item["bindings"] += [
                binding(r) for r in db.query(TelegramUserRole)
                .filter_by(role="admin", role_id=p.id, status="pending").all()
            ]
            out["admins"].append(item)

        item["profile_key"] = f"{p.role}:{p.id}"
        item["web"] = web_state(item["profile_key"], item.get("bindings", []))
        item["photo_ver"] = photo_vers.get(item["profile_key"])

    out["assigned_admin_count"] = sum(1 for a in admin_rows)
    # Full cell registry (unassigned rows included) for the admin cells tab.
    prof_names = {p.id: p.name for p in profiles}
    out["cells"] = [{
        "id": c.id, "verifix_code": c.verifix_code, "sap_code": c.sap_code,
        "name_workshop_uz": c.name_workshop_uz,
        "name_workshop_uz_cyrl": c.name_workshop_uz_cyrl,
        "name_workshop_ru": c.name_workshop_ru,
        "name_workshop_en": c.name_workshop_en,
        "manager_id": c.manager_id, "supervisor": mgr_names.get(c.manager_id),
        "leader_id": c.leader_id, "leader": prof_names.get(c.leader_id),
    } for c in cell_rows]
    return out


# ── Admin: create ─────────────────────────────────────────────────────────────

class CreateProfilePayload(BaseModel):
    role:       str
    name:       str
    # Per-language display names (columns on role_profiles; `name` = uz Latin)
    name_uz_cyrl: Optional[str] = None
    name_ru:      Optional[str] = None
    name_en:      Optional[str] = None
    shift:      Optional[int] = None        # shift-manager | supervisor
    manager_id: Optional[int] = None        # leader → supervisor unit
    cells:      Optional[list[str]] = None  # leader → owned cell codes (optional)
    verifix_id: Optional[int] = None        # supervisor → managers.id


@router.post("/admin")
def admin_create_profile(payload: CreateProfilePayload, db: Session = Depends(get_db),
                         caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    _deny_admin_profile(caller, payload.role)
    role = payload.role
    name = (payload.name or "").strip()
    if role not in PROFILE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid profile type")
    # Guest profiles are the registration flow's own creation — no admin path.
    if role == "guest":
        raise HTTPException(status_code=400, detail="Guest profiles are created at registration")
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")

    if role == "supervisor":
        if not payload.verifix_id or payload.verifix_id <= 0:
            raise HTTPException(status_code=400, detail="Verifix ID is required")
        if payload.shift not in (1, 2):
            raise HTTPException(status_code=400, detail="Shift must be 1 or 2")
        if db.query(Manager).filter_by(id=payload.verifix_id).first():
            raise HTTPException(status_code=409, detail="Verifix ID already in use")
        db.add(Manager(id=payload.verifix_id, name=name, shift=payload.shift, archived=False))
        db.commit()
        alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.created",
                        details=[("role", tv("role.supervisor")), ("name", name),
                                 ("shift", payload.shift),
                                 ("verifix_code", payload.verifix_id)])
        return {"ok": True, "id": payload.verifix_id}

    if role == "shift-manager":
        if payload.shift not in (1, 2):
            raise HTTPException(status_code=400, detail="Shift must be 1 or 2")
        # Registration resolves shift-managers by name — duplicates would be ambiguous.
        if db.query(RoleProfile).filter_by(role=role, name=name).first():
            raise HTTPException(status_code=409, detail="Profile with this name already exists")
        p = RoleProfile(role=role, name=name, shift=payload.shift)
    elif role == "leader":
        mgr = db.query(Manager).filter_by(id=payload.manager_id).first()
        if not mgr:
            raise HTTPException(status_code=400, detail="Supervisor unit not found")
        dup = db.query(RoleProfile).filter_by(role="leader", name=name,
                                              manager_id=payload.manager_id).first()
        if dup:
            raise HTTPException(status_code=409, detail="This leader already exists")
        p = RoleProfile(role=role, name=name, manager_id=payload.manager_id)
    else:  # top-manager | admin
        if db.query(RoleProfile).filter_by(role=role, name=name).first():
            raise HTTPException(status_code=409, detail="Profile with this name already exists")
        p = RoleProfile(role=role, name=name)

    _apply_name_columns(p, payload)
    db.add(p)
    if role == "leader" and payload.cells:
        db.flush()  # p.id must exist before cells can point at it
        _set_leader_cells(db, p.id, payload.cells)
    db.commit()
    create_details = [("role", tv("role." + role)), ("name", name)]
    if role == "shift-manager":
        create_details.append(("shift", payload.shift))
    if role == "leader":
        create_details.append(("unit", unit_name(db, payload.manager_id)))
        if payload.cells:
            create_details.append(("cells", ", ".join(payload.cells)))
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.created",
                    details=create_details)
    return {"ok": True, "id": p.id}


# ── Admin: cells registry ─────────────────────────────────────────────────────
# Registered BEFORE the /admin/{ptype}/{pid} routes: the static "cells" segment
# must win over the ptype catch-all.

class CellPayload(BaseModel):
    verifix_code:          Optional[str] = None   # required on create
    sap_code:              Optional[str] = None   # "" clears; None = untouched
    name_workshop_uz:      Optional[str] = None
    name_workshop_uz_cyrl: Optional[str] = None
    name_workshop_ru:      Optional[str] = None
    name_workshop_en:      Optional[str] = None
    manager_id:            Optional[int] = None   # supervisor unit; 0 = clear; None = untouched
    leader_id:             Optional[int] = None   # 0 = unassign; None = untouched


_CELL_TEXT_COLS = ("sap_code", "name_workshop_uz", "name_workshop_uz_cyrl",
                   "name_workshop_ru", "name_workshop_en")

# Workshop-name columns as they appear in a capability-alert diff. Every one is
# nullable and each is edited on its own tab, so each is audited separately —
# Russian is the register's authoritative name (the language every other one
# falls back to on display), so it carries the plain "name" label while the
# optional translations get an explicit language tag.
_CELL_NAME_DIFF = {
    "name":      "name_workshop_ru",
    "name · UZ": "name_workshop_uz",
    "name · ЎЗ": "name_workshop_uz_cyrl",
    "name · EN": "name_workshop_en",
}


def _apply_cell_fields(db: Session, row: Cell, payload: CellPayload) -> None:
    for col in _CELL_TEXT_COLS:
        val = getattr(payload, col)
        if val is not None:
            setattr(row, col, val.strip() or None)
    # Supervisor unit: 0 clears, a positive id must be a real managers row.
    # Applied before the leader so an owner (below) can override it.
    if payload.manager_id is not None:
        if payload.manager_id:
            if not db.query(Manager).filter_by(id=payload.manager_id).first():
                raise HTTPException(status_code=400, detail="Supervisor unit not found")
            row.manager_id = payload.manager_id
        else:
            row.manager_id = None
    if payload.leader_id is not None:
        if payload.leader_id:
            p = db.query(RoleProfile).filter_by(id=payload.leader_id, role="leader").first()
            if not p:
                raise HTTPException(status_code=400, detail="Leader profile not found")
            row.leader_id = payload.leader_id
            # A cell inherits its owning leader's supervisor — the leader is
            # authoritative, overriding any manager_id in the same payload.
            if p.manager_id:
                row.manager_id = p.manager_id
        else:
            row.leader_id = None


@router.get("/admin/cells")
def admin_list_cells(db: Session = Depends(get_db),
                     _: dict = Depends(require_page("cells"))):
    """The /cells page's read half. Gated by PAGE access, not by the
    cells-manage capability: anyone who may SEE the page (``page.view.cells``,
    the ``admin.cells.manage`` edit grant which implies it, a role ticked for
    "cells" on the Access matrix, or a real admin) must be able to load the
    register they are allowed to look at. Every write stays on the
    POST/PUT/DELETE endpoints below, each guarded by CAP_CELLS_MANAGE.

    Deliberately lighter than /admin/list: only the cells plus the supervisor
    and leader option lists the table and edit modal need — never the full
    profile roster or the Telegram bindings that endpoint exposes. Registered
    before the /admin/{ptype}/{pid} routes so the static "cells" segment wins
    (Starlette matches in registration order)."""
    mgr_names = {m.id: m.name for m in db.query(Manager).all()}
    leader_profiles = (db.query(RoleProfile)
                       .filter(RoleProfile.role == "leader")
                       .order_by(RoleProfile.id).all())
    prof_names = {p.id: p.name for p in leader_profiles}
    cell_rows = db.query(Cell).order_by(Cell.verifix_code).all()
    return {
        "supervisors": [
            {"id": m.id, "name": m.name, "shift": m.shift, "archived": bool(m.archived)}
            for m in db.query(Manager).order_by(Manager.id).all()
        ],
        "leaders": [
            {"id": p.id, "name": p.name, "manager_id": p.manager_id}
            for p in leader_profiles
        ],
        "cells": [{
            "id": c.id, "verifix_code": c.verifix_code, "sap_code": c.sap_code,
            "name_workshop_uz": c.name_workshop_uz,
            "name_workshop_uz_cyrl": c.name_workshop_uz_cyrl,
            "name_workshop_ru": c.name_workshop_ru,
            "name_workshop_en": c.name_workshop_en,
            "manager_id": c.manager_id, "supervisor": mgr_names.get(c.manager_id),
            "leader_id": c.leader_id, "leader": prof_names.get(c.leader_id),
        } for c in cell_rows],
    }


# ── /cells register → formatted Excel report ──────────────────────────────────
# The page ships what it is SHOWING (filtered + sorted, names already resolved
# to the viewer's language and transliterated by tl()), so the workbook mirrors
# the screen exactly instead of re-deriving a second, subtly different register.
# Unassigned slots arrive as "" and the placeholder label is applied here, so
# "no brigadir" stays visually distinct from a real name in the sheet.

_XLSX_TASHKENT = timezone(timedelta(hours=5))

# Report chrome, per language. Deliberately a small closed set: the workbook is
# a document, not a UI, so it needs only its own labels.
_CELLS_XLSX_T = {
    "uz": {
        "sheet": "Yacheykalar", "sum_sheet": "Umumiy",
        "title": "YACHEYKALAR REYESTRI", "sum_title": "BRIGADIRLAR BO'YICHA UMUMIY",
        "generated": "Shakllantirildi", "records": "Yozuvlar",
        "shown_of": "{n} ta ko'rsatilgan ({total} tadan)",
        "num": "№", "verifix": "Verifix kod", "sap": "SAP kod",
        "workshop": "Sex nomi", "brigadir": "Brigadir", "leader": "Lider",
        "no_brigadir": "Brigadir yo'q", "unassigned": "Biriktirilmagan",
        "cells_cnt": "Yacheykalar", "with_leader": "Lider bilan",
        "without_leader": "Lidersiz", "coverage": "Qamrov", "total": "JAMI",
    },
    "uz_cyrl": {
        "sheet": "Ячейкалар", "sum_sheet": "Умумий",
        "title": "ЯЧЕЙКАЛАР РЕЕСТРИ", "sum_title": "БРИГАДИРЛАР БЎЙИЧА УМУМИЙ",
        "generated": "Шакллантирилди", "records": "Ёзувлар",
        "shown_of": "{n} та кўрсатилган ({total} тадан)",
        "num": "№", "verifix": "Verifix код", "sap": "SAP код",
        "workshop": "Сех номи", "brigadir": "Бригадир", "leader": "Лидер",
        "no_brigadir": "Бригадир йўқ", "unassigned": "Бириктирилмаган",
        "cells_cnt": "Ячейкалар", "with_leader": "Лидер билан",
        "without_leader": "Лидерсиз", "coverage": "Қамров", "total": "ЖАМИ",
    },
    "ru": {
        "sheet": "Ячейки", "sum_sheet": "Сводка",
        "title": "РЕЕСТР ЯЧЕЕК", "sum_title": "СВОДКА ПО БРИГАДИРАМ",
        "generated": "Сформировано", "records": "Записей",
        "shown_of": "показано {n} из {total}",
        "num": "№", "verifix": "Verifix код", "sap": "SAP код",
        "workshop": "Название цеха", "brigadir": "Бригадир", "leader": "Лидер",
        "no_brigadir": "Бригадир не назначен", "unassigned": "Не закреплена",
        "cells_cnt": "Ячеек", "with_leader": "С лидером",
        "without_leader": "Без лидера", "coverage": "Покрытие", "total": "ИТОГО",
    },
    "en": {
        "sheet": "Cells", "sum_sheet": "Summary",
        "title": "CELLS REGISTER", "sum_title": "SUMMARY BY BRIGADIR",
        "generated": "Generated", "records": "Records",
        "shown_of": "{n} of {total} shown",
        "num": "#", "verifix": "Verifix code", "sap": "SAP code",
        "workshop": "Workshop", "brigadir": "Brigadir", "leader": "Leader",
        "no_brigadir": "No brigadir", "unassigned": "Unassigned",
        "cells_cnt": "Cells", "with_leader": "With leader",
        "without_leader": "No leader", "coverage": "Coverage", "total": "TOTAL",
    },
}

_XLSX_GOLD    = "C8973F"   # brand — title bands + header rows
_XLSX_ZEBRA   = "FAF9F7"   # even-row shading
_XLSX_RULE    = "E5E7EB"   # hairline grid
_XLSX_SUBBAND = "F5F2EC"   # subtitle strip under the title
_XLSX_MUTED   = "9CA3AF"   # placeholder text ("—", "not assigned")


class CellsExportRow(BaseModel):
    verifix_code: str = ""
    sap_code:     str = ""
    workshop:     str = ""
    supervisor:   str = ""   # "" = unassigned; the label is applied here
    leader:       str = ""   # "" = unassigned


class CellsExportBody(BaseModel):
    lang:  str = "ru"
    rows:  list[CellsExportRow] = []
    total: Optional[int] = None   # register size BEFORE filtering, for the subtitle


@router.post("/admin/cells/export.xlsx")
def admin_export_cells(request: Request, body: CellsExportBody, db: Session = Depends(get_db),
                       caller: dict = Depends(require_page("cells"))):
    """Send the cells register to the caller's Telegram chat as a formatted
    workbook. Read-only, so it rides PAGE access like the list endpoint — a
    view-only grantee may export exactly the register they can already read."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    L = _CELLS_XLSX_T.get(body.lang) or _CELLS_XLSX_T["ru"]
    rows = body.rows
    total = body.total if body.total is not None else len(rows)
    now = datetime.now(_XLSX_TASHKENT)

    hair = Side(style="thin", color=_XLSX_RULE)
    grid = Border(left=hair, right=hair, top=hair, bottom=hair)
    gold_fill = PatternFill("solid", fgColor=_XLSX_GOLD)
    zebra = PatternFill("solid", fgColor=_XLSX_ZEBRA)
    head_font = Font(bold=True, color="FFFFFF", size=10)
    muted = Font(color=_XLSX_MUTED, italic=True, size=10)
    body_font = Font(size=10, color="1F2937")
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    left_mid = Alignment(horizontal="left", vertical="center")

    def band(ws, span: str, text: str, *, height: int, fill: str,
             font: Font, align=None):
        """Merged full-width strip — the title/subtitle chrome both sheets share."""
        ws.merge_cells(span)
        c = ws[span.split(":")[0]]
        c.value = text
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = font
        c.alignment = align or Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[c.row].height = height

    def finish(ws, *, landscape: bool):
        """Report finish: no gridlines, gold tab, fit-to-width printing."""
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = _XLSX_GOLD
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.orientation = "landscape" if landscape else "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_options.horizontalCentered = True

    # ── Sheet 1: the register ────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = L["sheet"]

    headers = [L["num"], L["verifix"], L["sap"], L["workshop"], L["brigadir"], L["leader"]]
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    band(ws, f"A1:{last_col}1", L["title"], height=32, fill=_XLSX_GOLD,
         font=Font(bold=True, color="FFFFFF", size=14))
    subtitle = f"{L['generated']}: {now.strftime('%d.%m.%Y %H:%M')}   ·   {L['records']}: {len(rows)}"
    if len(rows) != total:
        subtitle += f"   ·   {L['shown_of'].format(n=len(rows), total=total)}"
    band(ws, f"A2:{last_col}2", subtitle, height=20, fill=_XLSX_SUBBAND,
         font=Font(size=9, color="6B7280"))
    ws.row_dimensions[3].height = 6          # breathing room, not an empty row

    HEAD_ROW = 4
    for i, h in enumerate(headers, 1):
        c = ws.cell(HEAD_ROW, i, h)
        c.fill, c.font, c.border = gold_fill, head_font, grid
        c.alignment = center
    ws.row_dimensions[HEAD_ROW].height = 26

    for n, r in enumerate(rows, 1):
        y = HEAD_ROW + n
        ws.cell(y, 1, n).alignment = center
        ws.cell(y, 2, r.verifix_code or "—").font = Font(bold=True, size=10, color="111827")
        ws.cell(y, 2).alignment = center
        ws.cell(y, 3, r.sap_code or "—").alignment = center
        ws.cell(y, 4, r.workshop or "—")
        ws.cell(y, 5, r.supervisor or L["no_brigadir"])
        ws.cell(y, 6, r.leader or L["unassigned"])
        for i in range(1, ncols + 1):
            c = ws.cell(y, i)
            c.border = grid
            if i in (1, 3):
                c.font = Font(size=10, color="6B7280")
            elif i >= 4:
                c.alignment = left_wrap if i == 4 else left_mid
                c.font = body_font
            if n % 2 == 0:
                c.fill = zebra
        # Placeholders read as absent data, never as a value.
        if not r.sap_code:
            ws.cell(y, 3).font = muted
        if not r.workshop:
            ws.cell(y, 4).font = muted
        if not r.supervisor:
            ws.cell(y, 5).font = muted
        if not r.leader:
            ws.cell(y, 6).font = muted

    for col, width in zip("ABCDEF", (5, 14, 13, 48, 28, 34)):
        ws.column_dimensions[col].width = width
    if rows:
        ws.auto_filter.ref = f"A{HEAD_ROW}:{last_col}{HEAD_ROW + len(rows)}"
    ws.freeze_panes = f"A{HEAD_ROW + 1}"
    ws.print_title_rows = f"{HEAD_ROW}:{HEAD_ROW}"
    finish(ws, landscape=True)

    # ── Sheet 2: coverage per brigadir ───────────────────────────────────────
    agg: dict[str, list[int]] = {}
    for r in rows:
        name = r.supervisor or L["no_brigadir"]
        slot = agg.setdefault(name, [0, 0])
        slot[0] += 1
        if r.leader:
            slot[1] += 1
    ordered = sorted(agg.items(), key=lambda kv: (-kv[1][0], kv[0]))

    ws2 = wb.create_sheet(L["sum_sheet"])
    sum_heads = [L["brigadir"], L["cells_cnt"], L["with_leader"], L["without_leader"], L["coverage"]]
    band(ws2, "A1:E1", L["sum_title"], height=32, fill=_XLSX_GOLD,
         font=Font(bold=True, color="FFFFFF", size=14))
    band(ws2, "A2:E2", f"{L['generated']}: {now.strftime('%d.%m.%Y %H:%M')}", height=20,
         fill=_XLSX_SUBBAND, font=Font(size=9, color="6B7280"))
    ws2.row_dimensions[3].height = 6
    for i, h in enumerate(sum_heads, 1):
        c = ws2.cell(4, i, h)
        c.fill, c.font, c.border, c.alignment = gold_fill, head_font, grid, center
    ws2.row_dimensions[4].height = 26

    # Coverage keeps the platform's traffic light: ≥80% green · ≥50% amber · red.
    tone = ((0.80, "DCFCE7", "166534"), (0.50, "FEF9C3", "854D0E"), (0.0, "FEE2E2", "991B1B"))
    for n, (name, (cnt, with_l)) in enumerate(ordered, 1):
        y = 4 + n
        pct = (with_l / cnt) if cnt else 0
        ws2.cell(y, 1, name).font = body_font
        ws2.cell(y, 1).alignment = left_mid
        for i, v in enumerate((cnt, with_l, cnt - with_l), 2):
            c = ws2.cell(y, i, v)
            c.font, c.alignment = body_font, center
        pc = ws2.cell(y, 5, pct)
        pc.number_format = "0%"
        pc.alignment = center
        bg, fg = next((b, f) for lim, b, f in tone if pct >= lim)
        pc.fill = PatternFill("solid", fgColor=bg)
        pc.font = Font(bold=True, size=10, color=fg)
        for i in range(1, 6):
            ws2.cell(y, i).border = grid
            if n % 2 == 0 and i < 5:
                ws2.cell(y, i).fill = zebra

    if ordered:
        y = 5 + len(ordered)
        tot_cells = sum(v[0] for v in agg.values())
        tot_with = sum(v[1] for v in agg.values())
        top = Side(style="medium", color=_XLSX_GOLD)
        vals = [L["total"], tot_cells, tot_with, tot_cells - tot_with,
                (tot_with / tot_cells) if tot_cells else 0]
        for i, v in enumerate(vals, 1):
            c = ws2.cell(y, i, v)
            c.font = Font(bold=True, size=10, color="111827")
            c.alignment = left_mid if i == 1 else center
            c.border = Border(left=hair, right=hair, top=top, bottom=hair)
            c.fill = PatternFill("solid", fgColor=_XLSX_SUBBAND)
        ws2.cell(y, 5).number_format = "0%"
        ws2.row_dimensions[y].height = 22

    for col, width in zip("ABCDE", (34, 12, 14, 14, 12)):
        ws2.column_dimensions[col].width = width
    ws2.freeze_panes = "A5"
    finish(ws2, landscape=False)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"cells_register_{now.strftime('%Y-%m-%d')}.xlsx"
    caption = (f"📋 {L['sheet']} — {now.strftime('%d.%m.%Y %H:%M')}  •  "
               f"{L['records']}: {len(rows)}")
    try:
        delivered = deliver_xlsx(request, caller, fname, buf.read(), caption)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Telegram send failed: {e}")
    if not isinstance(delivered, dict):
        return delivered
    return {"ok": True, "rows": len(rows)}


@router.post("/admin/cells")
def admin_create_cell(payload: CellPayload, db: Session = Depends(get_db),
                      caller: dict = Depends(require_cap(CAP_CELLS_MANAGE))):
    code = " ".join((payload.verifix_code or "").split())
    if not code:
        raise HTTPException(status_code=400, detail="Verifix code is required")
    if db.query(Cell).filter_by(verifix_code=code).first():
        raise HTTPException(status_code=409, detail=f"Cell {code} already exists")
    row = Cell(verifix_code=code)
    _apply_cell_fields(db, row, payload)
    db.add(row)
    cell_details = [("cell", code), ("unit", unit_name(db, row.manager_id))]
    db.commit()
    alert_grant_use(db, caller, CAP_CELLS_MANAGE, "cell.created",
                    details=cell_details)
    return {"ok": True, "id": row.id}


@router.put("/admin/cells/{cid}")
def admin_update_cell(cid: int, payload: CellPayload, db: Session = Depends(get_db),
                      caller: dict = Depends(require_cap(CAP_CELLS_MANAGE))):
    row = db.query(Cell).filter_by(id=cid).first()
    if not row:
        raise HTTPException(status_code=404, detail="Cell not found")
    old = {"verifix_code": row.verifix_code, "sap_code": row.sap_code,
           "manager_id": row.manager_id, "leader_id": row.leader_id,
           **{k: getattr(row, c) for k, c in _CELL_NAME_DIFF.items()}}
    if payload.verifix_code is not None:
        code = " ".join(payload.verifix_code.split())
        if not code:
            raise HTTPException(status_code=400, detail="Verifix code is required")
        dup = db.query(Cell).filter(Cell.verifix_code == code, Cell.id != cid).first()
        if dup:
            raise HTTPException(status_code=409, detail=f"Cell {code} already exists")
        row.verifix_code = code
    _apply_cell_fields(db, row, payload)
    new = {"verifix_code": row.verifix_code, "sap_code": row.sap_code,
           "manager_id": row.manager_id, "leader_id": row.leader_id,
           **{k: getattr(row, c) for k, c in _CELL_NAME_DIFF.items()}}
    db.commit()
    diff = [(k, old[k], new[k])
            for k in ("verifix_code", "sap_code", *_CELL_NAME_DIFF)
            if old[k] != new[k]]
    if old["manager_id"] != new["manager_id"]:
        diff.append(("unit", unit_name(db, old["manager_id"]),
                     unit_name(db, new["manager_id"])))
    if old["leader_id"] != new["leader_id"]:
        lnames = {p.id: p.name for p in db.query(RoleProfile).filter(
            RoleProfile.id.in_([i for i in (old["leader_id"], new["leader_id"]) if i]))}
        diff.append(("leader", lnames.get(old["leader_id"]), lnames.get(new["leader_id"])))
    if diff:
        alert_grant_use(db, caller, CAP_CELLS_MANAGE, "cell.updated",
                        details=[("cell", old["verifix_code"])],
                        changes=diff)
    return {"ok": True, "id": cid}


@router.delete("/admin/cells/{cid}")
def admin_delete_cell(cid: int, db: Session = Depends(get_db),
                      caller: dict = Depends(require_cap(CAP_CELLS_MANAGE))):
    row = db.query(Cell).filter_by(id=cid).first()
    if not row:
        raise HTTPException(status_code=404, detail="Cell not found")
    cell_details = [("cell", row.verifix_code),
                    ("unit", unit_name(db, row.manager_id))]
    db.delete(row)
    db.commit()
    alert_grant_use(db, caller, CAP_CELLS_MANAGE, "cell.deleted",
                    details=cell_details)
    return {"ok": True}


# ── Admin: update ─────────────────────────────────────────────────────────────

class UpdateProfilePayload(BaseModel):
    name:           Optional[str] = None
    # Per-language name columns ("" clears; None = leave untouched)
    name_uz_cyrl:   Optional[str] = None
    name_ru:        Optional[str] = None
    name_en:        Optional[str] = None
    shift:          Optional[int] = None
    manager_id:     Optional[int] = None        # leader → move to another unit
    cells:          Optional[list[str]] = None  # leader → replace owned cell codes
    new_verifix_id: Optional[int] = None        # supervisor → re-key managers.id
    archived:       Optional[bool] = None       # supervisor only
    overrides:      Optional[dict[str, str]] = None  # lang → display name ("" clears)


_NAME_LANG_COLS = ("name_uz_cyrl", "name_ru", "name_en")


def _apply_name_columns(p: RoleProfile, payload) -> None:
    for col in _NAME_LANG_COLS:
        val = getattr(payload, col)
        if val is not None:
            setattr(p, col, val.strip() or None)


def _apply_overrides(db: Session, canonical: str, overrides: dict[str, str]) -> None:
    # The session runs with autoflush=False; a rename in the same request keeps
    # its rekeyed override rows pending, invisible to the SELECTs below — flush
    # first or the inserts collide with them on uq_translation_lang_key.
    db.flush()
    key = f"name.{canonical}"
    for lang, value in overrides.items():
        row = db.query(Translation).filter_by(lang=lang, key=key).first()
        value = (value or "").strip()
        if value:
            if row:
                row.value = value
            else:
                db.add(Translation(lang=lang, key=key, value=value))
        elif row:
            db.delete(row)


@router.put("/admin/{ptype}/{pid}")
def admin_update_profile(ptype: str, pid: int, payload: UpdateProfilePayload,
                         db: Session = Depends(get_db),
                         caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    _deny_admin_profile(caller, ptype)
    if ptype not in PROFILE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid profile type")

    if ptype == "supervisor":
        mgr = db.query(Manager).filter_by(id=pid).first()
        if not mgr:
            raise HTTPException(status_code=404, detail="Unit not found")
        old = {"name": mgr.name, "shift": mgr.shift, "archived": bool(mgr.archived)}
        if payload.name is not None:
            _rename_profile(db, "supervisor", pid, payload.name)
        if payload.shift in (1, 2):
            mgr.shift = payload.shift
        if payload.archived is not None:
            mgr.archived = bool(payload.archived)
        if payload.overrides:
            _apply_overrides(db, mgr.name, payload.overrides)
        new_id = pid
        if payload.new_verifix_id and payload.new_verifix_id != pid:
            new_id = _rekey_manager_id(db, mgr, payload.new_verifix_id)
        # Read BEFORE commit: the rekey path deleted mgr's row, so a post-commit
        # attribute refresh would explode on the vanished id.
        new_vals = (("name", mgr.name), ("shift", mgr.shift), ("archived", bool(mgr.archived)))
        db.commit()
        diff = [(k, old[k], v) for k, v in new_vals if old[k] != v]
        if new_id != pid:
            diff.append(("verifix_code", pid, new_id))
        if diff:
            alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.updated",
                            details=[("role", tv("role.supervisor")),
                                     ("profile", old["name"])],
                            changes=diff)
        return {"ok": True, "id": new_id}

    p = db.query(RoleProfile).filter_by(id=pid).first()
    if not p or p.role != ptype:
        raise HTTPException(status_code=404, detail="Profile not found")

    old = {"name": p.name, "shift": p.shift, "manager_id": p.manager_id}
    if payload.name is not None:
        _rename_profile(db, ptype, pid, payload.name)
    _apply_name_columns(p, payload)
    if ptype == "shift-manager" and payload.shift in (1, 2):
        p.shift = payload.shift
    if ptype == "leader" and payload.manager_id and payload.manager_id != p.manager_id:
        mgr = db.query(Manager).filter_by(id=payload.manager_id).first()
        if not mgr:
            raise HTTPException(status_code=400, detail="Supervisor unit not found")
        # Bound leaders move with their profile — their role rows point at the unit.
        for r in _bound_role_rows(db, "leader", pid):
            r.role_id = payload.manager_id
        p.manager_id = payload.manager_id
    if ptype == "leader" and payload.cells is not None:
        _set_leader_cells(db, pid, payload.cells)
    if payload.overrides:
        _apply_overrides(db, p.name, payload.overrides)
    new_vals = {"name": p.name, "shift": p.shift, "manager_id": p.manager_id}
    db.commit()
    diff = [(k, old[k], new_vals[k]) for k in ("name", "shift") if old[k] != new_vals[k]]
    if old["manager_id"] != new_vals["manager_id"]:
        diff.append(("unit", unit_name(db, old["manager_id"]),
                     unit_name(db, new_vals["manager_id"])))
    if ptype == "leader" and payload.cells is not None:
        diff.append(("cells", None, ", ".join(payload.cells) or None))
    if diff:
        alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.updated",
                        details=[("role", tv("role." + ptype)),
                                 ("profile", old["name"])],
                        changes=diff)
    return {"ok": True, "id": pid}


def _rekey_manager_id(db: Session, mgr: Manager, new_id: int) -> int:
    """Change a unit's Verifix ID: insert a fresh managers row under the new id,
    re-point every referencing table, drop the old row. One transaction —
    the caller commits."""
    if new_id <= 0:
        raise HTTPException(status_code=400, detail="Invalid Verifix ID")
    if db.query(Manager).filter_by(id=new_id).first():
        raise HTTPException(status_code=409, detail="Verifix ID already in use")

    old_id = mgr.id
    db.add(Manager(id=new_id, name=mgr.name, shift=mgr.shift, archived=mgr.archived))
    db.flush()
    for table, col in _MANAGER_ID_REFS:
        db.execute(text(f"UPDATE {table} SET {col} = :new WHERE {col} = :old"),
                   {"new": new_id, "old": old_id})
    db.execute(text(
        "UPDATE telegram_user_roles SET role_id = :new "
        "WHERE role IN ('supervisor', 'leader') AND role_id = :old"
    ), {"new": new_id, "old": old_id})
    db.execute(text("DELETE FROM managers WHERE id = :old"), {"old": old_id})
    return new_id


# ── Admin: switch role ────────────────────────────────────────────────────────

class SwitchRolePayload(BaseModel):
    ptype:      str                    # current type
    pid:        int                    # current id (managers.id for supervisor)
    new_role:   str
    shift:      Optional[int] = None        # → shift-manager | supervisor
    manager_id: Optional[int] = None        # → leader (unit)
    cells:      Optional[list[str]] = None  # → leader (owned cell codes, optional)
    verifix_id: Optional[int] = None        # → supervisor (new managers.id)
    confirm:    bool = False                # acknowledge the impacts from the 409


def _migrate_role_row(db: Session, row: TelegramUserRole, new_role: str,
                      new_role_id: int, new_profile_id: int | None = None) -> None:
    """Re-point one binding at the profile's new role. If the user already holds
    exactly that binding, keep the stronger row and drop the other.

    "Exactly that binding" is the PROFILE, not (role, role_id): a leader's
    role_id is the unit, so a unit-wide match would delete this holder's OTHER
    leader profile as a phantom duplicate.

    The profile key moves with the binding — a stale key would leave the holder
    attached to the profile's OLD identity, i.e. invisible under the new one."""
    # None for a leader with no known profile id: "leader:<unit>" is not a key
    # any row carries, and passing it could collide with a leader PROFILE whose
    # id happens to equal the unit id.
    new_key = (f"{new_role}:{new_profile_id}" if new_profile_id
               else (f"{new_role}:{new_role_id}" if new_role != "leader" else None))
    dup = find_role_row(db, row.telegram_id, new_role, new_role_id,
                        key=new_key, name=row.full_name, exclude_id=row.id)
    if dup:
        if row.status == "approved" and dup.status != "approved":
            _remove_role_row(db, dup)
        else:
            _remove_role_row(db, row)
            return
        db.flush()  # the DELETE must land before the UPDATE re-uses the unique key
    row.role = new_role
    row.role_id = new_role_id
    row.profile_key = new_key


@router.post("/admin/switch-role")
def admin_switch_role(payload: SwitchRolePayload, db: Session = Depends(get_db),
                      caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    """Move a profile to another role. Only the name (and its per-language
    display overrides, keyed by name) moves along — every role-specific value
    comes from the payload. Holders migrate in place: their binding rows switch
    to the new role, so the next app open grants the new role's access.

    Side effects that lose or hide something (leader history staying behind,
    the unit being archived/deleted on a supervisor switch) are returned as a
    409 {"code": "confirm_required", ...} until the caller re-sends confirm."""
    ptype, new_role = payload.ptype, payload.new_role
    if ptype not in PROFILE_TYPES or new_role not in PROFILE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid profile type")
    # Both ends: a grantee may neither demote an admin nor promote anyone into
    # one — the latter is the self-promotion path this guard exists to close.
    _deny_admin_profile(caller, ptype, new_role)
    if ptype == new_role:
        raise HTTPException(status_code=400, detail="Profile already has this role")

    mgr = p = None
    if ptype == "supervisor":
        mgr = db.query(Manager).filter_by(id=payload.pid).first()
        if not mgr:
            raise HTTPException(status_code=404, detail="Unit not found")
        name = mgr.name
    else:
        p = db.query(RoleProfile).filter_by(id=payload.pid).first()
        if not p or p.role != ptype:
            raise HTTPException(status_code=404, detail="Profile not found")
        name = p.name

    # Target-specific values are never inherited — they must be explicit.
    if new_role in ("shift-manager", "supervisor") and payload.shift not in (1, 2):
        raise HTTPException(status_code=400, detail="Shift must be 1 or 2")
    if new_role == "supervisor":
        if not payload.verifix_id or payload.verifix_id <= 0:
            raise HTTPException(status_code=400, detail="Verifix ID is required")
        if db.query(Manager).filter_by(id=payload.verifix_id).first():
            raise HTTPException(status_code=409, detail="Verifix ID already in use")
    if new_role == "leader":
        if ptype == "supervisor" and payload.manager_id == payload.pid:
            raise HTTPException(status_code=400,
                                detail="The unit is being removed — pick another unit")
        if not db.query(Manager).filter_by(id=payload.manager_id).first():
            raise HTTPException(status_code=400, detail="Supervisor unit not found")

    # Registration resolves profiles by name — same duplicate rules as create
    # (guest names are not unique, so they are exempt).
    if new_role == "leader":
        if db.query(RoleProfile).filter_by(role="leader", name=name,
                                           manager_id=payload.manager_id).first():
            raise HTTPException(status_code=409, detail="This leader already exists")
    elif new_role not in ("supervisor", "guest"):
        if db.query(RoleProfile).filter_by(role=new_role, name=name).first():
            raise HTTPException(status_code=409, detail="Profile with this name already exists")

    if ptype == "admin":
        # Approved admins live in the admins table; role rows are pending /adminreg.
        rows = db.query(TelegramUserRole).filter_by(role="admin", role_id=payload.pid).all()
        admin_holders = db.query(Admin).filter_by(profile_id=payload.pid).all()
        if admin_holders and db.query(Admin).count() - len(admin_holders) < 1:
            raise HTTPException(status_code=400, detail="Cannot remove the last admin")
    else:
        rows = _bound_role_rows(db, ptype, payload.pid)
        admin_holders = []
    approved_rows = [r for r in rows if r.status == "approved"]

    if new_role == "admin":
        # One admin profile — one account (mirrors /adminreg semantics).
        if len(approved_rows) > 1:
            raise HTTPException(status_code=409,
                                detail="An admin profile can be held by one account only — "
                                       "unassign the extra holders first")
        if approved_rows and db.query(Admin).filter_by(
                telegram_id=approved_rows[0].telegram_id).first():
            raise HTTPException(status_code=409, detail="The holder is already an admin")

    # Impacts an admin must acknowledge before the switch goes through.
    impacts: dict = {}
    if ptype == "leader":
        refs = [r.id for r in rows]
        concerns = db.query(LeaderConcern).filter(
            LeaderConcern.leader_profile_id == payload.pid).count()
        tasks = 0
        if refs:
            concerns += (
                db.query(LeaderConcern)
                .filter(LeaderConcern.leader_profile_id.is_(None),
                        LeaderConcern.leader_role_ref.in_(refs)).count()
            )
            tasks = db.query(LeaderTask).filter(LeaderTask.leader_role_ref.in_(refs)).count()
        if concerns:
            impacts["concerns"] = concerns
        if tasks:
            impacts["tasks"] = tasks
    if ptype == "supervisor":
        if _manager_has_data(db, payload.pid):
            impacts["unit_archive"] = True
        else:
            impacts["unit_delete"] = True
            n_leaders = db.query(RoleProfile).filter_by(
                role="leader", manager_id=payload.pid).count()
            if n_leaders:
                impacts["unit_leaders"] = n_leaders
    if impacts and not payload.confirm:
        raise HTTPException(status_code=409, detail={"code": "confirm_required", **impacts})

    # ── target entity ──
    if new_role == "supervisor":
        db.add(Manager(id=payload.verifix_id, name=name, shift=payload.shift, archived=False))
        db.flush()
        target_profile = None
        target_role_id = payload.verifix_id
    else:
        if ptype == "supervisor":
            target_profile = RoleProfile(role=new_role, name=name)
            db.add(target_profile)
        else:
            target_profile = p
            target_profile.role = new_role
        target_profile.shift = payload.shift if new_role == "shift-manager" else None
        target_profile.manager_id = payload.manager_id if new_role == "leader" else None
        db.flush()
        if new_role == "leader":
            if payload.cells:
                _set_leader_cells(db, target_profile.id, payload.cells)
        else:
            # Leaving the leader role frees the profile's cells.
            _release_leader_cells(db, target_profile.id)
        target_role_id = payload.manager_id if new_role == "leader" else target_profile.id

    # ── migrate holders ──
    now = datetime.now(timezone.utc)
    if new_role == "admin":
        for r in rows:
            if r.status != "approved":
                _migrate_role_row(db, r, "admin", target_profile.id)
        for r in approved_rows:
            user = db.query(TelegramUser).filter_by(telegram_id=r.telegram_id).first()
            db.add(Admin(telegram_id=r.telegram_id, profile_id=target_profile.id,
                         language=(user.language if user else None) or "uz"))
            # The admins row is the binding — the role row goes away (see /adminreg).
            if user and user.active_role_ref == r.id:
                other = (
                    db.query(TelegramUserRole)
                    .filter(TelegramUserRole.telegram_id == r.telegram_id,
                            TelegramUserRole.status == "approved",
                            TelegramUserRole.id != r.id)
                    .first()
                )
                user.active_role_ref = other.id if other else None
            db.delete(r)
    else:
        for r in rows:
            _migrate_role_row(db, r, new_role, target_role_id,
                              target_role_id if new_role == "supervisor" else target_profile.id)
        for a in admin_holders:  # ptype == "admin": convert admins rows to role rows
            user = db.query(TelegramUser).filter_by(telegram_id=a.telegram_id).first()
            if not user:
                # Seeded admins have no telegram_users row — create the account shell.
                user = TelegramUser(telegram_id=a.telegram_id, full_name=name,
                                    role=new_role, role_id=target_role_id,
                                    status="approved", language=a.language or "uz")
                db.add(user)
            pkey = (f"{new_role}:"
                    f"{target_role_id if new_role == 'supervisor' else target_profile.id}")
            # By profile, not by (role, role_id): a unit-wide match would adopt
            # this holder's OTHER leader profile row and rename it to this one.
            row = find_role_row(db, a.telegram_id, new_role, target_role_id,
                                key=pkey, name=name)
            if row:
                row.full_name = name
                row.profile_key = pkey
                if row.status != "approved":
                    row.status, row.approved_at = "approved", now
            else:
                row = TelegramUserRole(telegram_id=a.telegram_id, role=new_role,
                                       role_id=target_role_id, full_name=name,
                                       profile_key=pkey,
                                       status="approved", approved_at=now)
                db.add(row)
            db.flush()
            if not user.active_role_ref:
                user.active_role_ref = row.id
            db.delete(a)

    # ── source cleanup ──
    if ptype == "supervisor":
        if impacts.get("unit_archive"):
            mgr.archived = True
        else:
            # No history: same cascade as delete — the unit's leader profiles go too.
            for lp in db.query(RoleProfile).filter_by(role="leader",
                                                      manager_id=payload.pid).all():
                for r in _bound_role_rows(db, "leader", lp.id):
                    _remove_role_row(db, r)
                _release_leader_cells(db, lp.id)
                db.delete(lp)
            db.flush()
            db.delete(mgr)
    elif new_role == "supervisor":
        if ptype == "leader":
            _release_leader_cells(db, p.id)
        db.delete(p)  # the identity now lives in the managers row

    new_id = target_role_id if new_role == "supervisor" else target_profile.id
    db.commit()
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.role_switched",
                    details=[("profile", name),
                             ("count", len(approved_rows))],
                    changes=[("role", tv("role." + ptype), tv("role." + new_role))])
    return {"ok": True, "role": new_role, "id": new_id}


# ── Admin: delete / unassign ──────────────────────────────────────────────────

@router.delete("/admin/{ptype}/{pid}")
def admin_delete_profile(ptype: str, pid: int, db: Session = Depends(get_db),
                         caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    _deny_admin_profile(caller, ptype)
    if ptype not in PROFILE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid profile type")

    if ptype == "supervisor":
        mgr = db.query(Manager).filter_by(id=pid).first()
        if not mgr:
            raise HTTPException(status_code=404, detail="Unit not found")
        mgr_name = mgr.name
        for r in _bound_role_rows(db, "supervisor", pid):
            _remove_role_row(db, r)
        if _manager_has_data(db, pid):
            mgr.archived = True   # history stays queryable; unit leaves pickers/dashboards
            db.commit()
            alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.deleted",
                            details=[("role", tv("role.supervisor")), ("profile", mgr_name)],
                            changes=[("status", None, tv("v.archived"))])
            return {"ok": True, "archived": True}
        # No history: take the unit's leader profiles (and their bindings) along.
        for lp in db.query(RoleProfile).filter_by(role="leader", manager_id=pid).all():
            for r in _bound_role_rows(db, "leader", lp.id):
                _remove_role_row(db, r)
            _release_leader_cells(db, lp.id)
            db.delete(lp)
        db.delete(mgr)
        db.commit()
        alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.deleted",
                        details=[("role", tv("role.supervisor")), ("profile", mgr_name)])
        return {"ok": True, "archived": False}

    p = db.query(RoleProfile).filter_by(id=pid).first()
    if not p or p.role != ptype:
        raise HTTPException(status_code=404, detail="Profile not found")

    if ptype == "admin":
        holder = db.query(Admin).filter_by(profile_id=pid).first()
        if holder and db.query(Admin).count() <= 1:
            raise HTTPException(status_code=400, detail="Cannot remove the last admin")
        if holder:
            db.delete(holder)
        # drop pending /adminreg requests for this profile
        for r in db.query(TelegramUserRole).filter_by(role="admin", role_id=pid).all():
            _remove_role_row(db, r)
    else:
        for r in _bound_role_rows(db, ptype, pid):
            _remove_role_row(db, r)

    if ptype == "leader":
        _release_leader_cells(db, pid)
    p_name = p.name
    db.delete(p)
    db.commit()
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.deleted",
                    details=[("role", tv("role." + ptype)), ("profile", p_name)])
    return {"ok": True}


class UnassignPayload(BaseModel):
    ptype:       str
    pid:         int
    role_ref:    Optional[int] = None  # non-admin bindings
    telegram_id: Optional[int] = None  # admin bindings


@router.post("/admin/unassign")
def admin_unassign_profile(payload: UnassignPayload, db: Session = Depends(get_db),
                           caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    _deny_admin_profile(caller, payload.ptype)
    if payload.ptype == "admin":
        holder = db.query(Admin).filter_by(telegram_id=payload.telegram_id,
                                           profile_id=payload.pid).first()
        if not holder:
            raise HTTPException(status_code=404, detail="Assignment not found")
        if db.query(Admin).count() <= 1:
            raise HTTPException(status_code=400, detail="Cannot remove the last admin")
        db.delete(holder)
        db.commit()
        return {"ok": True}

    row = db.query(TelegramUserRole).filter_by(id=payload.role_ref).first()
    if not row:
        raise HTTPException(status_code=404, detail="Assignment not found")
    alert_details = [("role", tv("role." + row.role)),
                     ("profile", row.full_name),
                     ("account", row.telegram_id)]
    _remove_role_row(db, row)
    db.commit()
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.unassigned",
                    details=alert_details)
    return {"ok": True}


# ── Web logins (username + password for the browser) ──────────────────────────
#
# A browser login belongs to the PROFILE, so it is managed here, in the row of
# the person it belongs to — not in a separate register that would drift out of
# step with the profiles it describes. A generated password is still DM'd to the
# profile's Telegram holders — that is how the person receives it — and no
# create/reset/bulk response carries it back.
#
# Reading one is a separate, deliberate act: `/admin/web-login/reveal`, one
# profile per call, ADMINS only, and audited. So running this tab still does not
# hand anyone a password list; asking for a specific person's password does, and
# leaves a line in the log saying who asked.

class WebLoginPayload(BaseModel):
    profile_key: str
    username: Optional[str] = None
    password: Optional[str] = None


class WebLoginBulkPayload(BaseModel):
    profile_keys: list[str]


def _web_guard(caller: dict, key: str) -> tuple[str, int]:
    """Resolve + authorize one profile key for web-login management.

    Reuses the escalation stop: someone running this tab through a grant may
    not mint a browser password for an ADMIN profile. Without it, a grantee
    could hand themselves an admin login by creating one and resetting it to a
    profile they hold."""
    role, ref = parse_profile_key(key)
    if not role or not ref or role not in PROFILE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid profile")
    _deny_admin_profile(caller, role)
    return role, ref


def _web_row(db: Session, key: str) -> WebCredential:
    cred = db.query(WebCredential).filter(WebCredential.profile_key == key).first()
    if not cred:
        raise HTTPException(status_code=404, detail="No web login for this profile")
    return cred


def _web_state(db: Session, cred: WebCredential, key: str) -> dict:
    """One profile's web-login state, including whether it can be DELIVERED.
    A profile with no Telegram holder cannot receive its own password, so the
    admin sees that instead of a login that would never arrive."""
    return {
        "username":      cred.username,
        "enabled":       bool(cred.enabled),
        "last_login_at": cred.last_login_at.isoformat() if cred.last_login_at else None,
        "locked":        web_auth.lock_seconds_left(cred) > 0,
        "deliverable":   bool(profile_holders(db, key)),
        "has_password":  bool(cred.password_enc),
    }


@router.post("/admin/web-login")
def admin_set_web_login(payload: WebLoginPayload, db: Session = Depends(get_db),
                        caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    """Create a browser login for a profile, or reset the password of an
    existing one. Either way the credential is DM'd to every holder."""
    key = payload.profile_key
    _web_guard(caller, key)

    name = profile_display_name(db, key)
    if not name:
        raise HTTPException(status_code=404, detail="Profile not found")
    if not web_auth.session_identity(db, key):
        raise HTTPException(status_code=400, detail="no_holder")

    cred = db.query(WebCredential).filter(WebCredential.profile_key == key).first()

    username = web_auth.normalize_username(payload.username or "") or (
        cred.username if cred else web_auth.suggest_username(db, name)
    )
    err = web_auth.username_error(username)
    if err:
        raise HTTPException(status_code=400, detail=err)
    clash = db.query(WebCredential).filter(WebCredential.username == username).first()
    if clash and (not cred or clash.id != cred.id):
        raise HTTPException(status_code=409, detail="username_taken")

    password = (payload.password or "").strip() or web_auth.generate_password()
    err = web_auth.password_error(password)
    if err:
        raise HTTPException(status_code=400, detail=err)

    if cred:
        cred.username = username
        web_auth.set_password(cred, password)
        # A reset ends every browser session that held the old password.
        cred.token_version = (cred.token_version or 1) + 1
        cred.enabled = True
        web_auth.clear_failures(db, cred)
        action = "weblogin.reset"
    else:
        cred = WebCredential(profile_key=key, username=username)
        web_auth.set_password(cred, password)
        db.add(cred)
        action = "weblogin.created"
    db.commit()

    sent = web_auth.dm_credentials(db, key, username, password)
    web_auth.audit(action.split(".")[1], caller, name, username, f"dm_sent={sent}")
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE, action,
                    details=[("profile", name), ("login", username)])
    return {"ok": True, "username": username, "sent": sent,
            "web": _web_state(db, cred, key)}


@router.post("/admin/web-login/reveal")
def admin_reveal_web_login(payload: WebLoginPayload, db: Session = Depends(get_db),
                           caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    """Read one profile's browser password back.

    ADMINS ONLY — a deliberate step narrower than the rest of this tab. Everything
    else here is grantable through ``admin.profiles.manage`` because it CHANGES a
    credential and is therefore visible to the person it belongs to (a reset
    arrives as a DM, a disable stops their login). Reading a password leaves no
    trace the owner can see, so it stays with the role that already answers for
    the whole platform rather than with a delegated one.

    Deliberately one profile per call, never a column on the register: a list
    endpoint would put every password in one response and in the browser's query
    cache, where this feature would stop being «look up a login» and start being
    «export the password file».

    Answers 200 with ``password: null`` when the credential predates the sealed
    copy (or SECRET_KEY was rotated) — the page then says «unknown» and offers a
    reset, which is the honest state. It never guesses.
    """
    key = payload.profile_key
    _web_guard(caller, key)
    if caller.get("role") != "admin":
        raise HTTPException(status_code=403, detail="admin_only")
    cred = _web_row(db, key)

    password = web_auth.open_password(cred.password_enc)
    # Audited like every other credential event — reading one is the only way to
    # use a login without touching it, so the trail has to record it too.
    web_auth.audit("revealed", caller, profile_display_name(db, key) or "",
                   cred.username, "known" if password else "unavailable")
    return {"ok": True, "username": cred.username, "password": password}


@router.post("/admin/web-login/bulk")
def admin_bulk_web_login(payload: WebLoginBulkPayload, db: Session = Depends(get_db),
                         caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    """Generate logins for everyone who does not have one yet — the rollout
    path. Profiles that already have a login are left ALONE (a bulk action must
    never silently reset a password someone is using), and profiles with no
    Telegram holder are reported back rather than skipped in silence."""
    created, skipped, no_holder = [], [], []
    # Names claimed within this batch are not in the DB yet, so collisions
    # between two new logins have to be tracked here as well.
    taken: set[str] = set()

    for key in payload.profile_keys:
        role, ref = parse_profile_key(key)
        if not role or not ref or role not in PROFILE_TYPES:
            continue
        if role == "admin" and caller.get("role") != "admin":
            continue
        name = profile_display_name(db, key)
        if not name:
            continue
        if db.query(WebCredential).filter(WebCredential.profile_key == key).first():
            skipped.append(name)
            continue
        if not web_auth.session_identity(db, key):
            no_holder.append(name)
            continue

        username = web_auth.suggest_username(db, name, taken)
        taken.add(username)
        password = web_auth.generate_password()
        cred = WebCredential(profile_key=key, username=username)
        web_auth.set_password(cred, password)
        db.add(cred)
        db.commit()
        web_auth.dm_credentials(db, key, username, password)
        web_auth.audit("created", caller, name, username, "bulk")
        created.append({"profile": name, "username": username})

    if created:
        alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "weblogin.bulk",
                        details=[("count", len(created))])
    return {"ok": True, "created": created, "skipped": skipped, "no_holder": no_holder}


@router.put("/admin/web-login")
def admin_rename_web_login(payload: WebLoginPayload, db: Session = Depends(get_db),
                           caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    """Change the username without touching the password."""
    key = payload.profile_key
    _web_guard(caller, key)
    cred = _web_row(db, key)

    username = web_auth.normalize_username(payload.username or "")
    err = web_auth.username_error(username)
    if err:
        raise HTTPException(status_code=400, detail=err)
    clash = db.query(WebCredential).filter(WebCredential.username == username).first()
    if clash and clash.id != cred.id:
        raise HTTPException(status_code=409, detail="username_taken")

    old = cred.username
    web_auth.audit("renamed", caller, profile_display_name(db, key) or "", old,
                   f"new_login={username}")
    cred.username = username
    # Sessions carry the username as their handle, so a rename must re-key them
    # — bumping the version signs the old ones out rather than orphaning them.
    cred.token_version = (cred.token_version or 1) + 1
    db.commit()
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "weblogin.renamed",
                    changes=[("login", old, username)])
    return {"ok": True, "web": _web_state(db, cred, key)}


@router.post("/admin/web-login/toggle")
def admin_toggle_web_login(payload: WebLoginPayload, db: Session = Depends(get_db),
                           caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    """Enable/disable browser access without destroying the credential — the
    reversible half of «this person should not be logging in from home»."""
    key = payload.profile_key
    _web_guard(caller, key)
    cred = _web_row(db, key)
    cred.enabled = not cred.enabled
    if not cred.enabled:
        cred.token_version = (cred.token_version or 1) + 1
    db.commit()
    web_auth.audit("enabled" if cred.enabled else "disabled", caller,
                   profile_display_name(db, key) or "", cred.username)
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE,
                    "weblogin.enabled" if cred.enabled else "weblogin.disabled",
                    details=[("login", cred.username)])
    return {"ok": True, "web": _web_state(db, cred, key)}


@router.post("/admin/web-login/revoke")
def admin_revoke_web_sessions(payload: WebLoginPayload, db: Session = Depends(get_db),
                              caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    """Sign out every browser holding this login, keeping the password valid —
    the answer to a laptop left logged in somewhere."""
    key = payload.profile_key
    _web_guard(caller, key)
    cred = _web_row(db, key)
    cred.token_version = (cred.token_version or 1) + 1
    db.commit()
    web_auth.audit("sessions_revoked", caller, profile_display_name(db, key) or "",
                   cred.username)
    return {"ok": True, "web": _web_state(db, cred, key)}


@router.delete("/admin/web-login")
def admin_delete_web_login(payload: WebLoginPayload, db: Session = Depends(get_db),
                           caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    key = payload.profile_key
    _web_guard(caller, key)
    cred = _web_row(db, key)
    username = cred.username
    web_auth.audit("deleted", caller, profile_display_name(db, key) or "", username)
    db.delete(cred)
    db.commit()
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "weblogin.deleted",
                    details=[("login", username)])
    return {"ok": True}


# ── Profile photos ────────────────────────────────────────────────────────────

_AVATAR_MAX_BYTES = 10 * 1024 * 1024
_AVATAR_SIDE = 512


def _photo_version(row: ProfilePhoto) -> int:
    return int(row.updated_at.timestamp()) if row.updated_at else 1


@router.get("/photo/{key}")
def profile_photo(key: str, db: Session = Depends(get_db),
                  _: dict = Depends(_caller)):
    """One profile's avatar bytes. Any approved session may read it — photos
    are org-internal exactly like the names they sit next to. The frontend
    fetches this as a blob (the auth headers ride along) and keys its cache by
    ``photo_ver``, so the long client cache here can never go stale."""
    row = db.query(ProfilePhoto).filter_by(profile_key=key).first()
    if not row:
        raise HTTPException(status_code=404, detail="No photo")
    return Response(
        content=row.data,
        media_type=row.mime or "image/jpeg",
        headers={"Cache-Control": "private, max-age=31536000, immutable"},
    )


@router.post("/admin/photo")
def admin_set_photo(profile_key: str = Form(...), file: UploadFile = File(...),
                    db: Session = Depends(get_db),
                    caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    """Set (or replace) a profile's photo. Admin surface only — users see their
    photo on /profile but the upload lives with the identity register.

    The stored bytes are NEVER the upload: the image is decoded, EXIF-rotated,
    centre-cropped square and re-encoded to a ≤512px JPEG, which both bounds the
    row size and strips anything a crafted file might carry."""
    role, ref = parse_profile_key(profile_key)
    if not role or not ref or role not in PROFILE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid profile")
    _deny_admin_profile(caller, role)
    name = profile_display_name(db, profile_key)
    if not name:
        raise HTTPException(status_code=404, detail="Profile not found")

    content = file.file.read(_AVATAR_MAX_BYTES + 1)
    if len(content) > _AVATAR_MAX_BYTES:
        raise HTTPException(status_code=400, detail="photo_too_large")
    validate_avatar(file, content)
    try:
        img = Image.open(BytesIO(content))
        img = ImageOps.exif_transpose(img)
        img = img.convert("RGB")
    except Exception:
        raise HTTPException(status_code=400, detail="invalid_image")

    w, h = img.size
    side = min(w, h)
    img = img.crop(((w - side) // 2, (h - side) // 2,
                    (w + side) // 2, (h + side) // 2))
    if side > _AVATAR_SIDE:
        img = img.resize((_AVATAR_SIDE, _AVATAR_SIDE), Image.Resampling.LANCZOS)
    buf = BytesIO()
    img.save(buf, "JPEG", quality=88)

    now = datetime.now(timezone.utc)
    row = db.query(ProfilePhoto).filter_by(profile_key=profile_key).first()
    if row:
        row.data, row.mime, row.updated_at = buf.getvalue(), "image/jpeg", now
    else:
        row = ProfilePhoto(profile_key=profile_key, mime="image/jpeg",
                           data=buf.getvalue(), updated_at=now)
        db.add(row)
    db.commit()
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.photo_set",
                    details=[("profile", name)])
    return {"ok": True, "photo_ver": _photo_version(row)}


class PhotoDeletePayload(BaseModel):
    profile_key: str


@router.delete("/admin/photo")
def admin_delete_photo(payload: PhotoDeletePayload, db: Session = Depends(get_db),
                       caller: dict = Depends(require_cap(CAP_PROFILES_MANAGE))):
    role, ref = parse_profile_key(payload.profile_key)
    if not role or not ref or role not in PROFILE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid profile")
    _deny_admin_profile(caller, role)
    row = db.query(ProfilePhoto).filter_by(profile_key=payload.profile_key).first()
    if not row:
        raise HTTPException(status_code=404, detail="No photo")
    db.delete(row)
    db.commit()
    alert_grant_use(db, caller, CAP_PROFILES_MANAGE, "profile.photo_removed",
                    details=[("profile", profile_display_name(db, payload.profile_key) or payload.profile_key)])
    return {"ok": True}


# ── Registration options (pre-login, Telegram-initData-gated) ─────────────────

class RegistrationOptionsPayload(BaseModel):
    init_data: str
    reg_token: Optional[str] = None  # bot-signed ?rt= from the register button


@router.post("/registration-options")
def registration_options(payload: RegistrationOptionsPayload, db: Session = Depends(get_db)):
    """Name lists for the registration pickers. Gated so anonymous web
    visitors get 401 (per the privacy decision). Two accepted credentials:
    Telegram initData (menu/inline/direct-link launches), or the bot-signed
    reg_token — the register page opens from a keyboard button (required for
    sendData()) and keyboard-button launches never receive initData."""
    if payload.init_data and payload.init_data != "__dev__":
        if not _validate_init_data(payload.init_data):
            raise HTTPException(status_code=401, detail="Invalid Telegram initData")
    elif payload.reg_token:
        if not validate_reg_token(payload.reg_token):
            raise HTTPException(status_code=401,
                                detail="Expired registration link — send /register to the bot again")
    elif not settings.dev_auth:
        raise HTTPException(status_code=401, detail="Missing Telegram initData")

    managers = (
        db.query(Manager)
        .filter(Manager.archived.is_(False))
        .order_by(Manager.shift, Manager.name)
        .all()
    )
    mgr_names = {m.id: m.name for m in managers}
    mgr_shifts = {m.id: m.shift for m in managers}

    # archived units keep their leaders out of the picker
    leader_profiles = [
        p for p in db.query(RoleProfile)
        .filter(RoleProfile.role == "leader")
        .order_by(RoleProfile.name).all()
        if p.manager_id in mgr_names
    ]
    leaders: dict[str, list[str]] = {}
    for p in leader_profiles:
        leaders.setdefault(mgr_names[p.manager_id], []).append(p.name)

    top_profiles = (
        db.query(RoleProfile).filter(RoleProfile.role == "top-manager")
        .order_by(RoleProfile.name).all()
    )
    shift_profiles = (
        db.query(RoleProfile).filter(RoleProfile.role == "shift-manager")
        .order_by(RoleProfile.shift, RoleProfile.name).all()
    )

    # Guest profiles without an approved holder are offered for re-claiming in
    # the registration picker. Guest names are NOT unique — a typed name always
    # gets its own fresh profile, so there is no taken-name list to check.
    guest_profiles = (
        db.query(RoleProfile).filter(RoleProfile.role == "guest")
        .order_by(RoleProfile.name).all()
    )

    # Every approved holder in ONE scan. profile_holders() is the canonical
    # predicate but costs a query or two per profile — a few hundred
    # round-trips here — so its rules are reproduced in bulk below.
    approved = (
        db.query(TelegramUserRole)
        .filter(TelegramUserRole.status == "approved").all()
    )
    approved_guest_ids = {r.role_id for r in approved if r.role == "guest"}
    held_direct: set[tuple[str, int]] = set()
    held_leader_keys: set[str] = set()
    held_leader_named: set[tuple[int, str]] = set()
    for r in approved:
        if r.role == "leader":
            # A leader row's role_id is the BRIGADIR's unit, shared by every
            # leader in it, so the profile is named by the stamped key — or,
            # on unstamped legacy rows, by full_name.
            if r.profile_key:
                held_leader_keys.add(r.profile_key)
            elif r.full_name:
                held_leader_named.add((r.role_id, r.full_name))
        elif r.role_id is not None:
            key = f"{r.role}:{r.role_id}"
            if not r.profile_key or r.profile_key == key:
                held_direct.add((r.role, r.role_id))

    # Per-language spellings, so the search matches whatever script is typed.
    # role_profiles carry name_* columns; managers have none, so a supervisor's
    # other spellings live only in the legacy name.<canonical> overrides.
    canonicals = ({m.name for m in managers} |
                  {p.name for p in leader_profiles + top_profiles + shift_profiles})
    tr_names: dict[str, dict[str, str]] = {}
    if canonicals:
        tr_keys = {f"name.{n}": n for n in canonicals}
        for t in db.query(Translation).filter(Translation.key.in_(list(tr_keys))).all():
            val = (t.value or "").strip()
            if val and t.lang in ("uz_cyrl", "ru", "en"):
                tr_names.setdefault(tr_keys[t.key], {})[t.lang] = val

    def _names(canonical: str, prof: Optional[RoleProfile] = None) -> dict[str, str]:
        """Columns first, legacy translation overrides second — the same
        resolution order tl() renders by."""
        out: dict[str, str] = {}
        if prof is not None:
            for lang, col in (("uz_cyrl", prof.name_uz_cyrl),
                              ("ru", prof.name_ru), ("en", prof.name_en)):
                if col and col.strip():
                    out[lang] = col.strip()
        for lang, val in tr_names.get(canonical, {}).items():
            out.setdefault(lang, val)
        return out

    # The name-first catalogue: one flat row per claimable profile. The page
    # searches names instead of asking for a role first, so somebody who
    # cannot name their own role can no longer file under the wrong one —
    # their name simply does not exist under any other role. Guests are
    # deliberately absent: that path is for people who are NOT staff.
    entries = [
        {
            "key": f"supervisor:{m.id}", "role": "supervisor", "name": m.name,
            "names": _names(m.name), "shift": m.shift, "supervisor": None,
            "taken": ("supervisor", m.id) in held_direct,
        }
        for m in managers
    ] + [
        {
            "key": f"leader:{p.id}", "role": "leader", "name": p.name,
            "names": _names(p.name, p), "shift": mgr_shifts.get(p.manager_id),
            "supervisor": mgr_names.get(p.manager_id),
            "taken": (f"leader:{p.id}" in held_leader_keys
                      or (p.manager_id, p.name) in held_leader_named),
        }
        for p in leader_profiles
    ] + [
        {
            "key": f"shift-manager:{p.id}", "role": "shift-manager", "name": p.name,
            "names": _names(p.name, p), "shift": p.shift, "supervisor": None,
            "taken": ("shift-manager", p.id) in held_direct,
        }
        for p in shift_profiles
    ] + [
        {
            "key": f"top-manager:{p.id}", "role": "top-manager", "name": p.name,
            "names": _names(p.name, p), "shift": None, "supervisor": None,
            "taken": ("top-manager", p.id) in held_direct,
        }
        for p in top_profiles
    ]

    return {
        # Legacy per-role lists — kept so a cached older bundle still renders
        # during the seconds between the static swap and the backend restart.
        "top_managers": [p.name for p in top_profiles],
        "shift_managers": [{"name": p.name, "shift": p.shift} for p in shift_profiles],
        "supervisors": [{"name": m.name, "shift": m.shift} for m in managers],
        "leaders": leaders,
        "guests": [
            {"id": p.id, "name": p.name}
            for p in guest_profiles if p.id not in approved_guest_ids
        ],
        "entries": entries,
    }


# ── Self-service: my profile names ────────────────────────────────────────────

@router.get("/mine")
def my_profiles(caller: dict = Depends(_caller), db: Session = Depends(get_db)):
    """The caller's approved profiles with their canonical names and stored
    per-language overrides — the settings-modal name editor's data source."""
    tid = int(caller["sub"])
    entries = []

    for r in (
        db.query(TelegramUserRole)
        .filter_by(telegram_id=tid, status="approved")
        .order_by(TelegramUserRole.id)
        .all()
    ):
        if r.role == "admin":
            continue  # surfaced via the admins-table entry below
        entries.append({
            "kind": "role", "role": r.role, "role_ref": r.id, "canonical": r.full_name,
        })

    admin_row = db.query(Admin).filter_by(telegram_id=tid).first()
    if admin_row and admin_row.profile_id:
        p = db.query(RoleProfile).filter_by(id=admin_row.profile_id, role="admin").first()
        if p:
            entries.append({"kind": "admin", "role": "admin",
                            "role_ref": None, "canonical": p.name})

    names = {e["canonical"] for e in entries if e["canonical"]}
    overrides: dict[str, dict[str, str]] = {n: {} for n in names}
    if names:
        keys = [f"name.{n}" for n in names]
        for t in db.query(Translation).filter(Translation.key.in_(keys)).all():
            overrides[t.key[len("name."):]][t.lang] = t.value
    for e in entries:
        e["overrides"] = overrides.get(e["canonical"], {})
    return {"profiles": entries}


class MyNamePayload(BaseModel):
    kind:      str                       # "role" | "admin"
    role_ref:  Optional[int] = None      # kind="role"
    name:      Optional[str] = None      # new canonical (uz) — rename cascade
    overrides: Optional[dict[str, str]] = None  # lang → display ("" clears)


@router.put("/mine")
def update_my_name(payload: MyNamePayload, caller: dict = Depends(_caller),
                   db: Session = Depends(get_db)):
    """Self-service rename: users edit their own profile's canonical name and
    per-language display variants once approved. Immediate, app-wide."""
    tid = int(caller["sub"])

    if payload.kind == "admin":
        admin_row = db.query(Admin).filter_by(telegram_id=tid).first()
        if not admin_row or not admin_row.profile_id:
            raise HTTPException(status_code=403, detail="No admin profile")
        p = db.query(RoleProfile).filter_by(id=admin_row.profile_id).first()
        if not p:
            raise HTTPException(status_code=404, detail="Profile not found")
        if payload.name is not None:
            _rename_profile(db, "admin", p.id, payload.name)
        if payload.overrides:
            _apply_overrides(db, p.name, payload.overrides)
        db.commit()
        return {"ok": True, "canonical": p.name}

    row = db.query(TelegramUserRole).filter_by(id=payload.role_ref, telegram_id=tid).first()
    if not row or row.status != "approved":
        raise HTTPException(status_code=403, detail="Not your approved profile")

    canonical = row.full_name
    if payload.name is not None and (payload.name or "").strip() != canonical:
        if row.role == "supervisor":
            _rename_profile(db, "supervisor", row.role_id, payload.name)
        elif row.role in ("shift-manager", "top-manager", "guest"):
            if row.role_id:
                _rename_profile(db, row.role, row.role_id, payload.name)
            else:  # legacy row without a profile — rename the row itself
                _rekey_name_overrides(db, canonical, payload.name.strip())
                row.full_name = payload.name.strip()
        else:  # leader — bind via (manager_id, name)
            p = db.query(RoleProfile).filter_by(
                role="leader", manager_id=row.role_id, name=canonical,
            ).first()
            if p:
                _rename_profile(db, "leader", p.id, payload.name)
            else:
                _rekey_name_overrides(db, canonical, payload.name.strip())
                row.full_name = payload.name.strip()
        db.refresh(row)
        canonical = row.full_name

    if payload.overrides:
        _apply_overrides(db, canonical, payload.overrides)
    db.commit()
    return {"ok": True, "canonical": canonical}


# ── Self-service: my profile card (/profile page) ─────────────────────────────

def _factory_dict(f: Optional[Factory]) -> Optional[dict]:
    if not f:
        return None
    return {"code": f.code, "name_uz": f.name_uz, "name_uz_cyrl": f.name_uz_cyrl,
            "name_ru": f.name_ru, "name_en": f.name_en}


def _cell_dict(c: Cell) -> dict:
    return {"verifix_code": c.verifix_code, "sap_code": c.sap_code,
            "name_workshop_uz": c.name_workshop_uz,
            "name_workshop_uz_cyrl": c.name_workshop_uz_cyrl,
            "name_workshop_ru": c.name_workshop_ru,
            "name_workshop_en": c.name_workshop_en}


@router.get("/me/details")
def my_profile_details(caller: dict = Depends(_caller), db: Session = Depends(get_db)):
    """Everything the /profile page shows about the caller's ACTIVE profile.

    Read-only by design: the page's edit surface is admin-only, and the one
    thing a user changes themselves (the web password) has its own endpoint.
    Resolved live from the profile so a rename or reassignment shows up on the
    next open, never from a JWT snapshot.
    """
    key = viewer_profile_key(db, caller)
    role = caller.get("role")
    out = {
        "role": role,
        "profile_key": key,
        "name": caller.get("full_name"),
        "names": {},
        "shift": None,
        "unit": None,
        "cells": [],
        "verifix_id": None,
        "factory": None,
        "holders": [],
        "web": None,
        "photo_ver": None,
    }
    if not key:
        return out  # legacy unresolvable identity — JWT basics only

    ptype, ref = parse_profile_key(key)
    canonical = out["name"]
    factory_id = None

    if ptype == "supervisor":
        m = db.query(Manager).filter_by(id=ref).first()
        if m:
            canonical = m.name
            out.update(shift=m.shift, verifix_id=m.id)
            factory_id = m.factory_id
    else:
        p = db.query(RoleProfile).filter_by(id=ref).first()
        if p:
            canonical = p.name
            out["names"] = {"uz_cyrl": p.name_uz_cyrl or "", "ru": p.name_ru or "",
                            "en": p.name_en or ""}
            if ptype == "shift-manager":
                out["shift"] = p.shift
            if ptype == "leader" and p.manager_id:
                m = db.query(Manager).filter_by(id=p.manager_id).first()
                if m:
                    out["unit"] = m.name
                    out["shift"] = m.shift
                    factory_id = m.factory_id
                out["cells"] = [
                    _cell_dict(c) for c in db.query(Cell)
                    .filter_by(leader_id=p.id).order_by(Cell.verifix_code).all()
                ]

    out["name"] = canonical

    # Per-language display names: profile columns first, the legacy
    # translation overrides second — the same resolution tl() renders by.
    if canonical:
        for t in db.query(Translation).filter(Translation.key == f"name.{canonical}").all():
            if t.lang in ("uz_cyrl", "ru", "en") and not out["names"].get(t.lang):
                out["names"][t.lang] = t.value

    if factory_id:
        out["factory"] = _factory_dict(db.query(Factory).filter_by(id=factory_id).first())

    me = int(caller["sub"])
    holder_ids = profile_holders(db, key)
    if holder_ids:
        users = {u.telegram_id: u for u in db.query(TelegramUser)
                 .filter(TelegramUser.telegram_id.in_(holder_ids)).all()}
        out["holders"] = [{
            "telegram_id": tid,
            "tg_name": users[tid].tg_name if tid in users else None,
            "username": users[tid].username if tid in users else None,
            "is_me": tid == me,
            # Same reachability flag the Profiles tab shows — a person sharing a
            # profile can see that their colleague's account gets no DMs.
            "dm_ok": users[tid].dm_failed_at is None if tid in users else True,
            "status": "approved",
        } for tid in holder_ids]

    cred = db.query(WebCredential).filter_by(profile_key=key).first()
    if cred:
        out["web"] = {
            "username": cred.username,
            "enabled": bool(cred.enabled),
            "locked": web_auth.lock_seconds_left(cred) > 0,
            "last_login_at": cred.last_login_at.isoformat() if cred.last_login_at else None,
        }

    photo = db.query(ProfilePhoto).filter_by(profile_key=key).first()
    if photo:
        out["photo_ver"] = _photo_version(photo)
    return out
