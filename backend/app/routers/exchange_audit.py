"""
Lost-worker audit — «Yo'qolgan xodimlar», the read-only report behind an
approved → supervisor people-exchange that left its workers on NO roster.

The failure it looks for (traced 2026-08-22, exchange doc of 19.08):

  A plain → supervisor exchange relocates the attendance row by setting
  ``manager_id`` to the receiving unit. Nothing on the row then records that it
  came from the sender's cells. Both upload doors — ``attendance_batch.
  _sync_manager`` and the legacy ``admin.upload_verifix`` — wipe
  ``(manager, date)`` WHOLESALE and re-insert only the rows of the cells routed
  to that manager, so any later save touching the RECEIVER's day deletes the
  transferred rows and puts nothing back: they belong to none of its cells.
  ``staff.reapply_task_exchanges`` is the one repair hook that runs afterwards
  and it skips → supervisor docs by design. If the sender's day is not
  re-projected after that (a closed day is skipped outright), the worker is
  gone from both units — present in the source workbook, absent everywhere the
  platform looks.

The document cannot repair itself: a plain → supervisor move stores no
``snapshot`` (``staff._build_exchange_payload`` writes one only for → task
moves and transfer-time splits), so the payload knows the worker's name, origin
unit, title and cell — and neither their clock nor their hours.

``AttendanceBatchRow`` does know. The batch is deliberately KEPT after Save so a
day can be re-projected rather than re-uploaded, which is what makes these rows
recoverable at all — and what this report measures, per worker:

  ``recoverable``  a batch row for (date, worker) still exists → re-projecting
                   the sender's day restores the full row, hours included.
  ``day_blocked``  it exists, but the sender's day is closed. ``_project``
                   skips a closed day, so it must be re-opened first.
  ``no_batch``     no batch row (the batch was dropped by «delete supervisor
                   day», or the date predates the single-file «Davomat» flow) →
                   only the original workbook for that date can restore it.

READ-ONLY on purpose. It writes nothing and repairs nothing: restoring these
rows moves historical numbers (загрузка, «came» counts, KPIs, the leaderboard
all rise on every repaired day), so the scope is measured and shown first and
the repair stays a separate, deliberate decision.
"""
import logging
from collections import defaultdict
from datetime import date as date_t, datetime, timedelta
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from app.database import get_db
from app.models import (
    Attendance, AttendanceBatch, AttendanceBatchRow, HrDocument, Manager,
)
from app.routers.admin import verify_admin
from app.services.day_state import day_state
from app.xlsx_delivery import deliver_xlsx

router = APIRouter(prefix="/api/admin/exchange-audit", tags=["exchange-audit"])

log = logging.getLogger(__name__)

# How far back a bare request looks. The whole history is reachable with an
# explicit ``from``; the default is a window an operator can actually act on.
DEFAULT_DAYS = 180


def _parse_date(s: Optional[str]) -> Optional[date_t]:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid date: {s}")


def _f(v) -> Optional[float]:
    return float(v) if v is not None else None


def _collect(db: Session, date_from: Optional[str], date_to: Optional[str]) -> dict:
    """Every worker an approved → supervisor exchange left on no roster.

    "On no roster" is deliberately the widest test there is: the name carries
    NO attendance row anywhere on that date. A worker moved on by a second
    document (A→B→C) still has a row and is not lost; only a name the platform
    cannot show at all is reported. That keeps the count honest — it is the
    number of people whose day is missing, not the number of odd-looking
    documents.
    """
    d_to   = _parse_date(date_to) or date_t.today()
    d_from = _parse_date(date_from) or (d_to - timedelta(days=DEFAULT_DAYS))
    if d_from > d_to:
        raise HTTPException(status_code=400, detail="from is after to")

    docs = db.query(HrDocument).filter(
        HrDocument.doc_type == "people_exchange",
        HrDocument.status   == "approved",
        HrDocument.date     >= d_from,
        HrDocument.date     <= d_to,
    ).order_by(HrDocument.date.desc(), HrDocument.id.desc()).all()

    # Claims: (date, worker_name) → the earliest document that moved them, so a
    # second hop reports the unit they actually started on (the same rule
    # `exchange_rewind` replays by).
    claims: dict[tuple, dict] = {}
    blanked: set = set()                # deliberately name-stripped, NOT lost
    docs_scanned = 0
    for doc in sorted(docs, key=lambda x: (x.date, x.id)):
        payload = doc.payload or {}
        if payload.get("target_type") != "supervisor" or not payload.get("target_manager_id"):
            continue
        docs_scanned += 1
        split = bool(payload.get("transfer_time"))
        for emp in payload.get("employees") or []:
            name = (emp or {}).get("worker_name")
            if not name:
                continue
            # A below-min transfer-time split BLANKS the worker's own row on
            # purpose (CLAUDE.md: cleared the bar on neither side ⇒ credited to
            # nobody). Their name is off the roster because the platform decided
            # it should be, not because an upload deleted them — reporting that
            # as a loss would invite "repairing" a rule that is working.
            if (emp.get("applied") or {}).get("task_blanked"):
                blanked.add((doc.date, name))
                continue
            key = (doc.date, name)
            if key in claims:
                # Later hop: the destination moves on, the origin does not.
                claims[key]["target_id"]   = payload["target_manager_id"]
                claims[key]["target_cell"] = payload.get("target_cell")
                claims[key]["doc_id"]      = doc.id
                claims[key]["split"]       = claims[key]["split"] or split
                claims[key]["hops"]       += 1
                continue
            claims[key] = {
                "doc_id":      doc.id,
                "date":        doc.date,
                "sender_id":   emp.get("old_manager_id") or doc.manager_id,
                "target_id":   payload["target_manager_id"],
                "target_cell": payload.get("target_cell"),
                "old_role":    emp.get("old_role") or None,
                "old_cell":    emp.get("old_verifix_code"),
                "split":       split,
                "created_by":  doc.created_by_name,
                "posted_by":   doc.approved_by_name,
                "hops":        1,
            }
    for key in blanked:
        claims.pop(key, None)
    if not claims:
        return _empty(d_from, d_to, docs_scanned)

    dates = {d for d, _n in claims}
    names = {n for _d, n in claims}

    # ── Who is still visible somewhere on their date? ─────────────────────────
    present = {
        (d, n) for d, n in db.query(Attendance.date, Attendance.worker_name).filter(
            Attendance.date.in_(list(dates)),
            Attendance.worker_name.in_(list(names)),
        ).distinct().all()
    }
    lost_keys = [k for k in claims if k not in present]
    if not lost_keys:
        return _empty(d_from, d_to, docs_scanned)

    # ── What the batch still holds for them (the recovery source) ────────────
    batch_ids = {
        b.date: b.id for b in db.query(AttendanceBatch).filter(
            AttendanceBatch.date.in_(sorted({d for d, _n in lost_keys})),
        ).all()
    }
    batch_rows: dict[tuple, AttendanceBatchRow] = {}
    if batch_ids:
        by_id = {v: k for k, v in batch_ids.items()}
        for r in db.query(AttendanceBatchRow).filter(
            AttendanceBatchRow.batch_id.in_(list(batch_ids.values())),
            AttendanceBatchRow.worker_name.in_(sorted({n for _d, n in lost_keys})),
        ).all():
            batch_rows.setdefault((by_id[r.batch_id], r.worker_name), r)

    mgr_names = {
        m.id: m.name for m in db.query(Manager).filter(
            Manager.id.in_({c["sender_id"] for c in claims.values()}
                           | {c["target_id"] for c in claims.values()}),
        ).all()
    }
    day_cache: dict[tuple, str] = {}

    def _day(mid, d):
        if not mid:
            return None
        k = (mid, d)
        if k not in day_cache:
            try:
                day_cache[k] = day_state(db, mid, d)[0]
            except Exception:                       # never let one odd day kill the report
                log.exception("day_state failed for manager %s on %s", mid, d)
                day_cache[k] = "unknown"
        return day_cache[k]

    rows, days, units, hours = [], set(), set(), 0.0
    counts = defaultdict(int)
    for key in lost_keys:
        d, name = key
        c   = claims[key]
        br  = batch_rows.get(key)
        s_day = _day(c["sender_id"], d)
        if br is None:
            state = "no_batch"
        elif s_day != "open":
            state = "day_blocked"
        else:
            state = "recoverable"
        counts[state] += 1
        h = _f(br.hours_worked) if br is not None else None
        if h:
            hours += h
        days.add(d)
        units.add(c["sender_id"])
        rows.append({
            "doc_id":       c["doc_id"],
            "date":         d.isoformat(),
            "worker_name":  name,
            "job_title":    (br.job_title if br is not None else None) or c["old_role"],
            "verifix_code": (br.verifix_code if br is not None else None) or c["old_cell"],
            "sender_id":    c["sender_id"],
            "sender_name":  mgr_names.get(c["sender_id"], str(c["sender_id"])),
            "target_id":    c["target_id"],
            "target_name":  mgr_names.get(c["target_id"], str(c["target_id"])),
            "target_cell":  c["target_cell"],
            "sender_day":   s_day,
            "target_day":   _day(c["target_id"], d),
            "hops":         c["hops"],
            "split":        c["split"],
            "created_by":   c["created_by"],
            "posted_by":    c["posted_by"],
            "state":        state,
            "clock_in_out": br.clock_in_out if br is not None else None,
            "hours_worked": h,
        })

    rows.sort(key=lambda r: (r["date"], r["sender_name"], r["worker_name"]), reverse=True)
    # Already-restored rows still missing their effective hours. They are NOT
    # losses and never appear in `rows` — a restored worker is on a roster
    # again — so the count rides on the summary, where the button can see it.
    heal_pending = len(_heal_query(db, d_from, d_to, names))
    return {
        "from": d_from.isoformat(),
        "to":   d_to.isoformat(),
        "docs_scanned": docs_scanned,
        "names": sorted(names),
        "rows": rows,
        "summary": {
            "workers":     len(rows),
            "days":        len(days),
            "units":       len(units),
            "hours":       round(hours, 2),
            "heal_pending": heal_pending,
            "recoverable": counts["recoverable"],
            "day_blocked": counts["day_blocked"],
            "no_batch":    counts["no_batch"],
        },
    }


def _empty(d_from, d_to, docs_scanned):
    return {
        "from": d_from.isoformat(), "to": d_to.isoformat(),
        "docs_scanned": docs_scanned, "rows": [], "names": [],
        "summary": {"workers": 0, "days": 0, "units": 0, "hours": 0.0,
                    "heal_pending": 0, "recoverable": 0, "day_blocked": 0,
                    "no_batch": 0},
    }


@router.get("")
def audit(
    date_from: Optional[str] = Query(None, alias="from"),
    date_to:   Optional[str] = Query(None, alias="to"),
    db: Session = Depends(get_db),
    _: dict = Depends(verify_admin),
):
    """The report the «Yo'qolgan xodimlar» tab renders."""
    return _collect(db, date_from, date_to)


# ─── Excel ───────────────────────────────────────────────────────────────────
# The file mirrors the SCREEN: the same period, the same state filter, the same
# search. A workbook that quietly carried more rows than the reader was looking
# at is how an operator ends up repairing days they never reviewed.

_HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEAD_FILL = PatternFill("solid", fgColor="C8973F")
_BODY_FONT = Font(size=10)
_LEFT   = Alignment(horizontal="left",   vertical="center")
_RIGHT  = Alignment(horizontal="right",  vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")

# key → (header fallback, width, alignment kind)
_COLS = [
    ("date",         "Sana",           12, "text"),
    ("worker_name",  "Xodim",          40, "text"),
    ("job_title",    "Lavozim",        24, "text"),
    ("verifix_code", "Yacheyka",       11, "text"),
    ("sender_name",  "Kimdan",         26, "text"),
    ("target_name",  "Kimga",          26, "text"),
    ("clock_in_out", "Kelgan-ketgan",  22, "text"),
    ("hours_worked", "Soat",           9,  "num"),
    ("state",        "Holati",         18, "text"),
    ("sender_day",   "Yuboruvchi kuni", 16, "text"),
    ("doc_id",       "Hujjat",         10, "int"),
    ("created_by",   "Yaratdi",        24, "text"),
    ("posted_by",    "Tasdiqladi",     24, "text"),
]


def _xl_text(v):
    """A worker name goes into a cell verbatim, so two guards: a control
    character makes openpyxl raise (the whole export 500s), and a leading
    = + - @ turns a name into a formula when the file opens."""
    if v is None or not isinstance(v, str):
        return v
    v = ILLEGAL_CHARACTERS_RE.sub("", v)
    if v[:1] in ("=", "+", "-", "@"):
        v = "'" + v
    return v


class AuditExportBody(BaseModel):
    """The page's own filter state, so the file is what the reader is looking
    at. ``labels`` carries the headers in the viewer's language — the sheet is
    read by the same person who read the screen, in the same words."""
    date_from: Optional[str] = None
    date_to:   Optional[str] = None
    state:     str = "all"
    q:         Optional[str] = None
    labels:    dict[str, str] = {}
    summary_labels: dict[str, str] = {}
    caption:   Optional[str] = None


def _matches(r: dict, state: str, q: str) -> bool:
    if state != "all" and r["state"] != state:
        return False
    if not q:
        return True
    return any(q in str(r.get(k) or "").lower()
               for k in ("worker_name", "sender_name", "target_name", "verifix_code", "date"))


@router.post("/export.xlsx")
def export_xlsx(
    request: Request,
    body: AuditExportBody,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_admin),
):
    """Excel of the lost-worker report as filtered on the page. A browser
    session downloads it; inside Telegram it lands in the caller's private
    chat (app/xlsx_delivery.py)."""
    data = _collect(db, body.date_from, body.date_to)
    q = (body.q or "").strip().lower()
    rows = [r for r in data["rows"] if _matches(r, body.state, q)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Lost workers"

    for i, (key, fallback, width, _kind) in enumerate(_COLS, 1):
        c = ws.cell(row=1, column=i, value=_xl_text(body.labels.get(key) or fallback))
        c.font, c.fill, c.alignment = _HEAD_FONT, _HEAD_FILL, _CENTER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, r in enumerate(rows, 2):
        for ci, (key, _f, _w, kind) in enumerate(_COLS, 1):
            v = r.get(key)
            if key == "state":
                v = body.labels.get(f"state.{v}") or v
            c = ws.cell(row=ri, column=ci, value=_xl_text(v) if kind == "text" else v)
            c.font = _BODY_FONT
            if kind == "num":
                c.number_format = "0.00"
                c.alignment = _RIGHT
            elif kind == "int":
                c.number_format = "0"
                c.alignment = _RIGHT
            else:
                c.alignment = _LEFT
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}{len(rows) + 1}"

    # A second sheet carrying the totals and the period. The counts on screen
    # are the whole point of the report; a file holding only the rows makes the
    # reader re-derive them, and «17 day-closed» is the number that decides
    # whether any of this can be repaired today.
    s = data["summary"]
    L = body.summary_labels or {}
    meta = wb.create_sheet("Info")
    meta.column_dimensions["A"].width = 34
    meta.column_dimensions["B"].width = 26
    pairs = [
        (L.get("period") or "Davr",           f'{data["from"]} — {data["to"]}'),
        (L.get("workers") or "Yo'qolgan xodim",  s["workers"]),
        (L.get("hours") or "Yo'qolgan soat",     s["hours"]),
        (L.get("days") or "Kun",                 s["days"]),
        (L.get("units") or "Brigada",            s["units"]),
        (L.get("recoverable") or "Tiklash mumkin", s["recoverable"]),
        (L.get("day_blocked") or "Kun yopiq",    s["day_blocked"]),
        (L.get("no_batch") or "Manba yo'q",      s["no_batch"]),
        (L.get("exported") or "Faylda qator",    len(rows)),
    ]
    for ri, (k, v) in enumerate(pairs, 1):
        a = meta.cell(row=ri, column=1, value=_xl_text(k)); a.font = Font(bold=True, size=10); a.alignment = _LEFT
        b = meta.cell(row=ri, column=2, value=_xl_text(v)); b.font = _BODY_FONT; b.alignment = _LEFT

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f'yoqolgan_xodimlar_{data["from"]}_{data["to"]}.xlsx'
    caption = body.caption or f'📄 {len(rows)} · {s["hours"]} h'
    try:
        return deliver_xlsx(request, payload, fname, bio.read(), caption)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Telegram send failed: {e}")


# ─── Repair ──────────────────────────────────────────────────────────────────
# Putting a lost worker back WHERE THE APPROVED DOCUMENT SAYS THEY BELONG — on
# the receiving unit, in the destination cell — by re-inserting the row from the
# batch the upload deleted it from.
#
# Deliberately NOT a re-projection of the sender's day. `_sync_manager` wipes
# `(manager, date)` and rebuilds it from the batch's ticked cells, which would
# also undo every other exchange, edit and split on that day. This writes ONE
# row per named worker and touches nothing else.
#
# It writes into CLOSED days on purpose (the operator's explicit instruction,
# 2026-08-22): a closed day is the normal state of a day old enough for this to
# have happened to it, and re-opening one to fix it would swing the supervisor's
# confirmed totals twice instead of once. The closure is left standing — only
# the missing rows are added. NOTHING is notified: no supervisor DM, no bell, no
# day-state change. Every insert is logged under `EXCHANGE-REPAIR` with its row
# id, so the write is auditable and individually reversible.

def _heal_query(db: Session, d_from: date_t, d_to: date_t, names):
    """Restored rows still missing their «Samarali soatlar».

    `effective_hours` is `hours_worked − early_arrival_min/60` — the batch row
    carries it and `_sync_manager` copies it across, but the first version of
    the plain restore below did not, so rows put back before v3.19.0 show a dash
    in that column. It is a pure function of two columns already ON the row, so
    the top-up needs no source and cannot disagree with anything.

    Three guards keep it to exactly those rows. Only workers an exchange
    actually named; only rows that HAVE both inputs — a no-show (`X`) has no
    hours and a nameless split leftover has no early arrival, and NULL is the
    right answer for both; and never a nameless row, which carries no name to
    match on in the first place.
    """
    if not names:
        return []
    return db.query(Attendance).filter(
        Attendance.date >= d_from,
        Attendance.date <= d_to,
        Attendance.worker_name.in_(list(names)),
        Attendance.effective_hours.is_(None),
        Attendance.hours_worked.isnot(None),
        Attendance.early_arrival_min.isnot(None),
    ).all()


def _restore_split(db: Session, d: date_t, name: str, r: dict, actor) -> dict:
    """Put back a worker whose exchange carried a TRANSFER TIME.

    A split day is not one row, so this cannot copy the batch row the way a
    plain move can: the document divided the day at T and the two halves belong
    to two different units. It therefore re-runs the ORIGINAL calculation —
    ``staff._compute_split`` over the snapshot the document stored when it was
    created — and writes exactly what ``_apply_split_exchange`` writes, so a
    restored day is arithmetically identical to one that was never broken.
    Re-deriving those hours here with a second formula is how the two halves
    would stop adding up to the day the person actually worked.

    Two rows may be involved and they are handled separately:
      * the NAMED row, on whichever side won the bigger half (``plan["stay"]``);
      * the nameless hours-only LEFTOVER on the other side, whose id the payload
        recorded at apply time. It is recreated ONLY if that id no longer
        resolves — the wipe took one unit's day, not both, so the surviving half
        must not be written twice.

    A `below_min` employee never reaches here: the audit drops `task_blanked`
    keys, because that name was stripped by the rule working, not by the bug.
    """
    from app.routers.staff import _compute_split          # deferred: cycle

    miss = lambda why: {"date": d.isoformat(), "worker_name": name, "reason": why}

    doc = db.query(HrDocument).filter(HrDocument.id == r["doc_id"]).first()
    if doc is None or doc.status != "approved":
        return miss("doc_gone")
    if r["hops"] > 1:
        # A→B→C with a split in the chain: which half landed where depends on
        # documents applied one after another over a row this one never saw.
        return miss("multi_hop")

    payload = doc.payload or {}
    emp = next((e for e in payload.get("employees") or []
                if (e or {}).get("worker_name") == name), None)
    if emp is None:
        return miss("doc_gone")
    applied = emp.get("applied")
    if not applied:
        # An approved split document always records what it did. Without it
        # there is no way to tell whether a leftover was ever written.
        return miss("no_applied")

    snap   = emp.get("snapshot") or {}
    target = payload.get("target_manager_id")
    sender = emp.get("old_manager_id") or doc.manager_id
    plan   = _compute_split(snap, payload.get("transfer_time"), payload.get("return_time"))

    if plan is None or applied.get("plain"):
        # The document could not split this worker and fell back to a plain full
        # move; restore it the same way — the whole snapshot, on the target.
        row = Attendance(
            manager_id=target, date=d, worker_name=name,
            job_title=snap.get("job_title"), schedule=snap.get("schedule"),
            clock_in_out=snap.get("clock_in_out"), hours_worked=snap.get("hours_worked"),
            early_arrival_min=snap.get("early_arrival_min"),
            effective_hours=snap.get("effective_hours"),
            verifix_code=payload.get("target_cell") or emp.get("old_verifix_code"),
        )
        db.add(row); db.flush()
        log.warning("EXCHANGE-REPAIR split(plain) id=%s worker=%r date=%s -> manager=%s "
                    "hours=%s doc=%s by=%s", row.id, name, d, target,
                    snap.get("hours_worked"), doc.id, actor)
        return {"date": d.isoformat(), "worker_name": name, "attendance_id": row.id,
                "manager_id": target, "manager_name": r["target_name"],
                "verifix_code": row.verifix_code,
                "hours_worked": _f(snap.get("hours_worked")), "side": "plain"}

    if plan["stay"]:
        # The home half won: the named row belongs to the SENDER, keeping its
        # early arrival, and the after-T hours sit on the target as a leftover.
        named = Attendance(
            manager_id=sender, date=d, worker_name=name,
            job_title=snap.get("job_title"), schedule=snap.get("schedule"),
            clock_in_out=plan.get("home_clock") or f'{plan["C"]}-{plan["T"]}',
            hours_worked=plan["part1"], effective_hours=plan["part1_eff"],
            early_arrival_min=snap.get("early_arrival_min"),
            verifix_code=emp.get("old_verifix_code"),
        )
        leftover_mgr, leftover_hrs = target, plan["part2"]
        leftover_cell = payload.get("target_cell")
        side, named_mgr = "stay", sender
    else:
        # The away half won: the named row moved to the TARGET carrying the
        # after-T hours (early belongs to the unit it was earned on, so 0 here),
        # and the before-T remainder stayed behind as the sender's leftover.
        named = Attendance(
            manager_id=target, date=d, worker_name=name,
            job_title=snap.get("job_title"), schedule=snap.get("schedule"),
            clock_in_out=plan.get("away_clock") or f'{plan["T"]}-{plan["O"]}',
            hours_worked=plan["part2"], effective_hours=plan["part2"],
            early_arrival_min=0,
            verifix_code=payload.get("target_cell") or emp.get("old_verifix_code"),
        )
        leftover_mgr, leftover_hrs = sender, plan["part1_eff"]
        leftover_cell = emp.get("old_verifix_code")
        side, named_mgr = "move", target

    db.add(named); db.flush()

    # The other half. Its id was recorded when the document was applied; recreate
    # it only if that row is really gone, or the surviving half doubles.
    lid = applied.get("leftover_id")
    alive = db.query(Attendance.id).filter(Attendance.id == lid).first() if lid else None
    leftover_id = None
    if leftover_hrs and leftover_hrs > 0 and alive is None:
        lo = Attendance(manager_id=leftover_mgr, date=d, worker_name=None,
                        hours_worked=leftover_hrs, verifix_code=leftover_cell)
        db.add(lo); db.flush()
        leftover_id = lo.id
        emp["applied"] = {**applied, "leftover_id": lo.id}
        flag_modified(doc, "payload")

    log.warning(
        "EXCHANGE-REPAIR split(%s) id=%s worker=%r date=%s -> manager=%s cell=%s "
        "hours=%s | leftover id=%s manager=%s hours=%s (was %s) doc=%s by=%s",
        side, named.id, name, d, named_mgr, named.verifix_code, named.hours_worked,
        leftover_id, leftover_mgr, leftover_hrs, lid, doc.id, actor,
    )
    return {"date": d.isoformat(), "worker_name": name, "attendance_id": named.id,
            "manager_id": named_mgr,
            "manager_name": r["target_name"] if side == "move" else r["sender_name"],
            "verifix_code": named.verifix_code, "hours_worked": _f(named.hours_worked),
            "side": side, "leftover_id": leftover_id}


class RepairBody(BaseModel):
    """The exact workers to restore, as (date, worker_name) pairs read off the
    report. Deliberately explicit rather than "repair everything in the filter":
    an operator repairs the rows they looked at, and a list computed minutes ago
    cannot silently grow between the read and the write."""
    keys: list[dict] = []
    date_from: Optional[str] = None
    date_to:   Optional[str] = None


@router.post("/repair")
def repair(
    body: RepairBody,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_admin),
):
    """Re-insert the named lost workers on their exchange's destination unit.

    Idempotent: a worker who already carries a row for that date is reported
    `already_present` and nothing is written. Refuses what it cannot do
    faithfully — `no_batch` (no source row to copy) and `split` (a transfer-time
    document's effect is not one whole row, and reproducing it without re-running
    the split would credit the wrong unit the wrong hours).
    """
    wanted = {(k.get("date"), k.get("worker_name")) for k in body.keys
              if k.get("date") and k.get("worker_name")}

    data  = _collect(db, body.date_from, body.date_to)
    by_key = {(r["date"], r["worker_name"]): r for r in data["rows"]}
    actor = payload.get("full_name") or payload.get("sub")

    # Top up rows an earlier press restored without their effective hours,
    # before writing anything new — it is a pure re-derivation from two columns
    # already on the row, so it can run unconditionally and cannot double.
    healed = 0
    for row in _heal_query(db, _parse_date(data["from"]), _parse_date(data["to"]),
                           data.get("names") or []):
        row.effective_hours = round(
            float(row.hours_worked) - float(row.early_arrival_min) / 60.0, 4)
        healed += 1
        log.warning("EXCHANGE-REPAIR healed effective_hours id=%s worker=%r date=%s -> %s by=%s",
                    row.id, row.worker_name, row.date, row.effective_hours, actor)

    done, skipped = [], []
    for d_iso, name in sorted(wanted):
        r = by_key.get((d_iso, name))
        if r is None:
            # Not in the current report: already restored by an earlier press,
            # or never lost. Either way there is nothing to write.
            skipped.append({"date": d_iso, "worker_name": name, "reason": "already_present"})
            continue
        if r["state"] == "no_batch":
            skipped.append({"date": d_iso, "worker_name": name, "reason": "no_batch"})
            continue
        d = datetime.strptime(d_iso, "%Y-%m-%d").date()

        if r["split"]:
            out = _restore_split(db, d, name, r, actor)
            (done if out.get("attendance_id") else skipped).append(out)
            continue

        batch = db.query(AttendanceBatch).filter(AttendanceBatch.date == d).first()
        br = db.query(AttendanceBatchRow).filter(
            AttendanceBatchRow.batch_id == batch.id,
            AttendanceBatchRow.worker_name == name,
        ).first() if batch else None
        if br is None:
            skipped.append({"date": d_iso, "worker_name": name, "reason": "no_batch"})
            continue

        row = Attendance(
            manager_id        = r["target_id"],
            date              = d,
            worker_name       = br.worker_name,
            job_title         = br.job_title,
            schedule          = br.schedule,
            clock_in_out      = br.clock_in_out,
            hours_worked      = br.hours_worked,
            early_arrival_min = br.early_arrival_min,
            # Copied, not re-derived — `_sync_manager` carries this straight
            # across from the batch row and a restored row must be
            # indistinguishable from a projected one.
            effective_hours   = br.effective_hours,
            # The destination cell the exchange named; the worker's own cell
            # when the document predates cell-level moves (a legacy no-cell day).
            verifix_code      = r["target_cell"] or br.verifix_code,
        )
        db.add(row)
        db.flush()
        log.warning(
            "EXCHANGE-REPAIR restored attendance id=%s worker=%r date=%s "
            "-> manager=%s cell=%s hours=%s doc=%s sender_day=%s by=%s",
            row.id, name, d_iso, r["target_id"], row.verifix_code,
            br.hours_worked, r["doc_id"], r["sender_day"], actor,
        )
        done.append({
            "date": d_iso, "worker_name": name, "attendance_id": row.id,
            "manager_id": r["target_id"], "manager_name": r["target_name"],
            "verifix_code": row.verifix_code, "hours_worked": _f(br.hours_worked),
        })

    db.commit()
    log.warning("EXCHANGE-REPAIR done by=%s restored=%d healed=%d skipped=%d",
                actor, len(done), healed, len(skipped))
    return {"restored": len(done), "healed": healed, "skipped": len(skipped),
            "rows": done, "skippedRows": skipped}
